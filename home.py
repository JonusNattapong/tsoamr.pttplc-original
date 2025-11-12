from flask import Blueprint, Flask, render_template, request, jsonify, redirect, url_for,send_from_directory,Response,render_template_string
from routes import main

import ast
import itertools
from xhtml2pdf import pisa
from io import StringIO
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, Reference,BarChart,PieChart
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart.series import SeriesLabel
import matplotlib.pyplot as plt
from io import BytesIO
from pptx import Presentation
import pytz
from bs4 import BeautifulSoup
import numpy as np
import pandas as pd
import cx_Oracle
from flask import flash
import logging
import datetime
import re
import json
from datetime import datetime
import pandas as pd
import sqlite3
import plotly.express as px
from flask import (
    Flask,
    render_template,
    request,
    session,
    send_file,
    redirect,
    url_for,
    jsonify,
    g
)
import socket
import struct
 
import math
from flask_wtf import FlaskForm
from werkzeug.security import generate_password_hash, check_password_hash
from wtforms import StringField, PasswordField, SubmitField, validators
from werkzeug.security import generate_password_hash
from sqlalchemy import desc
from flask import Flask, send_from_directory
from flask_migrate import Migrate
import hashlib
import os
import cx_Oracle
import plotly.subplots as sp
import plotly.graph_objs as go
import matplotlib as mpt
import time
import datetime 
import traceback
from flask import abort
#from flask import g, session
import tkinter as tk
from tkinter import messagebox
import win32com.client
import pythoncom
import mpld3
import sklearn

from matplotlib.ticker import MaxNLocator, FuncFormatter
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error
from sklearn.preprocessing import PolynomialFeatures
import matplotlib.dates as mdates
import joblib
import io
import base64
import plotly.graph_objs as go
import plotly.io as pio
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from functools import wraps
from task_schedule import update_task, get_all_tasks

### import customize function
from generate_modbus import computeCRC, build_request_message, format_tx_message

from convert_modbus import convert_raw_to_value

from create_SQL import create_SQL_text_delete_Billing
from create_SQL import create_SQL_text_insert_Billing
from create_SQL import create_SQL_text_insert_Billing_error
from all_function import is_valid_date, verifyNumericReturnNULL
app = Flask(__name__)

app.secret_key = os.environ.get("FLASK_SECRET_KEY", "fallback_secret_key")
# ปิดการแสดงผล log 
log = logging.getLogger('werkzeug') 
log.disabled = True 
app.logger.disabled = True

def md5_hash(input_string):
    return hashlib.md5(input_string.encode()).hexdigest()
app.secret_key = "AMR_AMR_SOMBAT"

app.register_blueprint(main)  # ลงทะเบียน Blueprint

### Global Variable 
QUANTITY_CONFIG_DATA = 20  
QUANTITY_BILLING_PER_DAY = 5

QUANTITY_RANGE_CONFIG_LIST = 10

COUNTER = 0
DAY_ROUND_REPEAT = 3
CONFIG_FILE = r"C:\AMRconfig.dat"

MAX_ADDRESS_LENGTH = 249

LIST_OF_CHANGE = ["CO2", "N2", "SG", "Pressure Base", "Pb", "Temperature Base", "Tb", "Pulse weight", "Imp.w"]  # highlight on change of config data


### Global variable declaration
def read_config(file_path):
    config = {}
    try:
        with open(file_path, "r") as file:
            for line in file:
                line = line.strip()  # ลบช่องว่างหรือ newline
                # ตรวจสอบว่าบรรทัดไม่ได้เป็น comment
                if not line.startswith("#") and "=" in line:
                    key, value = line.split("=", 1)
                    value = value.strip().strip('"')
                    config[key.strip()] = value.strip()
    except FileNotFoundError:
        print(f"File not found: {file_path}")
    return config


config_data = read_config(CONFIG_FILE)

# ประกาศตัวแปรจากค่าที่อ่านได้
username = config_data.get("username", "")
password = config_data.get("password", "")
hostname = config_data.get("hostname", "")
port = config_data.get("port", "")
service_name = config_data.get("service_name", "")
FILE_PATH = config_data.get("FILE_PATH", "")

# print(f"Username: {username}")
# print(f"Password: {password}")
# print(f"Hostname: {hostname}")
# print(f"Port: {port}")
# print(f"Service Name: {service_name}")
# print(f"File Path: {FILE_PATH}")

# username = "PTT_PIVOT"
# password = "PTT_PIVOT"
# hostname = "10.100.56.3"
# port = "1521"
# service_name = "PTTAMR_MST"

# username = "root"
# password = "amr2025"
# hostname = "192.168.102.192"
# port = "1521"
# service_name = "orcl"
#For popup Analyzer
#FILE_PATH = 'C:/Users/Administrator/Documents/GitHub/tsoamr.pttplc/txt/manual_popup/popupAnalys.txt'
#FILE_PATH = 'C:/Users/kingt/OneDrive/Desktop/AMR_TNAW/tsoamr.pttplc/txt/manual_popup/popupAnalys.txt'


communication_traffic = []
change_to_32bit_counter = 0 

## End Global
################

def convert_to_binary_string(value, bytes_per_value):
    binary_string = bin(value)[2:]  # Convert the value to binary string excluding the '0b' prefix
    return binary_string.zfill(bytes_per_value * 8)  # Zero-fill to fit the number of bits based on bytes_per_value

# Set the Flask secret key from the environment variable
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "fallback_secret_key")

def md5_hash(input_string):
    # เข้ารหัสรหัสผ่านโดยใช้ MD5
    return hashlib.md5(input_string.encode()).hexdigest()

############  connect database  #####################
active_connection = None  # Global variable to track the active connection


def connect_to_ptt_pivot_db():

    global active_connection
    # username = "PTT_PIVOT"
    # password = "PTT_PIVOT"
    # hostname = "10.100.56.3"
    # port = "1521"
    # service_name = "PTTAMR_MST"

    try:
        dsn = cx_Oracle.makedsn(hostname, port, service_name=service_name)
        connection = cx_Oracle.connect(username, password, dsn)
        active_connection = "PTT_PIVOT"
        print("Connected to PTT PIVOT database")
        return connection
    except cx_Oracle.Error as e:
        (error,) = e.args
        print("Oracle Error:", error)
        return None

def fetch_data(connection, query, params=None):
    try:
        with connection.cursor() as cursor:
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            results = cursor.fetchall()
        return results
    except cx_Oracle.Error as e:
        (error,) = e.args
        print("Oracle Error:", error)
        return []

def execute_query(connection, query, params=None):
    try:
        with connection.cursor() as cursor:
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            connection.commit()
        return True
    except cx_Oracle.Error as e:
        print("Oracle Error:", e)
        return False
    finally:
        connection.close()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
           
            return redirect(url_for('login'))  # Redirect to the login page
        return f(*args, **kwargs)
    return decorated_function

def get_data(username, password, hostname, port, service_name, filter_text=None, sort_column=None):
    try:
        connection = cx_Oracle.connect(user=username, password=password, dsn=f"{hostname}:{port}/{service_name}")
        cursor = connection.cursor()
        
        # คำสั่ง SQL พื้นฐาน
        query = (
            "SELECT description, USER_NAME, PASSWORD, USER_LEVEL FROM AMR_USER where description NOT like '%.remove%' "
        )

        # การกรองข้อมูล
        if filter_text:
            query += f" AND USER_NAME LIKE '%{filter_text}%'"

        # การเรียงลำดับข้อมูล
        if sort_column:
            query += f" ORDER BY {sort_column}"
        cursor.execute(query)

        # ดึงข้อมูลเป็นช่วงๆ (เช่น 100 แถวต่อครั้ง)
        chunk_size = 100
        data = []
        while True:
            rows = cursor.fetchmany(chunk_size)
            if not rows:
                break
            data.extend(
                [
                    {
                        "description": row[0],
                        "user_name": row[1],
                        "password": row[2],
                        "user_level": row[3],
                    }
                    for row in rows
                ]
            )
        return data
    except cx_Oracle.Error as e:
        (error,) = e.args
        print("Oracle Error:", error)
        return []
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()


# Example usage with filtering and sorting
filter_text = "example"  # Replace with your filter text or None for no filtering
sort_column = "USER_NAME"  # Replace with your desired column or None for no sorting
filtered_and_sorted_data = get_data(username, password, hostname, port, service_name, filter_text=filter_text, sort_column=sort_column)

############  /connect database  #####################

def get_autopoll_statistic(ptt_pivot_connection, tags_status="", date_system = datetime.datetime.now().strftime('%Y-%m-%d')):
    #date_system = datetime.datetime.now().strftime('%d-%m-%Y')
    data_autopoll_query = f"""
    SELECT DISTINCT amr_configured_data.meter_id, amr_field_id.tag_id
    FROM amr_configured_data, amr_field_id
    WHERE amr_configured_data.meter_id = amr_field_id.meter_id
    AND amr_configured_data.data_date = TO_DATE('{date_system}', 'YYYY-MM-DD')
    """
    data_autopoll = fetch_data(ptt_pivot_connection, data_autopoll_query)
    df_data_autopoll = pd.DataFrame(data_autopoll, columns=['meter_id', 'tag_id'])
    
    data_autopoll_error_query = f"""
    SELECT DISTINCT TAG_ID 
    FROM AMR_ERROR 
    WHERE TRUNC(DATA_DATE) = TO_DATE('{date_system}', 'YYYY-MM-DD') 
    AND REPEAT = {DAY_ROUND_REPEAT}
    """
    data_autopoll_error = fetch_data(ptt_pivot_connection, data_autopoll_error_query)
    df_data_autopoll_error = pd.DataFrame(data_autopoll_error, columns=['tag_id'])
    df_data_autopoll_cleaned = df_data_autopoll[~df_data_autopoll['tag_id'].isin(df_data_autopoll_error['tag_id'])]

    data_manual_recover_query = f"""
    SELECT DISTINCT TAG_ID 
    FROM AMR_ERROR 
    WHERE TRUNC(DATA_DATE) = TO_DATE('{date_system}', 'YYYY-MM-DD') 
    AND REPEAT = 99
    """
    data_manual_recover = fetch_data(ptt_pivot_connection, data_manual_recover_query)
    df_manual_recover = pd.DataFrame(data_manual_recover, columns=['tag_id'])
    df_data_autopoll_cleaned = df_data_autopoll_cleaned[~df_data_autopoll_cleaned['tag_id'].isin(df_manual_recover['tag_id'])]

    error_content = df_data_autopoll_error.shape[0]
    manual_content = df_manual_recover.shape[0]
    success_content = df_data_autopoll_cleaned.shape[0]
    
    allsite = success_content + error_content + manual_content
    error_content = error_content + manual_content
    
    if tags_status == "Success": 
        tag_ids = sorted(df_data_autopoll['tag_id'].tolist())
    else :
        tag_ids = sorted(df_data_autopoll_error['tag_id'].tolist())
    
    tags_manuals = sorted(df_manual_recover['tag_id'].tolist())
    
    return {
        'error_content': error_content,
        'manual_content': manual_content,
        'success_content': success_content,
        'allsite': allsite,
        'tag_ids': tag_ids,
        'tags_manuals': tags_manuals
    }

def fetch_user_data(username):
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        query = """
            SELECT "USER_NAME", "PASSWORD", "DESCRIPTION", "USER_LEVEL", "USER_GROUP"
            FROM AMR_USER
            WHERE "USER_NAME" = :username
        """
        params = {'username': username}
        user_data = fetch_data(ptt_pivot_connection, query, params)
        
        if user_data:
            username, password, description, user_level, user_group = user_data[0]
            return {'password': password, 'description': description, 'user_level': int(user_level), 'user_group': user_group}  
        return None

@app.route("/add_user", methods=["GET", "POST"])
@login_required
def add_user_route():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
   
        user_update = session['username']

        query = f"""SELECT user_level, user_group FROM amr_user_level"""
    
        name_group = fetch_data(ptt_pivot_connection,query)
    
    current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))).strftime('%d-%b-%y %I.%M.%S.%f %p ASIA/BANGKOK').upper()
    if request.method == "POST":
        description = request.form["description"]
        user_name = request.form["user_name"]
        password = request.form["password"]
        user_level = request.form["user_level"]

        # Hash the password using MD5
        hashed_password = md5_hash(password)
        USER_ENABLE = 1
        # Define the Oracle SQL query with RAWTOHEX and DBMS_OBFUSCATION_TOOLKIT
        query = """
            INSERT INTO AMR_USER (description, user_name, password, user_level, TIME_CREATE,USER_ENABLE,UPDATED_BY) 
            VALUES (:1, :2, RAWTOHEX(DBMS_OBFUSCATION_TOOLKIT.MD5(input_string => UTL_I18N.STRING_TO_RAW(:3, 'AL32UTF8'))), :4, :5,:6,:7)
        """

        # Prepare the parameters for the query
        params = (description, user_name, hashed_password, user_level, current_datetime,USER_ENABLE,user_update)
             
        if execute_query(connect_to_ptt_pivot_db(), query, params):
            flash('User added successfully', 'success') 
            
            return redirect(url_for("add_user_route"))  
        else:
            flash('Failed to add user', 'error')  

    return render_template("add_user.html", name_group=name_group,user_update=user_update)

@app.route("/get_data")
def get_data_route():
    username = "PTT_PIVOT"
    password = "PTT_PIVOT"
    hostname = "10.100.56.3"
    port = "1521"
    service_name = "PTTAMR_MST"
        
    filter_text = request.args.get("filter_text")  # สามารถดึงค่าการกรองจาก request
    sort_column = request.args.get("sort_column")  # สามารถดึงค่าคอลัมน์การเรียงจาก request

    data = get_data(username, password, hostname, port, service_name, filter_text=filter_text, sort_column=sort_column)
    return jsonify(data)

@app.route("/data_edit_user", methods=["POST"])
@login_required
def data_edit_user():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        data = request.get_json()  # Get JSON data from the request
        selected_level = data.get("Select_level")  # Get selected level
        selected_tag = data.get("selected_tag")  # Get selected tag
        username = session['username']
        
        if selected_level and not selected_tag:
            # This case handles the initial selection of the user level
            query = f"""SELECT description FROM amr_user
                        WHERE user_level = {selected_level}
                        AND user_enable = 1
                        ORDER BY description"""
            result = fetch_data(ptt_pivot_connection, query)
            df_tag = pd.DataFrame(result, columns=['description'])
            print("df_tag:",df_tag)
            # Return the list of tags as JSON to populate the dropdown
            return jsonify(df_tag['description'].tolist())
        
        elif selected_level and selected_tag:
            # This case handles when both level and tag are selected
            query = f"""SELECT description, USER_NAME, USER_LEVEL 
                        FROM amr_user  
                        WHERE description = '{selected_tag}'"""
            result = fetch_data(ptt_pivot_connection, query)
            data_tag = pd.DataFrame(result, columns=['description', 'USER_NAME', 'USER_LEVEL'])
            
            # Convert the DataFrame to a dictionary and return as JSON
            if not data_tag.empty:
                user_data = data_tag.iloc[0].to_dict()  # Get the first row of the DataFrame as a dictionary
                
                # Store description in session
                session['description'] = data_tag['description'].iloc[0]  # Store the first description value in session
                
                return jsonify(user_data)
            else:
                return jsonify({"error": "No data found"}), 404
    
@app.route("/edit_user", methods=["GET", "POST"])
@login_required
def edit_user_route():
    # ดึงข้อมูลผู้ใช้จากฐานข้อมูล
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        query = "SELECT DESCRIPTION, USER_NAME, PASSWORD, USER_LEVEL FROM AMR_USER"
        user_data = fetch_data(ptt_pivot_connection,query)
        # print(user_data)
        query_name = f"""SELECT user_level, user_group FROM amr_user_level"""
        name_group = fetch_data(ptt_pivot_connection,query_name)

    user_update = session['username']
    if not user_data:
        flash("User not found!", "error")
        return render_template("edit_user.html")
    current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))).strftime('%d-%b-%y %I.%M.%S.%f %p ASIA/BANGKOK').upper()
    # ถ้ามีการส่งค่า POST (คือการบันทึกการแก้ไข)
    if request.method == "POST":
        # ดึงข้อมูลจากฟอร์มแก้ไข
        description_row = session['description']
        description = request.form["description"]
        print("description",description)
        user_name = request.form["user_name"]
        password = request.form["password"]
        user_level = request.form["user_level"]

        # เข้ารหัสรหัสผ่านโดยใช้ MD5
        hashed_password = md5_hash(password)

        if password:
            # สร้างคำสั่ง SQL สำหรับการแก้ไขข้อมูลผู้ใช้
            update_query = """
                UPDATE AMR_USER 
                SET description = :1, 
                    user_name = :2, 
                    password = :3, 
                    user_level = :4, 
                    TIME_CREATE = :5,
                    UPDATED_BY = :6
                WHERE description = :7
            """

            # Prepare the parameters for the query
            update_params = (
                description,
                user_name,
                hashed_password,
                user_level,
                current_datetime,  # Add the current time to TIME_CREATE
                user_update,
                description_row  # Assuming you're matching on the original description
            )
            # print(update_params)
        else:
            update_query = """
                UPDATE AMR_USER 
                SET description = :1, 
                    user_name = :2, 
                    
                    user_level = :3, 
                    TIME_CREATE = :4,
                     UPDATED_BY = :5
                WHERE description = :6
            """

            # Prepare the parameters for the query
            update_params = (
                description,
                user_name,
                
                user_level,
                current_datetime,  # Add the current time to TIME_CREATE
                user_update,
                description_row  # Assuming you're matching on the original description
            )
            # print(update_params)
        # ทำการ execute คำสั่ง SQL และ commit การแก้ไข user_name
        if execute_query(connect_to_ptt_pivot_db(),update_query, update_params):
            flash("User updated successfully!", "success")
            return render_template("edit_user.html", user_data=user_data,name_group=name_group)
        else:
            flash("Failed to update user. Please try again.", "error")

    # กรณีไม่ใช่การส่งค่า POST ให้ส่งข้อมูลผู้ใช้ไปยัง HTML template หรือทำอย่างอื่นตามที่ต้องการ
    return render_template("edit_user.html",name_group=name_group, user_data=user_data,user_update=user_update)

@app.route("/data_remove_user", methods=["POST"])
@login_required
def data_remove_user():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        data = request.get_json()  # Get JSON data from the request
        selected_level = data.get("Select_level")  # Get selected level
        selected_tag = data.get("selected_tag")  # Get selected tag
        username = session['username']
        
        if selected_level and not selected_tag:
            # This case handles the initial selection of the user level
            query = f"""SELECT description FROM amr_user
                        WHERE user_level = {selected_level}
                        AND user_enable = 1
                        ORDER BY description"""
            result = fetch_data(ptt_pivot_connection, query)
            df_tag = pd.DataFrame(result, columns=['description'])
            print("df_tag:",df_tag)
            # Return the list of tags as JSON to populate the dropdown
            return jsonify(df_tag['description'].tolist())
        
        elif selected_level and selected_tag:
            # This case handles when both level and tag are selected
            query = f"""SELECT amr_user.description, amr_user.USER_NAME, amr_user_level.USER_Group 
                        FROM amr_user  ,amr_user_level
                        WHERE amr_user.user_level = amr_user_level.user_level
                        AND description = '{selected_tag}'"""
            result = fetch_data(ptt_pivot_connection, query)
            data_tag = pd.DataFrame(result, columns=['description', 'USER_NAME', 'USER_LEVEL'])
            
            # Convert the DataFrame to a dictionary and return as JSON
            if not data_tag.empty:
                user_data = data_tag.iloc[0].to_dict()  # Get the first row of the DataFrame as a dictionary
                
                # Store description in session
                session['description'] = data_tag['description'].iloc[0]  # Store the first description value in session
                
                return jsonify(user_data)
            else:
                return jsonify({"error": "No data found"}), 404
    
@app.route("/remove_user", methods=["GET", "POST"])
@login_required
def remove_user_route():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        query = "SELECT DESCRIPTION, USER_NAME, USER_LEVEL, USER_ENABLE FROM AMR_USER"
        user_data = fetch_data(ptt_pivot_connection,query)
        user_update = session['username']
        query_name = f"""SELECT user_level, user_group FROM amr_user_level"""
        
        name_group = fetch_data(ptt_pivot_connection,query_name)
        if not user_data:
            flash("Users not found!", "error")
            return redirect(url_for("index"))
    current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))).strftime('%d-%b-%y %I.%M.%S.%f %p ASIA/BANGKOK').upper()
    if request.method == "POST":
        # user= request.form.get("myDropdown")
        # print("user", user)
        # user_dict = json.loads(user)  
        # description = user_dict.get("description") 
        
      
        # user_name = request.form.get("user_name")
        description = request.form["description"]
        user_name = request.form["user_name"]
        print("description", description)

        status_mapping = 0
        user_remove = f"{description}.remove"
        
        update_query = "UPDATE AMR_USER SET USER_ENABLE = :1,TIME_CREATE = :2, DESCRIPTION = :3 ,UPDATED_BY = :4 WHERE USER_NAME = :5"
        update_params = (status_mapping, current_datetime, user_remove,user_update, user_name)
        
        if execute_query(connect_to_ptt_pivot_db(),update_query, update_params):
            flash("User status updated successfully!", "success")
            return redirect(url_for("remove_user_route"))

        else:
            flash("Failed to update user status. Please try again.", "error")

    return render_template("remove_user.html",name_group=name_group, user_data=user_data,user_update=user_update)
users = {}

@app.route('/login', methods=['GET', 'POST'])
def login():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        error_message = None  
        current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))).strftime('%d-%b-%y %I.%M.%S.%f %p ASIA/BANGKOK').upper()
        if request.method == 'POST':
            entered_username = request.form['username']
            entered_password = request.form['password']
            errors = []

            # Password validation criteria
            if len(entered_password) < 8:
                errors.append("Password must be at least 8 characters long.")
            if not re.search("[a-z]", entered_password):
                errors.append("Password must contain at least one lowercase letter.")
            if not re.search("[A-Z]", entered_password):
                errors.append("Password must contain at least one uppercase letter.")
            if not re.search("[0-9]", entered_password):
                errors.append("Password must contain at least one digit.")
            if not re.search("[@#$%^&+=]", entered_password):
                errors.append("Password must contain at least one special character (@, #, $, %, ^, &, +, =).")

            if errors:
                flash("⚠️ **Password is not secure! Please address the following issues:**<br>" + "<br>".join(errors), "error")
            else:
                # If the password passes all checks
                pass
            # Query to fetch user data from the database
            user_level = "3"
            query = """
                SELECT "USER_NAME", "PASSWORD", "DESCRIPTION", "USER_LEVEL", "USER_GROUP"
                FROM AMR_USER
                WHERE "USER_NAME" = :entered_username
                AND AMR_USER.user_enable like '1'
            """

            params = {'entered_username': entered_username}
            user_data = fetch_data(ptt_pivot_connection, query, params)
            
            query_time = f"""SELECT time_create FROM amr_user WHERE user_name = '{entered_username}'"""
            
            data_time = fetch_data(ptt_pivot_connection, query_time)
            
            df = pd.DataFrame(data_time,columns=['time_create'])
         

            if df['time_create'].isnull().all():
                if user_data:
                    stored_password, description, user_level, user_group = user_data[0][1:]

                    # ตรวจสอบว่ารหัสผ่านถูกเข้ารหัส MD5 หรือไม่
                    is_md5_hash = re.match(r'^[a-fA-F0-9]{32}$', stored_password)
                    
                    # เข้ารหัสรหัสผ่านที่ป้อน หากรหัสผ่านที่เก็บไว้เป็น MD5
                    if is_md5_hash:
                        entered_password = hashlib.md5(entered_password.encode()).hexdigest()

                    # ตรวจสอบว่ารหัสผ่านตรงกันหรือไม่
                    is_password_valid = entered_password == stored_password

                    if is_password_valid:
                        session['username'] = entered_username
                        users[entered_username] = {
                            'password': stored_password, 
                            'description': description, 
                            'user_level': user_level, 
                            'user_group': user_group
                        }

                        # สร้าง SQL สำหรับบันทึกประวัติการเข้าสู่ระบบ
                        user_history = f"""
                            INSERT INTO AMR_USER_HISTORY (USER_NAME, ACCESS_TIME, USER_LEVEL) 
                            VALUES ('{entered_username}', '{current_datetime}', '{user_level}')
                        """
                        update_sql(ptt_pivot_connection, user_history)

                        # ใช้ Dictionary Mapping เพื่อลดเงื่อนไขซ้ำ ๆ
                        redirect_mapping = {
                            '1': 'home_amr',
                            '4': 'home_amr',
                            '5': 'home_amr',
                            '2': 'home_user_group',
                            '3': 'home_user'
                        }

                        return redirect(url_for(redirect_mapping.get(user_level, 'login')))  # ถ้าไม่มี user_level ที่ตรงกันจะ redirect ไป login
                    else:
                        error_message = 'Incorrect password'
                else:
                    error_message = 'User not found'
                        
            else:
                
                    df['time_create'] = pd.to_datetime(df['time_create'], format='%d-%b-%y %I.%M.%S.%f %p %Z', errors='coerce')
                    date_only = df['time_create'].dt.strftime('%d-%m-%Y').iloc[0]
                    date_only_dt = datetime.datetime.strptime(date_only, '%d-%m-%Y')
                    date_system = datetime.datetime.now()
                    difference = (date_system - date_only_dt).days
                    # print(f"Date Only: {date_only}")
                    # print(f"Date System: {date_system.strftime('%d-%m-%Y')}")
                    # print(f"Difference in days: {difference}")

                    if difference > 90:
                        
                        flash("⚠️**Your password has been in use for over 90 days. Please change your password.")
                        if user_data:
                            stored_password, description, user_level, user_group = user_data[0][1:]
                            entered_password = hashlib.md5(entered_password.encode()).hexdigest()
                            if entered_password == stored_password:
                                session['username'] = entered_username

                                users[entered_username] = {
                                    'password': stored_password,
                                    'description': description,
                                    'user_level': user_level,
                                    'user_group': user_group
                                }
                                if user_level == '1' or user_level == '4' or user_level == '5':
                                    return redirect(url_for('home_amr'))
                                elif user_level == '2':
                                    return redirect(url_for('home_user_group'))
                                elif user_level == '3':
                                    return redirect(url_for('home_user'))
                            else:
                                error_message = 'Incorrect password!'
                        else:
                            error_message = 'User not found!'  
            
                    else:
                        print(f"ยังเหลือเวลาอีก {90 - difference} วันก่อนจะครบ 90 วัน")

                        if user_data:
                            stored_password, description, user_level, user_group = user_data[0][1:]
                            entered_password = hashlib.md5(entered_password.encode()).hexdigest()
                            if entered_password == stored_password:
                                session['username'] = entered_username

                                users[entered_username] = {'password': stored_password, 'description': description, 'user_level': user_level, 'user_group': user_group}
                                if user_level == '1' or user_level == '4' or user_level == '5':
                                    return redirect(url_for('home_amr'))
                                elif user_level == '2':
                                    return redirect(url_for('home_user_group'))
                                elif user_level == '3':
                                    return redirect(url_for('home_user'))
                            else:
                                error_message = 'Incorrect password!'
                        else:
                            error_message = 'User not found!' 
        
                    df= pd.DataFrame(data_time,columns=['time_create'])
                
        return render_template('login.html', error_message=error_message)
  
    
@app.route('/user_history')
@login_required
def user_history():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        username=session['username']
        # Get the selected date from the request or default to today's date
        selected_date = request.args.get("date_dropdown")
        
        if not selected_date:
            today = datetime.datetime.today()
            selected_date = today.strftime('%d-%m-%Y')  # Default to today's date if none is provided
        
        print("selected_date", selected_date)

        # Convert the date from '14-10-2024' to '07-OCT-24'
        date_object = datetime.datetime.strptime(selected_date, '%d-%m-%Y')  # Convert to datetime object
        formatted_date = date_object.strftime('%d-%b-%y').upper()  # Convert to '07-OCT-24'

        print("formatted_date", formatted_date)

        # Query to fetch user history based on the formatted date
        user_history = f"""
             SELECT AMR_USER_HISTORY.USER_NAME, amr_user_level.USER_Group, AMR_USER_HISTORY.ACCESS_TIME
            FROM AMR_USER_HISTORY ,amr_user_level
            WHERE amr_user_level.user_level = amr_user_history.user_level AND
            TRUNC(TO_DATE(SUBSTR(access_time, 1, 9), 'DD-MON-RR', 'NLS_DATE_LANGUAGE=ENGLISH')) 
                = TO_DATE('{formatted_date}', 'DD-MON-RR', 'NLS_DATE_LANGUAGE=ENGLISH')
            ORDER BY ACCESS_TIME DESC

        """
        user_data = fetch_data(ptt_pivot_connection, user_history)
        df_user_data = pd.DataFrame(user_data, columns=['USER_NAME', 'USER_LEVEL', 'ACCESS_TIME'])
        print(user_history)
       

        user_len = len(df_user_data)
        user_data_list = df_user_data.to_dict('records')
        
    return render_template('user_history.html', user_len=user_len, selected_date=selected_date, user_data_list=user_data_list,username=username)  
    
@app.route('/Change_Password_user', methods=['GET', 'POST'])
@login_required
def Change_Password_user():
    username=session['username']
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        if request.method == 'POST':
            entered_username = request.form['user_name']
            current_Password = request.form['password']
            new_Password = request.form['newpassword']
            confirm_NewPassword = request.form['confirmpassword']
   
            hashed_current_password = hashlib.md5(current_Password.encode()).hexdigest()
            print(f"Entered Username: {entered_username}")
            print(f"Current Password (plain): {current_Password}")
            print(f"Hashed Current Password: {hashed_current_password}")
            print(f"New Password: {new_Password}")
            print(f"Confirm New Password: {confirm_NewPassword}")
                
            query = f"""SELECT password FROM amr_user WHERE user_name = '{entered_username}'"""
            user_data = fetch_data(ptt_pivot_connection, query)
            stored_password_hash = user_data[0][0]  
            
            print(f"Stored Password Hash: {stored_password_hash}")
            if re.match(r'^[a-fA-F0-9]{32}$', stored_password_hash):
    
                print("Password is already an MD5 hash.")
            else:
                stored_password_hash = hashlib.md5(stored_password_hash.encode()).hexdigest()
                print(f"Converted Password to MD5: {stored_password_hash}")
            
            print(hashed_current_password,stored_password_hash)
            
            if hashed_current_password == stored_password_hash:
                print("Current password is correct. Proceed with password change.")
        
                pass
            else:
                print("Current password is incorrect.")
                
                return render_template('Change_Password_user.html',username=username, error_message="The current password is incorrect.")
                
                
                # Hash the password using MD5
            hashed_password = md5_hash(confirm_NewPassword)

            # Convert to RAWTOHEX before storing in Oracle
            hashed_password_hex = "RAWTOHEX(DBMS_OBFUSCATION_TOOLKIT.MD5(input_string => UTL_I18N.STRING_TO_RAW('{}', 'AL32UTF8')))".format(
                hashed_password
            )
            print(hashed_password_hex)
            current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))).strftime('%d-%b-%y %I.%M.%S.%f %p ASIA/BANGKOK').upper()
            query_password = f"""UPDATE amr_user SET password = '{hashed_password}' ,TIME_CREATE = '{current_datetime}',UPDATED_BY = '{username}' WHERE user_name = '{entered_username}'"""
            print(query_password)
            update_sql(ptt_pivot_connection,query_password)
            
            return redirect(url_for('login'))
    return render_template('Change_Password_user.html',username=username)


@app.route('/Change_Password_user_group', methods=['GET', 'POST'])
@login_required
def Change_Password_user_group():
    username=session['username']
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        if request.method == 'POST':
            entered_username = request.form['user_name']
            current_Password = request.form['password']
            new_Password = request.form['newpassword']
            confirm_NewPassword = request.form['confirmpassword']

            hashed_current_password = hashlib.md5(current_Password.encode()).hexdigest()
            print(f"Entered Username: {entered_username}")
            print(f"Current Password (plain): {current_Password}")
            print(f"Hashed Current Password: {hashed_current_password}")
            print(f"New Password: {new_Password}")
            print(f"Confirm New Password: {confirm_NewPassword}")
                
            query = f"""SELECT password FROM amr_user WHERE user_name = '{entered_username}'"""
            user_data = fetch_data(ptt_pivot_connection, query)
            stored_password_hash = user_data[0][0]  
            
            print(f"Stored Password Hash: {stored_password_hash}")
            if re.match(r'^[a-fA-F0-9]{32}$', stored_password_hash):
    
                print("Password is already an MD5 hash.")
            else:
                
                stored_password_hash = hashlib.md5(stored_password_hash.encode()).hexdigest()
                print(f"Converted Password to MD5: {stored_password_hash}")
            
            print(hashed_current_password,stored_password_hash)
            
            if hashed_current_password == stored_password_hash:
                print("Current password is correct. Proceed with password change.")
        
                pass
            else:
                print("Current password is incorrect.")
                
                return render_template('Change_Password_user_group.html',username=username, error_message="The current password is incorrect.")
                
                # Hash the password using MD5
            hashed_password = md5_hash(confirm_NewPassword)

            # Convert to RAWTOHEX before storing in Oracle
            hashed_password_hex = "RAWTOHEX(DBMS_OBFUSCATION_TOOLKIT.MD5(input_string => UTL_I18N.STRING_TO_RAW('{}', 'AL32UTF8')))".format(
                hashed_password
            )
            print(hashed_password_hex)
            current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))).strftime('%d-%b-%y %I.%M.%S.%f %p ASIA/BANGKOK').upper()
            query_password = f"""UPDATE amr_user SET password = '{hashed_password}' ,TIME_CREATE = '{current_datetime}',UPDATED_BY = '{username}' WHERE user_name = '{entered_username}'"""
            print(query_password)
            update_sql(ptt_pivot_connection,query_password)
            
            return redirect(url_for('login'))
    return render_template('Change_Password_user_group.html',username=username)
    

@app.route('/data_report')
@login_required
def data_report():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", ptt_pivot_connection)

        # Retrieve data from the session
        Data_overview = session['rows']
        Data_evctype = session['evc_type']

        # Convert Data_overview to DataFrame
        df_overview = pd.DataFrame(Data_overview, columns=['Date', 'Success', 'Error'])

        # Prepare evctype_counts with custom labels
        custom_labels = ['Elster','Actaris']  # Custom labels

        # Ensure the length of custom_labels matches the length of Data_evctype
        if len(custom_labels) != len(Data_evctype):
            raise ValueError("The length of custom_labels must match the length of Data_evctype")

        # Create DataFrame for evctype_counts with custom labels
        evctype_counts_df = pd.DataFrame({
            'Type': custom_labels,
            'Count': Data_evctype
        })

        # Save DataFrames to a BytesIO object
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Save Data_overview to the first sheet
            df_overview.to_excel(writer, index=False, sheet_name='Data Overview')

            # Access the workbook and the sheet
            workbook = writer.book
            sheet_overview = workbook['Data Overview']

            # Define positions for the data and charts
            start_row_overview = 1
            print("len",len(df_overview))
            start_row_evctype = len(df_overview) + 5  # Leave some space between sections

            # Apply header styles
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            bold_font = Font(bold=True)
            alignment = Alignment(horizontal="center")

            # Write headers and Data_overview to the sheet with styling
            for col_num, header in enumerate(df_overview.columns, 1):
                cell = sheet_overview.cell(row=start_row_overview, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = bold_font
                cell.alignment = alignment

            # Apply alternating row colors
            for r_idx, row in df_overview.iterrows():
                fill_color = "F2F2F2" if r_idx % 2 == 0 else "FFFFFF"
                for c_idx, value in enumerate(row):
                    cell = sheet_overview.cell(row=start_row_overview + r_idx + 1, column=c_idx + 1, value=value)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            # Write headers and evctype_counts_df to the sheet with styling
            for col_num, header in enumerate(evctype_counts_df.columns, 1):
                cell = sheet_overview.cell(row=start_row_evctype, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = bold_font
                cell.alignment = alignment

            for r_idx, row in evctype_counts_df.iterrows():
                fill_color = "F2F2F2" if r_idx % 2 == 0 else "FFFFFF"
                for c_idx, value in enumerate(row):
                    cell = sheet_overview.cell(row=start_row_evctype + r_idx + 1, column=c_idx + 1, value=value)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            # Create a bar chart for Data_overview
            chart = BarChart()
            chart.title = "Data Overview"
            chart.style = 13
            chart.x_axis.title = 'Date'
            chart.y_axis.title = 'Values'

            # Define the data for the chart
            data = Reference(sheet_overview, min_col=2, min_row=2, max_col=3, max_row=len(df_overview) + 1)
            chart.add_data(data, titles_from_data=False)  # Do not use titles from data
            chart.set_categories(Reference(sheet_overview, min_col=1, min_row=2, max_row=len(df_overview) + 1))

            # Manually set the series names (Success and Error)
            success_series = chart.series[0]
            success_series.tx = SeriesLabel(v="Success")  # Set custom title for success series

            # Set custom title for error series
            error_series = chart.series[1]
            error_series.tx = SeriesLabel(v="Error")  # Set custom title for error series

            # Set colors for each series
            success_series.graphicalProperties.solidFill = "4e73df"  # Light Blue
            error_series.graphicalProperties.solidFill = "e74a3b"    # Light Red

            # Add the chart to the sheet at the top of the Data Overview section
            sheet_overview.add_chart(chart, "E1")  # Adjust the position to start at row 2

            # Create a pie chart for evctype_counts_df
            pie_chart = PieChart()
            pie_chart.title = "EVC Type Distribution"

            # Define the data for the pie chart
           # Define the data for the pie chart
            data_pie = Reference(sheet_overview, min_col=2, min_row=start_row_evctype + 0, max_col=2, max_row=start_row_evctype + len(evctype_counts_df))
            labels = Reference(sheet_overview, min_col=1, min_row=start_row_evctype + 1, max_row=start_row_evctype + len(evctype_counts_df))

            print(data_pie)
            # Add data and labels to the pie chart
            pie_chart.add_data(data_pie, titles_from_data=True)
            pie_chart.set_categories(labels)

            # Add the pie chart to the sheet starting at row 17
            sheet_overview.add_chart(pie_chart, "E17")
            print("Data for pie chart:", evctype_counts_df)
            print("Data range for pie chart: min_row={}, max_row={}".format(start_row_evctype + 1, start_row_evctype + len(evctype_counts_df)))
        output.seek(0)  # Move the cursor to the beginning of the BytesIO object

        # Return the file as a downloadable response
        return send_file(
            output,
            as_attachment=True,
            download_name='data_report.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

@app.route('/data_report_Dashboard_pdf')
def data_report_Dashboard_pdf():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", ptt_pivot_connection)

        # Retrieve data from the session
        Data_overview = session['rows']
        Data_evctype = session['evc_type']

        # Convert Data_overview to DataFrame
        df_overview = pd.DataFrame(Data_overview, columns=['Date', 'Success', 'Error'])

        # Prepare evctype_counts with custom labels
        custom_labels = [ 'Elster','Actaris']

        # Ensure the length of custom_labels matches the length of Data_evctype
        if len(custom_labels) != len(Data_evctype):
            raise ValueError("The length of custom_labels must match the length of Data_evctype")

        evctype_counts_df = pd.DataFrame({
            'Type': custom_labels,
            'Count': Data_evctype
        })

        # Prepare an HTML template for the PDF with smaller, compact tables and page breaks
        html_template = """
        <html>
        <head>
            <style>
                body {
                    font-size: 12px;
                    margin: 20px;
                }
                table {
                    width: 100%;
                    border-collapse: collapse;
                    font-size: 10px;
                }
                th, td {
                    border: 1px solid black;
                    padding: 4px;
                    text-align: left;
                }
                th {
                    background-color: #f2f2f2;
                }
                h1, h2 {
                    font-size: 14px;
                }
                .page-break {
                    page-break-before: always;
                }
            </style>
        </head>
        <body>
            <h1>Data Report Overview</h1>

            <h2>Data Overview</h2>
            <table>
                <tr>
                    <th>Date</th>
                    <th>Success</th>
                    <th>Error</th>
                </tr>
                {% for row in df_overview %}
                <tr>
                    <td>{{ row['Date'] }}</td>
                    <td>{{ row['Success'] }}</td>
                    <td>{{ row['Error'] }}</td>
                </tr>
                {% endfor %}
            </table>

            <div class="page-break"></div>  <!-- Page break before the next section -->

            <h2>EVC Type Distribution</h2>
            <table>
                <tr>
                    <th>Type</th>
                    <th>Count</th>
                </tr>
                {% for row in evctype_counts_df %}
                <tr>
                    <td>{{ row['Type'] }}</td>
                    <td>{{ row['Count'] }}</td>
                </tr>
                {% endfor %}
            </table>
        </body>
        </html>
        """

        # Render the HTML template with the DataFrame data
        rendered_html = render_template_string(html_template, df_overview=df_overview.to_dict(orient='records'), evctype_counts_df=evctype_counts_df.to_dict(orient='records'))

        # Create a BytesIO object to hold the PDF
        pdf_output = BytesIO()

        # Convert the HTML to PDF using pisa
        pisa_status = pisa.CreatePDF(BytesIO(rendered_html.encode('utf-8')), dest=pdf_output)

        # Check if there was an error during PDF creation
        if pisa_status.err:
            return f"Error creating PDF: {pisa_status.err}"

        # Move the cursor to the beginning of the BytesIO object
        pdf_output.seek(0)

        # Return the PDF as a downloadable response
        return send_file(pdf_output, as_attachment=True, download_name='data_report_dashboard.pdf', mimetype='application/pdf')


@app.route('/data_report_billing_user_group')
def data_report_billing_user_group():
    with open('data_report_user_group.json', 'r') as json_file:
        combined_data = json.load(json_file)

   
    selected_date = session.get('selected_date_user_group', 'N/A')
    selected_tag = session.get('selected_tag_user_group', 'N/A')
    print("selected_date:", selected_date)
    print("selected_tag:", selected_tag)
    output = io.BytesIO()

    # Create an Excel writer object
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet("Billing Data Report")

        # Define formats
        title_format = workbook.add_format({'bold': True, 'font_size': 16, 'align': 'center', 'bg_color': '#D9D9D9'})
        header_format = workbook.add_format({'bold': True, 'align': 'center', 'bg_color': '#D9D9D9', 'border': 1})
        data_format = workbook.add_format({'align': 'center', 'border': 1})
        total_format = workbook.add_format({'bold': True, 'align': 'center', 'bg_color': '#FFCCCB', 'border': 1})

        # Add report header
        worksheet.merge_range('B1:H1', "Billing Data Report", title_format)
        worksheet.merge_range('B2:H2', f"Report Date: {selected_date}", header_format)
        worksheet.merge_range('B3:H3', f"Site: {selected_tag}", header_format)

        start_col = 1  # Start from column B (index 1)
        start_row = 4  # Start from row 5 (index 4)

        # Dictionary to store daily totals for CORRECTED_DIFF and UNCORRECTED_DIFF
        daily_totals = {}

        for i, data_dict in enumerate(combined_data):
            df = pd.DataFrame(data_dict)

            # Reorder columns
            df = df[['DATA_DATE', 'CORRECTED', 'UNCORRECTED', 'Pressure', 'Temperature']]

            # Convert 'CORRECTED' and 'UNCORRECTED' columns to numeric
            df['CORRECTED'] = pd.to_numeric(df['CORRECTED'], errors='coerce')
            df['UNCORRECTED'] = pd.to_numeric(df['UNCORRECTED'], errors='coerce')

            # Forward fill only for the purpose of calculating the differences, but don't modify the original columns
            df_filled = df.fillna(method='ffill')

            # Calculate the differences using the filled data
            df['CORRECTED_DIFF'] = df_filled['CORRECTED'].diff()
            df['UNCORRECTED_DIFF'] = df_filled['UNCORRECTED'].diff()

            # Where the original data had missing values, set the diffs to NaN (or keep them empty)
            df.loc[df['CORRECTED'].isna(), 'CORRECTED_DIFF'] = None
            df.loc[df['UNCORRECTED'].isna(), 'UNCORRECTED_DIFF'] = None

            # เขียนข้อมูลทั้งหมด (รวมถึง NaN) ลงใน Excel sheet
            df.to_excel(writer, sheet_name="Billing Data Report", startrow=start_row + 2, startcol=start_col, index=False, header=False)

            worksheet = writer.sheets["Billing Data Report"]

            # Format header for the run info
            worksheet.merge_range(start_row, start_col, start_row, start_col + len(df.columns) - 1, f'Run {i + 1} (Site: {selected_tag}, Date: {selected_date})', header_format)

            for col_num, value in enumerate(df.columns):
                worksheet.write(start_row + 1, start_col + col_num, value, header_format)

            # Format the data rows
            for row_num in range(start_row + 2, start_row + 2 + len(df)):
                worksheet.set_row(row_num, None, data_format)

            # Calculate totals for CORRECTED_DIFF and UNCORRECTED_DIFF
            total_corrected_diff = df['CORRECTED_DIFF'].sum()
            total_uncorrected_diff = df['UNCORRECTED_DIFF'].sum()

            # Accumulate daily totals
            for date, corrected_value, uncorrected_value in zip(df['DATA_DATE'], df['CORRECTED_DIFF'], df['UNCORRECTED_DIFF']):
                if pd.notna(date):  # Check that date is not NaN
                    if date not in daily_totals:
                        daily_totals[date] = [0, 0]  # [total_corrected, total_uncorrected]
                    daily_totals[date][0] += corrected_value if pd.notna(corrected_value) else 0
                    daily_totals[date][1] += uncorrected_value if pd.notna(uncorrected_value) else 0

            # Write totals to the worksheet
            total_row = start_row + 2 + len(df)  # Row for totals
            worksheet.write(total_row, start_col + 4, 'Total', total_format)
            worksheet.write(total_row, start_col + 5, total_corrected_diff, total_format)  # Total for CORRECTED_DIFF
            worksheet.write(total_row, start_col + 6, total_uncorrected_diff, total_format)  # Total for UNCORRECTED_DIFF

            # Update starting row for the next run
            start_row += len(df) + 8  # Update for space between data and charts

        # After processing all runs, create a new section for daily totals
        start_row += 2  # Adding a gap before daily totals
        worksheet.merge_range(start_row, 1, start_row, 3, "Daily Totals", header_format)

        # Write headers for daily totals, starting from column B
        worksheet.write(start_row + 1, 1, 'Date', header_format)  # Column B
        worksheet.write(start_row + 1, 2, 'TOTAL_DIFF_CORRECTED', header_format)  # Column C
        worksheet.write(start_row + 1, 3, 'TOTAL_DIFF_UNCORRECTED', header_format)  # Column D

        # Write daily totals to the new section, starting from the next row
        for row_num, (date, totals) in enumerate(daily_totals.items()):
            worksheet.write(start_row + 2 + row_num, 1, date)  # Date in column B
            worksheet.write(start_row + 2 + row_num, 2, totals[0] if not pd.isna(totals[0]) else 0)  # TOTAL_DIFF_CORRECTED in column C
            worksheet.write(start_row + 2 + row_num, 3, totals[1] if not pd.isna(totals[1]) else 0)  # TOTAL_DIFF_UNCORRECTED in column D

        # Calculate overall totals for both CORRECTED and UNCORRECTED
        overall_total_corrected = sum(total_corrected if not pd.isna(total_corrected) else 0 for total_corrected, _ in daily_totals.values())
        overall_total_uncorrected = sum(total_uncorrected if not pd.isna(total_uncorrected) else 0 for _, total_uncorrected in daily_totals.values())

        # Write the overall totals at the bottom of the daily totals
        overall_total_format = workbook.add_format({'bold': True, 'align': 'center', 'bg_color': '#FF9999', 'border': 1})

        # Write the overall totals at the bottom of the daily totals
        total_row_index = start_row + len(daily_totals) + 3  # Next row after the daily totals
        worksheet.write(total_row_index, 1, "Overall Total", overall_total_format)  # Label in column B with red format
        worksheet.write(total_row_index, 2, overall_total_corrected if not pd.isna(overall_total_corrected) else 0, overall_total_format)  # TOTAL_DIFF_CORRECTED in column C
        worksheet.write(total_row_index, 3, overall_total_uncorrected if not pd.isna(overall_total_uncorrected) else 0, overall_total_format)  # TOTAL_DIFF_UNCORRECTED in column D

        # Adjust column widths
        worksheet.set_column(1, 1, 15)  # Date column width (B)
        worksheet.set_column(2, 2, 20)  # TOTAL_DIFF_CORRECTED column width (C)
        worksheet.set_column(3, 3, 20)  # TOTAL_DIFF_UNCORRECTED column width (D)
        print("Daily Totals:", daily_totals)

    # Seek to the beginning of the BytesIO object
    output.seek(0)

    # Return the Excel file as an attachment
    return send_file(
        output,
        as_attachment=True,
        download_name='data_report_billing.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
@app.route('/data_report_billing_user')
def data_report_billing_user():
    with open('data_report_user.json', 'r') as json_file:
        combined_data = json.load(json_file)

   
    selected_date = session.get('selected_date_user', 'N/A')
    selected_tag = session.get('selected_tag_user', 'N/A')
    print("selected_date:", selected_date)
    print("selected_tag:", selected_tag)
    output = io.BytesIO()

    # Create an Excel writer object
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet("Billing Data Report")

        # Define formats
        title_format = workbook.add_format({'bold': True, 'font_size': 16, 'align': 'center', 'bg_color': '#D9D9D9'})
        header_format = workbook.add_format({'bold': True, 'align': 'center', 'bg_color': '#D9D9D9', 'border': 1})
        data_format = workbook.add_format({'align': 'center', 'border': 1})
        total_format = workbook.add_format({'bold': True, 'align': 'center', 'bg_color': '#FFCCCB', 'border': 1})

        # Add report header
        worksheet.merge_range('B1:H1', "Billing Data Report", title_format)
        worksheet.merge_range('B2:H2', f"Report Date: {selected_date}", header_format)
        worksheet.merge_range('B3:H3', f"Site: {selected_tag}", header_format)

        start_col = 1  # Start from column B (index 1)
        start_row = 4  # Start from row 5 (index 4)

        # Dictionary to store daily totals for CORRECTED_DIFF and UNCORRECTED_DIFF
        daily_totals = {}

        for i, data_dict in enumerate(combined_data):
            df = pd.DataFrame(data_dict)

            # Reorder columns
            df = df[['DATA_DATE', 'CORRECTED', 'UNCORRECTED', 'Pressure', 'Temperature']]

            # Convert 'CORRECTED' and 'UNCORRECTED' columns to numeric
            df['CORRECTED'] = pd.to_numeric(df['CORRECTED'], errors='coerce')
            df['UNCORRECTED'] = pd.to_numeric(df['UNCORRECTED'], errors='coerce')

            # Forward fill only for the purpose of calculating the differences, but don't modify the original columns
            df_filled = df.fillna(method='ffill')

            # Calculate the differences using the filled data
            df['CORRECTED_DIFF'] = df_filled['CORRECTED'].diff()
            df['UNCORRECTED_DIFF'] = df_filled['UNCORRECTED'].diff()

            # Where the original data had missing values, set the diffs to NaN (or keep them empty)
            df.loc[df['CORRECTED'].isna(), 'CORRECTED_DIFF'] = None
            df.loc[df['UNCORRECTED'].isna(), 'UNCORRECTED_DIFF'] = None

            # เขียนข้อมูลทั้งหมด (รวมถึง NaN) ลงใน Excel sheet
            df.to_excel(writer, sheet_name="Billing Data Report", startrow=start_row + 2, startcol=start_col, index=False, header=False)

            worksheet = writer.sheets["Billing Data Report"]

            # Format header for the run info
            worksheet.merge_range(start_row, start_col, start_row, start_col + len(df.columns) - 1, f'Run {i + 1} (Site: {selected_tag}, Date: {selected_date})', header_format)

            for col_num, value in enumerate(df.columns):
                worksheet.write(start_row + 1, start_col + col_num, value, header_format)

            # Format the data rows
            for row_num in range(start_row + 2, start_row + 2 + len(df)):
                worksheet.set_row(row_num, None, data_format)

            # Calculate totals for CORRECTED_DIFF and UNCORRECTED_DIFF
            total_corrected_diff = df['CORRECTED_DIFF'].sum()
            total_uncorrected_diff = df['UNCORRECTED_DIFF'].sum()

            # Accumulate daily totals
            for date, corrected_value, uncorrected_value in zip(df['DATA_DATE'], df['CORRECTED_DIFF'], df['UNCORRECTED_DIFF']):
                if pd.notna(date):  # Check that date is not NaN
                    if date not in daily_totals:
                        daily_totals[date] = [0, 0]  # [total_corrected, total_uncorrected]
                    daily_totals[date][0] += corrected_value if pd.notna(corrected_value) else 0
                    daily_totals[date][1] += uncorrected_value if pd.notna(uncorrected_value) else 0

            # Write totals to the worksheet
            total_row = start_row + 2 + len(df)  # Row for totals
            worksheet.write(total_row, start_col + 4, 'Total', total_format)
            worksheet.write(total_row, start_col + 5, total_corrected_diff, total_format)  # Total for CORRECTED_DIFF
            worksheet.write(total_row, start_col + 6, total_uncorrected_diff, total_format)  # Total for UNCORRECTED_DIFF

            # Update starting row for the next run
            start_row += len(df) + 8  # Update for space between data and charts

        # After processing all runs, create a new section for daily totals
        start_row += 2  # Adding a gap before daily totals
        worksheet.merge_range(start_row, 1, start_row, 3, "Daily Totals", header_format)

        # Write headers for daily totals, starting from column B
        worksheet.write(start_row + 1, 1, 'Date', header_format)  # Column B
        worksheet.write(start_row + 1, 2, 'TOTAL_DIFF_CORRECTED', header_format)  # Column C
        worksheet.write(start_row + 1, 3, 'TOTAL_DIFF_UNCORRECTED', header_format)  # Column D

        # Write daily totals to the new section, starting from the next row
        for row_num, (date, totals) in enumerate(daily_totals.items()):
            worksheet.write(start_row + 2 + row_num, 1, date)  # Date in column B
            worksheet.write(start_row + 2 + row_num, 2, totals[0] if not pd.isna(totals[0]) else 0)  # TOTAL_DIFF_CORRECTED in column C
            worksheet.write(start_row + 2 + row_num, 3, totals[1] if not pd.isna(totals[1]) else 0)  # TOTAL_DIFF_UNCORRECTED in column D

        # Calculate overall totals for both CORRECTED and UNCORRECTED
        overall_total_corrected = sum(total_corrected if not pd.isna(total_corrected) else 0 for total_corrected, _ in daily_totals.values())
        overall_total_uncorrected = sum(total_uncorrected if not pd.isna(total_uncorrected) else 0 for _, total_uncorrected in daily_totals.values())

        # Write the overall totals at the bottom of the daily totals
        overall_total_format = workbook.add_format({'bold': True, 'align': 'center', 'bg_color': '#FF9999', 'border': 1})

        # Write the overall totals at the bottom of the daily totals
        total_row_index = start_row + len(daily_totals) + 3  # Next row after the daily totals
        worksheet.write(total_row_index, 1, "Overall Total", overall_total_format)  # Label in column B with red format
        worksheet.write(total_row_index, 2, overall_total_corrected if not pd.isna(overall_total_corrected) else 0, overall_total_format)  # TOTAL_DIFF_CORRECTED in column C
        worksheet.write(total_row_index, 3, overall_total_uncorrected if not pd.isna(overall_total_uncorrected) else 0, overall_total_format)  # TOTAL_DIFF_UNCORRECTED in column D

        # Adjust column widths
        worksheet.set_column(1, 1, 15)  # Date column width (B)
        worksheet.set_column(2, 2, 20)  # TOTAL_DIFF_CORRECTED column width (C)
        worksheet.set_column(3, 3, 20)  # TOTAL_DIFF_UNCORRECTED column width (D)
        print("Daily Totals:", daily_totals)

    # Seek to the beginning of the BytesIO object
    output.seek(0)

    # Return the Excel file as an attachment
    return send_file(
        output,
        as_attachment=True,
        download_name='data_report_billing.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )  
    
@app.route('/data_report_billing_user_pdf')
def data_report_billing_user_pdf():
    with open('data_report_user.json', 'r') as json_file:
        combined_data = json.load(json_file)

   
    selected_date = session.get('selected_date_user', 'N/A')
    selected_tag = session.get('selected_tag_user', 'N/A')
    print("selected_date:", selected_date)
    print("selected_tag:", selected_tag)
      # Prepare HTML for the PDF
    html_content = f'''
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                h1, h2, h3 {{ text-align: center; }}
                table {{ width: 80%; border-collapse: collapse; margin: 20px auto; }}
                th, td {{ border: 1px solid #dddddd; text-align: center; padding: 4px; font-size: 12px; }}
                th {{ background-color: #D9D9D9; }}
                .total {{ background-color: #FFCCCB; font-weight: bold; }}
                .overall {{ background-color: #FF9999; font-weight: bold; }}
                @media print {{ .pagebreak {{ page-break-before: always; }} }}
            </style>
        </head>
        <body>
            <h1>Billing Data Report</h1>
            <h2>Report Date: {selected_date} / Site: {selected_tag}</h2>
    '''

    overall_daily_totals = [0, 0]  # Initialize overall totals

    for i, data_dict in enumerate(combined_data):
        df = pd.DataFrame(data_dict)

        # Reorder columns
        df = df[['DATA_DATE', 'CORRECTED', 'UNCORRECTED', 'Pressure', 'Temperature']]

        # Convert to numeric
        df['CORRECTED'] = pd.to_numeric(df['CORRECTED'], errors='coerce')
        df['UNCORRECTED'] = pd.to_numeric(df['UNCORRECTED'], errors='coerce')

        # Forward fill missing values and calculate differences
        df_filled = df.fillna(method='ffill')
        df['CORRECTED_DIFF'] = df_filled['CORRECTED'].diff()
        df['UNCORRECTED_DIFF'] = df_filled['UNCORRECTED'].diff()

        # Handle NaNs by replacing them with None for differences
        df.loc[df['CORRECTED'].isna(), 'CORRECTED_DIFF'] = None
        df.loc[df['UNCORRECTED'].isna(), 'UNCORRECTED_DIFF'] = None

        # Convert CORRECTED and UNCORRECTED to integers, handling NaNs by filling them with 0
        df['CORRECTED'] = df['CORRECTED'].fillna(0).astype(int)
        df['UNCORRECTED'] = df['UNCORRECTED'].fillna(0).astype(int)

        # Write data to HTML
        html_content += f'''
            <h3>Run: {i + 1}</h3>  <!-- Display the run number as a header -->
            <table>
                <tr>
                    <th>Data Date</th>
                    <th>Corrected</th>
                    <th>Uncorrected</th>
                    <th>Pressure</th>
                    <th>Temperature</th>
                    <th>Corrected Diff</th>
                    <th>Uncorrected Diff</th>
                </tr>
        '''

        daily_totals = [0, 0]  # Reset daily totals for this dataset

        for index, row in df.iterrows():
            html_content += f'''
            <tr>
                <td>{row['DATA_DATE']}</td>
                <td>{row['CORRECTED']}</td>
                <td>{row['UNCORRECTED']}</td>
                <td>{row['Pressure']}</td>
                <td>{row['Temperature']}</td>
                <td>{int(row['CORRECTED_DIFF']) if pd.notna(row['CORRECTED_DIFF']) else 0}</td>
                <td>{int(row['UNCORRECTED_DIFF']) if pd.notna(row['UNCORRECTED_DIFF']) else 0}</td>
            </tr>
            '''

            # Accumulate daily totals (as integers)
            if pd.notna(row['CORRECTED_DIFF']):
                daily_totals[0] += int(row['CORRECTED_DIFF'])
            if pd.notna(row['UNCORRECTED_DIFF']):
                daily_totals[1] += int(row['UNCORRECTED_DIFF'])

        # Add totals for this dataset
        html_content += f'''
            <tr class="total">
                <td colspan="5">Total</td>
                <td>{daily_totals[0]}</td>
                <td>{daily_totals[1]}</td>
            </tr>
        '''
        html_content += '</table>'  # End of the current dataset table

        # Accumulate overall totals
        overall_daily_totals[0] += daily_totals[0]
        overall_daily_totals[1] += daily_totals[1]

    # Overall totals table at the end
    html_content += '''
        <h3>Overall Totals</h3>
        <table>
            <tr>
                <th>Total Corrected Diff</th>
                <th>Total Uncorrected Diff</th>
            </tr>
            <tr class="overall">
                <td>{}</td>
                <td>{}</td>
            </tr>
        </table>
        </body>
        </html>
    '''.format(int(overall_daily_totals[0]), int(overall_daily_totals[1]))

    # Generate PDF
    pdf_file = io.BytesIO()
    pisa.CreatePDF(html_content, dest=pdf_file)
    pdf_file.seek(0)

    return send_file(
        pdf_file,
        as_attachment=True,
        download_name='data_report_billing.pdf',
        mimetype='application/pdf'
    )
    

@app.route('/data_report_config_user_group')
def data_report_config_user_group():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", ptt_pivot_connection)
        data_config = session.get('data_config', [])
        
        selected_date = session['selected_date_user_group']
        print("selected_date:", selected_date)
        selected_tag = session['selected_tag_user_group']
        print("selected_tag:", selected_tag)
        selected_type = session['vc_type_user_group']

        # Load JSON data
        with open('data_report_user_group.json', 'r') as json_file:
            data_config_html = json.load(json_file)
            print(data_config_html)
        
        # Create a Pandas DataFrame from the HTML
        tables = pd.read_html(''.join(data_config_html))

        # Create a BytesIO buffer to hold the Excel file
        output = BytesIO()

        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            # Access the XlsxWriter workbook and sheet objects
            workbook = writer.book
            worksheet = workbook.add_worksheet('Data Report')

            # Formatting options
            header_format = workbook.add_format({'bold': True, 'align': 'center', 'bg_color': '#D9D9D9', 'border': 1})
            data_format = workbook.add_format({'border': 1})

            # Report Title and Metadata
            report_title = "Billing Data Report"
            report_date = f"Report Date: {selected_date}"
            report_site = f"Site: {selected_tag}"

            worksheet.merge_range('C1:J1', report_title, header_format)
            worksheet.merge_range('C2:J2', report_date, header_format)
            worksheet.merge_range('C3:J3', report_site, header_format)

            row_offset = 4  # Starting from row 5 (index 4)

            for i, df in enumerate(tables):
                # Use the first row as the new header, and drop the first row entirely
                df.columns = df.iloc[0]  # Set the first row as column headers
                df = df[1:].reset_index(drop=True)  # Remove the first row and reset the index
                
                # Drop the first column (column with index 0)
                df = df.drop(columns=[df.columns[0]])

                # Replace inf, None with empty strings
                df.replace([float('inf'), float('-inf'), None], '', inplace=True)

                vc_name = selected_type.get(f'vc_name_list{i + 1}', 'N/A')  # Default to 'N/A' if not found
                worksheet.merge_range(row_offset, 0, row_offset, len(df.columns) - 1,
                                      f'Run {i + 1} (Site: {selected_tag}, Date: {selected_date}, Type: {vc_name})', header_format)
                row_offset += 1

                # Write the headers without gaps
                for col_num, value in enumerate(df.columns):
                    worksheet.write(row_offset, col_num, value, header_format)

                # Write the rest of the DataFrame
                for row_num, row_data in df.iterrows():
                    for col_num, cell_data in enumerate(row_data):
                        worksheet.write(row_offset + row_num + 1, col_num, cell_data, data_format)

                # Adjust column widths based on content
                for col_num, value in enumerate(df.columns):
                    max_length = max(df[value].astype(str).map(len).max(), len(value))
                    max_column_width = min(max_length + 1, 20)  # Max width of 20 characters for smaller size
                    worksheet.set_column(col_num, col_num, max_column_width)

                # Optional: Set the font size for all cells
                cell_format = workbook.add_format({'border': 1, 'font_size': 9})  # Set font size to 9

                # Write the rest of the DataFrame with the new cell format
                for row_num, row_data in df.iterrows():
                    for col_num, cell_data in enumerate(row_data):
                        worksheet.write(row_offset + row_num + 1, col_num, cell_data, cell_format)

                row_offset += len(df) + 2  # Add space after each table

        output.seek(0)

        # Send the file as an attachment
        return send_file(
            output,
            as_attachment=True,
            download_name='data_report_config.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        
        
           

@app.route('/data_report_billing_user_group_pdf')
def data_report_billing_user_group_pdf():
    with open('data_report_user_group.json', 'r') as json_file:
        combined_data = json.load(json_file)

   
    selected_date = session.get('selected_date_user_group', 'N/A')
    selected_tag = session.get('selected_tag_user_group', 'N/A')
    print("selected_date:", selected_date)
    print("selected_tag:", selected_tag)
    
      # Prepare HTML for the PDF
    html_content = f'''
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                h1, h2, h3 {{ text-align: center; }}
                table {{ width: 80%; border-collapse: collapse; margin: 20px auto; }}
                th, td {{ border: 1px solid #dddddd; text-align: center; padding: 4px; font-size: 12px; }}
                th {{ background-color: #D9D9D9; }}
                .total {{ background-color: #FFCCCB; font-weight: bold; }}
                .overall {{ background-color: #FF9999; font-weight: bold; }}
                @media print {{ .pagebreak {{ page-break-before: always; }} }}
            </style>
        </head>
        <body>
            <h1>Billing Data Report</h1>
            <h2>Report Date: {selected_date} / Site: {selected_tag}</h2>
    '''

    overall_daily_totals = [0, 0]  # Initialize overall totals

    for i, data_dict in enumerate(combined_data):
        df = pd.DataFrame(data_dict)

        # Reorder columns
        df = df[['DATA_DATE', 'CORRECTED', 'UNCORRECTED', 'Pressure', 'Temperature']]

        # Convert to numeric
        df['CORRECTED'] = pd.to_numeric(df['CORRECTED'], errors='coerce')
        df['UNCORRECTED'] = pd.to_numeric(df['UNCORRECTED'], errors='coerce')

        # Forward fill missing values and calculate differences
        df_filled = df.fillna(method='ffill')
        df['CORRECTED_DIFF'] = df_filled['CORRECTED'].diff()
        df['UNCORRECTED_DIFF'] = df_filled['UNCORRECTED'].diff()

        # Handle NaNs by replacing them with None for differences
        df.loc[df['CORRECTED'].isna(), 'CORRECTED_DIFF'] = None
        df.loc[df['UNCORRECTED'].isna(), 'UNCORRECTED_DIFF'] = None

        # Convert CORRECTED and UNCORRECTED to integers, handling NaNs by filling them with 0
        df['CORRECTED'] = df['CORRECTED'].fillna(0).astype(int)
        df['UNCORRECTED'] = df['UNCORRECTED'].fillna(0).astype(int)

        # Write data to HTML
        html_content += f'''
            <h3>Run: {i + 1}</h3>  <!-- Display the run number as a header -->
            <table>
                <tr>
                    <th>Data Date</th>
                    <th>Corrected</th>
                    <th>Uncorrected</th>
                    <th>Pressure</th>
                    <th>Temperature</th>
                    <th>Corrected Diff</th>
                    <th>Uncorrected Diff</th>
                </tr>
        '''

        daily_totals = [0, 0]  # Reset daily totals for this dataset

        for index, row in df.iterrows():
            html_content += f'''
            <tr>
                <td>{row['DATA_DATE']}</td>
                <td>{row['CORRECTED']}</td>
                <td>{row['UNCORRECTED']}</td>
                <td>{row['Pressure']}</td>
                <td>{row['Temperature']}</td>
                <td>{int(row['CORRECTED_DIFF']) if pd.notna(row['CORRECTED_DIFF']) else 0}</td>
                <td>{int(row['UNCORRECTED_DIFF']) if pd.notna(row['UNCORRECTED_DIFF']) else 0}</td>
            </tr>
            '''

            # Accumulate daily totals (as integers)
            if pd.notna(row['CORRECTED_DIFF']):
                daily_totals[0] += int(row['CORRECTED_DIFF'])
            if pd.notna(row['UNCORRECTED_DIFF']):
                daily_totals[1] += int(row['UNCORRECTED_DIFF'])

        # Add totals for this dataset
        html_content += f'''
            <tr class="total">
                <td colspan="5">Total</td>
                <td>{daily_totals[0]}</td>
                <td>{daily_totals[1]}</td>
            </tr>
        '''
        html_content += '</table>'  # End of the current dataset table

        # Accumulate overall totals
        overall_daily_totals[0] += daily_totals[0]
        overall_daily_totals[1] += daily_totals[1]

    # Overall totals table at the end
    html_content += '''
        <h3>Overall Totals</h3>
        <table>
            <tr>
                <th>Total Corrected Diff</th>
                <th>Total Uncorrected Diff</th>
            </tr>
            <tr class="overall">
                <td>{}</td>
                <td>{}</td>
            </tr>
        </table>
        </body>
        </html>
    '''.format(int(overall_daily_totals[0]), int(overall_daily_totals[1]))

    # Generate PDF
    pdf_file = io.BytesIO()
    pisa.CreatePDF(html_content, dest=pdf_file)
    pdf_file.seek(0)

    return send_file(
        pdf_file,
        as_attachment=True,
        download_name='data_report_billing.pdf',
        mimetype='application/pdf'
    )

@app.route('/data_report_config_user_group_pdf')
def data_report_config_user_group_pdf():
    # เชื่อมต่อฐานข้อมูล
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        data_config = session.get('data_config', [])
        selected_date = session['selected_date_user_group']
        selected_tag = session['selected_tag_user_group']
        selected_type = session['vc_type_user_group']
        # โหลด JSON data
        with open('data_report_user_group.json', 'r') as json_file:
            data_config_html = json.load(json_file)

        # สร้าง HTML content สำหรับ PDF
        html_content = f'''
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; font-size: 12px; }}
                h1 {{ font-size: 24px; text-align: center; margin-bottom: 10px; }}
                h2 {{ font-size: 15px; text-align: center; margin-bottom: 5px; }}
                h3 {{ font-size: 15px; text-align: center; margin-bottom: 5px; }}
                h4 {{ font-size: 10px; margin-top: 20px; }}
                table {{
                    width: auto; 
                    border-collapse: collapse; 
                    margin: 20px auto; 
                    table-layout: auto; 
                }}
                th, td {{
                    border: 1px solid #dddddd; 
                    text-align: center; 
                    padding: 4px; 
                    font-size: 7px; 
                    min-width: 60px; 
                    max-width: 300px; 
                    overflow: hidden; 
                    text-overflow: ellipsis; 
                    white-space: normal; 
                    word-wrap: break-word; 
                }}
                th {{ 
                    background-color: #D9D9D9; 
                    font-size: 6px; 
                }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                @media print {{ .pagebreak {{ page-break-before: always; }} }}
            </style>
        </head>
        <body>
            <h1>Billing Data Report</h1>
            <h2>Report Date: {selected_date}</h2>
            <h3>Site: {selected_tag}</h3>
        '''

        tables = pd.read_html(''.join(data_config_html))
        for i, df in enumerate(tables):
            # Use the first row as header
            df.columns = df.iloc[0]  # Set column names
            df = df[1:].reset_index(drop=True)  # Remove the first row and reset index
            
            # Remove the first column
            df = df.iloc[:, 1:]  # Drop the first column

            # Replace infinite values and None with np.nan
            df.replace([float('inf'), float('-inf'), None], np.nan, inplace=True)
            
            # Replace NaN with "N/A"
            df.fillna("N/A", inplace=True)

            # Add title for each table
            vc_name_key = f'vc_name_list{i + 1}'
            vc_name_value = selected_type.get(vc_name_key, 'N/A')  # Get corresponding VC name or default to "N/A"
            html_content += f'<h4>Run {i + 1} (Site: {selected_tag}, Date: {selected_date}, Type: {vc_name_value})</h4>'
            html_content += '<table><tr>'
            
            # Create header
            for col in df.columns:
                html_content += f'<th>{col}</th>'
            html_content += '</tr>'

            # Add data to the table
            for _, row in df.iterrows():
                html_content += '<tr>'
                for cell in row:
                    html_content += f'<td>{cell}</td>'
                html_content += '</tr>'
            html_content += '</table><br/>'
            
            # Add page break after each table
            html_content += '<div class="pagebreak"></div>'

        # Create PDF
        pdf_file = BytesIO()
        pisa.CreatePDF(html_content, dest=pdf_file)
        pdf_file.seek(0)

        return send_file(
            pdf_file,
            as_attachment=True,
            download_name='data_report_config.pdf',
            mimetype='application/pdf'
        )

@app.route('/data_report_billing')
def data_report_billing():
    with open('combined_data_report.json', 'r') as json_file:
        combined_data = json.load(json_file)

    selected_date = session.get('selected_date', 'N/A')
    selected_tag = session.get('selected_tag', 'N/A')
    
    
    print("selected_date:", selected_date)
    print("selected_tag:", selected_tag)

    # Prepare an output stream for the Excel file
    output = io.BytesIO()

    # Create an Excel writer object
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet("Billing Data Report")

        # Define formats
        title_format = workbook.add_format({'bold': True, 'font_size': 16, 'align': 'center', 'bg_color': '#D9D9D9'})
        header_format = workbook.add_format({'bold': True, 'align': 'center', 'bg_color': '#D9D9D9', 'border': 1})
        data_format = workbook.add_format({'align': 'center', 'border': 1})
        total_format = workbook.add_format({'bold': True, 'align': 'center', 'bg_color': '#FFCCCB', 'border': 1})

        # Add report header
        worksheet.merge_range('B1:H1', "Billing Data Report", title_format)
        worksheet.merge_range('B2:H2', f"Report Date: {selected_date}", header_format)
        worksheet.merge_range('B3:H3', f"Site: {selected_tag}", header_format)

        start_col = 1  # Start from column B (index 1)
        start_row = 4  # Start from row 5 (index 4)

        # Dictionary to store daily totals for CORRECTED_DIFF and UNCORRECTED_DIFF
        daily_totals = {}

        for i, data_dict in enumerate(combined_data):
            df = pd.DataFrame(data_dict)

            # Reorder columns
            df = df[['DATA_DATE', 'CORRECTED', 'UNCORRECTED', 'Pressure', 'Temperature']]

            # Convert 'CORRECTED' and 'UNCORRECTED' columns to numeric
            df['CORRECTED'] = pd.to_numeric(df['CORRECTED'], errors='coerce')
            df['UNCORRECTED'] = pd.to_numeric(df['UNCORRECTED'], errors='coerce')

            # Forward fill only for the purpose of calculating the differences, but don't modify the original columns
            df_filled = df.fillna(method='ffill')

            # Calculate the differences using the filled data
            df['CORRECTED_DIFF'] = df_filled['CORRECTED'].diff()
            df['UNCORRECTED_DIFF'] = df_filled['UNCORRECTED'].diff()

            # Where the original data had missing values, set the diffs to NaN (or keep them empty)
            df.loc[df['CORRECTED'].isna(), 'CORRECTED_DIFF'] = None
            df.loc[df['UNCORRECTED'].isna(), 'UNCORRECTED_DIFF'] = None

            # เขียนข้อมูลทั้งหมด (รวมถึง NaN) ลงใน Excel sheet
            df.to_excel(writer, sheet_name="Billing Data Report", startrow=start_row + 2, startcol=start_col, index=False, header=False)

            worksheet = writer.sheets["Billing Data Report"]

            # Format header for the run info
            worksheet.merge_range(start_row, start_col, start_row, start_col + len(df.columns) - 1, f'Run {i + 1} (Site: {selected_tag}, Date: {selected_date})', header_format)

            for col_num, value in enumerate(df.columns):
                worksheet.write(start_row + 1, start_col + col_num, value, header_format)

            # Format the data rows
            for row_num in range(start_row + 2, start_row + 2 + len(df)):
                worksheet.set_row(row_num, None, data_format)

            # Calculate totals for CORRECTED_DIFF and UNCORRECTED_DIFF
            total_corrected_diff = df['CORRECTED_DIFF'].sum()
            total_uncorrected_diff = df['UNCORRECTED_DIFF'].sum()

            # Accumulate daily totals
            for date, corrected_value, uncorrected_value in zip(df['DATA_DATE'], df['CORRECTED_DIFF'], df['UNCORRECTED_DIFF']):
                if pd.notna(date):  # Check that date is not NaN
                    if date not in daily_totals:
                        daily_totals[date] = [0, 0]  # [total_corrected, total_uncorrected]
                    daily_totals[date][0] += corrected_value if pd.notna(corrected_value) else 0
                    daily_totals[date][1] += uncorrected_value if pd.notna(uncorrected_value) else 0

            # Write totals to the worksheet
            total_row = start_row + 2 + len(df)  # Row for totals
            worksheet.write(total_row, start_col + 4, 'Total', total_format)
            worksheet.write(total_row, start_col + 5, total_corrected_diff, total_format)  # Total for CORRECTED_DIFF
            worksheet.write(total_row, start_col + 6, total_uncorrected_diff, total_format)  # Total for UNCORRECTED_DIFF

            # Update starting row for the next run
            start_row += len(df) + 8  # Update for space between data and charts

        # After processing all runs, create a new section for daily totals
        start_row += 2  # Adding a gap before daily totals
        worksheet.merge_range(start_row, 1, start_row, 3, "Daily Totals", header_format)

        # Write headers for daily totals, starting from column B
        worksheet.write(start_row + 1, 1, 'Date', header_format)  # Column B
        worksheet.write(start_row + 1, 2, 'TOTAL_DIFF_CORRECTED', header_format)  # Column C
        worksheet.write(start_row + 1, 3, 'TOTAL_DIFF_UNCORRECTED', header_format)  # Column D

        # Write daily totals to the new section, starting from the next row
        for row_num, (date, totals) in enumerate(daily_totals.items()):
            worksheet.write(start_row + 2 + row_num, 1, date)  # Date in column B
            worksheet.write(start_row + 2 + row_num, 2, totals[0] if not pd.isna(totals[0]) else 0)  # TOTAL_DIFF_CORRECTED in column C
            worksheet.write(start_row + 2 + row_num, 3, totals[1] if not pd.isna(totals[1]) else 0)  # TOTAL_DIFF_UNCORRECTED in column D

        # Calculate overall totals for both CORRECTED and UNCORRECTED
        overall_total_corrected = sum(total_corrected if not pd.isna(total_corrected) else 0 for total_corrected, _ in daily_totals.values())
        overall_total_uncorrected = sum(total_uncorrected if not pd.isna(total_uncorrected) else 0 for _, total_uncorrected in daily_totals.values())

        # Write the overall totals at the bottom of the daily totals
        overall_total_format = workbook.add_format({'bold': True, 'align': 'center', 'bg_color': '#FF9999', 'border': 1})

        # Write the overall totals at the bottom of the daily totals
        total_row_index = start_row + len(daily_totals) + 3  # Next row after the daily totals
        worksheet.write(total_row_index, 1, "Overall Total", overall_total_format)  # Label in column B with red format
        worksheet.write(total_row_index, 2, overall_total_corrected if not pd.isna(overall_total_corrected) else 0, overall_total_format)  # TOTAL_DIFF_CORRECTED in column C
        worksheet.write(total_row_index, 3, overall_total_uncorrected if not pd.isna(overall_total_uncorrected) else 0, overall_total_format)  # TOTAL_DIFF_UNCORRECTED in column D

        # Adjust column widths
        worksheet.set_column(1, 1, 15)  # Date column width (B)
        worksheet.set_column(2, 2, 20)  # TOTAL_DIFF_CORRECTED column width (C)
        worksheet.set_column(3, 3, 20)  # TOTAL_DIFF_UNCORRECTED column width (D)
        print("Daily Totals:", daily_totals)

    # Seek to the beginning of the BytesIO object
    output.seek(0)

    # Return the Excel file as an attachment
    return send_file(
        output,
        as_attachment=True,
        download_name='data_report_billing.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/data_report_billing_pdf')
def data_report_billing_pdf():
    with open('combined_data_report.json', 'r') as json_file:
        combined_data = json.load(json_file)

    selected_date = session.get('selected_date', 'N/A')
    selected_tag = session.get('selected_tag', 'N/A')

    # Prepare HTML for the PDF
    # Prepare HTML for the PDF
    html_content = f'''
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                h1, h2, h3 {{ text-align: center; }}
                table {{ width: 80%; border-collapse: collapse; margin: 20px auto; }}
                th, td {{ border: 1px solid #dddddd; text-align: center; padding: 4px; font-size: 12px; }}
                th {{ background-color: #D9D9D9; }}
                .total {{ background-color: #FFCCCB; font-weight: bold; }}
                .overall {{ background-color: #FF9999; font-weight: bold; }}
                @media print {{ .pagebreak {{ page-break-before: always; }} }}
            </style>
        </head>
        <body>
            <h1>Billing Data Report</h1>
            <h2>Report Date: {selected_date} / Site: {selected_tag}</h2>
    '''

    overall_daily_totals = [0, 0]  # Initialize overall totals

    for i, data_dict in enumerate(combined_data):
        df = pd.DataFrame(data_dict)

        # Reorder columns
        df = df[['DATA_DATE', 'CORRECTED', 'UNCORRECTED', 'Pressure', 'Temperature']]

        # Convert to numeric
        df['CORRECTED'] = pd.to_numeric(df['CORRECTED'], errors='coerce')
        df['UNCORRECTED'] = pd.to_numeric(df['UNCORRECTED'], errors='coerce')

        # Forward fill missing values and calculate differences
        df_filled = df.fillna(method='ffill')
        df['CORRECTED_DIFF'] = df_filled['CORRECTED'].diff()
        df['UNCORRECTED_DIFF'] = df_filled['UNCORRECTED'].diff()

        # Handle NaNs by replacing them with None for differences
        df.loc[df['CORRECTED'].isna(), 'CORRECTED_DIFF'] = None
        df.loc[df['UNCORRECTED'].isna(), 'UNCORRECTED_DIFF'] = None

        # Convert CORRECTED and UNCORRECTED to integers, handling NaNs by filling them with 0
        df['CORRECTED'] = df['CORRECTED'].fillna(0).astype(int)
        df['UNCORRECTED'] = df['UNCORRECTED'].fillna(0).astype(int)

        # Write data to HTML
        html_content += f'''
            <h3>Run: {i + 1}</h3>  <!-- Display the run number as a header -->
            <table>
                <tr>
                    <th>Data Date</th>
                    <th>Corrected</th>
                    <th>Uncorrected</th>
                    <th>Pressure</th>
                    <th>Temperature</th>
                    <th>Corrected Diff</th>
                    <th>Uncorrected Diff</th>
                </tr>
        '''

        daily_totals = [0, 0]  # Reset daily totals for this dataset

        for index, row in df.iterrows():
            html_content += f'''
            <tr>
                <td>{row['DATA_DATE']}</td>
                <td>{row['CORRECTED']}</td>
                <td>{row['UNCORRECTED']}</td>
                <td>{row['Pressure']}</td>
                <td>{row['Temperature']}</td>
                <td>{int(row['CORRECTED_DIFF']) if pd.notna(row['CORRECTED_DIFF']) else 0}</td>
                <td>{int(row['UNCORRECTED_DIFF']) if pd.notna(row['UNCORRECTED_DIFF']) else 0}</td>
            </tr>
            '''

            # Accumulate daily totals (as integers)
            if pd.notna(row['CORRECTED_DIFF']):
                daily_totals[0] += int(row['CORRECTED_DIFF'])
            if pd.notna(row['UNCORRECTED_DIFF']):
                daily_totals[1] += int(row['UNCORRECTED_DIFF'])

        # Add totals for this dataset
        html_content += f'''
            <tr class="total">
                <td colspan="5">Total</td>
                <td>{daily_totals[0]}</td>
                <td>{daily_totals[1]}</td>
            </tr>
        '''
        html_content += '</table>'  # End of the current dataset table

        # Accumulate overall totals
        overall_daily_totals[0] += daily_totals[0]
        overall_daily_totals[1] += daily_totals[1]

    # Overall totals table at the end
    html_content += '''
        <h3>Overall Totals</h3>
        <table>
            <tr>
                <th>Total Corrected Diff</th>
                <th>Total Uncorrected Diff</th>
            </tr>
            <tr class="overall">
                <td>{}</td>
                <td>{}</td>
            </tr>
        </table>
        </body>
        </html>
    '''.format(int(overall_daily_totals[0]), int(overall_daily_totals[1]))

    # Generate PDF
    pdf_file = io.BytesIO()
    pisa.CreatePDF(html_content, dest=pdf_file)
    pdf_file.seek(0)

    return send_file(
        pdf_file,
        as_attachment=True,
        download_name='data_report_billing.pdf',
        mimetype='application/pdf'
    )


@app.route('/data_report_config')
def data_report_config():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", ptt_pivot_connection)
        data_config = session.get('data_config', [])
        
        selected_date = session['selected_date']
        print("selected_date:", selected_date)
        selected_tag = session['selected_tag']
        print("selected_tag:", selected_tag)
        selected_type = session['vc_type']
        print("selected_type:", selected_type)  # Log selected_type for debugging

        # Load JSON data
        with open('combined_data_report.json', 'r') as json_file:
            data_config_html = json.load(json_file)
            print(data_config_html)
        
        # Create a Pandas DataFrame from the HTML
        tables = pd.read_html(''.join(data_config_html))

        # Create a BytesIO buffer to hold the Excel file
        output = BytesIO()

        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            # Access the XlsxWriter workbook and sheet objects
            workbook = writer.book
            worksheet = workbook.add_worksheet('Data Report')

            # Formatting options
            header_format = workbook.add_format({'bold': True, 'align': 'center', 'bg_color': '#D9D9D9', 'border': 1})
            data_format = workbook.add_format({'border': 1})

            # Report Title and Metadata
            report_title = "Billing Data Report"
            report_date = f"Report Date: {selected_date}"
            report_site = f"Site: {selected_tag}"

            worksheet.merge_range('C1:L1', report_title, header_format)
            worksheet.merge_range('C2:L2', report_date, header_format)
            worksheet.merge_range('C3:L3', report_site, header_format)

            row_offset = 4  # Starting from row 5 (index 4)

            for i, df in enumerate(tables):
                # Use the first row as the new header, and drop the first row entirely
                df.columns = df.iloc[0]  # Set the first row as column headers
                df = df[1:].reset_index(drop=True)  # Remove the first row and reset the index
                
                # Drop the first column (column with index 0)
                df = df.drop(columns=[df.columns[0]])

                # Replace inf, None with empty strings
                df.replace([float('inf'), float('-inf'), None], '', inplace=True)

                # Get the corresponding vc_name for the current run
                vc_name = selected_type.get(f'vc_name_list{i + 1}', 'N/A')  # Default to 'N/A' if not found
                worksheet.merge_range(row_offset, 0, row_offset, len(df.columns) - 1,
                                      f'Run {i + 1} (Site: {selected_tag}, Date: {selected_date}, Type: {vc_name})', header_format)
                row_offset += 1

                # Write the headers without gaps
                for col_num, value in enumerate(df.columns):
                    worksheet.write(row_offset, col_num, value, header_format)

                # Write the rest of the DataFrame
                for row_num, row_data in df.iterrows():
                    for col_num, cell_data in enumerate(row_data):
                        worksheet.write(row_offset + row_num + 1, col_num, cell_data, data_format)

                # Adjust column widths based on content
                for col_num, value in enumerate(df.columns):
                    max_length = max(df[value].astype(str).map(len).max(), len(value))
                    max_column_width = min(max_length + 1, 20)  # Max width of 20 characters for smaller size
                    worksheet.set_column(col_num, col_num, max_column_width)

                # Optional: Set the font size for all cells
                cell_format = workbook.add_format({'border': 1, 'font_size': 9})  # Set font size to 9

                # Write the rest of the DataFrame with the new cell format
                for row_num, row_data in df.iterrows():
                    for col_num, cell_data in enumerate(row_data):
                        worksheet.write(row_offset + row_num + 1, col_num, cell_data, cell_format)

                row_offset += len(df) + 2  # Add space after each table

        output.seek(0)

        # Send the file as an attachment
        return send_file(
            output,
            as_attachment=True,
            download_name='data_report_config.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )


@app.route('/data_report_config_pdf')
def data_report_config_pdf():
    # Connect to the database
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        data_config = session.get('data_config', [])
        selected_date = session['selected_date']
        selected_tag = session['selected_tag']
        selected_type = session['vc_type']
        
        # Load JSON data
        with open('combined_data_report.json', 'r') as json_file:
            data_config_html = json.load(json_file)

        # Create HTML content for the PDF
        html_content = f'''
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; font-size: 12px; }}
                h1 {{ font-size: 24px; text-align: center; margin-bottom: 10px; }}
                h2 {{ font-size: 15px; text-align: center; margin-bottom: 5px; }}
                h3 {{ font-size: 15px; text-align: center; margin-bottom: 5px; }}
                h4 {{ font-size: 10px; margin-top: 20px; }}
                table {{
                    width: auto; 
                    border-collapse: collapse; 
                    margin: 20px auto; 
                    table-layout: auto; 
                }}
                th, td {{
                    border: 1px solid #dddddd; 
                    text-align: center; 
                    padding: 4px; 
                    font-size: 7px; 
                    min-width: 60px; 
                    max-width: 300px; 
                    overflow: hidden; 
                    text-overflow: ellipsis; 
                    white-space: normal; 
                    word-wrap: break-word; 
                }}
                th {{ 
                    background-color: #D9D9D9; 
                    font-size: 6px; 
                }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                @media print {{ .pagebreak {{ page-break-before: always; }} }}
            </style>
        </head>
        <body>
            <h1>Billing Data Report</h1>
            <h2>Report Date: {selected_date}</h2>
            <h3>Site: {selected_tag}</h3>
        '''

        tables = pd.read_html(''.join(data_config_html))
        for i, df in enumerate(tables):
            # Use the first row as header
            df.columns = df.iloc[0]  # Set column names
            df = df[1:].reset_index(drop=True)  # Remove the first row and reset index
            
            # Remove the first column
            df = df.iloc[:, 1:]  # Drop the first column

            # Replace infinite values and None with np.nan
            df.replace([float('inf'), float('-inf'), None], np.nan, inplace=True)
            
            # Replace NaN with "N/A"
            df.fillna("N/A", inplace=True)

            # Add title for each table
            vc_name_key = f'vc_name_list{i + 1}'
            vc_name_value = selected_type.get(vc_name_key, 'N/A')  # Get corresponding VC name or default to "N/A"
            html_content += f'<h4>Run {i + 1} (Site: {selected_tag}, Date: {selected_date}, Type: {vc_name_value})</h4>'
            html_content += '<table><tr>'
            
            # Create header
            for col in df.columns:
                html_content += f'<th>{col}</th>'
            html_content += '</tr>'

            # Add data to the table
            for _, row in df.iterrows():
                html_content += '<tr>'
                for cell in row:
                    html_content += f'<td>{cell}</td>'
                html_content += '</tr>'
            html_content += '</table><br/>'
            
            # Add page break after each table
            html_content += '<div class="pagebreak"></div>'

        # Create PDF
        pdf_file = BytesIO()
        pisa.CreatePDF(html_content, dest=pdf_file)
        pdf_file.seek(0)

        return send_file(
            pdf_file,
            as_attachment=True,
            download_name='data_report_config.pdf',
            mimetype='application/pdf'
        )



@app.route('/data_report_hourly')
def data_report_hourly():
    with open('combined_data_hourly_report.json', 'r') as json_file:
        data_hourly = json.load(json_file)
    
    selected_date = session.get('selected_date', 'N/A')  # Use 'N/A' as a default if session value is not found
    selected_tag = session.get('selected_tag', 'N/A')    # Use 'N/A' as a default if session value is not found
    
    print("selected_date:", selected_date)
    print("selected_tag:", selected_tag) 
    
    # Prepare an output stream for the Excel file
    output = io.BytesIO()

    # Create an Excel writer object
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet("Hourly Data Report")

        # Define formats
        header_format = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'bg_color': '#D9D9D9', 'border': 1})
        table_header_format = workbook.add_format({'bold': True, 'bg_color': '#D9D9D9', 'border': 1})
        data_format = workbook.add_format({'border': 1})

        # Add report header
        report_title = "Hourly Data Report"
        report_date = f"Report Date: {selected_date} "
        report_site = f"Site: {selected_tag}"
        worksheet.merge_range('F1:O1', report_title, header_format)
        worksheet.merge_range('F2:O2', report_date, header_format)
        worksheet.merge_range('F3:O3', report_site, header_format)

        start_col = 1  # Start from column B (index 1)
        start_row = 4  # Start from row 5 (index 4)

        for i, data_dict in enumerate(data_hourly):
            # Convert dictionary to DataFrame
            df = pd.DataFrame(data_dict)
            print("df1",df)
            # Ensure DataFrame is not sorted and retains the original order
            df = df[['Date_or_Hour', 'CORRECTED', 'UNCORRECTED', 'Pressure', 'Temperature']]

            # Write DataFrame to the worksheet
            df.to_excel(writer, sheet_name="Hourly Data Report", startrow=start_row + 1, startcol=start_col, index=False, header=False)

            # Access the worksheet
            worksheet = writer.sheets["Hourly Data Report"]

            # Format header row
            worksheet.merge_range(start_row, start_col, start_row, start_col + len(df.columns) - 1, f'Run {i + 1} (Site: {selected_tag}, Date: {selected_date})', header_format)
            
            for col_num, value in enumerate(df.columns):
                worksheet.write(start_row + 1, start_col + col_num, value, table_header_format)

            # Format the data rows
            for row_num in range(start_row + 2, start_row + 2 + len(df)):
                worksheet.set_row(row_num, None, data_format)

            # Add a table to the worksheet for each DataFrame
            worksheet.add_table(start_row + 1, start_col, start_row + len(df), start_col + len(df.columns) - 1, 
                     {'name': f'Table{i+1}', 
                      'header_row': True, 
                      'columns': [{'header': col} for col in df.columns],
                      'banded_rows': False})  # Set to False for no banding

            # Update the starting column for the next DataFrame
            start_col += len(df.columns) + 3  # Adjust spacing between tables

    # Seek to the beginning of the BytesIO object
    output.seek(0)

    # Return the Excel file as an attachment
    return send_file(
        output,
        as_attachment=True,
        download_name='data_report_hourly.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/data_report_hourly_pdf')
def data_report_hourly_pdf():
    with open('combined_data_hourly_report.json', 'r') as json_file:
        data_hourly = json.load(json_file)

    selected_date = session.get('selected_date', 'N/A')
    selected_tag = session.get('selected_tag', 'N/A')

    # Prepare HTML for the PDF
    html_content = f'''
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                h1, h2, h3 {{ text-align: center; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px auto; }}
                th, td {{ border: 1px solid #dddddd; text-align: center; padding: 4px; }}
                th {{ background-color: #D9D9D9; }}
                .page-break {{ page-break-after: always; }}
            </style>
        </head>
        <body>
            <h1>Hourly Data Report</h1>
            <h2>Report Date: {selected_date}</h2>
            <h3>Site: {selected_tag}</h3>
    '''

    for i, data_dict in enumerate(data_hourly):
        df = pd.DataFrame(data_dict)

        # Ensure DataFrame has the correct order
        df = df[['Date_or_Hour', 'CORRECTED', 'UNCORRECTED', 'Pressure', 'Temperature']]

        # Write data to HTML
        html_content += f'<div class="page-break"><h3>Run {i + 1} (Site: {selected_tag}, Date: {selected_date})</h3>'
        html_content += '<table><tr><th>Data Date</th><th>Corrected</th><th>Uncorrected</th><th>Pressure</th><th>Temperature</th></tr>'

        for index, row in df.iterrows():
            html_content += f'''
                <tr>
                    <td>{row['Date_or_Hour']}</td>
                   
                    <td>{row['CORRECTED']}</td>
                    <td>{row['UNCORRECTED']}</td>
                    <td>{row['Pressure']}</td>
                    <td>{row['Temperature']}</td>
                </tr>
            '''

        html_content += '</table></div>'  # Close div for page break

    html_content += '</body></html>'

    # Generate PDF
    pdf_file = io.BytesIO()
    pisa.CreatePDF(html_content, dest=pdf_file)
    pdf_file.seek(0)

    return send_file(
        pdf_file,
        as_attachment=True,
        download_name='data_report_hourly.pdf',
        mimetype='application/pdf'
    )


@app.route('/data')
def get_data_data():
    connection = connect_to_ptt_pivot_db()
    if connection is None:
        return jsonify({"error": "Failed to connect to database"}), 500
    date_system = request.args.get('date', datetime.datetime.now().strftime('%Y-%m'))
    # print("date_system",date_system)
    query = f"""SELECT TO_CHAR(data_date, 'DD-MM-YY ') AS data_date, ALL_SUCCESS, all_error
            FROM amr_data_autopoll
            WHERE TO_CHAR(data_date, 'YYYY-MM') = '{date_system}'
            ORDER BY data_date ASC"""
    rows = fetch_data(connection, query)
    # print("rows",rows)
    session['rows'] = rows
    # print(session['rows'])
    if not rows:
        data = [{'data_date': '', 'ALL_SUCCESS': None, 'all_error': None}]
    else:
        data = [{'data_date': row[0], 'ALL_SUCCESS': row[1], 'all_error': row[2]} for row in rows]
    
    print("data",data)
    connection.close()
    return jsonify(data)

@app.route('/repeat1')
def get_repeat1():
    connection = connect_to_ptt_pivot_db()
    date_system = request.args.get('date', datetime.datetime.now().strftime('%Y-%m'))
    if connection is None:
        return jsonify({"error": "Failed to connect to database"}), 500

    query = f"""SELECT TO_CHAR(data_date, 'DD-MM-YY ') AS data_date, METER_POLL_REPEAT1, ERROR_REPEAT1
               FROM amr_data_autopoll
               WHERE TO_CHAR(data_date, 'YYYY-MM') = '{date_system}'
               ORDER BY data_date ASC"""
    rows = fetch_data(connection, query)
    if not rows:
        data = [{'data_date': '', 'ALL_SUCCESS': None, 'all_error': None}]
    else:
        data = [{ 'data_date': row[0], 'METER_POLL_REPEAT1': row[1], 'ERROR_REPEAT1': row[2] } for row in rows]
    connection.close()
    return jsonify(data)

@app.route('/repeat2')
def get_repeat2():
    connection = connect_to_ptt_pivot_db()
    date_system = request.args.get('date', datetime.datetime.now().strftime('%Y-%m'))
    if connection is None:
        return jsonify({"error": "Failed to connect to database"}), 500

    query = f"""SELECT TO_CHAR(data_date, 'DD-MM-YY ') AS data_date, METER_POLL_REPEAT2, ERROR_REPEAT2
               FROM amr_data_autopoll
               WHERE TO_CHAR(data_date, 'YYYY-MM') = '{date_system}'
               ORDER BY data_date ASC"""
    rows = fetch_data(connection, query)
    if not rows:
        data = [{'data_date': '', 'ALL_SUCCESS': None, 'all_error': None}]
    else:

        data = [{ 'data_date': row[0], 'METER_POLL_REPEAT1': row[1], 'ERROR_REPEAT1': row[2] } for row in rows]
    connection.close()
    return jsonify(data)

@app.route('/repeat3')
def get_repeat3():
    connection = connect_to_ptt_pivot_db()
    date_system = request.args.get('date', datetime.datetime.now().strftime('%Y-%m'))
    if connection is None:
        return jsonify({"error": "Failed to connect to database"}), 500

    query = f"""SELECT TO_CHAR(data_date, 'DD-MM-YY ') AS data_date, METER_POLL_REPEAT3, ERROR_REPEAT3
               FROM amr_data_autopoll
               WHERE TO_CHAR(data_date, 'YYYY-MM') = '{date_system}'
               ORDER BY data_date ASC"""
    rows = fetch_data(connection, query)
    if not rows:
        data = [{'data_date': '', 'ALL_SUCCESS': None, 'all_error': None}]
    else:

        data = [{ 'data_date': row[0], 'METER_POLL_REPEAT3': row[1], 'ERROR_REPEAT3': row[2] } for row in rows]
    connection.close()
    return jsonify(data)

@app.route('/data_evctype')
@login_required
def data_evctype():
    connection = connect_to_ptt_pivot_db()
    
    if connection is None:
        return jsonify({"error": "Failed to connect to database"}), 500
    
    date_system = datetime.datetime.now().strftime('%Y-%m')
    data_values = [13]
    data_actaris = [5, 8, 9, 10, 12]
    
    all_rows = []
    data_values_lengths = []  # To store lengths for data_values
    data_actaris_lengths = []  # To store lengths for data_actaris
    
    # First for loop (data_values)
    for i in data_values:
        query = f""" SELECT
                    AMR_PL_GROUP.PL_REGION_ID as region,
                    AMR_FIELD_ID.TAG_ID as Sitename,
                    MIN(AMR_FIELD_METER.METER_NO_STREAM) as NoRun,
                    MIN(AMR_FIELD_METER.METER_STREAM_NO) as RunNo,
                    MIN(AMR_FIELD_METER.METER_ID) as METERID,
                    MIN(AMR_VC_TYPE.VC_NAME) as VCtype,
                    MIN(AMR_FIELD_ID.SIM_IP) as IPAddress,
                    MIN(AMR_PORT_INFO.PORT_NO) as port,
                    MIN(AMR_POLL_RANGE.poll_config) as poll_config,
                    MIN(AMR_POLL_RANGE.poll_billing) as poll_billing,
                    MIN(AMR_POLL_RANGE.POLL_CONFIG_ENABLE) as POLL_CONFIG_ENABLE,
                    MIN(AMR_POLL_RANGE.POLL_BILLING_ENABLE) as POLL_BILLING_ENABLE,
                    MIN(AMR_VC_TYPE.id) as evctype
                FROM
                    AMR_POLL_RANGE
                    JOIN AMR_FIELD_ID ON AMR_FIELD_ID.FIELD_ID = AMR_FIELD_ID.FIELD_ID
                    JOIN AMR_USER ON AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP
                    JOIN AMR_FIELD_CUSTOMER ON AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID
                    JOIN AMR_FIELD_METER ON AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID
                    JOIN AMR_PL_GROUP ON AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID
                    JOIN AMR_VC_TYPE ON AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE
                    JOIN AMR_PORT_INFO ON AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID
                WHERE
                    AMR_FIELD_METER.METER_AUTO_ENABLE = 1
                    AND AMR_POLL_RANGE.evc_type = AMR_VC_TYPE.id
                    AND AMR_VC_TYPE.id = {i}
                GROUP BY
                    AMR_PL_GROUP.PL_REGION_ID,
                    AMR_FIELD_ID.TAG_ID
                ORDER BY
                    Sitename ASC, port """
        
        rows = fetch_data(connection, query)
        all_rows.extend(rows)
        data_values_lengths.append(len(rows))  # Store each length in the separate list
    
    # Second for loop (data_actaris)
    for i in data_actaris:
        query = f""" SELECT
                AMR_PL_GROUP.PL_REGION_ID as region,
                AMR_FIELD_ID.TAG_ID as Sitename,
                MIN(AMR_FIELD_METER.METER_NO_STREAM) as NoRun,
                MIN(AMR_FIELD_METER.METER_STREAM_NO) as RunNo,
                MIN(AMR_FIELD_METER.METER_ID) as METERID,
                MIN(AMR_VC_TYPE.VC_NAME) as VCtype,
                MIN(AMR_FIELD_ID.SIM_IP) as IPAddress,
                MIN(AMR_PORT_INFO.PORT_NO) as port,
                MIN(AMR_POLL_RANGE.poll_config) as poll_config,
                MIN(AMR_POLL_RANGE.poll_billing) as poll_billing,
                MIN(AMR_POLL_RANGE.POLL_CONFIG_ENABLE) as POLL_CONFIG_ENABLE,
                MIN(AMR_POLL_RANGE.POLL_BILLING_ENABLE) as POLL_BILLING_ENABLE,
                MIN(AMR_VC_TYPE.id) as evctype
            FROM
                AMR_POLL_RANGE
                JOIN AMR_FIELD_ID ON AMR_FIELD_ID.FIELD_ID = AMR_FIELD_ID.FIELD_ID
                JOIN AMR_USER ON AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP
                JOIN AMR_FIELD_CUSTOMER ON AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID
                JOIN AMR_FIELD_METER ON AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID
                JOIN AMR_PL_GROUP ON AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID
                JOIN AMR_VC_TYPE ON AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE
                JOIN AMR_PORT_INFO ON AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID
            WHERE
                AMR_FIELD_METER.METER_AUTO_ENABLE = 1
                AND AMR_POLL_RANGE.evc_type = AMR_VC_TYPE.id
                AND AMR_VC_TYPE.id = {i}
            GROUP BY
                AMR_PL_GROUP.PL_REGION_ID,
                AMR_FIELD_ID.TAG_ID
            ORDER BY
                Sitename ASC, port """
        
        rows = fetch_data(connection, query)
        all_rows.extend(rows)
        data_actaris_lengths.append(len(rows))  # Store each length in the separate list
    
    connection.close()  # Ensure the database connection is closed
    
    # Combine the total lengths for both loops into a list
    total_lengths = [sum(data_values_lengths), sum(data_actaris_lengths)]
    session['evc_type'] = total_lengths
    #print(session['evc_type'],"test")
    response = {
        "data": all_rows,
        "lengths": total_lengths  # Display the sum as two separate values in a list
    }
    return jsonify(response)

@app.route('/')
@login_required
def home_amr():  
        
        if 'username' in session:
            username = session['username']
            user = users.get(username)  # Using get method to handle KeyError
            if user:
                with connect_to_ptt_pivot_db() as ptt_pivot_connection:
                    data_result = get_autopoll_statistic(ptt_pivot_connection, "", datetime.datetime.now().strftime('%Y-%m-%d'))
                    error_content = data_result['error_content']
                    success_content = data_result['success_content']
                    allsite = data_result['allsite']
                            
                return render_template('home.html', username=username, description=user['description'], user_level=user['user_level'], success_content=success_content, error_content=error_content,allsite=allsite)
            else:
                # Handle the case where the user is not in the users dictionary
                
                return redirect(url_for('login'))
        else:
           
            return redirect(url_for('login'))

@app.route('/home_user_group')
@login_required
def home_user_group():  
        
        if 'username' in session:
            username = session['username']
            user = users.get(username)  # Using get method to handle KeyError
            if user:
                with connect_to_ptt_pivot_db() as ptt_pivot_connection:
              
                    print("Active Connection:", ptt_pivot_connection)
                    print("username:",username)
                    date_system = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%d-%m-%Y')
                    # date_system = '01-07-2024'
                    # date_system = datetime.datetime.now().strftime('%d-%m-%Y')
                    # Adjust the query to use bind variables
                    selected_region = session['username']
                    region_query = f"""
                    SELECT amr_region.REGION_NAME from amr_user,amr_region WHERE amr_user.user_group=amr_region.id AND amr_user.user_name like '{selected_region}'
                    """
                    
                    # Fetch unique region values
                    region_results = fetch_data(ptt_pivot_connection, region_query)
                    df_detail = pd.DataFrame(region_results, columns=['REGION_NAME'])

                    region_options = df_detail['REGION_NAME'].iloc[0]
                    
                    # region_query = """
                    # SELECT ID,REGION_NAME FROM AMR_REGION 
                    # """
                    # region_results = fetch_data(ptt_pivot_connection, region_query)
                    # result_data = pd.DataFrame(region_results, columns=['ID', 'REGION_NAME'])
                    # region_names = result_data['REGION_NAME'].tolist()
                    # print(result_data['REGION_NAME'])
                ##########################################
                    
                    
                    error_data_db = f""" SELECT DISTINCT amr_error.meter_id, amr_error.tag_id
                    FROM amr_region, amr_pl_group, amr_field_id, amr_error 
                    WHERE amr_region.id = amr_pl_group.PL_REGION_ID 
                    AND amr_pl_group.field_id = amr_field_id.field_id
                    AND amr_field_id.meter_id = amr_error.meter_id
                    AND amr_region.region_name like '{region_options}'
                    AND TRUNC(amr_error.DATA_DATE) = TO_DATE('{date_system}', 'DD-MM-YYYY')"""
                    error_data_db_results  =  fetch_data(ptt_pivot_connection, error_data_db)
                    result_error_data_db = pd.DataFrame(error_data_db_results, columns=['meter_id', 'tag_id'])
                    #print(result_error_data_db)
                    
                    error_content = result_error_data_db.shape[0]
                    
                ############################################

                    success_data_db = f"""SELECT DISTINCT  amr_field_id.tag_id,amr_configured_data.meter_id
                    FROM amr_region, amr_pl_group, amr_field_id, amr_configured_data
                    WHERE amr_region.id = amr_pl_group.PL_REGION_ID 
                    AND amr_pl_group.field_id = amr_field_id.field_id
                    AND amr_field_id.meter_id = amr_configured_data.meter_id
                    AND amr_region.region_name like '{region_options}'
                    AND amr_configured_data.data_date = TO_DATE('{date_system}', 'DD-MM-YYYY')"""
                    success_data_db_results  =  fetch_data(ptt_pivot_connection, success_data_db)
                    result_success_data_db = pd.DataFrame(success_data_db_results, columns=['tag_id', 'meter_id'])
                    # print("result_success_data_db",result_success_data_db)
                    
                    
                    df_data_cleaned = result_success_data_db[~result_success_data_db['tag_id'].isin(result_error_data_db['tag_id'])]
                    success_content = df_data_cleaned.shape[0]
                    # print("success_content",success_content)
                    tag_ids_success = sorted(df_data_cleaned['tag_id'].tolist())
                    tag_ids_error = sorted(result_error_data_db['tag_id'].tolist())
                    combined_tags = sorted(tag_ids_success + tag_ids_error)
                    allsite = len(combined_tags)
                #########################################################################
                    
                    
                return render_template('home_user_group.html',allsite=allsite,combined_tags=combined_tags,tag_ids_success=tag_ids_success,tag_ids_error=tag_ids_error,region_options=region_options,error_content=error_content,success_content=success_content,selected_region=selected_region,username=username)
            else:
                # Handle the case where the user is not in the users dictionary
                
                return redirect(url_for('login'))
        else:
        
            return redirect(url_for('login'))
        
@app.route('/home_user')
@login_required
def home_user():  
        if 'username' in session:
            username = session['username']
            user = users.get(username)  # Using get method to handle KeyError
            if user:
                print(user)
                with connect_to_ptt_pivot_db() as ptt_pivot_connection:
                #with connect_to_ptt_pivot_db() as ptt_pivot_connection:
                    print("Active Connection:", active_connection)

                    # Fetch user information from the session
                    logged_in_user = session['username']
                    if logged_in_user not in users:
                        return redirect(url_for('login'))
                
                    
                    user_info = users[logged_in_user]
                    user_level = user_info.get('user_level')
                    print("user_level",user_level)
                    description = user_info.get('description')
                    print("description", description)
                    
                    logged_in_user = logged_in_user
                    print("user:", logged_in_user)

                    query_type = request.args.get("query_type")


                    # SQL query to fetch unique PL_REGION_ID values
                    region_query = """
                    SELECT amr_region.REGION_NAME
                    FROM AMR_REGION,AMR_user,amr_field_id,amr_pl_group
                    WHERE amr_user.user_group = amr_field_id.meter_id
                        AND amr_field_id.field_id = amr_pl_group.field_id
                        AND amr_pl_group.pl_region_id = amr_region.id
                        AND user_name = :logged_in_user
                        AND amr_user.user_enable like '1'
                    """

                    tag_query = """
                    SELECT amr_field_id.tag_id
                    FROM AMR_REGION,AMR_user,amr_field_id,amr_pl_group
                    WHERE amr_user.user_group = amr_field_id.meter_id
                        AND amr_field_id.field_id = amr_pl_group.field_id
                        AND amr_pl_group.pl_region_id = amr_region.id
                        AND user_name = :logged_in_user
                        AND amr_user.user_enable like '1'
                        ORDER BY AMR_FIELD_ID.TAG_ID
                    """

                    # Fetch unique region values
                    region_results = fetch_data(ptt_pivot_connection, region_query, params={'logged_in_user': logged_in_user})
                    region_options = [str(region[0]) for region in region_results]
                    for region in region_results:
                        region_results = fetch_data(ptt_pivot_connection, region_query, params={'logged_in_user': logged_in_user})
                        region_options = str(region[0]) 
                        print("region:", region_options)
                    tag_options = None
                    tag_results = fetch_data(ptt_pivot_connection, tag_query, params={'logged_in_user': logged_in_user})
                    for tag in tag_results:
                        tag_options = str(tag[0])
                        print("site:", tag_options)
                    selected_tag = request.args.get("tag_dropdown")
                    selected_region = request.args.get("region_dropdown")
                    selected_day = request.args.get("day_dropdown")

                    selected_date = request.args.get("date_dropdown_daily")
                    session['selected_date_user'] = selected_date
                    session['selected_tag_user'] = selected_tag
                    if selected_date:
                        # Check if the selected_date is in the format of a single day (dd/mm/yyyy)
                        if "/" in selected_date and len(selected_date) == 10:
                            # Example: 25/08/2024
                            # Enter the first if block
                            # Your code for handling a single date goes here
                            pass

                        # Check if the selected_date is a range of days (dd/mm/yyyy to dd/mm/yyyy)
                        elif "to" in selected_date:
                            # Example: 25/08/2024 to 30/08/2024
                            # Enter the else block
                            # Your code for handling a date range goes here
                            start_date_str, end_date_str = selected_date.split(" to ")

                    # Convert the string dates to datetime objects
                            start_date = datetime.datetime.strptime(start_date_str, '%d/%m/%Y')
                            end_date = datetime.datetime.strptime(end_date_str, '%d/%m/%Y')
                            
                            formatted_start_date = start_date.strftime('%d/%m/%Y')
                            formatted_end_date = end_date.strftime('%d/%m/%Y')

                        # Check if the selected_date is a month (mm/yyyy)
                        elif "/" in selected_date and len(selected_date) == 7:
                            # Example: 08/2024
                            # Enter the third else block
                            # Your code for handling a month goes here
                            pass
                    

                    query = ""
        
                    if query_type == "daily_data":
                        # SQL query for main data
                        if "/" in selected_date and len(selected_date) == 10:

                            query = f"""
                                    SELECT DISTINCT
                                AMR_PL_GROUP.PL_REGION_ID,
                                AMR_FIELD_ID.TAG_ID,
                                AMR_FIELD_ID.METER_ID,
                                TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD-MM-YYYY')  AS DATA_DATE,
                                AMR_BILLING_DATA.CORRECTED_VOL as CORRECTED,
                                AMR_BILLING_DATA.UNCORRECTED_VOL as UNCORRECTED,
                                AMR_BILLING_DATA.AVR_PF as Pressure,
                                AMR_BILLING_DATA.AVR_TF as Temperature,
                                AMR_BILLING_DATA.METER_STREAM_NO  
                            FROM
                                AMR_FIELD_ID, AMR_PL_group, AMR_BILLING_DATA,amr_region
                            WHERE
                                AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                                AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                                AND AMR_BILLING_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                                AND AMR_BILLING_DATA.METER_STREAM_NO IS NOT NULL
                                AND TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD/MM/YYYY') = '{selected_date}'
                                AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                                AND amr_region.REGION_NAME = '{selected_region}'
                            ORDER BY
                                TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD-MM-YYYY') DESC
                                """
                            
                        elif "to" in selected_date:
                            query = f"""
                                    SELECT DISTINCT
                                    AMR_PL_GROUP.PL_REGION_ID,
                                    AMR_FIELD_ID.TAG_ID,
                                    AMR_FIELD_ID.METER_ID,
                                    TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD/MM/YY') AS DATA_DATE,
                                    AMR_BILLING_DATA.CORRECTED_VOL AS CORRECTED,
                                    AMR_BILLING_DATA.UNCORRECTED_VOL AS UNCORRECTED,
                                    AMR_BILLING_DATA.AVR_PF AS Pressure,
                                    AMR_BILLING_DATA.AVR_TF AS Temperature,
                                    AMR_BILLING_DATA.METER_STREAM_NO  
                                FROM
                                    AMR_FIELD_ID
                                JOIN AMR_PL_GROUP 
                                    ON AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
                                JOIN AMR_BILLING_DATA 
                                    ON AMR_BILLING_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                                JOIN AMR_REGION 
                                    ON AMR_PL_GROUP.PL_REGION_ID = AMR_REGION.ID
                                WHERE
                                    AMR_BILLING_DATA.METER_STREAM_NO IS NOT NULL
                                    AND AMR_BILLING_DATA.DATA_DATE BETWEEN TO_DATE('{formatted_start_date}', 'DD/MM/YYYY') 
                                                                        AND TO_DATE('{formatted_end_date}', 'DD/MM/YYYY')
                                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                                    AND AMR_REGION.REGION_NAME = '{selected_region}'
                                ORDER BY
                                    TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD/MM/YY') DESC
                                """
                            
                        elif "/" in selected_date and len(selected_date) == 7:

                            print("One or both dates are missing.")
                            # Your code here for missing dates
                            # Fetch tag options based on the selected region
                            query = f"""
                            SELECT DISTINCT
                                AMR_PL_GROUP.PL_REGION_ID,
                                AMR_FIELD_ID.TAG_ID,
                                AMR_FIELD_ID.METER_ID,
                                TO_CHAR(AMR_BILLING_DATA.DATA_DATE)  AS DATA_DATE,
                                AMR_BILLING_DATA.CORRECTED_VOL as CORRECTED,
                                AMR_BILLING_DATA.UNCORRECTED_VOL as UNCORRECTED,
                                AMR_BILLING_DATA.AVR_PF as Pressure,
                                AMR_BILLING_DATA.AVR_TF as Temperature,
                                AMR_BILLING_DATA.METER_STREAM_NO  
                            FROM
                                AMR_FIELD_ID, AMR_PL_group, AMR_BILLING_DATA,amr_region
                            WHERE
                                AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                                AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                                AND AMR_BILLING_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                                AND AMR_BILLING_DATA.METER_STREAM_NO IS NOT NULL
                                AND TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'MM/YYYY') = '{selected_date}'
                                AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                                AND amr_region.REGION_NAME = '{selected_region}'
                            ORDER BY
                                TO_CHAR(AMR_BILLING_DATA.DATA_DATE) DESC
                            """
                        results = fetch_data(ptt_pivot_connection, query)
                        
                        df = pd.DataFrame(
                            results,
                            columns=[
                                "PL_REGION_ID",
                                "TAG_ID",
                                "METER_ID",
                                "DATA_DATE",
                                "CORRECTED",
                                "UNCORRECTED",
                                "Pressure",
                                "Temperature",
                                "METER_STREAM_NO",
                            ],
                        )
                        df["Pressure"] = df["Pressure"].astype(float)
                        df["Temperature"] = df["Temperature"].astype(float)

                        df["Pressure"] = df["Pressure"].round(4)
                        df["Temperature"] = df["Temperature"].round(4)
                        df["CORRECTED"] = pd.to_numeric(df["CORRECTED"], errors='coerce').fillna(0).astype(int).astype(str)
                        df["UNCORRECTED"] = pd.to_numeric(df["UNCORRECTED"], errors='coerce').fillna(0).astype(int).astype(str)



                        df["METER_STREAM_NO"] = df["METER_STREAM_NO"].astype(int)
                        meter_id_list = df["METER_ID"].tolist()
 
                        if not meter_id_list:
                            # If empty, set the vc_name_dict with default values
                            vc_name_dict = {f'vc_name_list{i}': None for i in range(1, 7)}
                        else:
                            query_vc_name = f"""SELECT amr_vc_type.vc_name, amr_field_meter.METER_STREAM_NO 
                                                FROM amr_vc_type, amr_field_meter 
                                                WHERE amr_field_meter.METER_STREAM_TYPE = amr_vc_type.id 
                                                AND amr_field_meter.meter_id LIKE '{meter_id_list[0]}' 
                                                ORDER BY amr_field_meter.METER_STREAM_NO"""
                            
                            results_vc_name = fetch_data(ptt_pivot_connection, query_vc_name)
                            df_vc_name = pd.DataFrame(
                                results_vc_name,
                                columns=["vc_name", "METER_STREAM_NO"]
                            )

                            unique_meter_streams = df_vc_name.drop_duplicates(subset=["METER_STREAM_NO"])[["METER_STREAM_NO", "vc_name"]]

                            meter_stream_no_list = unique_meter_streams["METER_STREAM_NO"].tolist()
                            vc_name_list = unique_meter_streams["vc_name"].tolist()

                            vc_name_dict = {}
                            for i in range(1, 7):
                                if len(vc_name_list) >= i:
                                    vc_name_dict[f'vc_name_list{i}'] = vc_name_list[i - 1]
                                else:
                                    vc_name_dict[f'vc_name_list{i}'] = None  # or "nong", depending on your preference

                            print(vc_name_dict)
                                    
                        
                            query_type = f"""
                                SELECT DISTINCT  amr_vc_type.id as METER_ID , amr_field_meter.METER_STREAM_NO FROM amr_billing_data,amr_field_meter,amr_vc_type 
                                WHERE amr_billing_data.meter_id = amr_field_meter.meter_id 
                                AND amr_field_meter.METER_STREAM_TYPE = amr_vc_type.id
                                AND amr_billing_data.meter_id like '{meter_id_list[0]}'
                                AND  amr_field_meter.METER_STREAM_NO IS NOT NULL
                                ORDER BY amr_field_meter.METER_STREAM_NO """
                                
                            results_type = fetch_data(ptt_pivot_connection, query_type)
                            df_type = pd.DataFrame(
                                results_type,
                                columns=[
                                
                                    "METER_ID",
                                    "METER_STREAM_NO",
                    
                                ]
                            )
                            meter_stream_no_list = df_type["METER_STREAM_NO"].tolist()
                            amr_vc_type_list = df_type["METER_ID"].tolist()
                            # print("test11:",amr_vc_type_list)
                    
                    
                        
                        
                        # print(data_type)
                        
                        
                        
                        # Get the selected Meter ID before removing it from the DataFrame
                        # selected_meter_id = df["METER_ID"].iloc[0]
                        selected_meter_id = None

                        # Check if 'METER_ID' column exists and the DataFrame is not empty
                        if not df.empty and 'METER_ID' in df.columns and len(df['METER_ID']) > 0:
                            selected_meter_id = df['METER_ID'].iloc[0]
                            print(f"Selected Meter ID: {selected_meter_id}")
                        else:
                            print("DataFrame is empty or 'METER_ID' column doesn't exist.")

                        # Now, remove the "METER_ID" column from the DataFrame
                        df = df.drop(["PL_REGION_ID", "TAG_ID", "METER_ID"], axis=1)
                        
                        # Remove newline characters
                        df = df.apply(lambda x: x.str.replace("\n", "") if x.dtype == "object" else x)

                        df = df.drop_duplicates(subset=["DATA_DATE", "METER_STREAM_NO"], keep="first")
                        # Sort DataFrame by 'DATA_DATE'
                        df = df.sort_values(by="DATA_DATE")
                        # print("df",df)
                        # Assuming 'df' is the DataFrame created from the query results
                        num_streams = 6
                        df_runs = {}

                        # Loop to create DataFrames for each METER_STREAM_NO
                        for i in range(1, num_streams + 1):
                            df_runs[f'df_run{i}'] = df[df['METER_STREAM_NO'] == int(i)]
                            
                        # Check if each DataFrame has data before including in the tables dictionary
                        tables = {
                            "config_data": None,
                        }

                        graphs = {
                            "corrected": None,
                            "uncorrected": None,
                            "pressure": None,
                            "temperature": None
                        }

                        if "/" in selected_date and len(selected_date) == 10:
                            # print("selected_date:",selected_date)
                            if selected_date:
                                # Convert selected_date to 'YYYY-MM-DD' format for consistency
                                selected_date_formatted = pd.to_datetime(selected_date, format='%d/%m/%Y').strftime('%d/%m/%Y')

                                # Get the current date in 'YYYY-MM-DD' format
                                current_date = datetime.datetime.now().strftime('%d-%m-%Y')

                                # Determine if the selected date is the current date
                                is_current_day = selected_date_formatted == current_date

                                # Update the query to use the selected date
                                if is_current_day:
                                    # If the selected date is today, show only today's data
                                    query_day = f"""
                                        SELECT TO_CHAR(TO_DATE('{current_date}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                                        FROM DUAL
                                    """

                                else:
                                    # If the selected date is not today, show only the selected day's data
                                    query_day = f"""
                                        SELECT TO_CHAR(TO_DATE('{selected_date_formatted}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                                        FROM DUAL
                                    """

                        elif "to" in selected_date:
                            if formatted_start_date and formatted_end_date:
                                # แปลงวันที่ให้เป็นรูปแบบ 'YYYY-MM-DD' สำหรับใช้ใน SQL
                                start_date_formatted = pd.to_datetime(formatted_start_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                                end_date_formatted = pd.to_datetime(formatted_end_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                                
                                print("Start Date:", formatted_start_date)
                                print("End Date:", formatted_end_date)
                                
                                # สร้างรายการวันที่ในช่วงที่เลือกโดยใช้ SQL
                                query_day = f"""
                                    SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{start_date_formatted}', 'YY-MM-DD'), 'DD/MM/YY') AS DATA_DATE
                                    FROM DUAL
                                    CONNECT BY LEVEL <= TO_DATE('{end_date_formatted}', 'YY-MM-DD') - TO_DATE('{start_date_formatted}', 'YY-MM-DD') + 1
                                """

                        elif "/" in selected_date and len(selected_date) == 7:
                            df_month_list = pd.DataFrame(columns=['DATA_DATE'])
                            
                            # Check if 'selected_date' is available
                            if selected_date:
                                # Convert selected_date to 'YYYY-MM' format for consistency
                                selected_date_formatted = pd.to_datetime(selected_date, format='%m/%Y').strftime('%Y-%m')
                                # print("date2:", selected_date_formatted)

                                # Get the current month and year
                                current_month_year = datetime.datetime.now().strftime('%Y-%m')

                                # Determine if the selected date is in the current month
                                is_current_month = selected_date_formatted == current_month_year
                                
                                # Update the query to use the selected date
                                if is_current_month:
                                    # If the selected date is in the current month, show all days up to the current date
                                    query_day = f"""
                                        SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{current_month_year}-01', 'YYYY-MM-DD')) AS Date1
                                        FROM DUAL
                                        CONNECT BY LEVEL <= TO_DATE('{datetime.datetime.now().strftime('%Y-%m-%d')}', 'YYYY-MM-DD') - TO_DATE('{current_month_year}-01', 'YYYY-MM-DD') + 1
                                    """

                                else:
                                    # If the selected date is in a previous month, show all days of the selected month
                                    query_day = f"""
                                        SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) AS Date1
                                        FROM DUAL
                                        CONNECT BY LEVEL <= LAST_DAY(TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) - TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD') + 1
                                    """

                        query_day_result = fetch_data(ptt_pivot_connection, query_day)
                        df_month_list = pd.DataFrame(query_day_result, columns=['DATA_DATE'])
                        # print("df_month_list",df_month_list)
                        def make_link(row, run_number):
                            if row['CORRECTED'] == 'N/A' and row['UNCORRECTED'] == 'N/A':
                                return f'<a href="#" onclick="sendDataDate(\'{row["DATA_DATE"]}\', {run_number}); openPopup(); return false;">{row["DATA_DATE"]}</a>'
                                
                            return row['DATA_DATE']
                        
             
                        for i in range(1, num_streams + 1):
                            df_run = df_runs[f'df_run{i}']
                            
                            if not df_run.empty:
                                merged_df = pd.merge(df_month_list, df_run, on='DATA_DATE', how='outer')
                                df_runs[f'df_run{i}'] = merged_df
                                html_style = """
                                        <style>
                                        .data_daily th, .data_daily td .data_daily tr{
                                            text-align: center; 
                                        }
                                        </style>
                                        """
                        # print("df_run_no_last",  session['data_billing'])




                        html_style = '''
                            <style>
                                .data_daily {
                                    border-collapse: collapse;
                                    width: 100%;
                                }
                                .data_daily th, .data_daily td {
                                    border: 1px solid #ddd;
                                    padding: 8px;
                                    color: #000000;
                                }
                                .data_daily th {
                                    padding-top: 12px;
                                    padding-bottom: 12px;
                                    text-align: left;
                                    background-color: #f0f0f0;
                                    
                                }
                            </style>
                            '''
                            
                            
                        data_type = []
                        if not meter_id_list:
                            pass
                        else :
                            for amr_vc_type in amr_vc_type_list:
                                amr_vc_type_table = f"""
                                
                                
                                SELECT DESCRIPTION 
                                FROM amr_mapping_billing 
                                WHERE evc_type LIKE '{amr_vc_type}' 
                                AND DAILY like '1' 
                                ORDER BY OR_DER


                                """
                                
                                amr_vc_type_table_result = fetch_data(ptt_pivot_connection, amr_vc_type_table)
                                filtered_result = pd.DataFrame(amr_vc_type_table_result)
                                
                        
                                transposed_result = filtered_result.T
                                
                                
                                selected_data = transposed_result.iloc[0, 0:5]  # Extract values to get a flat array
                                data_type.append(selected_data)
                        # Print the values in a horizontal format
                        # for i in range(4):
                        #     print(data_type[i])
                                
                        
                            
                        df_graphs = []  # สร้างลิสต์เพื่อเก็บข้อมูล
                        combined_data = []
                        all_totals = []  # สร้างลิสต์เพื่อเก็บค่า totals ของแต่ละรอบ

                        for i in range(num_streams):
                            df_run = df_runs.get(f'df_run{i+1}', pd.DataFrame())

                            df_graphs_run = pd.DataFrame(df_run)
                            df_graphs.append(df_graphs_run)  # Store DataFrame in the list

                            if not df_run.empty:
                                df_run = df_run.drop('METER_STREAM_NO', axis=1, errors='ignore')
                                df_run = df_run.fillna("N/A")
                                combined_data.append(df_run.to_dict())

                               

                                # Prepare for constructing HTML
                                df_run_no_last = df_run
                                hedder = data_type[i]
                                header_df = pd.DataFrame(hedder).T


                                # Create the HTML manually as before
                                html_output = '<table class="data_daily">'

                                # Extract the first row to use as header
                                header_row = header_df.iloc[0]
                                html_output += '<thead><tr>'
                                for item in header_row:
                                    html_output += f'<th style="text-align: center;">{item}</th>'
                                html_output += '</tr></thead>'

                                # Now create the tbody
                                html_output += '<tbody>'
                                for index, row in df_run_no_last.iloc[0:].iterrows():  # Skip the first row
                                    html_output += '<tr>'
                                    for item in row:
                                        html_output += f'<td>{item}</td>'
                                    html_output += '</tr>'

                                # Add the total row (calculated from the daily differences)
                                html_output += '<tr><td style="text-align: center; background-color: #1cc88a33;"">Total</td>'
                                

                                                # Convert 'CORRECTED' and 'UNCORRECTED' to numeric
                                df_run_no_last['CORRECTED'] = pd.to_numeric(df_run_no_last['CORRECTED'], errors='coerce')
                                df_run_no_last['UNCORRECTED'] = pd.to_numeric(df_run_no_last['UNCORRECTED'], errors='coerce')

                                # Forward fill missing values to ensure calculation continuity
                                df_run_no_last[['CORRECTED', 'UNCORRECTED']] = df_run_no_last[['CORRECTED', 'UNCORRECTED']].ffill()

                                # Calculate the difference for only 'CORRECTED' and 'UNCORRECTED' columns after filling missing values
                                df_run_diff = df_run_no_last[['CORRECTED', 'UNCORRECTED']].diff().fillna(0)

                                # Filter out zero values
                                df_run_diff_filtered = df_run_diff[df_run_diff != 0]

                                # Now calculate the totals for the non-zero values (daily difference totals)
                                totals = df_run_diff_filtered.sum()

                                # Add this round's totals to all_totals list
                                all_totals.append(totals)
                                
                                
                                
                                for col in df_run_no_last.columns[1:]:  # Skip the first column (assumed non-numeric)
                                    if col in totals:
                                        total_value = int(totals[col])
                                        html_output += f'<td  style="background-color: #1cc88a33;">{total_value}</td>'
                                    # else:
                                    #     html_output += '<td ></td>'  # Leave empty for non-numeric columns
                                html_output += '</tr>'

                                html_output += '</tbody>'
                                html_output += '</table>'

                                # Store the constructed HTML in the tables dictionary
                                tables[f"daily_data_run{i+1}"] = html_style + html_output
                                
                        # After looping through all runs, sum all totals
                        final_totals = pd.DataFrame(all_totals).sum()

                        html_final_totals = '<table class="data_totals">'
                        html_final_totals += '<thead><tr><th>Daily Corrected Volume</th><th>Daily Uncorrected Volume</th></tr></thead>'
                        html_final_totals += '<tbody><tr>'  # Start the body and first row

                        # Iterate over the final_totals index to populate the total values
                        for col in final_totals.index:
                            html_final_totals += f'<td>{int(final_totals[col])}</td>'
                        html_final_totals += '</tr></tbody></table>'

                        # Store the final combined totals table
                        tables["combined_totals"] = html_style + html_final_totals

                        with open('data_report_user.json', 'w') as json_file:
                            json.dump(combined_data, json_file, indent=4)

                        df = df.sort_values(by="DATA_DATE", ascending=True)
                        # ส่ง graph_html ไปยัง HTML template ของ Flask
                        return render_template(
                            "billingdata_user.html",
                            tables=tables,
                            titles=df.columns.values,
                            selected_date=selected_date,
                            selected_tag=selected_tag,
                            selected_region=selected_region,
                            region_options=region_options,
                            tag_options=tag_options,
                            selected_meter_id=selected_meter_id,
                            vc_name_dict=vc_name_dict,username=logged_in_user
                            # graph_corrected=graph_corrected,
                            # graph_uncorrected=graph_uncorrected,
                            # graph_pressure=graph_pressure,
                            # graph_temperature=graph_temperature,
                        ) 
                    else:
                        # Render the template without executing the query
                        return render_template(
                            "billingdata_user.html",
                            selected_date=selected_date,
                            selected_region=selected_region,
                            selected_tag=selected_tag,
                            region_options=region_options,
                            tag_options=tag_options,
                            tables={},
                            username=logged_in_user,  
                            description=description,
                            user_level=user_level
                            
                        )   
            else:
                # Handle the case where the user is not in the users dictionary
                flash('User information not found. Please log in again.', 'error')
                return redirect(url_for('login'))
        else:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        
############  View Billing Data   #####################

# Get List of sitename on selected Region
@app.route("/get_tags", methods=["GET"])
@login_required
def get_tags():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        selected_region = request.args.get("selected_region")

        tag_query = """
         SELECT DISTINCT AMR_FIELD_ID.TAG_ID
        FROM AMR_FIELD_ID,amr_region,AMR_PL_GROUP
        
        WHERE AMR_PL_GROUP.PL_REGION_ID = amr_region.id
        AND AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
        AND amr_region.REGION_NAME = :region_id
        AND AMR_FIELD_ID.tag_id NOT like '%.remove%'
        ORDER BY  TAG_ID
        """

        tag_results = fetch_data(ptt_pivot_connection,tag_query, params={"region_id": selected_region})
        tag_options = [str(tag[0]) for tag in tag_results]
        tag_options.sort()
        return jsonify({"tag_options": tag_options})
  
@app.route("/billing_data")
@login_required
def billing_data():
    username = session.get('username')
    print("username",username)
    
    if 'username' not in session:
        
        return redirect(url_for('login'))
    
    # connect_to_amr_db
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
    #with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        
        # print("Active Connection:", active_connection)

        query_type = request.args.get("query_type")
       
        print(query_type)
        # SQL query to fetch unique PL_REGION_ID values
        region_query = """
        SELECT * FROM AMR_REGION 
        """

        tag_query = """
        
        SELECT DISTINCT AMR_FIELD_ID.TAG_ID
        FROM AMR_FIELD_ID,amr_region,AMR_PL_GROUP
        
        WHERE AMR_PL_GROUP.PL_REGION_ID = amr_region.id
        AND AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
        AND amr_region.REGION_NAME = :region_id
        AND AMR_FIELD_ID.tag_id NOT like '%.remove%'
        ORDER BY  TAG_ID
        """

        # Fetch unique region values
        region_results = fetch_data(ptt_pivot_connection, region_query)
        region_options = [str(region[1]) for region in region_results]
        
        selected_tag = request.args.get("tag_dropdown")
        selected_region = request.args.get("region_dropdown")
        selected_day = request.args.get("day_dropdown")

        selected_date = request.args.get("date_dropdown_daily")
        session['selected_date'] = selected_date
        session['selected_tag'] = selected_tag
        session['selected_region'] = selected_region
        session['selected_day'] = selected_day

        ## check billing_date
        query = f"""
                    SELECT 
                    AMR_FIELD_ID.BILLING_DATE
                FROM
                    AMR_FIELD_ID
                WHERE
                    AMR_FIELD_ID.TAG_ID LIKE '{selected_tag}'
                """
        
        results = fetch_data(ptt_pivot_connection, query)
        if not results:
            billing_date = '0'
        else:
            billing_date = results[0][0]

        if selected_date:
            # Check if the selected_date is in the format of a single day (dd/mm/yyyy)
            if "/" in selected_date and len(selected_date) == 10:
                pass

            # Check if the selected_date is a range of days (dd/mm/yyyy to dd/mm/yyyy)
            elif "to" in selected_date:
                start_date_str, end_date_str = selected_date.split(" to ")

        # Convert the string dates to datetime objects
                start_date = datetime.datetime.strptime(start_date_str, '%d/%m/%Y')
                end_date = datetime.datetime.strptime(end_date_str, '%d/%m/%Y')
                
                formatted_start_date = start_date.strftime('%d/%m/%Y')
                formatted_end_date = end_date.strftime('%d/%m/%Y')

            # Check if the selected_date is a month (mm/yyyy)
            elif "/" in selected_date and len(selected_date) == 7:
                # Example: 08/2024
                # Enter the third else block
                # Your code for handling a month goes here

                ## Tul 2025-02-02 fix report    
                selected_dt = datetime.datetime.strptime(selected_date, "%m/%Y")
                first_day_of_month = selected_dt.replace(day=1)
                end_of_month = first_day_of_month + pd.offsets.MonthEnd(0)
                today = datetime.datetime.today()
                current_month = today.strftime("%m/%Y")

                def get_previous_month_27():
                    today = datetime.date.today()
                    first_day_of_current_month = today.replace(day=1)
                    last_day_of_previous_month = first_day_of_current_month - datetime.timedelta(days=1)
                    date_27_previous_month = last_day_of_previous_month.replace(day=27)
                    return date_27_previous_month
                
                if billing_date == '0':
                    start_date = first_day_of_month - datetime.timedelta(days=1)
                    if selected_date == current_month:
                        # ถ้าเป็นเดือนปัจจุบัน: end_date = เมื่อวาน
                        end_date = today - datetime.timedelta(days=1)
                    else:
                        # ถ้าเป็นเดือนอื่น: end_date = วันสุดท้ายของเดือนนั้น
                        end_date = end_of_month  # วันสุดท้ายของเดือนที่เลือก
                else:  # billing_date = 1
                    current_month_chk = today.month
                    current_year_chk = today.year

                    if selected_dt.month == current_month_chk and selected_dt.year == current_year_chk:

                        previous_month = today.replace(day=1) - datetime.timedelta(days=1)
                        start_date = previous_month.replace(day=27)
                        end_date = selected_dt.replace(day=27)

                        # if today.day >= 29:
                        #     start_date = today.replace(day=27)  # start_date = วันที่ 27 ของเดือนปัจจุบัน
                        # else:
                        #     previous_month = today.replace(day=1) - datetime.timedelta(days=1)  # หาเดือนก่อนหน้า
                        #     start_date = previous_month.replace(day=27)  # start_date = วันที่ 27 ของเดือนก่อนหน้า

                        # end_date = today - datetime.timedelta(days=1)  # end_date = เมื่อวาน

                    else:
                        previous_month = selected_dt.replace(day=1) - datetime.timedelta(days=1)  # หาเดือนก่อนหน้า
                        start_date = previous_month.replace(day=27)  # start_date
                        end_date = selected_dt.replace(day=27)  # end_date

                    yesterday = today - datetime.timedelta(days=1)
                    if end_date > yesterday:
                        end_date = yesterday
                
                formatted_start_date = start_date.strftime('%d/%m/%Y')
                formatted_end_date = end_date.strftime('%d/%m/%Y')
            ### end case
                
        tag_results = fetch_data(ptt_pivot_connection, tag_query, params={"region_id": selected_region})
        tag_options = [str(tag[0]) for tag in tag_results]
        
        query = ""
        
        if query_type == "daily_data":
            # SQL query for main data
            if "/" in selected_date and len(selected_date) == 10:

                query = f"""
                        SELECT DISTINCT
                    AMR_PL_GROUP.PL_REGION_ID,
                    AMR_FIELD_ID.TAG_ID,
                    AMR_FIELD_ID.METER_ID,
                    TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD-MM-YYYY')  AS DATA_DATE,
                    
                    AMR_BILLING_DATA.CORRECTED_VOL as CORRECTED,
                    AMR_BILLING_DATA.UNCORRECTED_VOL as UNCORRECTED,
                    AMR_BILLING_DATA.AVR_PF as Pressure,
                    AMR_BILLING_DATA.AVR_TF as Temperature,
                    AMR_BILLING_DATA.METER_STREAM_NO
                FROM
                    AMR_FIELD_ID, AMR_PL_group, AMR_BILLING_DATA,amr_region
                WHERE
                    AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                    AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                    AND AMR_BILLING_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    AND AMR_BILLING_DATA.METER_STREAM_NO IS NOT NULL
                    AND TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD/MM/YYYY') = '{selected_date}'
                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                    AND amr_region.REGION_NAME = '{selected_region}'
                ORDER BY
                    TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD-MM-YYYY') DESC
                    """
                
            elif "to" in selected_date or ("/" in selected_date and len(selected_date) == 7):
                query = f"""
                        SELECT DISTINCT
                        AMR_PL_GROUP.PL_REGION_ID,
                        AMR_FIELD_ID.TAG_ID,
                        AMR_FIELD_ID.METER_ID,
                        TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD/MM/YY') AS DATA_DATE,
                        AMR_BILLING_DATA.CORRECTED_VOL AS CORRECTED,
                        AMR_BILLING_DATA.UNCORRECTED_VOL AS UNCORRECTED,
                        AMR_BILLING_DATA.AVR_PF as Pressure,
                        AMR_BILLING_DATA.AVR_TF as Temperature,

                        AMR_BILLING_DATA.METER_STREAM_NO
                    FROM
                        AMR_FIELD_ID
                    JOIN AMR_PL_GROUP 
                        ON AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
                    JOIN AMR_BILLING_DATA 
                        ON AMR_BILLING_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    JOIN AMR_REGION 
                        ON AMR_PL_GROUP.PL_REGION_ID = AMR_REGION.ID
                    WHERE
                        AMR_BILLING_DATA.METER_STREAM_NO IS NOT NULL
                        AND AMR_BILLING_DATA.DATA_DATE BETWEEN TO_DATE('{formatted_start_date}', 'DD/MM/YYYY') 
                                                            AND TO_DATE('{formatted_end_date}', 'DD/MM/YYYY')
                        AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                        AND AMR_REGION.REGION_NAME = '{selected_region}'
                    ORDER BY
                        TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD/MM/YY') DESC
                    """
            results = fetch_data(ptt_pivot_connection, query)

            df = pd.DataFrame(
                results,
                columns=[
                    "PL_REGION_ID",
                    "TAG_ID",
                    "METER_ID",
                    "DATA_DATE",
                    "CORRECTED",
                    "UNCORRECTED",
                    "Pressure",
                    "Temperature",
                    "METER_STREAM_NO",
                ],
            )
            
            ### HARD CODE decimal 4 place

            df["Pressure"] = df["Pressure"].astype(float)
            df["Temperature"] = df["Temperature"].astype(float)

            df["Pressure"] = df["Pressure"].round(4)
            df["Temperature"] = df["Temperature"].round(4)
            df["CORRECTED"] = pd.to_numeric(df["CORRECTED"], errors='coerce').fillna(0).astype(int).astype(str)
            df["UNCORRECTED"] = pd.to_numeric(df["UNCORRECTED"], errors='coerce').fillna(0).astype(int).astype(str)

            df["METER_STREAM_NO"] = df["METER_STREAM_NO"].astype(int)
            meter_id_list = df["METER_ID"].tolist()

            if not meter_id_list:
                # If empty, set the vc_name_dict with default values
                vc_name_dict = {f'vc_name_list{i}': None for i in range(1, 7)}
            else:
                query_vc_name = f"""SELECT amr_vc_type.vc_name, amr_field_meter.METER_STREAM_NO 
                                    FROM amr_vc_type, amr_field_meter 
                                    WHERE amr_field_meter.METER_STREAM_TYPE = amr_vc_type.id 
                                    AND amr_field_meter.meter_id LIKE '{meter_id_list[0]}' 
                                    ORDER BY amr_field_meter.METER_STREAM_NO"""
                
                results_vc_name = fetch_data(ptt_pivot_connection, query_vc_name)
                df_vc_name = pd.DataFrame(
                    results_vc_name,
                    columns=["vc_name", "METER_STREAM_NO"]
                )

                unique_meter_streams = df_vc_name.drop_duplicates(subset=["METER_STREAM_NO"])[["METER_STREAM_NO", "vc_name"]]

                meter_stream_no_list = unique_meter_streams["METER_STREAM_NO"].tolist()
                vc_name_list = unique_meter_streams["vc_name"].tolist()

                vc_name_dict = {}
                for i in range(1, 7):
                    if len(vc_name_list) >= i:
                        vc_name_dict[f'vc_name_list{i}'] = vc_name_list[i - 1]
                    else:
                        vc_name_dict[f'vc_name_list{i}'] = None  # or "nong", depending on your preference

                #print(vc_name_dict)
                        
                query_type = f"""
                    SELECT DISTINCT  amr_vc_type.id as METER_ID , amr_field_meter.METER_STREAM_NO FROM amr_billing_data,amr_field_meter,amr_vc_type 
                    WHERE amr_billing_data.meter_id = amr_field_meter.meter_id 
                    AND amr_field_meter.METER_STREAM_TYPE = amr_vc_type.id
                    AND amr_billing_data.meter_id like '{meter_id_list[0]}'
                    AND  amr_field_meter.METER_STREAM_NO IS NOT NULL
                    ORDER BY amr_field_meter.METER_STREAM_NO """
                    
                results_type = fetch_data(ptt_pivot_connection, query_type)
                df_type = pd.DataFrame(
                    results_type,
                    columns=[
                        "METER_ID",
                        "METER_STREAM_NO",
                    ]
                )
                meter_stream_no_list = df_type["METER_STREAM_NO"].tolist()
                amr_vc_type_list = df_type["METER_ID"].tolist()
            
            # Get the selected Meter ID before removing it from the DataFrame
            # selected_meter_id = df["METER_ID"].iloc[0]
            selected_meter_id = None

            # Check if 'METER_ID' column exists and the DataFrame is not empty
            if not df.empty and 'METER_ID' in df.columns and len(df['METER_ID']) > 0:
                selected_meter_id = df['METER_ID'].iloc[0]
                print(f"Selected Meter ID: {selected_meter_id}")
            else:
                print("DataFrame is empty or 'METER_ID' column doesn't exist.")

            # Now, remove the "METER_ID" column from the DataFrame
            df = df.drop(["PL_REGION_ID", "TAG_ID", "METER_ID"], axis=1)
            
            # Remove newline characters
            df = df.apply(lambda x: x.str.replace("\n", "") if x.dtype == "object" else x)

            df = df.drop_duplicates(subset=["DATA_DATE", "METER_STREAM_NO"], keep="first")
            # Sort DataFrame by 'DATA_DATE'
            
            num_streams = 6
            df_runs = {}

            # Loop to create DataFrames for each METER_STREAM_NO
            for i in range(1, num_streams + 1):
                df_runs[f'df_run{i}'] = df[df['METER_STREAM_NO'] == int(i)]
            
            ## Keep data to session for graph GGGG
            # session["df_runs"] = json.dumps({key: df_run.to_json() for key, df_run in df_runs.items()})

                
            # Check if each DataFrame has data before including in the tables dictionary
            tables = {
                "config_data": None,
            }

            graphs = {
                "corrected": None,
                "uncorrected": None,
                "pressure": None,
                "temperature": None
            }
            #################### DF Month list
  
            df_month_list = pd.DataFrame({"DATA_DATE": pd.date_range(start=start_date, end=end_date).strftime("%d/%m/%y")})
            
            def make_link(row, run_number):
                if row['CORRECTED'] == 'N/A' and row['UNCORRECTED'] == 'N/A':
                    print("HHere")
                    return f'<a href="#" onclick="sendDataDate(\'{row["DATA_DATE"]}\', {run_number}); openPopup(); return false;">{row["DATA_DATE"]}</a>'
                return row['DATA_DATE']

            for i in range(1, num_streams + 1):
                df_run = df_runs[f'df_run{i}']
                if not df_run.empty:
                    
                    merged_df = pd.merge(df_month_list, df_run, on='DATA_DATE', how='outer')
                    ## sort by date
                    merged_df["DATA_DATE"] = pd.to_datetime(merged_df["DATA_DATE"], format="%d/%m/%y")
                    merged_df = merged_df.sort_values(by="DATA_DATE")
                    merged_df["DATA_DATE"] = merged_df["DATA_DATE"].dt.strftime("%d/%m/%y")

                    merged_df['DATA_DATE'] = merged_df['DATA_DATE'].apply(lambda x: x + ' 24:00')

                    # ตรวจสอบว่าคอลัมน์ 'Pressure' เป็นตัวเลข (float/int) และไม่ใช่ NaN ก่อนแปลง
                    merged_df['Pressure'] = merged_df['Pressure'].apply(
                        lambda x: f"{x:.4f}" if isinstance(x, (int, float)) and not pd.isna(x) else x)
                    
                    df_runs[f'df_run{i}'] = merged_df
                    
                    html_style = """
                            <style>
                            .data_daily th, .data_daily td .data_daily tr{
                                text-align: center; 
                            }
                            </style>
                            """
            # print("df_run_no_last",  session['data_billing'])

            html_style = '''
                <style>
                    .data_daily {
                        border-collapse: collapse;
                        width: 100%;
                    }
                    .data_daily th, .data_daily td {
                        border: 1px solid #ddd;
                        padding: 8px;
                        color: #000000;
                    }
                    .data_daily th {
                        padding-top: 12px;
                        padding-bottom: 12px;
                        text-align: left;
                        background-color: #f0f0f0;   
                    }
                </style>
                '''
            data_type = []
            if not meter_id_list:
                pass
            else :
                for amr_vc_type in amr_vc_type_list:
                    amr_vc_type_table = f"""

                    SELECT DESCRIPTION 
                    FROM amr_mapping_billing 
                    WHERE evc_type LIKE '{amr_vc_type}' 
                    AND DAILY like '1' 
                    ORDER BY OR_DER
                    """
                    
                    amr_vc_type_table_result = fetch_data(ptt_pivot_connection, amr_vc_type_table)
                    filtered_result = pd.DataFrame(amr_vc_type_table_result)
                                
                    transposed_result = filtered_result.T
                    
                    selected_data = transposed_result.iloc[0, 0:5]  # Extract values to get a flat array
                    data_type.append(selected_data)
            # Print the values in a horizontal format
            # for i in range(4):
            #     print(data_type[i])
            
            df_graphs = []  # สร้างลิสต์เพื่อเก็บข้อมูล
            combined_data = []
            all_totals = []  # สร้างลิสต์เพื่อเก็บค่า totals ของแต่ละรอบ

            for i in range(num_streams):
                df_run = df_runs.get(f'df_run{i+1}', pd.DataFrame())

                df_graphs_run = pd.DataFrame(df_run)
                df_graphs.append(df_graphs_run)  # Store DataFrame in the list

                if not df_run.empty:
                    df_run = df_run.drop('METER_STREAM_NO', axis=1, errors='ignore')
                    df_run = df_run.fillna("N/A")
                    
                    combined_data.append(df_run.to_dict())
                    
                    df_run['DATA_DATE'] = df_run.apply(lambda row: make_link(row, i+1), axis=1)

                    # Prepare for constructing HTML
                    df_run_no_last = df_run
                    hedder = data_type[i]
                    header_df = pd.DataFrame(hedder).T

                    # Create the HTML manually as before
                    html_output = '<table class="data_daily">'

                    # Extract the first row to use as header
                    header_row = header_df.iloc[0]
                    html_output += '<thead><tr>'
                    for item in header_row:
                        html_output += f'<th style="text-align: center;">{item}</th>'
                    html_output += '</tr></thead>'

                    # Now create the tbody
                    html_output += '<tbody>'
                    for index, row in df_run_no_last.iloc[0:].iterrows():  # Skip the first row
                        html_output += '<tr>'
                        for item in row:
                            html_output += f'<td>{item}</td>'
                        html_output += '</tr>'

                    # Add the total row (calculated from the daily differences)
                    html_output += '<tr><td style="text-align: center; background-color: #1cc88a33;"">Total</td>'
                    
                                     # Convert 'CORRECTED' and 'UNCORRECTED' to numeric
                    df_run_no_last['CORRECTED'] = pd.to_numeric(df_run_no_last['CORRECTED'], errors='coerce')
                    df_run_no_last['UNCORRECTED'] = pd.to_numeric(df_run_no_last['UNCORRECTED'], errors='coerce')

                    # Forward fill missing values to ensure calculation continuity
                    df_run_no_last[['CORRECTED', 'UNCORRECTED']] = df_run_no_last[['CORRECTED', 'UNCORRECTED']].ffill()

                    # Calculate the difference for only 'CORRECTED' and 'UNCORRECTED' columns after filling missing values
                    df_run_diff = df_run_no_last[['CORRECTED', 'UNCORRECTED']].diff().fillna(0)

                    # Filter out zero values
                    df_run_diff_filtered = df_run_diff[df_run_diff != 0]

                    # Now calculate the totals for the non-zero values (daily difference totals)
                    totals = df_run_diff_filtered.sum()

                    # Add this round's totals to all_totals list
                    all_totals.append(totals)
 
                    for col in df_run_no_last.columns[1:]:  # Skip the first column (assumed non-numeric)
                        if col in totals:
                            total_value = int(totals[col])
                            html_output += f'<td  style="background-color: #1cc88a33;">{total_value}</td>'
                        # else:
                        #     html_output += '<td ></td>'  # Leave empty for non-numeric columns
                    html_output += '</tr>'

                    html_output += '</tbody>'
                    html_output += '</table>'

                    # Store the constructed HTML in the tables dictionary
                    tables[f"daily_data_run{i+1}"] = html_style + html_output
                    
            # After looping through all runs, sum all totals
            final_totals = pd.DataFrame(all_totals).sum()

            html_final_totals = '<table class="data_totals">'
            html_final_totals += '<thead><tr><th>Daily Corrected Volume</th><th>Daily Uncorrected Volume</th></tr></thead>'
            html_final_totals += '<tbody><tr>'  # Start the body and first row

            # Iterate over the final_totals index to populate the total values
            for col in final_totals.index:
                html_final_totals += f'<td>{int(final_totals[col])}</td>'
            html_final_totals += '</tr></tbody></table>'

            # Store the final combined totals table
            tables["combined_totals"] = html_style + html_final_totals
            with open('combined_data_report.json', 'w') as json_file:
                json.dump(combined_data, json_file, indent=4)
                           
            return render_template(
                "billingdata.html",
                tables=tables,
                titles=df.columns.values,
                selected_date=selected_date,
                selected_tag=selected_tag,
                selected_region=selected_region,
                region_options=region_options,
                tag_options=tag_options,
                selected_meter_id=selected_meter_id,
                
                vc_name_dict=vc_name_dict,
                username=username
            )

            # Return the template with the DataFrame

######################## config_data

        elif query_type == "config_data":
            if "/" in selected_date and len(selected_date) == 10:

                query = f"""
                    SELECT DISTINCT
                        AMR_PL_GROUP.PL_REGION_ID,
                        AMR_FIELD_ID.TAG_ID,
                        amr_field_id.meter_id,
                        AMR_CONFIGURED_DATA.METER_STREAM_NO,
                        AMR_CONFIGURED_DATA.amr_vc_type,
                        amr_vc_type.vc_name,
                        TO_CHAR(AMR_CONFIGURED_DATA.DATA_DATE, 'DD-MM-YYYY'),
                                                
                        amr_configured_data.amr_config1,
                        amr_configured_data.amr_config2,
                        amr_configured_data.amr_config3,
                        amr_configured_data.amr_config4,
                        amr_configured_data.amr_config5,
                        amr_configured_data.amr_config6,
                        amr_configured_data.amr_config7,
                        amr_configured_data.amr_config8,
                        amr_configured_data.amr_config9,
                        amr_configured_data.amr_config10,
                        amr_configured_data.amr_config11,
                        amr_configured_data.amr_config12,
                        amr_configured_data.amr_config13,
                        amr_configured_data.amr_config14,
                        amr_configured_data.amr_config15,
                        amr_configured_data.amr_config16,
                        amr_configured_data.amr_config17,
                        amr_configured_data.amr_config18,
                        amr_configured_data.amr_config19,
                        amr_configured_data.amr_config20
                                            
                        FROM
                    AMR_FIELD_ID, AMR_PL_group, AMR_CONFIGURED_DATA,amr_field_meter,amr_vc_type,amr_region
                WHERE
                    AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                    AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                    AND  AMR_CONFIGURED_DATA.amr_vc_type = amr_vc_type.id
                    AND amr_field_meter.meter_id =  amr_field_id.meter_id
                    AND AMR_CONFIGURED_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    AND AMR_CONFIGURED_DATA.METER_STREAM_NO IS NOT NULL
                    AND TO_CHAR(AMR_CONFIGURED_DATA.DATA_DATE, 'DD/MM/YYYY') = '{selected_date}'
                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                    AND amr_region.REGION_NAME = '{selected_region}'
                    
                ORDER BY 
                    AMR_CONFIGURED_DATA.METER_STREAM_NO
                """

            elif "to" in selected_date:
                query = f"""
                            SELECT DISTINCT
                        AMR_PL_GROUP.PL_REGION_ID,
                        AMR_FIELD_ID.TAG_ID,
                        amr_field_id.meter_id,
                        AMR_CONFIGURED_DATA.METER_STREAM_NO,
                        AMR_CONFIGURED_DATA.amr_vc_type,
                        amr_vc_type.vc_name,
                        TO_CHAR(AMR_CONFIGURED_DATA.DATA_DATE, 'DD-MM-YYYY'),
                        
                        amr_configured_data.amr_config1,
                        amr_configured_data.amr_config2,
                        amr_configured_data.amr_config3,
                        amr_configured_data.amr_config4,
                        amr_configured_data.amr_config5,
                        amr_configured_data.amr_config6,
                        amr_configured_data.amr_config7,
                        amr_configured_data.amr_config8,
                        amr_configured_data.amr_config9,
                        amr_configured_data.amr_config10,
                        amr_configured_data.amr_config11,
                        amr_configured_data.amr_config12,
                        amr_configured_data.amr_config13,
                        amr_configured_data.amr_config14,
                        amr_configured_data.amr_config15,
                        amr_configured_data.amr_config16,
                        amr_configured_data.amr_config17,
                        amr_configured_data.amr_config18,
                        amr_configured_data.amr_config19,
                        amr_configured_data.amr_config20
                        
                        FROM
                    AMR_FIELD_ID, AMR_PL_group, AMR_CONFIGURED_DATA,amr_field_meter,amr_vc_type,amr_region
                WHERE
                    AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                    AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                    AND  AMR_CONFIGURED_DATA.amr_vc_type = amr_vc_type.id
                    AND amr_field_meter.meter_id =  amr_field_id.meter_id
                    AND AMR_CONFIGURED_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    AND AMR_CONFIGURED_DATA.METER_STREAM_NO IS NOT NULL 
                    AND AMR_CONFIGURED_DATA.DATA_DATE BETWEEN TO_DATE('{formatted_start_date}', 'DD/MM/YYYY') AND TO_DATE('{formatted_end_date}', 'DD/MM/YYYY')
                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                    AND amr_region.REGION_NAME = '{selected_region}'
                    
                ORDER BY 
                    AMR_CONFIGURED_DATA.METER_STREAM_NO
                    """

            elif "/" in selected_date and len(selected_date) == 7:

                query = f"""
                    SELECT DISTINCT
                        AMR_PL_GROUP.PL_REGION_ID,
                        AMR_FIELD_ID.TAG_ID,
                        amr_field_id.meter_id,
                        AMR_CONFIGURED_DATA.METER_STREAM_NO,
                        AMR_CONFIGURED_DATA.amr_vc_type,
                        amr_vc_type.vc_name,
                        TO_CHAR(AMR_CONFIGURED_DATA.DATA_DATE) as DATA_DATE,                   
                        amr_configured_data.amr_config1,
                        amr_configured_data.amr_config2,
                        amr_configured_data.amr_config3,
                        amr_configured_data.amr_config4,
                        amr_configured_data.amr_config5,
                        amr_configured_data.amr_config6,
                        amr_configured_data.amr_config7,
                        amr_configured_data.amr_config8,
                        amr_configured_data.amr_config9,
                        amr_configured_data.amr_config10,
                        amr_configured_data.amr_config11,
                        amr_configured_data.amr_config12,
                        amr_configured_data.amr_config13,
                        amr_configured_data.amr_config14,
                        amr_configured_data.amr_config15,
                        amr_configured_data.amr_config16,
                        amr_configured_data.amr_config17,
                        amr_configured_data.amr_config18,
                        amr_configured_data.amr_config19,
                        amr_configured_data.amr_config20
                        FROM
                    AMR_FIELD_ID, AMR_PL_group, AMR_CONFIGURED_DATA,amr_field_meter,amr_vc_type,amr_region
                WHERE
                    AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                    AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                    AND  AMR_CONFIGURED_DATA.amr_vc_type = amr_vc_type.id
                    AND amr_field_meter.meter_id =  amr_field_id.meter_id
                    AND AMR_CONFIGURED_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    AND AMR_CONFIGURED_DATA.METER_STREAM_NO IS NOT NULL
                    AND TO_CHAR(AMR_CONFIGURED_DATA.DATA_DATE, 'MM/YYYY') = '{selected_date}'
                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                    AND amr_region.REGION_NAME = '{selected_region}'
                ORDER BY 
                    AMR_CONFIGURED_DATA.METER_STREAM_NO      
                """
            results = fetch_data(ptt_pivot_connection, query)
            #print(que)
            df = pd.DataFrame(
                results,
                columns=[
                    "PL_REGION_ID",
                    "TAG_ID",
                    "METER_ID",
                    "METER_STREAM_NO",
                    "amr_vc_type",
                    "vc_name",
                    "DATA_DATE",
                    "AMR_CONFIG1",
                    "AMR_CONFIG2",
                    "AMR_CONFIG3",
                    "AMR_CONFIG4",
                    "AMR_CONFIG5",
                    "AMR_CONFIG6",
                    "AMR_CONFIG7",
                    "AMR_CONFIG8",
                    "AMR_CONFIG9",
                    "AMR_CONFIG10",
                    "AMR_CONFIG11",
                    "AMR_CONFIG12",
                    "AMR_CONFIG13",
                    "AMR_CONFIG14",
                    "AMR_CONFIG15",
                    "AMR_CONFIG16",
                    "AMR_CONFIG17",
                    "AMR_CONFIG18",
                    "AMR_CONFIG19",
                    "AMR_CONFIG20",
                ]
            )

            df = df.sort_values(by=['METER_STREAM_NO', 'DATA_DATE'])
            
            df_configured = df

            ##

            # unique_meter_streams = df.drop_duplicates(subset=["METER_STREAM_NO"])[["METER_STREAM_NO", "amr_vc_type","vc_name"]]
            # meter_stream_no_list = unique_meter_streams["METER_STREAM_NO"].tolist()
            # amr_vc_type_list = unique_meter_streams["amr_vc_type"].tolist()
            # vc_name_list = unique_meter_streams["vc_name"].tolist()
            
            query = f"""SELECT afm.METER_STREAM_NO AS meter_stream_no , afm.METER_STREAM_TYPE AS meter_stream_type, avt.VC_NAME AS vc_name FROM 
            AMR_FIELD_METER afm , AMR_VC_TYPE avt, AMR_FIELD_ID afi 
            WHERE 
            afm.METER_ID = afi.METER_ID 
            AND afm.METER_STREAM_TYPE = AVT.ID 
            AND afi.TAG_ID LIKE '{selected_tag}' 
            ORDER BY meter_stream_no"""
            results = fetch_data(ptt_pivot_connection, query)
            meter_stream_no_list = [item[0] for item in results]
            amr_vc_type_list = [item[1] for item in results]
            vc_name_list = [item[2] for item in results]
            
            vc_name_dict = {}
            for i in range(1, 7):
                if len(vc_name_list) >= i:
                    vc_name_dict[f'vc_name_list{i}'] = vc_name_list[i-1]
                else:
                    vc_name_dict[f'vc_name_list{i}'] = None
            
            session['vc_type'] = vc_name_dict

            if "/" in selected_date and len(selected_date) == 10:   
                if selected_date:
                    # Convert selected_date to 'YYYY-MM-DD' format for consistency
                    selected_date_formatted = pd.to_datetime(selected_date, format='%d/%m/%Y').strftime('%d/%m/%Y')

                    # Get the current date in 'YYYY-MM-DD' format
                    current_date = datetime.datetime.now().strftime('%d-%m-%Y')

                    # Determine if the selected date is the current date
                    is_current_day = selected_date_formatted == current_date

                    # Update the query to use the selected date
                    if is_current_day:
                        # If the selected date is today, show only today's data
                        query_day = f"""
                            SELECT TO_CHAR(TO_DATE('{current_date}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                            FROM DUAL
                        """

                    else:
                        # If the selected date is not today, show only the selected day's data
                        query_day = f"""
                            SELECT TO_CHAR(TO_DATE('{selected_date_formatted}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                            FROM DUAL
                        """

            elif "to" in selected_date:

                if formatted_start_date and formatted_end_date:
                    # แปลงวันที่ให้เป็นรูปแบบ 'YYYY-MM-DD' สำหรับใช้ใน SQL
                    start_date_formatted = pd.to_datetime(formatted_start_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                    end_date_formatted = pd.to_datetime(formatted_end_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                    
                    print("Start Date:", formatted_start_date)
                    print("End Date:", formatted_end_date)
                    
                    # สร้างรายการวันที่ในช่วงที่เลือกโดยใช้ SQL
                    query_day = f"""
                        SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{start_date_formatted}', 'YY-MM-DD'), 'DD/MM/YY') AS DATA_DATE
                        FROM DUAL
                        CONNECT BY LEVEL <= TO_DATE('{end_date_formatted}', 'YY-MM-DD') - TO_DATE('{start_date_formatted}', 'YY-MM-DD') + 1
                    """

            elif "/" in selected_date and len(selected_date) == 7:

                selected_date = request.args.get("date_dropdown_daily")

                # Initialize df_month_list outside the if statement
                df_month_list = pd.DataFrame(columns=['DATA_DATE'])
                
                # Check if 'selected_date' is available
                if selected_date:
                    # Convert selected_date to 'YYYY-MM' format for consistency
                    selected_date_formatted = pd.to_datetime(selected_date, format='%m/%Y').strftime('%Y-%m')
                    # print("date2:", selected_date_formatted)

                    # Get the current month and year
                    current_month_year = datetime.datetime.now().strftime('%Y-%m')

                    # Determine if the selected date is in the current month
                    is_current_month = selected_date_formatted == current_month_year
                    
                    # Update the query to use the selected date
                    if is_current_month:
                        # If the selected date is in the current month, show all days up to the current date
                        query_day = f"""
                            SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{current_month_year}-01', 'YYYY-MM-DD')) AS Date1
                            FROM DUAL
                            CONNECT BY LEVEL <= TO_DATE('{datetime.datetime.now().strftime('%Y-%m-%d')}', 'YYYY-MM-DD') - TO_DATE('{current_month_year}-01', 'YYYY-MM-DD') + 1
                        """

                    else:
                        # If the selected date is in a previous month, show all days of the selected month
                        query_day = f"""
                            SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) AS Date1
                            FROM DUAL
                            CONNECT BY LEVEL <= LAST_DAY(TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) - TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD') + 1
                        """

            query_day_result = fetch_data(ptt_pivot_connection, query_day)
            df_month_list = pd.DataFrame(query_day_result, columns=['DATA_DATE'])

            # Convert the DATA_DATE to datetime format if not already
            df_month_list['DATA_DATE'] = pd.to_datetime(df_month_list['DATA_DATE'], errors='coerce', dayfirst=True)
            
            combined_data = []
            
            for amr_vc_type in amr_vc_type_list:
                amr_vc_type_table = f"""
                
                SELECT DESCRIPTION
                FROM (
                    SELECT 'Date' AS DESCRIPTION, 0 AS OR_DER 
                    FROM DUAL
                    UNION ALL
                    SELECT DESCRIPTION, OR_DER
                    FROM amr_mapping_config
                    WHERE evc_type LIKE '{amr_vc_type}'
                    AND DESCRIPTION IS NOT NULL
                )
                ORDER BY OR_DER

                """
                amr_vc_type_table_result = fetch_data(ptt_pivot_connection, amr_vc_type_table)
                filtered_result = pd.DataFrame(amr_vc_type_table_result)
                
                transposed_result = filtered_result.T
                ## HARDCODE
                # adjust column add dummy column to be date meter at column 0
                # because after this the column 0 will be dropped 
                transposed_result.insert(0,'','DATE_DUMMY')
                ## re-order column
                if amr_vc_type in ("5", "8", "9", "10"): 
                    col_to_move = transposed_result.pop(transposed_result.columns[3])  # นำคอลัมน์ที่ 2 ออก SG
                    transposed_result.insert(9, col_to_move.name, col_to_move)  # แทรกที่ index 10 (หลังคอลัมน์ที่ 9)

                    col_to_move = transposed_result.pop(transposed_result.columns[10])  # นำคอลัมน์ที่  ออก Date
                    transposed_result.insert(2, col_to_move.name, col_to_move)  # แทรกที่ index 3
                    # print(transposed_result.iloc[:, 11])
                    # Remove the 11th column (index 11) Battery Status
                    col_to_move = transposed_result.pop(transposed_result.columns[11])
                    # Add the column back at the end
                    transposed_result[col_to_move.name] = col_to_move
                
                elif amr_vc_type in ("13"):
                    transposed_result.drop(transposed_result.columns[2], axis=1, inplace=True) # blank column
                
                combined_data.append(transposed_result)
                # HARD #############################
            for meter_stream_no in meter_stream_no_list:
                
                df_meter_stream_no_table = df[df_configured['METER_STREAM_NO'] == meter_stream_no].drop(['PL_REGION_ID', 'TAG_ID', 'METER_ID', 'METER_STREAM_NO','amr_vc_type','vc_name'], axis=1)
                
                # แปลง 'DATA_DATE' ให้เป็น datetime
                df_meter_stream_no_table['DATA_DATE'] = pd.to_datetime(df_meter_stream_no_table['DATA_DATE'], errors='coerce', dayfirst=True)
                #df_meter_stream_no_table['DATA_DATE'] = pd.to_datetime(df_meter_stream_no_table['DATA_DATE'], format='%d-%m-%Y', errors='coerce')

                # เปลี่ยนชื่อคอลัมน์ 'DATA_DATE' เป็นชื่ออื่นเพื่อให้ไม่ซ้ำกับ df_month_list
                df_meter_stream_no_table.rename(columns={'DATA_DATE': 'DATA_DATE_METER'}, inplace=True)

                # รวม DataFrame บนคอลัมน์ 'DATA_DATE' ของ df_month_list และ 'DATA_DATE_METER' ของ df_meter_stream_no_table
                df_combined = pd.merge(df_month_list, df_meter_stream_no_table, left_on='DATA_DATE', right_on='DATA_DATE_METER', how='left')
                
                # ลบแถวที่มีค่าวันที่ซ้ำกัน โดยให้เหลือเพียงแถวแรกที่พบ
                df_combined = df_combined.drop_duplicates(subset='DATA_DATE')

                # แปลงรูปแบบวันที่ในคอลัมน์ 'DATA_DATE' และ 'DATA_DATE_METER'
                df_combined['DATA_DATE'] = df_combined['DATA_DATE'].dt.strftime('%d-%b-%Y')
                df_combined['DATA_DATE_METER'] = df_combined['DATA_DATE_METER'].dt.strftime('%d-%m-%Y')

                # ลบคอลัมน์ที่มีค่าเป็น None หรือ NaN ทั้งหมด
                df_combined.dropna(axis=1, how='all', inplace=True)

                # ลบแถวที่มีค่าเป็น None หรือ NaN ทั้งหมด (ถ้ามี)
                df_combined.dropna(axis=0, how='all', inplace=True)

                # swap between ActualDate-1> and MeterDate->0
                # because after this the column 0 will be dropped 
                df_combined = df_combined[df_combined.columns[[1, 0] + list(range(2, len(df_combined.columns)))]]
                df_combined = df_combined.fillna("N/A")

                ## HARDCODE
                amr_vc_type = amr_vc_type_list[int(meter_stream_no)-1]
                if amr_vc_type in ("5", "8", "9", "10"): 
                    col_to_move = df_combined.pop(df_combined.columns[3])  # นำคอลัมน์ที่ 2 ออก
                    df_combined.insert(9, col_to_move.name, col_to_move)  # แทรกที่ index 10 (หลังคอลัมน์ที่ 9)
                    
                    col_to_move = df_combined.pop(df_combined.columns[10])  # นำคอลัมน์ที่ 2 ออก
                    df_combined.insert(2, col_to_move.name, col_to_move)  # แทรกที่ index 3 
                    
                    columns = list(df_combined.columns)
                    columns.append(columns.pop(columns.index('AMR_CONFIG10')))
                    df_combined = df_combined[columns]

                    df_combined['AMR_CONFIG10'] = df_combined['AMR_CONFIG10'].apply(lambda x: 'Battery Low' if x != '0' else 'Normal')
                    #print(df_combined.columns)

                elif amr_vc_type in ("13"):        
                    df_combined.drop(df_combined.columns[2], axis=1, inplace=True)
                    #Alarm Battery EK280
                    df_combined['AMR_CONFIG20'] = df_combined['AMR_CONFIG20'].apply(lambda x: 'Battery Low' if x == '0' else 'Normal')
                ## HARD ######################

                combined_data.append(df_combined)

            def replace_thead_with_first_row(html_table):
                soup = BeautifulSoup(html_table, 'html.parser')
                
                # Extract the first row from tbody
                first_row = soup.tbody.find('tr')
                
                # Find the thead and replace its content with the first row
                soup.thead.clear()  # Remove existing content in thead
                soup.thead.append(first_row)  # Add the first row as the new thead content
                
                return str(soup)
            
            def style_first_row(html_table, color, font_size):
                """
                Modify the HTML table to make the first row bold, add a background color,
                increase font size, and set equal column widths.
                Additionally, hide the first column in all rows.
                """
                styled_table = html_table.replace(
                    '<tr>',
                    f'<tr style="font-weight: bold; background-color: {color}; font-size: {font_size}; text-align: center;">',
                    1
                ).replace(
                    '<td>',
                    '<td style="width: 100px; text-align: center;color: #000000;">'
                )
                
                # Add a <style> block to hide the first td in every tr
                styled_table += '''
                <style>
                    td:first-child { display: none; }
                    th:first-child { display: none; }
                </style>
                '''
                return styled_table
            
            background_color = '#f0f0f0'  
            font_size = 'larger'  

            def highlight_qm_qb(html_table, col_list):
                """ทำให้ค่าต่างจากแถวก่อนหน้าในคอลัมน์ที่กำหนดเป็นตัวหนา  และ Yellow highlight"""
                soup = BeautifulSoup(html_table, 'html.parser')

                # เก็บค่าของแถวก่อนหน้า
                prev_values = {col: None for col in col_list}  
                count = 0
                for row in soup.find_all('tr')[2:]:  # ข้าม header row and first row
                    cells = row.find_all('td')

                    # เช็คว่ามีคอลัมน์เพียงพอหรือไม่
                    if max(col_list) >= len(cells):
                        continue  # ข้ามแถวที่ไม่มีคอลัมน์เพียงพอ

                    current_values = {col: cells[col].text.strip() for col in col_list}
                    count+=1

                    if all(value != 'N/A' for value in current_values.values()):
                        for col in col_list:
                            
                            if prev_values[col] is not None and float(current_values[col]) != float(prev_values[col]):
                                cells[col].string = ""
                                cells[col].append(BeautifulSoup(
                                    f"<b style='background-color: yellow; color: black;'>{current_values[col]}</b>",
                                    "html.parser"
                                ))
                        prev_values = current_values.copy()
                    # อัปเดตค่าแถวก่อนหน้า
                    #prev_values = current_values.copy()
                return str(soup)
            
            def generate_html_tables(amr_vc_type_list, combined_data, df_month_list, background_color, font_size, classes=None):
                html_tables = []
                data_html = []
                if 'data_config' in session:
                    session.pop('data_config', None)

                for i in range(len(amr_vc_type_list)):
                    # Extract header and data DataFrames
                    header_row_df = combined_data[i]
                    data_df = combined_data[i + len(amr_vc_type_list)]
                    # Concatenate header_row_df and data_df if needed
                    if len(data_df.columns) != len(header_row_df.columns):
                        if len(data_df.columns) < len(header_row_df.columns):
                            for _ in range(len(header_row_df.columns) - len(data_df.columns)):
                                data_df[len(data_df.columns)] = ""
                        else:
                            for _ in range(len(data_df.columns) - len(header_row_df.columns)):
                                header_row_df[len(header_row_df.columns)] = ""
                    
                    # Ensure columns are aligned
                    data_df.columns = header_row_df.columns
                    
                    # Convert DataFrame to HTML with optional class
                    
                    html_content = pd.concat([header_row_df, data_df]).to_html(index=False, classes=classes)
                    
                    data_html.append(html_content)
                    # Style the first row of HTML table
                    html_table = style_first_row(html_content, background_color, font_size)
                    
                    header_row_df.columns = header_row_df.iloc[0]  # ใช้แถวแรกเป็นชื่อคอลัมน์

                    # HARD CODE
                    list_check = LIST_OF_CHANGE
                    header_list = []  # สร้างลิสต์เก็บ index ของคอลัมน์

                    # วนลูปเพื่อดึง index ของคอลัมน์จาก header_row_df
                    for j in list_check:
                        if j in header_row_df.columns:  # เช็คว่าคอลัมน์มีอยู่จริง
                            header_list.append(header_row_df.columns.get_loc(j))

                    # ถ้ามีคอลัมน์ครบตามที่กำหนด ให้เรียกฟังก์ชัน highlight_qm_qb
                    print("LLLL")
                    print(header_list)
                    if header_list:
                        html_table = highlight_qm_qb(html_table, header_list)

                    # Replace the thead with the first row
                    updated_html_table = replace_thead_with_first_row(html_table)
                    # Append the updated table to the list
                    html_tables.append(updated_html_table)
                
                with open('combined_data_report.json', 'w') as json_file:
                    json.dump(data_html, json_file, indent=4)
                return html_tables
            # session['data_config']=html_tables
            # Assuming `style_first_row` is defined elsewhere
            html_tables = generate_html_tables(amr_vc_type_list, combined_data, df_month_list, background_color, font_size, classes='data_config')
            html_tables_dict = {}
            vc_names_dict = {}

            # Assign values to html_tables_dict
            for i in range(len(html_tables)):
                if i < 6:  # Ensure we don't go beyond the number of expected tables
                    html_tables_dict[f'html_table{i+1}'] = html_tables[i]
                                                
            for i in range(len(amr_vc_type_list)):
                if i < 6:  # Ensure we don't go beyond the number of expected lists
                    vc_names_dict[f'vc_name_list{i+1}'] = globals().get(f'vc_name_list{i+1}', None)

            if "/" in selected_date and len(selected_date) == 10:   
                if selected_date:
                    # Convert selected_date to 'YYYY-MM-DD' format for consistency
                    selected_date_formatted = pd.to_datetime(selected_date, format='%d/%m/%Y').strftime('%d/%m/%Y')

                    # Get the current date in 'YYYY-MM-DD' format
                    current_date = datetime.datetime.now().strftime('%d-%m-%Y')

                    # Determine if the selected date is the current date
                    is_current_day = selected_date_formatted == current_date

                    # Update the query to use the selected date
                    if is_current_day:
                        # If the selected date is today, show only today's data
                        query_day = f"""
                            SELECT TO_CHAR(TO_DATE('{current_date}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                            FROM DUAL
                        """
                    else:
                        # If the selected date is not today, show only the selected day's data
                        query_day = f"""
                            SELECT TO_CHAR(TO_DATE('{selected_date_formatted}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                            FROM DUAL
                        """

            elif "to" in selected_date:

                if formatted_start_date and formatted_end_date:
                    # แปลงวันที่ให้เป็นรูปแบบ 'YYYY-MM-DD' สำหรับใช้ใน SQL
                    start_date_formatted = pd.to_datetime(formatted_start_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                    end_date_formatted = pd.to_datetime(formatted_end_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                    
                    # สร้างรายการวันที่ในช่วงที่เลือกโดยใช้ SQL
                    query_day = f"""
                        SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{start_date_formatted}', 'YY-MM-DD'), 'DD/MM/YY') AS DATA_DATE
                        FROM DUAL
                        CONNECT BY LEVEL <= TO_DATE('{end_date_formatted}', 'YY-MM-DD') - TO_DATE('{start_date_formatted}', 'YY-MM-DD') + 1
                    """

            elif "/" in selected_date and len(selected_date) == 7:

                selected_date = request.args.get("date_dropdown_daily")
                
                # Initialize df_month_list outside the if statement
                df_month_list = pd.DataFrame(columns=['DATA_DATE'])
                
                # Check if 'selected_date' is available
                if selected_date:
                    # Convert selected_date to 'YYYY-MM' format for consistency
                    selected_date_formatted = pd.to_datetime(selected_date, format='%m/%Y').strftime('%Y-%m')
                    # print("date2:", selected_date_formatted)

                    # Get the current month and year
                    current_month_year = datetime.datetime.now().strftime('%Y-%m')

                    # Determine if the selected date is in the current month
                    is_current_month = selected_date_formatted == current_month_year
                    
                    # Update the query to use the selected date
                    if is_current_month:
                        # If the selected date is in the current month, show all days up to the current date
                        query_day = f"""
                            SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{current_month_year}-01', 'YYYY-MM-DD')) AS Date1
                            FROM DUAL
                            CONNECT BY LEVEL <= TO_DATE('{datetime.datetime.now().strftime('%Y-%m-%d')}', 'YYYY-MM-DD') - TO_DATE('{current_month_year}-01', 'YYYY-MM-DD') + 1
                        """
                        
                    else:
                        # If the selected date is in a previous month, show all days of the selected month
                        query_day = f"""
                            SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) AS Date1
                            FROM DUAL
                            CONNECT BY LEVEL <= LAST_DAY(TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) - TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD') + 1
                        """
            # Fetch data for the month list
            query_day_result = fetch_data(ptt_pivot_connection, query_day)
            df_month_list = pd.DataFrame(query_day_result, columns=['DATA_DATE'])
                
            selected_meter_id = None

            # Check if 'METER_ID' column exists and the DataFrame is not empty
            if not df.empty and 'METER_ID' in df.columns and len(df['METER_ID']) > 0:
                selected_meter_id = df['METER_ID'].iloc[0]
                print(f"Selected Meter ID: {selected_meter_id}")
            else:
                print("DataFrame is empty or 'METER_ID' column doesn't exist.")

            # Now, remove the "METER_ID" column from the DataFrame
            df = df.drop(["PL_REGION_ID", "TAG_ID", "METER_ID"], axis=1)

            # Remove newline characters
            df = df.apply(lambda x: x.str.replace("\n", "") if x.dtype == "object" else x)

            df = df.drop_duplicates(subset=["DATA_DATE", "METER_STREAM_NO"], keep="first")
            # Sort DataFrame by 'DATA_DATE'
            df = df.sort_values(by="DATA_DATE")

            num_streams = 6
            df_runs = {}
            #print(df)
            # Loop to create DataFrames for each METER_STREAM_NO
            for i in range(1, num_streams + 1):
                df_runs[f'df_run{i}'] = df[df['METER_STREAM_NO'] == str(i)]

            # Check if each DataFrame has data before including in the tables dictionary
            # Create Full table of selectedMonth
            tables = {
                "daily_data": None,   
            }
    
            # Merge DataFrames using a loop
            for i in range(1, num_streams + 1):
                df_run = df_runs[f'df_run{i}']
                
                if not df_run.empty:
                    merged_df = pd.merge(df_month_list, df_run, on='DATA_DATE', how='outer')
                    df_runs[f'df_run{i}'] = merged_df

            for i in range(1, num_streams + 1):
                df_run = df_runs[f'df_run{i}']
                
                if not df_run.empty:
                    df_run = df_run.drop('METER_STREAM_NO', axis=1, errors='ignore')
                    df_run = df_run.fillna("N/A")
                    
                    tables[f"config_data_run{i}"] = df_run.to_html(classes='data_config',index=False, border=0,header=None,na_rep="N/A")
                    # print(tables)

            return render_template(
                "billingdata.html",
                tables=tables,
                titles=df.columns.values,
                selected_date=selected_date,
                selected_tag=selected_tag,
                selected_region=selected_region,
                region_options=region_options,
                tag_options=tag_options,
                **{k: html_tables_dict.get(k, None) for k in [f'html_table{i+1}' for i in range(6)]},
                vc_name_dict=vc_name_dict,
                selected_meter_id=selected_meter_id,
                username=username
            )
####################################hourly_data

        elif query_type == "hourly_data":
            if "/" in selected_date and len(selected_date) == 10:
                query = f"""
                SELECT DISTINCT
                    AMR_PL_GROUP.PL_REGION_ID,
                    AMR_FIELD_ID.TAG_ID,
                    AMR_FIELD_ID.METER_ID,
                    TO_CHAR(AMR_BILLING_HOURLY_DATA.DATA_DATE),
                    AMR_BILLING_HOURLY_DATA.data_hour as Hour,
                    AMR_BILLING_HOURLY_DATA.CORRECTED_VOL as CORRECTED,
                AMR_BILLING_HOURLY_DATA.UNCORRECTED_VOL as UNCORRECTED,
                AMR_BILLING_HOURLY_DATA.AVR_PF as Pressure,
                    AMR_BILLING_HOURLY_DATA.AVR_TF as Temperature,
                    AMR_BILLING_HOURLY_DATA.METER_STREAM_NO as METER_STREAM_NO
                    
                FROM
                    AMR_FIELD_ID, AMR_PL_group, AMR_BILLING_HOURLY_DATA,amr_region
                WHERE
                    AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                    AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                    AND AMR_BILLING_HOURLY_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    AND AMR_BILLING_HOURLY_DATA.METER_STREAM_NO IS NOT NULL
                  AND TO_CHAR(AMR_BILLING_HOURLY_DATA.DATA_DATE, 'DD/MM/YYYY') = '{selected_date}'
                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                    AND amr_region.REGION_NAME = '{selected_region}'
                ORDER BY data_hour
                """
            elif "to" in selected_date:
               
                query = f"""
                SELECT DISTINCT
                    AMR_PL_GROUP.PL_REGION_ID,
                    AMR_FIELD_ID.TAG_ID,
                    AMR_FIELD_ID.METER_ID,
                    TO_CHAR(AMR_BILLING_HOURLY_DATA.DATA_DATE),
                    AMR_BILLING_HOURLY_DATA.data_hour as Hour,
                    AMR_BILLING_HOURLY_DATA.CORRECTED_VOL as CORRECTED,
                AMR_BILLING_HOURLY_DATA.UNCORRECTED_VOL as UNCORRECTED,
                AMR_BILLING_HOURLY_DATA.AVR_PF as Pressure,
                    AMR_BILLING_HOURLY_DATA.AVR_TF as Temperature,
                    AMR_BILLING_HOURLY_DATA.METER_STREAM_NO as METER_STREAM_NO
                    
                FROM
                    AMR_FIELD_ID, AMR_PL_group, AMR_BILLING_HOURLY_DATA,amr_region
                WHERE
                    AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                    AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                    AND AMR_BILLING_HOURLY_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    AND AMR_BILLING_HOURLY_DATA.METER_STREAM_NO IS NOT NULL
                 AND AMR_BILLING_HOURLY_DATA.DATA_DATE BETWEEN TO_DATE('{formatted_start_date}', 'DD/MM/YYYY') AND TO_DATE('{formatted_end_date}', 'DD/MM/YYYY')
                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                    AND amr_region.REGION_NAME = '{selected_region}'
                ORDER BY data_hour
                """
            elif "/" in selected_date and len(selected_date) == 7:
                query = f"""
                SELECT DISTINCT
                    AMR_PL_GROUP.PL_REGION_ID,
                    AMR_FIELD_ID.TAG_ID,
                    AMR_FIELD_ID.METER_ID,
                    TO_CHAR(AMR_BILLING_HOURLY_DATA.DATA_DATE),
                    AMR_BILLING_HOURLY_DATA.data_hour as Hour,
                    AMR_BILLING_HOURLY_DATA.CORRECTED_VOL as CORRECTED,
                AMR_BILLING_HOURLY_DATA.UNCORRECTED_VOL as UNCORRECTED,
                AMR_BILLING_HOURLY_DATA.AVR_PF as Pressure,
                    AMR_BILLING_HOURLY_DATA.AVR_TF as Temperature,
                    AMR_BILLING_HOURLY_DATA.METER_STREAM_NO as METER_STREAM_NO
                    
                FROM
                    AMR_FIELD_ID, AMR_PL_group, AMR_BILLING_HOURLY_DATA,amr_region
                WHERE
                    AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                    AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                    AND AMR_BILLING_HOURLY_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    AND AMR_BILLING_HOURLY_DATA.METER_STREAM_NO IS NOT NULL
                    AND TO_CHAR(AMR_BILLING_HOURLY_DATA.DATA_DATE, 'MM/YYYY') = '{selected_date}'
                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                    AND amr_region.REGION_NAME = '{selected_region}'
                ORDER BY data_hour
                """
            results = fetch_data(ptt_pivot_connection, query)
            df = pd.DataFrame(
                results,
                columns=[
                    "PL_REGION_ID",
                    "TAG_ID",
                    "METER_ID",
                    "DATA_DATE",
                    "Hour",
                    "CORRECTED",
                    "UNCORRECTED",
                    "Pressure",
                    "Temperature",
                    "METER_STREAM_NO", 
                ],
            )
            df["Pressure"] = df["Pressure"].astype(float)
            df["Temperature"] = df["Temperature"].astype(float)

            df["Pressure"] = df["Pressure"].round(4)
            df["Temperature"] = df["Temperature"].round(4)
            df["CORRECTED"] = pd.to_numeric(df["CORRECTED"], errors='coerce').fillna(0).astype(int).astype(str)
            df["UNCORRECTED"] = pd.to_numeric(df["UNCORRECTED"], errors='coerce').fillna(0).astype(int).astype(str)

            df["METER_STREAM_NO"] = df["METER_STREAM_NO"].astype(int)
            df["Hour"] = pd.to_numeric(df["Hour"], errors='coerce').fillna(0).astype(int).astype(str)
            
              # selected_meter_id = df["METER_ID"].iloc[0]
            # print(df)
            df["Date_or_Hour"] = df["DATA_DATE"].astype(str) + ' or ' + df["Hour"].astype(str)


            # Select the columns with the combined Date_Hour and the rest of the columns in their original order
            df_selected = df[["Date_or_Hour", "CORRECTED", "UNCORRECTED", "Pressure", "Temperature", "METER_STREAM_NO"]]
            # print(df)
            # Initialize the tables dictionary
            meter_id_list = df["METER_ID"].tolist()
      
            if not meter_id_list:
                # If empty, set the vc_name_dict with default values
                vc_name_dict = {f'vc_name_list{i}': None for i in range(1, 7)}
            else:

                
                query_vc_name = f"""SELECT amr_vc_type.vc_name,amr_field_meter.METER_STREAM_NO from amr_vc_type ,amr_field_meter 
                                    WHERE amr_field_meter.METER_STREAM_TYPE = amr_vc_type.id 
                                    AND amr_field_meter.meter_id like '{meter_id_list[0]}'
                                    order by amr_field_meter.METER_STREAM_NO """
                                    
                                    
                results_vc_name = fetch_data(ptt_pivot_connection, query_vc_name)
                df_vc_name = pd.DataFrame(
                    results_vc_name,
                    columns=[
                        "vc_name",
                        "METER_STREAM_NO",
                        
                    
                    ]
                )
                
                unique_meter_streams = df_vc_name.drop_duplicates(subset=["METER_STREAM_NO"])[["METER_STREAM_NO","vc_name"]]

                meter_stream_no_list = unique_meter_streams["METER_STREAM_NO"].tolist()
            
                vc_name_list = unique_meter_streams["vc_name"].tolist()


                vc_name_dict = {}
                for i in range(1, 7):
                    if len(vc_name_list) >= i:
                        vc_name_dict[f'vc_name_list{i}'] = vc_name_list[i-1]
                    else:
                        vc_name_dict[f'vc_name_list{i}'] = None
                print(vc_name_dict)

                
                query_type = f"""
                    SELECT DISTINCT  amr_vc_type.id as METER_ID , amr_field_meter.METER_STREAM_NO FROM amr_billing_hourly_data,amr_field_meter,amr_vc_type 
                    WHERE amr_billing_hourly_data.meter_id = amr_field_meter.meter_id 
                    AND amr_field_meter.METER_STREAM_TYPE = amr_vc_type.id
                    AND amr_billing_hourly_data.meter_id like '{meter_id_list[0]}'
                    AND  amr_field_meter.METER_STREAM_NO IS NOT NULL
                    ORDER BY amr_field_meter.METER_STREAM_NO """
                    
                results_type = fetch_data(ptt_pivot_connection, query_type)
                df_type = pd.DataFrame(
                    results_type,
                    columns=[
                    
                        "METER_ID",
                        "METER_STREAM_NO",
        
                    ]
                )
                meter_stream_no_list = df_type["METER_STREAM_NO"].tolist()
                amr_vc_type_list = df_type["METER_ID"].tolist()

            data_type = []
            if not meter_id_list:
                pass
            else :
                for amr_vc_type in amr_vc_type_list:
                    amr_vc_type_table = f"""
                    
                    
                    SELECT DESCRIPTION 
                    FROM amr_mapping_hourly 
                    WHERE evc_type LIKE '{amr_vc_type}' 
                    AND hourly like '1' 
                    ORDER BY OR_DER


                    """
                    
                    amr_vc_type_table_result = fetch_data(ptt_pivot_connection, amr_vc_type_table)
                    filtered_result = pd.DataFrame(amr_vc_type_table_result)
                    
            
                    transposed_result = filtered_result.T
                    
                    
                    selected_data = transposed_result.iloc[0, 0:5]  # Extract values to get a flat array
                    data_type.append(selected_data)
           
            # print("data_type_hourly",data_type)
            tables = {}
            html_style = '''
                <style>
                    .data_daily {
                        border-collapse: collapse;
                        width: 100%;
                    }
                    .data_daily th, .data_daily td {
                        border: 1px solid #ddd;
                        padding: 8px;
                        color: #000000;
                    }
                    .data_daily th {
                        padding-top: 12px;
                        padding-bottom: 12px;
                        text-align: left;
                        background-color: #f0f0f0;
                    }
                  
                </style>
            '''

            meter_streams = [1, 2, 3, 4, 5, 6]
            data_hourly = []

            # Loop through each stream number and process data
            for stream_no in meter_streams:
                df_run = df_selected[df_selected["METER_STREAM_NO"] == stream_no]
                df_run = df_run.drop('METER_STREAM_NO', axis=1, errors='ignore')
                if not df_run.empty:
                    data_hourly.append(df_run.to_dict())
                    df_run_no_last = df_run
                    hedder = data_type[stream_no - 1]  # Assuming `data_type` is indexed to match stream numbers
                    header_df = pd.DataFrame(hedder).T

                    # Create the HTML table with both data_daily and data_hourly classes
                    html_output = '<table class="data_hourly">'
                    
                    # Extract the first row to use as the header
                    header_row = header_df.iloc[0]
                    html_output += '<thead><tr>'
                    for item in header_row:
                        html_output += f'<th style="text-align: center;">{item}</th>'
                    html_output += '</tr></thead>'
                    
                    # Create the tbody
                    html_output += '<tbody>'
                    for index, row in df_run_no_last.iterrows():
                        html_output += '<tr>'
                        for item in row:
                            html_output += f'<td>{item}</td>'
                        html_output += '</tr>'
                    html_output += '</tbody>'
                    
                    html_output += '</table>'
                    
                    # Store the constructed HTML in the tables dictionary
                    tables[f"daily_data_run{stream_no}"] = html_style + html_output
            
           
            with open('combined_data_hourly_report.json', 'w') as json_file:
                json.dump(data_hourly, json_file, indent=4)
            
          
            return render_template(
                "billingdata.html",
                tables=tables,
                titles=df.columns.values,
                selected_date=selected_date,
                selected_tag=selected_tag,
                selected_region=selected_region,
                region_options=region_options,
                tag_options=tag_options,
                vc_name_dict=vc_name_dict,
                username=username

                )

    return render_template(
                "billingdata.html",
                selected_date=selected_date,
                selected_region=selected_region,
                selected_tag=selected_tag,
                region_options=region_options,
                tag_options=tag_options,
                tables={},username=username
            )
       

@app.route("/billing_show_graph")
@login_required
def billing_show_graph():
    username = session.get('username')
    print("username",username)
    
    if 'username' not in session:
        
        return redirect(url_for('login'))
    
    
    # # Restore df from session GGGG
    # if "df_runs" not in session:
    #     return "No data available", 400
    # df_runs = {key: pd.read_json(value) for key, value in json.loads(session["df_runs"]).items()}    
    # site_run = 0
    # for i in range(1, 7):
    #     #df_run = df_runs[f'df_run{i}']
    #     if not df_runs[f'df_run{i}'].empty:
    #         site_run+=1


    # connect_to_amr_db
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        
        # SQL query to fetch unique PL_REGION_ID values
        region_query = """
        SELECT * FROM AMR_REGION 
        """

        tag_query = """
        
        SELECT DISTINCT AMR_FIELD_ID.TAG_ID
        FROM AMR_FIELD_ID,amr_region,AMR_PL_GROUP
        
        WHERE AMR_PL_GROUP.PL_REGION_ID = amr_region.id
        AND AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
        AND amr_region.REGION_NAME = :region_id
        AND AMR_FIELD_ID.tag_id NOT like '%.remove%'
        ORDER BY  TAG_ID
        """

        # Fetch unique region values
        region_results = fetch_data(ptt_pivot_connection, region_query)
        region_options = [str(region[1]) for region in region_results]
        
        selected_date = session['selected_date']
        selected_tag = session['selected_tag'] 
        selected_region = session['selected_region']
        selected_day = session['selected_day'] 
        print("selected_date:",selected_date)
        print("selected_tag:",selected_tag)
        print("selected_region:",selected_region)
        print("selected_day:",selected_day)
        
        if selected_date:
            # Check if the selected_date is in the format of a single day (dd/mm/yyyy)
            if "/" in selected_date and len(selected_date) == 10:
                # Example: 25/08/2024
                # Enter the first if block
                # Your code for handling a single date goes here
                pass

            # Check if the selected_date is a range of days (dd/mm/yyyy to dd/mm/yyyy)
            elif "to" in selected_date:
                # Example: 25/08/2024 to 30/08/2024
                # Enter the else block
                # Your code for handling a date range goes here
                start_date_str, end_date_str = selected_date.split(" to ")

        # Convert the string dates to datetime objects
                start_date = datetime.datetime.strptime(start_date_str, '%d/%m/%Y')
                end_date = datetime.datetime.strptime(end_date_str, '%d/%m/%Y')

                formatted_start_date = start_date.strftime('%d/%m/%Y')
                formatted_end_date = end_date.strftime('%d/%m/%Y')

            # Check if the selected_date is a month (mm/yyyy)
            elif "/" in selected_date and len(selected_date) == 7:
                # Example: 08/2024
                pass
        
        tag_results = fetch_data(ptt_pivot_connection, tag_query, params={"region_id": selected_region})
        tag_options = [str(tag[0]) for tag in tag_results]
        
        query = ""
      
        # SQL query for main data
        if "/" in selected_date and len(selected_date) == 10:

            query = f"""
                    SELECT DISTINCT
                AMR_PL_GROUP.PL_REGION_ID,
                AMR_FIELD_ID.TAG_ID,
                AMR_FIELD_ID.METER_ID,
                TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD-MM-YYYY')  AS DATA_DATE,
                
                AMR_BILLING_DATA.CORRECTED_VOL as CORRECTED,
                AMR_BILLING_DATA.UNCORRECTED_VOL as UNCORRECTED,
                AMR_BILLING_DATA.AVR_PF as Pressure,
                AMR_BILLING_DATA.AVR_TF as Temperature,
                AMR_BILLING_DATA.METER_STREAM_NO  
            FROM
                AMR_FIELD_ID, AMR_PL_group, AMR_BILLING_DATA,amr_region
            WHERE
                AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                AND AMR_BILLING_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                AND AMR_BILLING_DATA.METER_STREAM_NO IS NOT NULL
                AND TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD/MM/YYYY') = '{selected_date}'
                AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                AND amr_region.REGION_NAME = '{selected_region}'
            ORDER BY
                TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD-MM-YYYY') DESC
                """
            
        elif "to" in selected_date:
            query = f"""
                    SELECT DISTINCT
                    AMR_PL_GROUP.PL_REGION_ID,
                    AMR_FIELD_ID.TAG_ID,
                    AMR_FIELD_ID.METER_ID,
                    TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD/MM/YY') AS DATA_DATE,
                    AMR_BILLING_DATA.CORRECTED_VOL AS CORRECTED,
                    AMR_BILLING_DATA.UNCORRECTED_VOL AS UNCORRECTED,
                    AMR_BILLING_DATA.AVR_PF AS Pressure,
                    AMR_BILLING_DATA.AVR_TF AS Temperature,
                    AMR_BILLING_DATA.METER_STREAM_NO  
                FROM
                    AMR_FIELD_ID
                JOIN AMR_PL_GROUP 
                    ON AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
                JOIN AMR_BILLING_DATA 
                    ON AMR_BILLING_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                JOIN AMR_REGION 
                    ON AMR_PL_GROUP.PL_REGION_ID = AMR_REGION.ID
                WHERE
                    AMR_BILLING_DATA.METER_STREAM_NO IS NOT NULL
                    AND AMR_BILLING_DATA.DATA_DATE BETWEEN TO_DATE('{formatted_start_date}', 'DD/MM/YYYY') 
                                                        AND TO_DATE('{formatted_end_date}', 'DD/MM/YYYY')
                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                    AND AMR_REGION.REGION_NAME = '{selected_region}'
                ORDER BY
                    TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD/MM/YY') DESC
                """
            
        elif "/" in selected_date and len(selected_date) == 7:

            print("One or both dates are missing.")
            # Your code here for missing dates
            # Fetch tag options based on the selected region
            query = f"""
            SELECT DISTINCT
                AMR_PL_GROUP.PL_REGION_ID,
                AMR_FIELD_ID.TAG_ID,
                AMR_FIELD_ID.METER_ID,
                TO_CHAR(AMR_BILLING_DATA.DATA_DATE)  AS DATA_DATE,
                
                AMR_BILLING_DATA.CORRECTED_VOL as CORRECTED,
                AMR_BILLING_DATA.UNCORRECTED_VOL as UNCORRECTED,
                AMR_BILLING_DATA.AVR_PF as Pressure,
                AMR_BILLING_DATA.AVR_TF as Temperature,
                AMR_BILLING_DATA.METER_STREAM_NO  
            FROM
                AMR_FIELD_ID, AMR_PL_group, AMR_BILLING_DATA,amr_region
            WHERE
                AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                AND AMR_BILLING_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                AND AMR_BILLING_DATA.METER_STREAM_NO IS NOT NULL
                AND TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'MM/YYYY') = '{selected_date}'
                AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                AND amr_region.REGION_NAME = '{selected_region}'
            ORDER BY
                TO_CHAR(AMR_BILLING_DATA.DATA_DATE) DESC
            """
        results = fetch_data(ptt_pivot_connection, query)
        
        df = pd.DataFrame(
            results,
            columns=[
                "PL_REGION_ID",
                "TAG_ID",
                "METER_ID",
                "DATA_DATE",
                "CORRECTED",
                "UNCORRECTED",
                "Pressure",
                "Temperature",
                "METER_STREAM_NO",
            ],
        )
        df["Pressure"] = df["Pressure"].astype(float)
        df["Temperature"] = df["Temperature"].astype(float)

        df["Pressure"] = df["Pressure"].round(4)
        df["Temperature"] = df["Temperature"].round(4)
        df["CORRECTED"] = pd.to_numeric(df["CORRECTED"], errors='coerce').fillna(0).astype(int).astype(str)
        df["UNCORRECTED"] = pd.to_numeric(df["UNCORRECTED"], errors='coerce').fillna(0).astype(int).astype(str)



        df["METER_STREAM_NO"] = df["METER_STREAM_NO"].astype(int)
        meter_id_list = df["METER_ID"].tolist()
        
        
        if not meter_id_list:
            # If empty, set the vc_name_dict with default values
            vc_name_dict = {f'vc_name_list{i}': None for i in range(1, 7)}
        else:
            query_vc_name = f"""SELECT amr_vc_type.vc_name, amr_field_meter.METER_STREAM_NO 
                                FROM amr_vc_type, amr_field_meter 
                                WHERE amr_field_meter.METER_STREAM_TYPE = amr_vc_type.id 
                                AND amr_field_meter.meter_id LIKE '{meter_id_list[0]}' 
                                ORDER BY amr_field_meter.METER_STREAM_NO"""
            
            results_vc_name = fetch_data(ptt_pivot_connection, query_vc_name)
            df_vc_name = pd.DataFrame(
                results_vc_name,
                columns=["vc_name", "METER_STREAM_NO"]
            )

            unique_meter_streams = df_vc_name.drop_duplicates(subset=["METER_STREAM_NO"])[["METER_STREAM_NO", "vc_name"]]

            meter_stream_no_list = unique_meter_streams["METER_STREAM_NO"].tolist()
            vc_name_list = unique_meter_streams["vc_name"].tolist()

            vc_name_dict = {}
            for i in range(1, 7):
                if len(vc_name_list) >= i:
                    vc_name_dict[f'vc_name_list{i}'] = vc_name_list[i - 1]
                else:
                    vc_name_dict[f'vc_name_list{i}'] = None  # or "nong", depending on your preference
                    
            query_type = f"""
                SELECT DISTINCT  amr_vc_type.id as METER_ID , amr_field_meter.METER_STREAM_NO FROM amr_billing_data,amr_field_meter,amr_vc_type 
                WHERE amr_billing_data.meter_id = amr_field_meter.meter_id 
                AND amr_field_meter.METER_STREAM_TYPE = amr_vc_type.id
                AND amr_billing_data.meter_id like '{meter_id_list[0]}'
                AND  amr_field_meter.METER_STREAM_NO IS NOT NULL
                ORDER BY amr_field_meter.METER_STREAM_NO """
                
            results_type = fetch_data(ptt_pivot_connection, query_type)
            df_type = pd.DataFrame(
                results_type,
                columns=[
                    "METER_ID",
                    "METER_STREAM_NO",
                ]
            )
            meter_stream_no_list = df_type["METER_STREAM_NO"].tolist()
            amr_vc_type_list = df_type["METER_ID"].tolist()
        
        
        # Get the selected Meter ID before removing it from the DataFrame
        # selected_meter_id = df["METER_ID"].iloc[0]
        selected_meter_id = None

        # Check if 'METER_ID' column exists and the DataFrame is not empty
        if not df.empty and 'METER_ID' in df.columns and len(df['METER_ID']) > 0:
            selected_meter_id = df['METER_ID'].iloc[0]
            print(f"Selected Meter ID: {selected_meter_id}")
        else:
            print("DataFrame is empty or 'METER_ID' column doesn't exist.")

        # Now, remove the "METER_ID" column from the DataFrame
        df = df.drop(["PL_REGION_ID", "TAG_ID", "METER_ID"], axis=1)
        
        # Remove newline characters
        df = df.apply(lambda x: x.str.replace("\n", "") if x.dtype == "object" else x)

        df = df.drop_duplicates(subset=["DATA_DATE", "METER_STREAM_NO"], keep="first")
        # Sort DataFrame by 'DATA_DATE'
        df = df.sort_values(by="DATA_DATE")
        # print("df",df)
        # Assuming 'df' is the DataFrame created from the query results
        num_streams = 6
        df_runs = {}

        # Loop to create DataFrames for each METER_STREAM_NO
        for i in range(1, num_streams + 1):
            df_runs[f'df_run{i}'] = df[df['METER_STREAM_NO'] == int(i)]
            
        # Check if each DataFrame has data before including in the tables dictionary
        tables = {
            "config_data": None,
        }

        graphs = {
            "corrected": None,
            "uncorrected": None,
            "pressure": None,
            "temperature": None
        }
            
        if "/" in selected_date and len(selected_date) == 10:
            # print("selected_date:",selected_date)
            if selected_date:
                # Convert selected_date to 'YYYY-MM-DD' format for consistency
                selected_date_formatted = pd.to_datetime(selected_date, format='%d/%m/%Y').strftime('%d/%m/%Y')

                # Get the current date in 'YYYY-MM-DD' format
                current_date = datetime.datetime.now().strftime('%d-%m-%Y')

                # Determine if the selected date is the current date
                is_current_day = selected_date_formatted == current_date

                # Update the query to use the selected date
                if is_current_day:
                    # If the selected date is today, show only today's data
                    query_day = f"""
                        SELECT TO_CHAR(TO_DATE('{current_date}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                        FROM DUAL
                    """
                else:
                    # If the selected date is not today, show only the selected day's data
                    query_day = f"""
                        SELECT TO_CHAR(TO_DATE('{selected_date_formatted}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                        FROM DUAL
                    """

        elif "to" in selected_date:
            if formatted_start_date and formatted_end_date:
                # แปลงวันที่ให้เป็นรูปแบบ 'YYYY-MM-DD' สำหรับใช้ใน SQL
                start_date_formatted = pd.to_datetime(formatted_start_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                end_date_formatted = pd.to_datetime(formatted_end_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                
                print("Start Date:", formatted_start_date)
                print("End Date:", formatted_end_date)
                
                # สร้างรายการวันที่ในช่วงที่เลือกโดยใช้ SQL
                query_day = f"""
                    SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{start_date_formatted}', 'YY-MM-DD'), 'DD/MM/YY') AS DATA_DATE
                    FROM DUAL
                    CONNECT BY LEVEL <= TO_DATE('{end_date_formatted}', 'YY-MM-DD') - TO_DATE('{start_date_formatted}', 'YY-MM-DD') + 1
                """

        elif "/" in selected_date and len(selected_date) == 7:
            df_month_list = pd.DataFrame(columns=['DATA_DATE'])
            
            # Check if 'selected_date' is available
            if selected_date:
                # Convert selected_date to 'YYYY-MM' format for consistency
                selected_date_formatted = pd.to_datetime(selected_date, format='%m/%Y').strftime('%Y-%m')
                # print("date2:", selected_date_formatted)

                # Get the current month and year
                current_month_year = datetime.datetime.now().strftime('%Y-%m')

                # Determine if the selected date is in the current month
                is_current_month = selected_date_formatted == current_month_year
                
                # Update the query to use the selected date
                if is_current_month:
                    # If the selected date is in the current month, show all days up to the current date
                    query_day = f"""
                        SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{current_month_year}-01', 'YYYY-MM-DD')) AS Date1
                        FROM DUAL
                        CONNECT BY LEVEL <= TO_DATE('{datetime.datetime.now().strftime('%Y-%m-%d')}', 'YYYY-MM-DD') - TO_DATE('{current_month_year}-01', 'YYYY-MM-DD') + 1
                    """

                else:
                    # If the selected date is in a previous month, show all days of the selected month
                    query_day = f"""
                        SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) AS Date1
                        FROM DUAL
                        CONNECT BY LEVEL <= LAST_DAY(TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) - TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD') + 1
                    """

        query_day_result = fetch_data(ptt_pivot_connection, query_day)
        df_month_list = pd.DataFrame(query_day_result, columns=['DATA_DATE'])
        # print("df_month_list",df_month_list)
        def make_link(row, run_number):
            if row['CORRECTED'] == 'N/A' and row['UNCORRECTED'] == 'N/A':
                return f'<a href="#" onclick="sendDataDate(\'{row["DATA_DATE"]}\', {run_number}); openPopup(); return false;">{row["DATA_DATE"]}</a>'
                
            return row['DATA_DATE']
        
        for i in range(1, num_streams + 1):
            df_run = df_runs[f'df_run{i}']
            ### HHHHH html
            if not df_run.empty:
                merged_df = pd.merge(df_month_list, df_run, on='DATA_DATE', how='outer')
                df_runs[f'df_run{i}'] = merged_df
                html_style = """
                        <style>
                        .data_daily th, .data_daily td .data_daily tr{
                            text-align: center; 
                        }
                        </style>
                        """

        html_style = '''
            <style>
                .data_daily {
                    border-collapse: collapse;
                    width: 100%;
                }
                .data_daily th, .data_daily td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    color: #000000;
                }
                .data_daily th {
                    padding-top: 12px;
                    padding-bottom: 12px;
                    text-align: left;
                    background-color: #f0f0f0;
                    
                }
            </style>
            ''' 
            
        data_type = []
        if not meter_id_list:
            pass
        else :
            for amr_vc_type in amr_vc_type_list:
                amr_vc_type_table = f"""
                
                
                SELECT DESCRIPTION 
                FROM amr_mapping_billing 
                WHERE evc_type LIKE '{amr_vc_type}' 
                AND DAILY like '1' 
                ORDER BY OR_DER
                """                
                amr_vc_type_table_result = fetch_data(ptt_pivot_connection, amr_vc_type_table)
                filtered_result = pd.DataFrame(amr_vc_type_table_result)
                transposed_result = filtered_result.T
                selected_data = transposed_result.iloc[0, 0:5]  # Extract values to get a flat array
                data_type.append(selected_data)
        # Print the values in a horizontal format
        # for i in range(4):
        #     print(data_type[i])
                

        df_graphs = []  # สร้างลิสต์เพื่อเก็บข้อมูล
        combined_data = []
        all_totals = []  # สร้างลิสต์เพื่อเก็บค่า totals ของแต่ละรอบ

        for i in range(num_streams):
            df_run = df_runs.get(f'df_run{i+1}', pd.DataFrame())

            df_graphs_run = pd.DataFrame(df_run)
            df_graphs.append(df_graphs_run)  # Store DataFrame in the list

            if not df_run.empty:
                df_run = df_run.drop('METER_STREAM_NO', axis=1, errors='ignore')
                df_run = df_run.fillna("N/A")
                combined_data.append(df_run.to_dict())

                df_run['DATA_DATE'] = df_run.apply(lambda row: make_link(row, i+1), axis=1)

                # Prepare for constructing HTML
                df_run_no_last = df_run
                hedder = data_type[i]
                header_df = pd.DataFrame(hedder).T

                # Create the HTML manually as before
                html_output = '<table class="data_daily">'

                # Extract the first row to use as header
                header_row = header_df.iloc[0]
                html_output += '<thead><tr>'
                for item in header_row:
                    html_output += f'<th style="text-align: center;">{item}</th>'
                html_output += '</tr></thead>'

                # Now create the tbody
                html_output += '<tbody>'
                for index, row in df_run_no_last.iloc[0:].iterrows():  # Skip the first row
                    html_output += '<tr>'
                    for item in row:
                        html_output += f'<td>{item}</td>'
                    html_output += '</tr>'

                # Add the total row (calculated from the daily differences)
                html_output += '<tr><td style="text-align: center; background-color: #1cc88a33;"">Total</td>'
                
                                    # Convert 'CORRECTED' and 'UNCORRECTED' to numeric
                df_run_no_last['CORRECTED'] = pd.to_numeric(df_run_no_last['CORRECTED'], errors='coerce')
                df_run_no_last['UNCORRECTED'] = pd.to_numeric(df_run_no_last['UNCORRECTED'], errors='coerce')

                # Forward fill missing values to ensure calculation continuity
                df_run_no_last[['CORRECTED', 'UNCORRECTED']] = df_run_no_last[['CORRECTED', 'UNCORRECTED']].ffill()

                # Calculate the difference for only 'CORRECTED' and 'UNCORRECTED' columns after filling missing values
                df_run_diff = df_run_no_last[['CORRECTED', 'UNCORRECTED']].diff().fillna(0)

                # Filter out zero values
                df_run_diff_filtered = df_run_diff[df_run_diff != 0]

                # Now calculate the totals for the non-zero values (daily difference totals)
                totals = df_run_diff_filtered.sum()

                # Add this round's totals to all_totals list
                all_totals.append(totals)
                
                
                for col in df_run_no_last.columns[1:]:  # Skip the first column (assumed non-numeric)
                    if col in totals:
                        total_value = int(totals[col])
                        html_output += f'<td  style="background-color: #1cc88a33;">{total_value}</td>'
                    # else:
                    #     html_output += '<td ></td>'  # Leave empty for non-numeric columns
                html_output += '</tr>'

                html_output += '</tbody>'
                html_output += '</table>'

                # Store the constructed HTML in the tables dictionary
                tables[f"daily_data_run{i+1}"] = html_style + html_output
                
        # After looping through all runs, sum all totals
        final_totals = pd.DataFrame(all_totals).sum()

        html_final_totals = '<table class="data_totals">'
        html_final_totals += '<thead><tr><th>Daily Corrected Volume</th><th>Daily Uncorrected Volume</th></tr></thead>'
        html_final_totals += '<tbody><tr>'  # Start the body and first row

        # Iterate over the final_totals index to populate the total values
        for col in final_totals.index:
            html_final_totals += f'<td>{int(final_totals[col])}</td>'
        html_final_totals += '</tr></tbody></table>'

        # Store the final combined totals table
        tables["combined_totals"] = html_style + html_final_totals
        with open('combined_data_report.json', 'w') as json_file:
            json.dump(combined_data, json_file, indent=4)
    
        graphs_corrected = []
        daily_sum_corrected = pd.DataFrame()   
        graphs_uncorrected = []
     
        daily_sum_uncorrected = pd.DataFrame()
        fig_pressure = go.Figure()
        fig_temperature = go.Figure()
        daily_sum = pd.DataFrame()
        
        daily_diff_corrected_data=[]
        colors = ['blue', 'orange', 'green', 'purple', 'yellow']
        
        for j in range(num_streams):
            test_df = df_graphs[j]
            df_data_date = test_df['DATA_DATE']
            df_corrected = pd.to_numeric(test_df['CORRECTED'], errors='coerce')  # แปลงให้เป็นตัวเลข
            df_valid = test_df[~df_corrected.isna()]  # แทนค่า NaN ด้วย 0
            
            # คำนวณความต่างรายวัน
            daily_diff_corrected = df_corrected.diff()
        
            if not daily_diff_corrected.empty:
                daily_diff_corrected_data.append(daily_diff_corrected.to_dict())

            if not df_corrected.isna().all():
                # trace_nan_corrected = go.Scatter(
                #     x=df_data_date,
                #     y=daily_diff_corrected,
                #     mode="lines",
                #     name=f"Run {j+1} - N/A",
                #     line=dict(color="red", width=2),
                #     connectgaps=True
                # )
                trace_corrected = go.Scatter(
                    x=df_data_date,
                    y=daily_diff_corrected,
                    mode="lines+markers",
                    name=f"Run {j+1} - Corrected",
                    line=dict(color="blue", width=2),
                    connectgaps=False
                )
                fig_corrected = sp.make_subplots(rows=1, cols=1, subplot_titles=[f"Run {j+1} - Corrected"])
                # fig_corrected.add_trace(trace_nan_corrected)
                fig_corrected.add_trace(trace_corrected)
                fig_corrected.update_layout(legend=dict(x=0.6, y=1.25, orientation="h"))
                graph_corrected = fig_corrected.to_html(full_html=False)
                graphs_corrected.append(graph_corrected)

            temp_df = pd.DataFrame({
                'DATA_DATE': df_data_date,
                'DAILY_DIFF': daily_diff_corrected
            })
            
            # รวมผลรวม daily_diff ตาม DATA_DATE สำหรับ Run ปัจจุบัน
            result_df = temp_df.groupby('DATA_DATE', as_index=False)['DAILY_DIFF'].sum()

            # ถ้า daily_sum_corrected ยังไม่มีข้อมูล ก็เริ่มต้นด้วย result_df
            if daily_sum_corrected.empty:
                daily_sum_corrected = result_df
            else:
                # รวมผลรวม daily_diff ของแต่ละ Run กับ daily_sum_corrected
                daily_sum_corrected = pd.merge(daily_sum_corrected, result_df, on='DATA_DATE', how='outer', suffixes=('', f'_Run{j+1}'))

                # บวกค่า DAILY_DIFF ของ Run ปัจจุบันเข้ากับค่าใน daily_sum_corrected
                daily_sum_corrected['DAILY_DIFF'] = daily_sum_corrected['DAILY_DIFF'].fillna(0) + daily_sum_corrected[f'DAILY_DIFF_Run{j+1}'].fillna(0)
                
                # ลบคอลัมน์ที่ไม่ต้องการ
                daily_sum_corrected.drop(columns=[f'DAILY_DIFF_Run{j+1}'], inplace=True)

        
        # Graph Sum
        trace_sum_corrected = go.Scatter(
            x=daily_sum_corrected['DATA_DATE'],
            y=daily_sum_corrected['DAILY_DIFF'],
            mode="lines+markers",
            name="Sum of Corrected Daily Diff",
            line=dict(color=" green", width=2)
        )

        fig_sum_corrected = sp.make_subplots(rows=1, cols=1, subplot_titles=["Sum of Corrected Daily Diff"])
        fig_sum_corrected.add_trace(trace_sum_corrected)
        fig_sum_corrected.update_layout(legend=dict(x=0.6, y=1.25, orientation="h"))

        # สร้างกราฟเป็น HTML
        graph_sum_corrected = fig_sum_corrected.to_html(full_html=False)
        graphs_corrected.append(graph_sum_corrected)

                    
        daily_diff_uncorrected_data = []
        # sum_uncorrected = []
        for j in range(num_streams):
            test_df = df_graphs[j]
            df_data_date = test_df['DATA_DATE']
            df_uncorrected = pd.to_numeric(test_df['UNCORRECTED'], errors='coerce')  # แปลงให้เป็นตัวเลข
            df_valid = test_df[~df_uncorrected.isna()]  # แทนค่า NaN ด้วย 0

            # คำนวณความต่างรายวัน
            daily_diff_uncorrected = df_uncorrected.diff()
            if not daily_diff_uncorrected.empty:
                daily_diff_uncorrected_data.append(daily_diff_uncorrected.to_dict())
                
                
            if not df_uncorrected.isna().all():
                # trace_nan_uncorrected = go.Scatter(
                #     x=df_data_date,
                #     y=daily_diff_uncorrected,
                #     mode="lines",
                #     name=f"Run {j+1} - N/A",
                #     line=dict(color="red", width=2),
                #     connectgaps=True
                # )
                trace_uncorrected = go.Scatter(
                    x=df_data_date,
                    y=daily_diff_uncorrected,
                    mode="lines+markers",
                    name=f"Run {j+1} - Uncorrected",
                    line=dict(color="orange", width=2),
                    connectgaps=False
                )
                fig_uncorrected = sp.make_subplots(rows=1, cols=1, subplot_titles=[f"Run {j+1} - Uncorrected"])
                # fig_uncorrected.add_trace(trace_nan_uncorrected)
                fig_uncorrected.add_trace(trace_uncorrected)
                fig_uncorrected.update_layout(legend=dict(x=0.6, y=1.25, orientation="h"))
                graph_uncorrected = fig_uncorrected.to_html(full_html=False)
                graphs_uncorrected.append(graph_uncorrected)
                        
            # สร้าง DataFrame สำหรับสะสมผลรวม


            # สร้าง DataFrame สำหรับ Run ปัจจุบัน
            temp_df = pd.DataFrame({
                'DATA_DATE': df_data_date,
                'DAILY_DIFF': daily_diff_uncorrected
            })
            
            # รวมผลรวม daily_diff ตาม DATA_DATE สำหรับ Run ปัจจุบัน
            result_df = temp_df.groupby('DATA_DATE', as_index=False)['DAILY_DIFF'].sum()

            # ถ้า daily_sum_uncorrected ยังไม่มีข้อมูล ก็เริ่มต้นด้วย result_df
            if daily_sum_uncorrected.empty:
                daily_sum_uncorrected = result_df
            else:
                # รวมผลรวม daily_diff ของแต่ละ Run กับ daily_sum_uncorrected
                daily_sum_uncorrected = pd.merge(daily_sum_uncorrected, result_df, on='DATA_DATE', how='outer', suffixes=('', f'_Run{j+1}'))

                # บวกค่า DAILY_DIFF ของ Run ปัจจุบันเข้ากับค่าใน daily_sum_uncorrected
                daily_sum_uncorrected['DAILY_DIFF'] = daily_sum_uncorrected['DAILY_DIFF'].fillna(0) + daily_sum_uncorrected[f'DAILY_DIFF_Run{j+1}'].fillna(0)

                # ลบคอลัมน์ที่ไม่ต้องการ
                daily_sum_uncorrected.drop(columns=[f'DAILY_DIFF_Run{j+1}'], inplace=True)
                
        trace_sum_uncorrected = go.Scatter(
            x=daily_sum_uncorrected['DATA_DATE'],
            y=daily_sum_uncorrected['DAILY_DIFF'],
            mode="lines+markers",
            name="Sum of Uncorrected Daily Diff",
            line=dict(color=" green", width=2)
        )

        fig_sum_uncorrected = sp.make_subplots(rows=1, cols=1, subplot_titles=["Sum of Uncorrected Daily Diff"])
        fig_sum_uncorrected.add_trace(trace_sum_uncorrected)
        fig_sum_uncorrected.update_layout(legend=dict(x=0.6, y=1.25, orientation="h"))

        # สร้างกราฟเป็น HTML
        graph_sum_uncorrected = fig_sum_uncorrected.to_html(full_html=False)
        graphs_uncorrected.append(graph_sum_uncorrected)
        
                
        for j in range(num_streams):
            test_df = df_graphs[j]
            df_data_date = test_df['DATA_DATE']   
            
            if not test_df["Pressure"].isna().all():
                fig_pressure.add_trace(go.Scatter(
                    x=df_data_date,
                    y=test_df["Pressure"],
                    mode="lines+markers",
                    name=f"Run {j+1} - Pressure",
                    line=dict(color=colors[j], width=2),
                    connectgaps=False  # เชื่อมจุดที่มีค่า NaN
                ))

            if not test_df["Temperature"].isna().all():
                fig_temperature.add_trace(go.Scatter(
                    x=df_data_date,
                    y=test_df["Temperature"],
                    mode="lines+markers",
                    name=f"Run {j+1} - Temperature",
                    line=dict(color=colors[j], width=2),
                    connectgaps=False  # เชื่อมจุดที่มีค่า NaN
                ))

        # ตั้งค่าการแสดงผลของกราฟ Pressure และ Temperature
        fig_pressure.update_layout(
            xaxis_title="Date",
            yaxis_title="Pressure",
            legend=dict(x=0.5, y=1.20, orientation="h", xanchor="center")
        )
        graph_pressure = fig_pressure.to_html(full_html=False)
        graphs_corrected.append(graph_pressure)

        fig_temperature.update_layout(
            xaxis_title="Date",
            yaxis_title="Temperature",
            legend=dict(x=0.5, y=1.20, orientation="h", xanchor="center")
        )
        graph_temperature= fig_temperature.to_html(full_html=False)
        graphs_uncorrected.append(graph_temperature)


        return {
        'graphs_corrected': graphs_corrected,
        'graphs_uncorrected': graphs_uncorrected,
       
    }



 
# @app.route("/save_date", methods=["POST"])
# def save_date():
#     data_date = request.json.get('data_date')
#     print(f"Received data date: {data_date}")
#     return jsonify({"message": "Date received successfully", "data_date": data_date})          

@app.route("/billing_data_error", methods=["POST"])
@login_required
def billing_data_error():
    response = {}  # Initialize response dictionary

    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        try:
            data = request.json
            data_date = data.get('data_date')
            selected_region = data.get('selected_region')
            selected_tag = data.get('selected_tag')
            run_number = data.get('run_number')
            
            date_only = data_date.split(' ')[0]  # แยกที่ช่องว่างและเลือกเฉพาะส่วนวันที่

            billing_data_error_query = f"""SELECT AMR_BILLING_DATA_ERROR.DATA_DATE,
                        AMR_BILLING_DATA_ERROR.METER_ID,
                        AMR_BILLING_DATA_ERROR.METER_STREAM_NO,
                        AMR_BILLING_DATA_ERROR.CORRECTED_VOL,
                        AMR_BILLING_DATA_ERROR.UNCORRECTED_VOL,
                        AMR_BILLING_DATA_ERROR.AVR_PF,
                        AMR_BILLING_DATA_ERROR.AVR_TF,
                        AMR_BILLING_DATA_ERROR.TIME_CREATE
                        FROM AMR_BILLING_DATA_ERROR  , amr_field_id
                        WHERE amr_field_id.meter_id = amr_billing_data_error.meter_id
                        AND amr_billing_data_error.meter_stream_no like '{run_number}'
                        AND amr_billing_data_error.DATA_DATE like '{date_only}' 
                        AND amr_field_id.TAG_ID like '{selected_tag}' """
            results = fetch_data(ptt_pivot_connection, billing_data_error_query)
            
            df = pd.DataFrame(results, columns=[
                "DATA_DATE",
                "METER_ID",
                "RUN",
                "CORRECTED_VOL",
                "UNCORRECTED_VOL",
                "AVR_PF",
                "AVR_TF",
                "TIME_CREATE",
            ])
            # print(df)
            response = {
                "status": "success",
                "data": df.to_dict(orient='records')
            }
            
        except Exception as e:
            print(f"Error: {e}")
            response = {
                "status": "error",
                "message": f"An error occurred while fetching data: {str(e)}"
            }
    
    return jsonify(response)

@app.route("/selected_data", methods=["POST"])
@login_required
def selected_data():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        try:
            data = request.json

            # Define the column names
            column_names = [
                "DATA_DATE",
                "METER_ID",
                "RUN",
                "CORRECTED_VOL",
                "UNCORRECTED_VOL",
                "AVR_PF",
                "AVR_TF"
            ]

            # Convert data to DataFrame
            df = pd.DataFrame([data], columns=column_names)  # Assuming 'data' is a single dictionary

            # Convert DATA_DATE to the desired format
            df['DATA_DATE'] = pd.to_datetime(df['DATA_DATE']).dt.date

            # Ensure numeric fields are treated as numbers and convert them to native Python types
            df['CORRECTED_VOL'] = df['CORRECTED_VOL'].astype(int)
            df['UNCORRECTED_VOL'] = df['UNCORRECTED_VOL'].astype(int)
            df[["AVR_PF", "AVR_TF"]] = df[["AVR_PF", "AVR_TF"]].astype(str).apply(lambda x: x.str.replace(",", ".", regex=False))
            df['AVR_PF'] = df['AVR_PF'].astype(float)
            df['AVR_TF'] = df['AVR_TF'].astype(float)

            data_date = df.at[0, 'DATA_DATE']
            meter_id = df.at[0, 'METER_ID']
            run = df.at[0, 'RUN']
            corrected_vol = int(df.at[0, 'CORRECTED_VOL'])
            uncorrected_vol = int(df.at[0, 'UNCORRECTED_VOL'])
            avr_pf = float(df.at[0, 'AVR_PF'])
            avr_tf = float(df.at[0, 'AVR_TF'])
            #print(df)

            # Create the current datetime with the required format
            current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))).strftime('%d-%b-%y %I.%M.%S.%f %p ASIA/BANGKOK').upper()

            # Create the SQL query string using bind variables to handle data types correctly
            sql_text_billing_insert = """
            INSERT INTO AMR_BILLING_DATA (DATA_DATE, METER_ID, METER_STREAM_NO, CORRECTED_VOL, UNCORRECTED_VOL, AVR_PF, AVR_TF, TIME_CREATE) 
            VALUES (:data_date, :meter_id, :run, :corrected_vol, :uncorrected_vol, :avr_pf, :avr_tf, :current_datetime)
            """
            # delete_databilling_error = """
            # DELETE FROM AMR_BILLING_DATA_ERROR 
            # WHERE DATA_DATE = :data_date AND METER_ID = :meter_id AND METER_STREAM_NO = :run
            # """

            with cx_Oracle.connect(username, password, f"{hostname}:{port}/{service_name}") as connection:
                with connection.cursor() as cursor:
                    cursor.execute(sql_text_billing_insert, {
                        'data_date': data_date,
                        'meter_id': meter_id,
                        'run': run,
                        'corrected_vol': corrected_vol,
                        'uncorrected_vol': uncorrected_vol,
                        'avr_pf': avr_pf,
                        'avr_tf': avr_tf,
                        'current_datetime': current_datetime
                    })
                    connection.commit()
                    
                    # cursor.execute(delete_databilling_error, {
                    #     'data_date': data_date,
                    #     'meter_id': meter_id,
                    #     'run': run
                    # })
                    # connection.commit()
                    print("Billing successfully")
                    
            
                                                
            response = {
                "status": "success",
                "message": "Data received successfully"
            }
        except Exception as e:
            print(f"Error: {e}")
            response = {
                "status": "error",
                "message": f"An error occurred while processing selected data: {str(e)}"
            }
    
    return jsonify(response)

            

@app.route("/billing_data_user_group")
@login_required
def billing_data_user_group():
    if 'username' not in session:
        return redirect(url_for('login'))
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
    #with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        query_type = request.args.get("query_type")
        username = session['username']

        # Fetch user information from the session
        logged_in_user = session['username']
        if logged_in_user not in users:
            return redirect(url_for('login'))
        
        user_info = users[logged_in_user]
        user_level = user_info.get('user_level')
        description = user_info.get('description')
        print("description", description)
        
        logged_in_user = logged_in_user
        print("user:", logged_in_user)


    # SQL query to fetch unique PL_REGION_ID values
        region_query = """
        SELECT amr_user.user_group 
        FROM AMR_user 
        WHERE user_name = :logged_in_user
        AND amr_user.user_enable like '1' 
        """
        print("ree", region_query)
        # Fetch unique region values
        region_results = fetch_data(ptt_pivot_connection, region_query, params={'logged_in_user': logged_in_user})
        
        
        region_options_tmp = [str(region[0]) for region in region_results]
        
        
        region_options = region_options_tmp[0]


        region_query_options = f"""SELECT REGION_NAME FROM AMR_REGION  where id like '{region_options}' """
        region_results_options = fetch_data(ptt_pivot_connection, region_query_options)
        region_names = [row[0] for row in region_results_options]
        region_options = region_names[0]
        print("Region Names:", region_options)
        
        
        
        # tag_query = """
        # SELECT DISTINCT TAG_ID
        # FROM AMR_FIELD_ID
        # JOIN AMR_PL_GROUP ON AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID 
        # WHERE AMR_PL_GROUP.PL_REGION_ID = :region_id
        # """
        
        tag_query = """
        SELECT DISTINCT AMR_FIELD_ID.TAG_ID 
        FROM AMR_FIELD_ID, amr_user, amr_pl_group,amr_region
        WHERE
            amr_user.user_group = amr_pl_group.pl_region_id
            AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
            and AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID
            and AMR_FIELD_ID.tag_id NOT like '%.remove%'
            and amr_region.REGION_NAME = :region_options
            ORDER BY AMR_FIELD_ID.TAG_ID
        """
        
        print(tag_query)
        tag_results = fetch_data(ptt_pivot_connection, tag_query, params={"region_options": region_options})
        tag_options = [str(tag[0]) for tag in tag_results]
        print("tag:", tag_options)
        
        selected_tag = request.args.get("tag_dropdown")
        selected_region = request.args.get("region_dropdown")
        selected_day = request.args.get("day_dropdown")

        selected_date = request.args.get("date_dropdown_daily")
        session['selected_date_user_group'] = selected_date
        session['selected_tag_user_group'] = selected_tag
        print("test",selected_date)
        print("test2",selected_tag)
        if selected_date:
            # Check if the selected_date is in the format of a single day (dd/mm/yyyy)
            if "/" in selected_date and len(selected_date) == 10:
                # Example: 25/08/2024
                # Enter the first if block
                # Your code for handling a single date goes here
                pass

            # Check if the selected_date is a range of days (dd/mm/yyyy to dd/mm/yyyy)
            elif "to" in selected_date:
                # Example: 25/08/2024 to 30/08/2024
                # Enter the else block
                # Your code for handling a date range goes here
                start_date_str, end_date_str = selected_date.split(" to ")

        # Convert the string dates to datetime objects
                start_date = datetime.datetime.strptime(start_date_str, '%d/%m/%Y')
                end_date = datetime.datetime.strptime(end_date_str, '%d/%m/%Y')

                
                formatted_start_date = start_date.strftime('%d/%m/%Y')
                formatted_end_date = end_date.strftime('%d/%m/%Y')

            # Check if the selected_date is a month (mm/yyyy)
            elif "/" in selected_date and len(selected_date) == 7:
                # Example: 08/2024
                # Enter the third else block
                # Your code for handling a month goes here
                pass
        
   
        query = ""

        if query_type == "daily_data":
            # SQL query for main data
            if "/" in selected_date and len(selected_date) == 10:

                query = f"""
                        SELECT DISTINCT
                    AMR_PL_GROUP.PL_REGION_ID,
                    AMR_FIELD_ID.TAG_ID,
                    AMR_FIELD_ID.METER_ID,
                    TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD-MM-YYYY')  AS DATA_DATE,
                    
                    AMR_BILLING_DATA.CORRECTED_VOL as CORRECTED,
                    AMR_BILLING_DATA.UNCORRECTED_VOL as UNCORRECTED,
                    AMR_BILLING_DATA.AVR_PF as Pressure,
                    AMR_BILLING_DATA.AVR_TF as Temperature,
                    AMR_BILLING_DATA.METER_STREAM_NO  
                FROM
                    AMR_FIELD_ID, AMR_PL_group, AMR_BILLING_DATA,amr_region
                WHERE
                    AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                    AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                    AND AMR_BILLING_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    AND AMR_BILLING_DATA.METER_STREAM_NO IS NOT NULL
                    AND TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD/MM/YYYY') = '{selected_date}'
                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                    AND amr_region.REGION_NAME = '{selected_region}'
                ORDER BY
                    TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD-MM-YYYY') DESC
                    """
                
            elif "to" in selected_date:
                query = f"""
                        SELECT DISTINCT
                        AMR_PL_GROUP.PL_REGION_ID,
                        AMR_FIELD_ID.TAG_ID,
                        AMR_FIELD_ID.METER_ID,
                        TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD/MM/YY') AS DATA_DATE,
                        AMR_BILLING_DATA.CORRECTED_VOL AS CORRECTED,
                        AMR_BILLING_DATA.UNCORRECTED_VOL AS UNCORRECTED,
                        AMR_BILLING_DATA.AVR_PF AS Pressure,
                        AMR_BILLING_DATA.AVR_TF AS Temperature,
                        AMR_BILLING_DATA.METER_STREAM_NO  
                    FROM
                        AMR_FIELD_ID
                    JOIN AMR_PL_GROUP 
                        ON AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
                    JOIN AMR_BILLING_DATA 
                        ON AMR_BILLING_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    JOIN AMR_REGION 
                        ON AMR_PL_GROUP.PL_REGION_ID = AMR_REGION.ID
                    WHERE
                        AMR_BILLING_DATA.METER_STREAM_NO IS NOT NULL
                        AND AMR_BILLING_DATA.DATA_DATE BETWEEN TO_DATE('{formatted_start_date}', 'DD/MM/YYYY') 
                                                            AND TO_DATE('{formatted_end_date}', 'DD/MM/YYYY')
                        AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                        AND AMR_REGION.REGION_NAME = '{selected_region}'
                    ORDER BY
                        TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'DD/MM/YY') DESC
                    """
                
            elif "/" in selected_date and len(selected_date) == 7:

                print("One or both dates are missing.")
                # Your code here for missing dates
                # Fetch tag options based on the selected region
                query = f"""
                SELECT DISTINCT
                    AMR_PL_GROUP.PL_REGION_ID,
                    AMR_FIELD_ID.TAG_ID,
                    AMR_FIELD_ID.METER_ID,
                    TO_CHAR(AMR_BILLING_DATA.DATA_DATE)  AS DATA_DATE,
                    
                    AMR_BILLING_DATA.CORRECTED_VOL as CORRECTED,
                    AMR_BILLING_DATA.UNCORRECTED_VOL as UNCORRECTED,
                    AMR_BILLING_DATA.AVR_PF as Pressure,
                    AMR_BILLING_DATA.AVR_TF as Temperature,
                    AMR_BILLING_DATA.METER_STREAM_NO  
                FROM
                    AMR_FIELD_ID, AMR_PL_group, AMR_BILLING_DATA,amr_region
                WHERE
                    AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                    AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                    AND AMR_BILLING_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    AND AMR_BILLING_DATA.METER_STREAM_NO IS NOT NULL
                    AND TO_CHAR(AMR_BILLING_DATA.DATA_DATE, 'MM/YYYY') = '{selected_date}'
                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                    AND amr_region.REGION_NAME = '{selected_region}'
                ORDER BY
                    TO_CHAR(AMR_BILLING_DATA.DATA_DATE) DESC
                """
            results = fetch_data(ptt_pivot_connection, query)
            
            df = pd.DataFrame(
                results,
                columns=[
                    "PL_REGION_ID",
                    "TAG_ID",
                    "METER_ID",
                    "DATA_DATE",
                    "CORRECTED",
                    "UNCORRECTED",
                    "Pressure",
                    "Temperature",
                    "METER_STREAM_NO",
                ],
            )
            df["Pressure"] = df["Pressure"].astype(float)
            df["Temperature"] = df["Temperature"].astype(float)

            df["Pressure"] = df["Pressure"].round(4)
            df["Temperature"] = df["Temperature"].round(4)
            df["CORRECTED"] = pd.to_numeric(df["CORRECTED"], errors='coerce').fillna(0).astype(int).astype(str)
            df["UNCORRECTED"] = pd.to_numeric(df["UNCORRECTED"], errors='coerce').fillna(0).astype(int).astype(str)



            df["METER_STREAM_NO"] = df["METER_STREAM_NO"].astype(int)
            meter_id_list = df["METER_ID"].tolist()


  
            if not meter_id_list:
                # If empty, set the vc_name_dict with default values
                vc_name_dict = {f'vc_name_list{i}': None for i in range(1, 7)}
            else:
                query_vc_name = f"""SELECT amr_vc_type.vc_name, amr_field_meter.METER_STREAM_NO 
                                    FROM amr_vc_type, amr_field_meter 
                                    WHERE amr_field_meter.METER_STREAM_TYPE = amr_vc_type.id 
                                    AND amr_field_meter.meter_id LIKE '{meter_id_list[0]}' 
                                    ORDER BY amr_field_meter.METER_STREAM_NO"""
                
                results_vc_name = fetch_data(ptt_pivot_connection, query_vc_name)
                df_vc_name = pd.DataFrame(
                    results_vc_name,
                    columns=["vc_name", "METER_STREAM_NO"]
                )

                unique_meter_streams = df_vc_name.drop_duplicates(subset=["METER_STREAM_NO"])[["METER_STREAM_NO", "vc_name"]]

                meter_stream_no_list = unique_meter_streams["METER_STREAM_NO"].tolist()
                vc_name_list = unique_meter_streams["vc_name"].tolist()

                vc_name_dict = {}
                for i in range(1, 7):
                    if len(vc_name_list) >= i:
                        vc_name_dict[f'vc_name_list{i}'] = vc_name_list[i - 1]
                    else:
                        vc_name_dict[f'vc_name_list{i}'] = None  # or "nong", depending on your preference

                        
            
                query_type = f"""
                    SELECT DISTINCT  amr_vc_type.id as METER_ID , amr_field_meter.METER_STREAM_NO FROM amr_billing_data,amr_field_meter,amr_vc_type 
                    WHERE amr_billing_data.meter_id = amr_field_meter.meter_id 
                    AND amr_field_meter.METER_STREAM_TYPE = amr_vc_type.id
                    AND amr_billing_data.meter_id like '{meter_id_list[0]}'
                    AND  amr_field_meter.METER_STREAM_NO IS NOT NULL
                    ORDER BY amr_field_meter.METER_STREAM_NO """
                    
                results_type = fetch_data(ptt_pivot_connection, query_type)
                df_type = pd.DataFrame(
                    results_type,
                    columns=[
                        "METER_ID",
                        "METER_STREAM_NO",
                    ]
                )
                meter_stream_no_list = df_type["METER_STREAM_NO"].tolist()
                amr_vc_type_list = df_type["METER_ID"].tolist()
            
            # Get the selected Meter ID before removing it from the DataFrame
            # selected_meter_id = df["METER_ID"].iloc[0]
            selected_meter_id = None

            # Check if 'METER_ID' column exists and the DataFrame is not empty
            if not df.empty and 'METER_ID' in df.columns and len(df['METER_ID']) > 0:
                selected_meter_id = df['METER_ID'].iloc[0]
                print(f"Selected Meter ID (graph): {selected_meter_id}")
            else:
                print("DataFrame is empty or 'METER_ID' column doesn't exist.")

            # Now, remove the "METER_ID" column from the DataFrame
            df = df.drop(["PL_REGION_ID", "TAG_ID", "METER_ID"], axis=1)
            
            # Remove newline characters
            df = df.apply(lambda x: x.str.replace("\n", "") if x.dtype == "object" else x)

            df = df.drop_duplicates(subset=["DATA_DATE", "METER_STREAM_NO"], keep="first")
            # Sort DataFrame by 'DATA_DATE'
            df = df.sort_values(by="DATA_DATE")
            # print("df",df)
            # Assuming 'df' is the DataFrame created from the query results
            num_streams = 6
            df_runs = {}

            # Loop to create DataFrames for each METER_STREAM_NO
            for i in range(1, num_streams + 1):
                df_runs[f'df_run{i}'] = df[df['METER_STREAM_NO'] == int(i)]
            
            # Check if each DataFrame has data before including in the tables dictionary
            tables = {
                "config_data": None,
            }

            graphs = {
                "corrected": None,
                "uncorrected": None,
                "pressure": None,
                "temperature": None
            }
                         
            if "/" in selected_date and len(selected_date) == 10:
                # print("selected_date:",selected_date)
                if selected_date:
                    # Convert selected_date to 'YYYY-MM-DD' format for consistency
                    selected_date_formatted = pd.to_datetime(selected_date, format='%d/%m/%Y').strftime('%d/%m/%Y')

                    # Get the current date in 'YYYY-MM-DD' format
                    current_date = datetime.datetime.now().strftime('%d-%m-%Y')

                    # Determine if the selected date is the current date
                    is_current_day = selected_date_formatted == current_date

                    # Update the query to use the selected date
                    if is_current_day:
                        # If the selected date is today, show only today's data
                        query_day = f"""
                            SELECT TO_CHAR(TO_DATE('{current_date}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                            FROM DUAL
                        """

                    else:
                        # If the selected date is not today, show only the selected day's data
                        query_day = f"""
                            SELECT TO_CHAR(TO_DATE('{selected_date_formatted}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                            FROM DUAL
                        """

            elif "to" in selected_date:
                if formatted_start_date and formatted_end_date:
                    # แปลงวันที่ให้เป็นรูปแบบ 'YYYY-MM-DD' สำหรับใช้ใน SQL
                    start_date_formatted = pd.to_datetime(formatted_start_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                    end_date_formatted = pd.to_datetime(formatted_end_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                    
                    print("Start Date:", formatted_start_date)
                    print("End Date:", formatted_end_date)
                    
                    # สร้างรายการวันที่ในช่วงที่เลือกโดยใช้ SQL
                    query_day = f"""
                        SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{start_date_formatted}', 'YY-MM-DD'), 'DD/MM/YY') AS DATA_DATE
                        FROM DUAL
                        CONNECT BY LEVEL <= TO_DATE('{end_date_formatted}', 'YY-MM-DD') - TO_DATE('{start_date_formatted}', 'YY-MM-DD') + 1
                    """

            elif "/" in selected_date and len(selected_date) == 7:
                df_month_list = pd.DataFrame(columns=['DATA_DATE'])
                
                # Check if 'selected_date' is available
                if selected_date:
                    # Convert selected_date to 'YYYY-MM' format for consistency
                    selected_date_formatted = pd.to_datetime(selected_date, format='%m/%Y').strftime('%Y-%m')
                    # print("date2:", selected_date_formatted)

                    # Get the current month and year
                    current_month_year = datetime.datetime.now().strftime('%Y-%m')

                    # Determine if the selected date is in the current month
                    is_current_month = selected_date_formatted == current_month_year
                    
                    # Update the query to use the selected date
                    if is_current_month:
                        # If the selected date is in the current month, show all days up to the current date
                        query_day = f"""
                            SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{current_month_year}-01', 'YYYY-MM-DD')) AS Date1
                            FROM DUAL
                            CONNECT BY LEVEL <= TO_DATE('{datetime.datetime.now().strftime('%Y-%m-%d')}', 'YYYY-MM-DD') - TO_DATE('{current_month_year}-01', 'YYYY-MM-DD') + 1
                        """
                    else:
                        # If the selected date is in a previous month, show all days of the selected month
                        query_day = f"""
                            SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) AS Date1
                            FROM DUAL
                            CONNECT BY LEVEL <= LAST_DAY(TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) - TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD') + 1
                        """

            query_day_result = fetch_data(ptt_pivot_connection, query_day)
            df_month_list = pd.DataFrame(query_day_result, columns=['DATA_DATE'])
            # print("df_month_list",df_month_list)
            def make_link(row, run_number):
                if row['CORRECTED'] == 'N/A' and row['UNCORRECTED'] == 'N/A':
                    return f'<a href="#" onclick="sendDataDate(\'{row["DATA_DATE"]}\', {run_number}); openPopup(); return false;">{row["DATA_DATE"]}</a>'
                    
                return row['DATA_DATE']
            
            for i in range(1, num_streams + 1):
                df_run = df_runs[f'df_run{i}']
                
                if not df_run.empty:
                    merged_df = pd.merge(df_month_list, df_run, on='DATA_DATE', how='outer')
                    df_runs[f'df_run{i}'] = merged_df
                    html_style = """
                            <style>
                            .data_daily th, .data_daily td .data_daily tr{
                                text-align: center; 
                            }
                            </style>
                            """
            # print("df_run_no_last",  session['data_billing'])


            html_style = '''
                <style>
                    .data_daily {
                        border-collapse: collapse;
                        width: 100%;
                    }
                    .data_daily th, .data_daily td {
                        border: 1px solid #ddd;
                        padding: 8px;
                        color: #000000;
                    }
                    .data_daily th {
                        padding-top: 12px;
                        padding-bottom: 12px;
                        text-align: left;
                        background-color: #f0f0f0;
                        
                    }
                </style>
                '''
                

            data_type = []
            if not meter_id_list:
                pass
            else :
                for amr_vc_type in amr_vc_type_list:
                    amr_vc_type_table = f"""
                    
                    
                    SELECT DESCRIPTION 
                    FROM amr_mapping_billing 
                    WHERE evc_type LIKE '{amr_vc_type}' 
                    AND DAILY like '1' 
                    ORDER BY OR_DER


                    """
                    
                    amr_vc_type_table_result = fetch_data(ptt_pivot_connection, amr_vc_type_table)
                    filtered_result = pd.DataFrame(amr_vc_type_table_result)
                    
            
                    transposed_result = filtered_result.T
                    
                    
                    selected_data = transposed_result.iloc[0, 0:5]  # Extract values to get a flat array
                    data_type.append(selected_data)
            # Print the values in a horizontal format
            # for i in range(4):
            #     print(data_type[i])
                    
            
                
            df_graphs = []  # สร้างลิสต์เพื่อเก็บข้อมูล
            combined_data = []
            all_totals = []  # สร้างลิสต์เพื่อเก็บค่า totals ของแต่ละรอบ

            for i in range(num_streams):
                df_run = df_runs.get(f'df_run{i+1}', pd.DataFrame())

                df_graphs_run = pd.DataFrame(df_run)
                df_graphs.append(df_graphs_run)  # Store DataFrame in the list

                if not df_run.empty:
                    df_run = df_run.drop('METER_STREAM_NO', axis=1, errors='ignore')
                    df_run = df_run.fillna("N/A")
                    combined_data.append(df_run.to_dict())

                   

                    # Prepare for constructing HTML
                    df_run_no_last = df_run
                    hedder = data_type[i]
                    header_df = pd.DataFrame(hedder).T

                   


                    # Create the HTML manually as before
                    html_output = '<table class="data_daily">'

                    # Extract the first row to use as header
                    header_row = header_df.iloc[0]
                    html_output += '<thead><tr>'
                    for item in header_row:
                        html_output += f'<th style="text-align: center;">{item}</th>'
                    html_output += '</tr></thead>'

                    # Now create the tbody
                    html_output += '<tbody>'
                    for index, row in df_run_no_last.iloc[0:].iterrows():  # Skip the first row
                        html_output += '<tr>'
                        for item in row:
                            html_output += f'<td>{item}</td>'
                        html_output += '</tr>'

                    # Add the total row (calculated from the daily differences)
                    html_output += '<tr><td style="text-align: center; background-color: #1cc88a33;"">Total</td>'
                    
                    
                    
                    
                    
                                     # Convert 'CORRECTED' and 'UNCORRECTED' to numeric
                    df_run_no_last['CORRECTED'] = pd.to_numeric(df_run_no_last['CORRECTED'], errors='coerce')
                    df_run_no_last['UNCORRECTED'] = pd.to_numeric(df_run_no_last['UNCORRECTED'], errors='coerce')

                    # Forward fill missing values to ensure calculation continuity
                    df_run_no_last[['CORRECTED', 'UNCORRECTED']] = df_run_no_last[['CORRECTED', 'UNCORRECTED']].ffill()

                    # Calculate the difference for only 'CORRECTED' and 'UNCORRECTED' columns after filling missing values
                    df_run_diff = df_run_no_last[['CORRECTED', 'UNCORRECTED']].diff().fillna(0)

                    # Filter out zero values
                    df_run_diff_filtered = df_run_diff[df_run_diff != 0]

                    # Now calculate the totals for the non-zero values (daily difference totals)
                    totals = df_run_diff_filtered.sum()

                    # Add this round's totals to all_totals list
                    all_totals.append(totals)
                    
                    
                    
                    for col in df_run_no_last.columns[1:]:  # Skip the first column (assumed non-numeric)
                        if col in totals:
                            total_value = int(totals[col])
                            html_output += f'<td  style="background-color: #1cc88a33;">{total_value}</td>'
                        # else:
                        #     html_output += '<td ></td>'  # Leave empty for non-numeric columns
                    html_output += '</tr>'

                    html_output += '</tbody>'
                    html_output += '</table>'

                    # Store the constructed HTML in the tables dictionary
                    tables[f"daily_data_run{i+1}"] = html_style + html_output
                    
            # After looping through all runs, sum all totals
            final_totals = pd.DataFrame(all_totals).sum()

            html_final_totals = '<table class="data_totals">'
            html_final_totals += '<thead><tr><th>Daily Corrected Volume</th><th>Daily Uncorrected Volume</th></tr></thead>'
            html_final_totals += '<tbody><tr>'  # Start the body and first row

            # Iterate over the final_totals index to populate the total values
            for col in final_totals.index:
                html_final_totals += f'<td>{int(final_totals[col])}</td>'
            html_final_totals += '</tr></tbody></table>'

            # Store the final combined totals table
            tables["combined_totals"] = html_style + html_final_totals
                

            with open('data_report_user_group.json', 'w') as json_file:
                json.dump(combined_data, json_file, indent=4)

            df = df.sort_values(by="DATA_DATE", ascending=True)
            # ส่ง graph_html ไปยัง HTML template ของ Flask
            return render_template(
                "billingdata_user_group.html",
                tables=tables,
                titles=df.columns.values,
                selected_date=selected_date,
                selected_tag=selected_tag,
                selected_region=selected_region,
                region_options=region_options,
                tag_options=tag_options,
                selected_meter_id=selected_meter_id,
                vc_name_dict=vc_name_dict,
                username=username
                # graph_corrected=graph_corrected,
                # graph_uncorrected=graph_uncorrected,
                # graph_pressure=graph_pressure,
                # graph_temperature=graph_temperature,
                

            )


        elif query_type == "config_data":


            if "/" in selected_date and len(selected_date) == 10:

                query = f"""
                    SELECT DISTINCT
                        AMR_PL_GROUP.PL_REGION_ID,
                        AMR_FIELD_ID.TAG_ID,
                        amr_field_id.meter_id,
                        AMR_CONFIGURED_DATA.METER_STREAM_NO,
                        AMR_CONFIGURED_DATA.amr_vc_type,
                        amr_vc_type.vc_name,
                        TO_CHAR(AMR_CONFIGURED_DATA.DATA_DATE, 'DD-MM-YYYY'),
                        
                        
                        amr_configured_data.amr_config1,
                        amr_configured_data.amr_config2,
                        amr_configured_data.amr_config3,
                        amr_configured_data.amr_config4,
                        amr_configured_data.amr_config5,
                        amr_configured_data.amr_config6,
                        amr_configured_data.amr_config7,
                        amr_configured_data.amr_config8,
                        amr_configured_data.amr_config9,
                        amr_configured_data.amr_config10,
                        amr_configured_data.amr_config11,
                        amr_configured_data.amr_config12,
                        amr_configured_data.amr_config13,
                        amr_configured_data.amr_config14,
                        amr_configured_data.amr_config15,
                        amr_configured_data.amr_config16,
                        amr_configured_data.amr_config17,
                        amr_configured_data.amr_config18,
                        amr_configured_data.amr_config19,
                        amr_configured_data.amr_config20
                        
                        
                    
                        
                        FROM
                    AMR_FIELD_ID, AMR_PL_group, AMR_CONFIGURED_DATA,amr_field_meter,amr_vc_type,amr_region
                WHERE
                    AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                    AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                    AND  AMR_CONFIGURED_DATA.amr_vc_type = amr_vc_type.id
                    AND amr_field_meter.meter_id =  amr_field_id.meter_id
                    AND AMR_CONFIGURED_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    AND AMR_CONFIGURED_DATA.METER_STREAM_NO IS NOT NULL
                    AND TO_CHAR(AMR_CONFIGURED_DATA.DATA_DATE, 'DD/MM/YYYY') = '{selected_date}'
                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                    AND amr_region.REGION_NAME = '{selected_region}'
                    
                ORDER BY 
                    AMR_CONFIGURED_DATA.METER_STREAM_NO
                    
                    
                """


            elif "to" in selected_date:
                query = f"""
                        



                            SELECT DISTINCT
                        AMR_PL_GROUP.PL_REGION_ID,
                        AMR_FIELD_ID.TAG_ID,
                        amr_field_id.meter_id,
                        AMR_CONFIGURED_DATA.METER_STREAM_NO,
                        AMR_CONFIGURED_DATA.amr_vc_type,
                        amr_vc_type.vc_name,
                        TO_CHAR(AMR_CONFIGURED_DATA.DATA_DATE, 'DD-MM-YYYY'),
                        
                        
                        amr_configured_data.amr_config1,
                        amr_configured_data.amr_config2,
                        amr_configured_data.amr_config3,
                        amr_configured_data.amr_config4,
                        amr_configured_data.amr_config5,
                        amr_configured_data.amr_config6,
                        amr_configured_data.amr_config7,
                        amr_configured_data.amr_config8,
                        amr_configured_data.amr_config9,
                        amr_configured_data.amr_config10,
                        amr_configured_data.amr_config11,
                        amr_configured_data.amr_config12,
                        amr_configured_data.amr_config13,
                        amr_configured_data.amr_config14,
                        amr_configured_data.amr_config15,
                        amr_configured_data.amr_config16,
                        amr_configured_data.amr_config17,
                        amr_configured_data.amr_config18,
                        amr_configured_data.amr_config19,
                        amr_configured_data.amr_config20
                        
                        
                    
                        
                        FROM
                    AMR_FIELD_ID, AMR_PL_group, AMR_CONFIGURED_DATA,amr_field_meter,amr_vc_type,amr_region
                WHERE
                    AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                    AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                    AND  AMR_CONFIGURED_DATA.amr_vc_type = amr_vc_type.id
                    AND amr_field_meter.meter_id =  amr_field_id.meter_id
                    AND AMR_CONFIGURED_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    AND AMR_CONFIGURED_DATA.METER_STREAM_NO IS NOT NULL
                    
                    AND AMR_CONFIGURED_DATA.DATA_DATE BETWEEN TO_DATE('{formatted_start_date}', 'DD/MM/YYYY') AND TO_DATE('{formatted_end_date}', 'DD/MM/YYYY')
                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                    AND amr_region.REGION_NAME = '{selected_region}'
                    
                ORDER BY 
                    AMR_CONFIGURED_DATA.METER_STREAM_NO
                    """


            elif "/" in selected_date and len(selected_date) == 7:

                query = f"""
                    SELECT DISTINCT
                        AMR_PL_GROUP.PL_REGION_ID,
                        AMR_FIELD_ID.TAG_ID,
                        amr_field_id.meter_id,
                        AMR_CONFIGURED_DATA.METER_STREAM_NO,
                        AMR_CONFIGURED_DATA.amr_vc_type,
                        amr_vc_type.vc_name,
                        TO_CHAR(AMR_CONFIGURED_DATA.DATA_DATE),
                        
                        
                        amr_configured_data.amr_config1,
                        amr_configured_data.amr_config2,
                        amr_configured_data.amr_config3,
                        amr_configured_data.amr_config4,
                        amr_configured_data.amr_config5,
                        amr_configured_data.amr_config6,
                        amr_configured_data.amr_config7,
                        amr_configured_data.amr_config8,
                        amr_configured_data.amr_config9,
                        amr_configured_data.amr_config10,
                        amr_configured_data.amr_config11,
                        amr_configured_data.amr_config12,
                        amr_configured_data.amr_config13,
                        amr_configured_data.amr_config14,
                        amr_configured_data.amr_config15,
                        amr_configured_data.amr_config16,
                        amr_configured_data.amr_config17,
                        amr_configured_data.amr_config18,
                        amr_configured_data.amr_config19,
                        amr_configured_data.amr_config20
                        
                        
                    
                        
                        FROM
                    AMR_FIELD_ID, AMR_PL_group, AMR_CONFIGURED_DATA,amr_field_meter,amr_vc_type,amr_region
                WHERE
                    AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                    AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                    AND  AMR_CONFIGURED_DATA.amr_vc_type = amr_vc_type.id
                    AND amr_field_meter.meter_id =  amr_field_id.meter_id
                    AND AMR_CONFIGURED_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                    AND AMR_CONFIGURED_DATA.METER_STREAM_NO IS NOT NULL
                    AND TO_CHAR(AMR_CONFIGURED_DATA.DATA_DATE, 'MM/YYYY') = '{selected_date}'
                    AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                    AND amr_region.REGION_NAME = '{selected_region}'
                    
                ORDER BY 
                    AMR_CONFIGURED_DATA.METER_STREAM_NO
                    
                    
                """

            
            results = fetch_data(ptt_pivot_connection, query)
            
            df = pd.DataFrame(
                results,
                columns=[
                    "PL_REGION_ID",
                    "TAG_ID",
                    "METER_ID",
                    "METER_STREAM_NO",
                    "amr_vc_type",
                    "vc_name",
                    "DATA_DATE",
                    "AMR_CONFIG1",
                    "AMR_CONFIG2",
                    "AMR_CONFIG3",
                    "AMR_CONFIG4",
                    "AMR_CONFIG5",
                    "AMR_CONFIG6",
                    "AMR_CONFIG7",
                    "AMR_CONFIG8",
                    "AMR_CONFIG9",
                    "AMR_CONFIG10",
                    "AMR_CONFIG11",
                    "AMR_CONFIG12",
                    "AMR_CONFIG13",
                    "AMR_CONFIG14",
                    "AMR_CONFIG15",
                    "AMR_CONFIG16",
                    "AMR_CONFIG17",
                    "AMR_CONFIG18",
                    "AMR_CONFIG19",
                    "AMR_CONFIG20",
                    
                    
                ]
            )
            
        
        
            unique_meter_streams = df.drop_duplicates(subset=["METER_STREAM_NO"])[["METER_STREAM_NO", "amr_vc_type","vc_name"]]

            meter_stream_no_list = unique_meter_streams["METER_STREAM_NO"].tolist()
            amr_vc_type_list = unique_meter_streams["amr_vc_type"].tolist()
            vc_name_list = unique_meter_streams["vc_name"].tolist()


            vc_name_dict = {}
            for i in range(1, 7):
                if len(vc_name_list) >= i:
                    vc_name_dict[f'vc_name_list{i}'] = vc_name_list[i-1]
                else:
                    vc_name_dict[f'vc_name_list{i}'] = None
            
        
            session['vc_type_user_group']= vc_name_dict
           

            if "/" in selected_date and len(selected_date) == 10:   
                # print("METER_STREAM_NO:", meter_stream_no_list)
                # print("amr_vc_type:", amr_vc_type_list)
                print("selected_date:",selected_date)
                if selected_date:
                    # Convert selected_date to 'YYYY-MM-DD' format for consistency
                    selected_date_formatted = pd.to_datetime(selected_date, format='%d/%m/%Y').strftime('%d/%m/%Y')

                    # Get the current date in 'YYYY-MM-DD' format
                    current_date = datetime.datetime.now().strftime('%d-%m-%Y')

                    # Determine if the selected date is the current date
                    is_current_day = selected_date_formatted == current_date

                    # Update the query to use the selected date
                    if is_current_day:
                        # If the selected date is today, show only today's data
                        query_day = f"""
                            SELECT TO_CHAR(TO_DATE('{current_date}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                            FROM DUAL
                        """

                    else:
                        # If the selected date is not today, show only the selected day's data
                        query_day = f"""
                            SELECT TO_CHAR(TO_DATE('{selected_date_formatted}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                            FROM DUAL
                        """

            elif "to" in selected_date:

                if formatted_start_date and formatted_end_date:
                    # แปลงวันที่ให้เป็นรูปแบบ 'YYYY-MM-DD' สำหรับใช้ใน SQL
                    start_date_formatted = pd.to_datetime(formatted_start_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                    end_date_formatted = pd.to_datetime(formatted_end_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                    
                    print("Start Date:", formatted_start_date)
                    print("End Date:", formatted_end_date)
                    
                    # สร้างรายการวันที่ในช่วงที่เลือกโดยใช้ SQL
                    query_day = f"""
                        SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{start_date_formatted}', 'YY-MM-DD'), 'DD/MM/YY') AS DATA_DATE
                        FROM DUAL
                        CONNECT BY LEVEL <= TO_DATE('{end_date_formatted}', 'YY-MM-DD') - TO_DATE('{start_date_formatted}', 'YY-MM-DD') + 1
                    """

            elif "/" in selected_date and len(selected_date) == 7:


                selected_date = request.args.get("date_dropdown_daily")
                

                # Initialize df_month_list outside the if statement
                df_month_list = pd.DataFrame(columns=['DATA_DATE'])
                
                # Check if 'selected_date' is available
                if selected_date:
                    # Convert selected_date to 'YYYY-MM' format for consistency
                    selected_date_formatted = pd.to_datetime(selected_date, format='%m/%Y').strftime('%Y-%m')
                    # print("date2:", selected_date_formatted)

                    # Get the current month and year
                    current_month_year = datetime.datetime.now().strftime('%Y-%m')

                    # Determine if the selected date is in the current month
                    is_current_month = selected_date_formatted == current_month_year
                    
                    # Update the query to use the selected date
                    if is_current_month:
                        # If the selected date is in the current month, show all days up to the current date
                        query_day = f"""
                            SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{current_month_year}-01', 'YYYY-MM-DD')) AS Date1
                            FROM DUAL
                            CONNECT BY LEVEL <= TO_DATE('{datetime.datetime.now().strftime('%Y-%m-%d')}', 'YYYY-MM-DD') - TO_DATE('{current_month_year}-01', 'YYYY-MM-DD') + 1
                        """

                        
                    else:
                        # If the selected date is in a previous month, show all days of the selected month
                        query_day = f"""
                            SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) AS Date1
                            FROM DUAL
                            CONNECT BY LEVEL <= LAST_DAY(TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) - TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD') + 1
                        """

            query_day_result = fetch_data(ptt_pivot_connection, query_day)
            df_month_list = pd.DataFrame(query_day_result, columns=['DATA_DATE'])

            # Convert the DATA_DATE to datetime format if not already
            df_month_list['DATA_DATE'] = pd.to_datetime(df_month_list['DATA_DATE'], errors='coerce', dayfirst=True)
            
            combined_data = []
            
            for amr_vc_type in amr_vc_type_list:
                amr_vc_type_table = f"""
                
                
                SELECT DESCRIPTION
                FROM (
                    SELECT 'date' AS DESCRIPTION, 0 AS OR_DER 
                    FROM DUAL
                    UNION ALL
                    SELECT DESCRIPTION, OR_DER
                    FROM amr_mapping_config
                    WHERE evc_type LIKE '{amr_vc_type}'
                    AND DESCRIPTION IS NOT NULL
                )
                ORDER BY OR_DER


                """
                
                amr_vc_type_table_result = fetch_data(ptt_pivot_connection, amr_vc_type_table)
                filtered_result = pd.DataFrame(amr_vc_type_table_result)
                
        
                transposed_result = filtered_result.T

                combined_data.append(transposed_result)
        

            if "/" in selected_date and len(selected_date) == 10: 
                for meter_stream_no in meter_stream_no_list:
                    meter_stream_no_table = f"""
                    SELECT DISTINCT
                        amr_configured_data.amr_config1 AS DATA_DATE,
                        amr_configured_data.amr_config2,
                        amr_configured_data.amr_config3,
                        amr_configured_data.amr_config4,
                        amr_configured_data.amr_config5,
                        amr_configured_data.amr_config6,
                        amr_configured_data.amr_config7,
                        amr_configured_data.amr_config8,
                        amr_configured_data.amr_config9,
                        amr_configured_data.amr_config10,
                        amr_configured_data.amr_config11,
                        amr_configured_data.amr_config12,
                        amr_configured_data.amr_config13,
                        amr_configured_data.amr_config14,
                        amr_configured_data.amr_config15,
                        amr_configured_data.amr_config16,
                        amr_configured_data.amr_config17,
                        amr_configured_data.amr_config18,
                        amr_configured_data.amr_config19,
                        amr_configured_data.amr_config20
                    FROM
                        AMR_FIELD_ID, AMR_PL_group, AMR_CONFIGURED_DATA, amr_field_meter,amr_region
                    WHERE
                        AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                        AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                        AND amr_field_meter.meter_id = amr_field_id.meter_id
                        AND AMR_CONFIGURED_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                        AND AMR_CONFIGURED_DATA.METER_STREAM_NO LIKE '{meter_stream_no}'
                        AND TO_CHAR(AMR_CONFIGURED_DATA.DATA_DATE, 'DD/MM/YYYY') = '{selected_date}'
                        AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                        AND amr_region.REGION_NAME = '{selected_region}'
                    ORDER BY 
                        amr_configured_data.amr_config1
                    """
                    meter_stream_no_table_result = fetch_data(ptt_pivot_connection, meter_stream_no_table)
            
                    df_meter_stream_no_table = pd.DataFrame(meter_stream_no_table_result, columns=[
                        'DATA_DATE', 'amr_config2', 'amr_config3', 'amr_config4', 'amr_config5', 
                        'amr_config6', 'amr_config7', 'amr_config8', 'amr_config9', 'amr_config10',
                        'amr_config11', 'amr_config12', 'amr_config13', 'amr_config14', 'amr_config15', 
                        'amr_config16', 'amr_config17', 'amr_config18', 'amr_config19', 'amr_config20'
                    ])

                    # แปลง 'DATA_DATE' ให้เป็น datetime
                    df_meter_stream_no_table['DATA_DATE'] = pd.to_datetime(df_meter_stream_no_table['DATA_DATE'], errors='coerce', dayfirst=True)

                    # เปลี่ยนชื่อคอลัมน์ 'DATA_DATE' เป็นชื่ออื่นเพื่อให้ไม่ซ้ำกับ df_month_list
                    df_meter_stream_no_table.rename(columns={'DATA_DATE': 'DATA_DATE_METER'}, inplace=True)

                    # รวม DataFrame บนคอลัมน์ 'DATA_DATE' ของ df_month_list และ 'DATA_DATE_METER' ของ df_meter_stream_no_table
                    df_combined = pd.merge(df_month_list, df_meter_stream_no_table, left_on='DATA_DATE', right_on='DATA_DATE_METER', how='left')

                    # ลบแถวที่มีค่าวันที่ซ้ำกัน โดยให้เหลือเพียงแถวแรกที่พบ
                    df_combined = df_combined.drop_duplicates(subset='DATA_DATE')

                    # แปลงรูปแบบวันที่ในคอลัมน์ 'DATA_DATE' และ 'DATA_DATE_METER'
                    df_combined['DATA_DATE'] = df_combined['DATA_DATE'].dt.strftime('%d-%b-%Y')
                    df_combined['DATA_DATE_METER'] = df_combined['DATA_DATE_METER'].dt.strftime('%d-%m-%Y')

                    # ลบคอลัมน์ที่มีค่าเป็น None หรือ NaN ทั้งหมด
                    df_combined.dropna(axis=1, how='all', inplace=True)

                    # ลบแถวที่มีค่าเป็น None หรือ NaN ทั้งหมด (ถ้ามี)
                    df_combined.dropna(axis=0, how='all', inplace=True)

                    # แสดงผล DataFrame ที่รวมแล้ว
                    # print("Combined DataFrame without empty columns and rows:")
                    # print(df_combined)
                    combined_data.append(df_combined)
                # print("combined_data",combined_data)


            elif "to" in selected_date:
                for meter_stream_no in meter_stream_no_list:
                    meter_stream_no_table = f"""
                    SELECT DISTINCT
                        amr_configured_data.amr_config1 AS DATA_DATE,
                        amr_configured_data.amr_config2,
                        amr_configured_data.amr_config3,
                        amr_configured_data.amr_config4,
                        amr_configured_data.amr_config5,
                        amr_configured_data.amr_config6,
                        amr_configured_data.amr_config7,
                        amr_configured_data.amr_config8,
                        amr_configured_data.amr_config9,
                        amr_configured_data.amr_config10,
                        amr_configured_data.amr_config11,
                        amr_configured_data.amr_config12,
                        amr_configured_data.amr_config13,
                        amr_configured_data.amr_config14,
                        amr_configured_data.amr_config15,
                        amr_configured_data.amr_config16,
                        amr_configured_data.amr_config17,
                        amr_configured_data.amr_config18,
                        amr_configured_data.amr_config19,
                        amr_configured_data.amr_config20
                    FROM
                        AMR_FIELD_ID, AMR_PL_group, AMR_CONFIGURED_DATA, amr_field_meter,amr_region
                    WHERE
                        AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                        AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                        AND amr_field_meter.meter_id = amr_field_id.meter_id
                        AND AMR_CONFIGURED_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                        AND AMR_CONFIGURED_DATA.METER_STREAM_NO LIKE '{meter_stream_no}'
                        AND AMR_CONFIGURED_DATA.DATA_DATE BETWEEN TO_DATE('{formatted_start_date}', 'DD/MM/YYYY') AND TO_DATE('{formatted_end_date}', 'DD/MM/YYYY')
                        AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                        AND amr_region.REGION_NAME = '{selected_region}'
                    ORDER BY 
                        amr_configured_data.amr_config1
                    """
                    meter_stream_no_table_result = fetch_data(ptt_pivot_connection, meter_stream_no_table)
            
                    df_meter_stream_no_table = pd.DataFrame(meter_stream_no_table_result, columns=[
                        'DATA_DATE', 'amr_config2', 'amr_config3', 'amr_config4', 'amr_config5', 
                        'amr_config6', 'amr_config7', 'amr_config8', 'amr_config9', 'amr_config10',
                        'amr_config11', 'amr_config12', 'amr_config13', 'amr_config14', 'amr_config15', 
                        'amr_config16', 'amr_config17', 'amr_config18', 'amr_config19', 'amr_config20'
                    ])

                    # แปลง 'DATA_DATE' ให้เป็น datetime
                    df_meter_stream_no_table['DATA_DATE'] = pd.to_datetime(df_meter_stream_no_table['DATA_DATE'], errors='coerce', dayfirst=True)

                    # เปลี่ยนชื่อคอลัมน์ 'DATA_DATE' เป็นชื่ออื่นเพื่อให้ไม่ซ้ำกับ df_month_list
                    df_meter_stream_no_table.rename(columns={'DATA_DATE': 'DATA_DATE_METER'}, inplace=True)

                    # รวม DataFrame บนคอลัมน์ 'DATA_DATE' ของ df_month_list และ 'DATA_DATE_METER' ของ df_meter_stream_no_table
                    df_combined = pd.merge(df_month_list, df_meter_stream_no_table, left_on='DATA_DATE', right_on='DATA_DATE_METER', how='left')

                    # ลบแถวที่มีค่าวันที่ซ้ำกัน โดยให้เหลือเพียงแถวแรกที่พบ
                    df_combined = df_combined.drop_duplicates(subset='DATA_DATE')

                    # แปลงรูปแบบวันที่ในคอลัมน์ 'DATA_DATE' และ 'DATA_DATE_METER'
                    df_combined['DATA_DATE'] = df_combined['DATA_DATE'].dt.strftime('%d-%b-%Y')
                    df_combined['DATA_DATE_METER'] = df_combined['DATA_DATE_METER'].dt.strftime('%d-%m-%Y')

                    # ลบคอลัมน์ที่มีค่าเป็น None หรือ NaN ทั้งหมด
                    df_combined.dropna(axis=1, how='all', inplace=True)

                    # ลบแถวที่มีค่าเป็น None หรือ NaN ทั้งหมด (ถ้ามี)
                    df_combined.dropna(axis=0, how='all', inplace=True)

                    # แสดงผล DataFrame ที่รวมแล้ว
                    # print("Combined DataFrame without empty columns and rows:")
                    # print(df_combined)
                    combined_data.append(df_combined)
                # print("combined_data",combined_data)
                    
            elif "/" in selected_date and len(selected_date) == 7:
                    # Fetch data from the database
                    for meter_stream_no in meter_stream_no_list:
                        meter_stream_no_table = f"""
                    SELECT DISTINCT
                        amr_configured_data.amr_config1 AS DATA_DATE,
                        amr_configured_data.amr_config2,
                        amr_configured_data.amr_config3,
                        amr_configured_data.amr_config4,
                        amr_configured_data.amr_config5,
                        amr_configured_data.amr_config6,
                        amr_configured_data.amr_config7,
                        amr_configured_data.amr_config8,
                        amr_configured_data.amr_config9,
                        amr_configured_data.amr_config10,
                        amr_configured_data.amr_config11,
                        amr_configured_data.amr_config12,
                        amr_configured_data.amr_config13,
                        amr_configured_data.amr_config14,
                        amr_configured_data.amr_config15,
                        amr_configured_data.amr_config16,
                        amr_configured_data.amr_config17,
                        amr_configured_data.amr_config18,
                        amr_configured_data.amr_config19,
                        amr_configured_data.amr_config20
                    FROM
                        AMR_FIELD_ID, AMR_PL_group, AMR_CONFIGURED_DATA, amr_field_meter,amr_region
                    WHERE
                        AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
                        AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                        AND amr_field_meter.meter_id = amr_field_id.meter_id
                        AND AMR_CONFIGURED_DATA.METER_ID = AMR_FIELD_ID.METER_ID
                        AND AMR_CONFIGURED_DATA.METER_STREAM_NO LIKE '{meter_stream_no}'
                        AND TO_CHAR(AMR_CONFIGURED_DATA.DATA_DATE, 'MM/YYYY') = '{selected_date}'
                        AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
                        AND amr_region.REGION_NAME = '{selected_region}'
                    ORDER BY 
                        amr_configured_data.amr_config1
                    """
                        meter_stream_no_table_result = fetch_data(ptt_pivot_connection, meter_stream_no_table)
                
                        df_meter_stream_no_table = pd.DataFrame(meter_stream_no_table_result, columns=[
                            'DATA_DATE', 'amr_config2', 'amr_config3', 'amr_config4', 'amr_config5', 
                            'amr_config6', 'amr_config7', 'amr_config8', 'amr_config9', 'amr_config10',
                            'amr_config11', 'amr_config12', 'amr_config13', 'amr_config14', 'amr_config15', 
                            'amr_config16', 'amr_config17', 'amr_config18', 'amr_config19', 'amr_config20'
                        ])



                        # แปลง 'DATA_DATE' ให้เป็น datetime
                        df_meter_stream_no_table['DATA_DATE'] = pd.to_datetime(df_meter_stream_no_table['DATA_DATE'], errors='coerce', dayfirst=True)

                        # เปลี่ยนชื่อคอลัมน์ 'DATA_DATE' เป็นชื่ออื่นเพื่อให้ไม่ซ้ำกับ df_month_list
                        df_meter_stream_no_table.rename(columns={'DATA_DATE': 'DATA_DATE_METER'}, inplace=True)

                        # รวม DataFrame บนคอลัมน์ 'DATA_DATE' ของ df_month_list และ 'DATA_DATE_METER' ของ df_meter_stream_no_table
                        df_combined = pd.merge(df_month_list, df_meter_stream_no_table, left_on='DATA_DATE', right_on='DATA_DATE_METER', how='left')

                        # ลบแถวที่มีค่าวันที่ซ้ำกัน โดยให้เหลือเพียงแถวแรกที่พบ
                        df_combined = df_combined.drop_duplicates(subset='DATA_DATE')

                        # แปลงรูปแบบวันที่ในคอลัมน์ 'DATA_DATE' และ 'DATA_DATE_METER'
                        df_combined['DATA_DATE'] = df_combined['DATA_DATE'].dt.strftime('%d-%b-%Y')
                        df_combined['DATA_DATE_METER'] = df_combined['DATA_DATE_METER'].dt.strftime('%d-%m-%Y')

                        # ลบคอลัมน์ที่มีค่าเป็น None หรือ NaN ทั้งหมด
                        df_combined.dropna(axis=1, how='all', inplace=True)

                        # ลบแถวที่มีค่าเป็น None หรือ NaN ทั้งหมด (ถ้ามี)
                        df_combined.dropna(axis=0, how='all', inplace=True)

                        # แสดงผล DataFrame ที่รวมแล้ว
                        # print("Combined DataFrame without empty columns and rows:")
                        # print(df_combined)
                        combined_data.append(df_combined)
                    # print("combined_data",combined_data)
                 
                
            def replace_thead_with_first_row(html_table):
                soup = BeautifulSoup(html_table, 'html.parser')
                
                # Extract the first row from tbody
                first_row = soup.tbody.find('tr')
                
                # Find the thead and replace its content with the first row
                soup.thead.clear()  # Remove existing content in thead
                soup.thead.append(first_row)  # Add the first row as the new thead content
                
                return str(soup)
            
            def style_first_row(html_table, color, font_size):
                """
                Modify the HTML table to make the first row bold, add a background color,
                increase font size, and set equal column widths.
                Additionally, hide the first column in all rows.
                """
                styled_table = html_table.replace(
                    '<tr>',
                    f'<tr style="font-weight: bold; background-color: {color}; font-size: {font_size}; text-align: center;">',
                    1
                ).replace(
                    '<td>',
                    '<td style="width: 100px; text-align: center;">'
                )
                
                # Add a <style> block to hide the first td in every tr
                styled_table += '''
                <style>
                    td:first-child { display: none; }
                    th:first-child { display: none; }
                </style>
                '''
                
                return styled_table

            background_color = '#f0f0f0'  
            font_size = 'larger'  

            
            def generate_html_tables(amr_vc_type_list, combined_data, df_month_list, background_color, font_size, classes=None):
                html_tables = []
                data_hourly = []
                if 'data_config' in session:
                    session.pop('data_config', None)

                for i in range(len(amr_vc_type_list)):
                    # Extract header and data DataFrames
                    header_row_df = combined_data[i]
                    data_df = combined_data[i + len(amr_vc_type_list)]

                    # Concatenate header_row_df and data_df if needed
                    if len(data_df.columns) != len(header_row_df.columns):
                        if len(data_df.columns) < len(header_row_df.columns):
                            for _ in range(len(header_row_df.columns) - len(data_df.columns)):
                                data_df[len(data_df.columns)] = ""
                        else:
                            for _ in range(len(data_df.columns) - len(header_row_df.columns)):
                                header_row_df[len(header_row_df.columns)] = ""

                    # Ensure columns are aligned
                    data_df.columns = header_row_df.columns

                    # Convert DataFrame to HTML with optional class
                    html_content = pd.concat([header_row_df, data_df]).to_html(index=False, classes=classes)
                    # print(html_content)
                    data_hourly.append(html_content)
                    # print("test",data_hourly)
                    # Style the first row of HTML table
                    html_table = style_first_row(html_content, background_color, font_size)

                    # Replace the thead with the first row
                    updated_html_table = replace_thead_with_first_row(html_table)
                    
                    # Append the updated table to the list
                    html_tables.append(updated_html_table)
                
                
                with open('data_report_user_group.json', 'w') as json_file:
                    json.dump(data_hourly, json_file, indent=4)
                return html_tables
            # session['data_config']=html_tables
            
            # Assuming `style_first_row` is defined elsewhere
            html_tables = generate_html_tables(amr_vc_type_list, combined_data, df_month_list, background_color, font_size, classes='data_config')
            
            html_tables_dict = {}
            vc_names_dict = {}

            # Assign values to html_tables_dict
            for i in range(len(html_tables)):
                if i < 6:  # Ensure we don't go beyond the number of expected tables
                    html_tables_dict[f'html_table{i+1}'] = html_tables[i]
                                                
            for i in range(len(amr_vc_type_list)):
                if i < 6:  # Ensure we don't go beyond the number of expected lists
                    vc_names_dict[f'vc_name_list{i+1}'] = globals().get(f'vc_name_list{i+1}', None)
                    
 
            if "/" in selected_date and len(selected_date) == 10:   
                # print("METER_STREAM_NO:", meter_stream_no_list)
                # print("amr_vc_type:", amr_vc_type_list)
                print("selected_date:",selected_date)
                if selected_date:
                    # Convert selected_date to 'YYYY-MM-DD' format for consistency
                    selected_date_formatted = pd.to_datetime(selected_date, format='%d/%m/%Y').strftime('%d/%m/%Y')

                    # Get the current date in 'YYYY-MM-DD' format
                    current_date = datetime.datetime.now().strftime('%d-%m-%Y')

                    # Determine if the selected date is the current date
                    is_current_day = selected_date_formatted == current_date

                    # Update the query to use the selected date
                    if is_current_day:
                        # If the selected date is today, show only today's data
                        query_day = f"""
                            SELECT TO_CHAR(TO_DATE('{current_date}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                            FROM DUAL
                        """

                    else:
                        # If the selected date is not today, show only the selected day's data
                        query_day = f"""
                            SELECT TO_CHAR(TO_DATE('{selected_date_formatted}', 'DD-MM-YYYY'), 'DD-MM-YYYY') AS Date1
                            FROM DUAL
                        """

            elif "to" in selected_date:

                if formatted_start_date and formatted_end_date:
                    # แปลงวันที่ให้เป็นรูปแบบ 'YYYY-MM-DD' สำหรับใช้ใน SQL
                    start_date_formatted = pd.to_datetime(formatted_start_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                    end_date_formatted = pd.to_datetime(formatted_end_date, format='%d/%m/%Y').strftime('%Y-%m-%d')
                    
                    print("Start Date:", formatted_start_date)
                    print("End Date:", formatted_end_date)
                    
                    # สร้างรายการวันที่ในช่วงที่เลือกโดยใช้ SQL
                    query_day = f"""
                        SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{start_date_formatted}', 'YY-MM-DD'), 'DD/MM/YY') AS DATA_DATE
                        FROM DUAL
                        CONNECT BY LEVEL <= TO_DATE('{end_date_formatted}', 'YY-MM-DD') - TO_DATE('{start_date_formatted}', 'YY-MM-DD') + 1
                    """

            elif "/" in selected_date and len(selected_date) == 7:


                selected_date = request.args.get("date_dropdown_daily")
                

                # Initialize df_month_list outside the if statement
                df_month_list = pd.DataFrame(columns=['DATA_DATE'])
                
                # Check if 'selected_date' is available
                if selected_date:
                    # Convert selected_date to 'YYYY-MM' format for consistency
                    selected_date_formatted = pd.to_datetime(selected_date, format='%m/%Y').strftime('%Y-%m')
                    # print("date2:", selected_date_formatted)

                    # Get the current month and year
                    current_month_year = datetime.datetime.now().strftime('%Y-%m')

                    # Determine if the selected date is in the current month
                    is_current_month = selected_date_formatted == current_month_year
                    
                    # Update the query to use the selected date
                    if is_current_month:
                        # If the selected date is in the current month, show all days up to the current date
                        query_day = f"""
                            SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{current_month_year}-01', 'YYYY-MM-DD')) AS Date1
                            FROM DUAL
                            CONNECT BY LEVEL <= TO_DATE('{datetime.datetime.now().strftime('%Y-%m-%d')}', 'YYYY-MM-DD') - TO_DATE('{current_month_year}-01', 'YYYY-MM-DD') + 1
                        """

                        
                    else:
                        # If the selected date is in a previous month, show all days of the selected month
                        query_day = f"""
                            SELECT TO_CHAR(TRUNC(LEVEL - 1) + TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) AS Date1
                            FROM DUAL
                            CONNECT BY LEVEL <= LAST_DAY(TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD')) - TO_DATE('{selected_date_formatted}-01', 'YYYY-MM-DD') + 1
                        """
            # Fetch data for the month list
            query_day_result = fetch_data(ptt_pivot_connection, query_day)
            df_month_list = pd.DataFrame(query_day_result, columns=['DATA_DATE'])
                
            selected_meter_id = None

            # Check if 'METER_ID' column exists and the DataFrame is not empty
            if not df.empty and 'METER_ID' in df.columns and len(df['METER_ID']) > 0:
                selected_meter_id = df['METER_ID'].iloc[0]
                print(f"Selected Meter ID: {selected_meter_id}")
            else:
                print("DataFrame is empty or 'METER_ID' column doesn't exist.")

            # Now, remove the "METER_ID" column from the DataFrame
            df = df.drop(["PL_REGION_ID", "TAG_ID", "METER_ID"], axis=1)

            # Remove newline characters
            df = df.apply(lambda x: x.str.replace("\n", "") if x.dtype == "object" else x)

            df = df.drop_duplicates(subset=["DATA_DATE", "METER_STREAM_NO"], keep="first")
            # Sort DataFrame by 'DATA_DATE'
            df = df.sort_values(by="DATA_DATE")
            

            num_streams = 6
            df_runs = {}

            # Loop to create DataFrames for each METER_STREAM_NO
            for i in range(1, num_streams + 1):
                df_runs[f'df_run{i}'] = df[df['METER_STREAM_NO'] == str(i)]

            # Check if each DataFrame has data before including in the tables dictionary
            # Create Full table of selectedMonth
            tables = {
                "daily_data": None,
                
            }
            
            
                
            

            # Merge DataFrames using a loop
            for i in range(1, num_streams + 1):
                df_run = df_runs[f'df_run{i}']
                
                if not df_run.empty:
                    merged_df = pd.merge(df_month_list, df_run, on='DATA_DATE', how='outer')
                    df_runs[f'df_run{i}'] = merged_df
                
            # common_table_properties = {"classes": "data", "index": False, "header": None, "na_rep": "N/A"}
            
            # print(common_table_properties)

            for i in range(1, num_streams + 1):
                df_run = df_runs[f'df_run{i}']
                
                if not df_run.empty:
                    df_run = df_run.drop('METER_STREAM_NO', axis=1, errors='ignore')
                    df_run = df_run.fillna("N/A")
                    # print("df_run",df_run)
                    tables[f"config_data_run{i}"] = df_run.to_html(classes='data_config',index=False, border=0,header=None,na_rep="N/A")
                    # print(tables)

            return render_template(
                "billingdata_user_group.html",
                tables=tables,
                titles=df.columns.values,
                selected_date=selected_date,
                selected_tag=selected_tag,
                selected_region=selected_region,
                region_options=region_options,
                tag_options=tag_options,
                **{k: html_tables_dict.get(k, None) for k in [f'html_table{i+1}' for i in range(6)]},
                vc_name_dict=vc_name_dict,
                selected_meter_id=selected_meter_id,username=username
            )
        else:
            # Render the template without executing the query
            return render_template(
                "billingdata_user_group.html",
                selected_date=selected_date,
                selected_region=selected_region,
                selected_tag=selected_tag,
                region_options=region_options,
                tag_options=tag_options,
                tables={},
                username=username,  
                description=description,
                user_level=user_level
            )
                        
          

def process_form_data( selected_tag, selected_region, poll_type):
    # Here you can process or store your data as needed
    
    print("Tag Dropdown:", selected_tag)
    print("Region Dropdown:", selected_region)
    print("Poll type Dropdown:", poll_type)
    # Return or use these values as needed
    return  selected_tag, selected_region, poll_type


############ Manualpoll_data  #####################
@app.route("/Manualpoll_data", methods=["GET", "POST"])
@login_required
def Manualpoll_data():
    if 'username' not in session:
        return redirect(url_for('login'))
            
    selected_tag = request.args.get("tag_dropdown")
    selected_region = request.args.get("region_dropdown")
    poll_type = request.args.get("display_dropdown")
        
    session['selected_tag'] = selected_tag
    # print(session['selected_tag'])
    session['selected_region'] = selected_region
    session['poll_type'] = poll_type

    
    username = session['username']
    process_form_data(selected_tag, selected_region, poll_type)
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:

        region_query = """
            SELECT * FROM AMR_REGION 
        """
        
        tag_query = """
            SELECT DISTINCT AMR_FIELD_ID.TAG_ID
        FROM AMR_FIELD_ID,amr_region,AMR_PL_GROUP
        
        WHERE AMR_PL_GROUP.PL_REGION_ID = amr_region.id
        AND AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
        AND amr_region.REGION_NAME = :region_id
        AND AMR_FIELD_ID.tag_id NOT like '%.remove%'
        ORDER BY  TAG_ID
            """
        region_results = fetch_data(ptt_pivot_connection, region_query)
        region_options = [str(region[1]) for region in region_results]

        if poll_type == "1":
            query = """
                SELECT
                    AMR_FIELD_METER.METER_STREAM_NO as RunNo,
                    AMR_PL_GROUP.PL_REGION_ID as region,
                    AMR_FIELD_ID.TAG_ID as Sitename,
                    AMR_FIELD_METER.METER_NO_STREAM as NoRun,
                    AMR_FIELD_METER.METER_ID as METERID,
                    AMR_VC_TYPE.VC_NAME as VCtype,
                    AMR_FIELD_ID.SIM_IP as IPAddress,
                    AMR_PORT_INFO.PORT_NO as port,
                    amr_poll_range.evc_type as evc_type,
                    amr_vc_type.vc_name as vc_name,
                    amr_poll_range.poll_billing as poll_billing,
                    amr_poll_range.poll_config as poll_config,
                    amr_poll_range.poll_billing_enable as poll_billing_enable,
                    amr_poll_range.poll_config_enable as poll_config_enable,
                    amr_field_meter.modbus_id as modbus_id,
                    AMR_PORT_INFO.ID as port_id
                FROM
                    AMR_FIELD_ID,
                    AMR_USER,
                    AMR_FIELD_CUSTOMER,
                    AMR_FIELD_METER,
                    AMR_PL_GROUP,
                    AMR_VC_TYPE,
                    AMR_PORT_INFO,
                    amr_poll_range,
                    amr_region
                WHERE
                    AMR_PL_GROUP.PL_REGION_ID = amr_region.id AND
                    AMR_USER.USER_ENABLE=1 AND
                    AMR_VC_TYPE.ID=AMR_POLL_RANGE.EVC_TYPE AND
                    AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID AND
                    AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP AND
                    AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID AND
                    AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID AND
                    AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE AND
                    AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID
                    {tag_condition}
                    {region_condition}
            """
            tag_condition = "AND AMR_FIELD_ID.TAG_ID IS NOT NULL"
            region_condition = "AND amr_region.REGION_NAME = 'default_region_id'"
            region_results = fetch_data(ptt_pivot_connection, region_query)
            region_options = [str(region[1]) for region in region_results]
            tag_results = fetch_data(ptt_pivot_connection, tag_query, params={"region_id": selected_region})
            tag_options = [str(tag[0]) for tag in tag_results]
            tag_options.sort()
            if selected_tag:
                tag_condition = f"AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'"
            if selected_region:
                region_condition = f"AND amr_region.REGION_NAME = '{selected_region}'"
            query = query.format(tag_condition=tag_condition, region_condition=region_condition)
            results = fetch_data(ptt_pivot_connection, query)
            #print(query)

            df = pd.DataFrame(
                results,
                columns=[
                    "RUN",
                    "Region",
                    "Sitename",
                    "NoRun",
                    "METERID",
                    "VCtype",
                    "IPAddress",
                    "Port",
                    "evc_type",
                    "vc_name",
                    "poll_billing",
                    "poll_config",
                    "poll_billing_enable",
                    "poll_config_enable",
                    "modbus_id",
                    "port_id"
                ],
            )
            columns_to_drop = [
                "NoRun",
                "vc_name",
                "poll_billing",
                "poll_config",
                "poll_billing_enable",
                "poll_config_enable",
                "modbus_id",
                "port_id"
            ]

            # Drop the columns from the DataFrame
            df = df.drop(columns=columns_to_drop)
            # print("RTTRTRTRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRRR")
            # print(df)
        else:
            query = """
                SELECT
                    AMR_FIELD_METER.METER_STREAM_NO as RunNo,
                    AMR_PL_GROUP.PL_REGION_ID as region,
                    AMR_FIELD_ID.TAG_ID as Sitename,
                    AMR_FIELD_METER.METER_NO_STREAM as NoRun,
                    AMR_FIELD_METER.METER_ID as METERID,
                    AMR_VC_TYPE.VC_NAME as VCtype,
                    AMR_FIELD_ID.SIM_IP as IPAddress,
                    AMR_PORT_INFO.PORT_NO as port,
                    amr_poll_range_hourly.evc_type as evc_type,
                    amr_vc_type.vc_name as vc_name,
                    amr_poll_range_hourly.poll_hourly as poll_hourly,
                    amr_poll_range_hourly.poll_hourly_enable as poll_hourly_enable,
                    amr_field_meter.modbus_id as modbus_id,
                    AMR_PORT_INFO.ID as port_id
                FROM
                    AMR_FIELD_ID,
                    AMR_USER,
                    AMR_FIELD_CUSTOMER,
                    AMR_FIELD_METER,
                    AMR_PL_GROUP,
                    AMR_VC_TYPE,
                    AMR_PORT_INFO,
                    amr_poll_range_hourly,
                    amr_region
                WHERE
                    AMR_PL_GROUP.PL_REGION_ID = amr_region.id AND
                    AMR_USER.USER_ENABLE=1 AND
                    AMR_VC_TYPE.id=AMR_POLL_RANGE.EVC_TYPE AND
                    AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID AND
                    AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP AND
                    AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID AND
                    AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID AND
                    AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE AND
                    AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID
                    {tag_condition}
                    {region_condition}
            """
            print("hourly")
            tag_condition = "AND AMR_FIELD_ID.TAG_ID IS NOT NULL"
            region_condition = "AND amr_region.REGION_NAME = 'default_region_id'"
            region_results = fetch_data(ptt_pivot_connection, region_query)
            region_options = [str(region[1]) for region in region_results]
            tag_results = fetch_data(ptt_pivot_connection, tag_query, params={"region_id": selected_region})
            tag_options = [str(tag[0]) for tag in tag_results]
            tag_options.sort()
            if selected_tag:
                tag_condition = f"AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'"
            if selected_region:
                region_condition = f"AND amr_region.REGION_NAME = '{selected_region}'"
            query = query.format(tag_condition=tag_condition, region_condition=region_condition)
            results = fetch_data(ptt_pivot_connection, query)
           
            df = pd.DataFrame(
                results,
                columns=[
                    "RUN",
                    "Region",
                    "Sitename",
                    "NoRun",
                    "METERID",
                    "VCtype",
                    "IPAddress",
                    "Port",
                    "evc_type",
                    "vc_name",
                    "poll_hourly",
                    "poll_hourly_enable",
                    "modbus_id",
                    "port_id"
                ],
            )
            columns_to_drop = [
                "NoRun",
                "vc_name",
                "poll_hourly",
                "poll_hourly_enable",
                "modbus_id",
                "port_id"
            ]

            # Drop the columns from the DataFrame
            df = df.drop(columns=columns_to_drop)
        test_poll_type = "Select Poll Type"
        if poll_type == '1':
            test_poll_type = "Daily"
        elif poll_type == '2':
            test_poll_type = "Hourly"
        
        return render_template(
            "Manual poll.html",
            tables=[df.to_html(classes="data", index=False)],
            titles=df.columns.values.tolist(),
            
            selected_tag=selected_tag,
            selected_region=selected_region,
            region_options=region_options,
            tag_options=tag_options,
            df=df,test_poll_type=test_poll_type,poll_type=poll_type,username=username
        )
    
################## popup analys ###################################

@app.route('/get_file')
def get_file():
    if os.path.exists(FILE_PATH):
        with open(FILE_PATH, 'r') as file:
            data = file.read()
        return jsonify(data=data)
    else:
        return jsonify(data="File not found"), 404
    
@app.route('/clear_data', methods=['POST'])
def clear_data():
    if os.path.exists(FILE_PATH):
        with open(FILE_PATH, 'w') as file:
            file.write("")
        return jsonify(message="File data cleared"), 200
    else:
        return jsonify(message="File not found"), 404


   
#################### manualpoll ####################
def process_form_data_selected_row( selected_row):
    # Here you can process or store your data as needed
    return  selected_row

# read site data when select region-> site -> mode
@app.route("/read_data", methods=["POST"])
@login_required
def read_data():
    try:
        
        # if 'username' not in session:
        #     return redirect(url_for('login'))
        
        selected_row = request.form.get("selected_row")
        
        session['selected_row'] = selected_row
        selected_tag = session.get('selected_tag')
        selected_region = session.get('selected_region')
        poll_type = session.get('poll_type')
        username = session['username']
         
        process_form_data( selected_tag, selected_region, poll_type)
        if not os.path.exists(FILE_PATH):
            with open(FILE_PATH, 'w') as file:
                file.write("")

        
        with connect_to_ptt_pivot_db() as ptt_pivot_connection:
           
            global change_to_32bit_counter  # Use the global variable
            slave_id = int(1)
            function_code = int(3)
            
                       
            region_query = """
                SELECT * FROM AMR_REGION 
            """
            
            tag_query = """
                SELECT DISTINCT AMR_FIELD_ID.TAG_ID
                FROM AMR_FIELD_ID,amr_region,AMR_PL_GROUP
                
                WHERE AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                AND AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
                AND amr_region.REGION_NAME = :region_id
                AND AMR_FIELD_ID.tag_id NOT like '%.remove%'
                ORDER BY  TAG_ID
                """
            region_results = fetch_data(ptt_pivot_connection, region_query)
            region_options = [str(region[1]) for region in region_results]

##################################################

            if poll_type == "1":  #### daily
                query = """
                    SELECT
                        AMR_FIELD_METER.METER_STREAM_NO as RunNo,
                        AMR_PL_GROUP.PL_REGION_ID as region,
                        AMR_FIELD_ID.TAG_ID as Sitename,
                        AMR_FIELD_METER.METER_NO_STREAM as NoRun,
                        AMR_FIELD_METER.METER_ID as METERID,
                        AMR_VC_TYPE.VC_NAME as VCtype,
                        AMR_FIELD_ID.SIM_IP as IPAddress,
                        AMR_PORT_INFO.PORT_NO as port,
                        amr_poll_range.evc_type as evc_type,
                        amr_vc_type.vc_name as vc_name,
                        amr_poll_range.poll_billing as poll_billing,
                        amr_poll_range.poll_config as poll_config,
                        amr_poll_range.poll_billing_enable as poll_billing_enable,
                        amr_poll_range.poll_config_enable as poll_config_enable,
                        amr_field_meter.modbus_id as modbus_id,
                        AMR_PORT_INFO.ID as port_id
                    FROM
                        AMR_FIELD_ID,
                        AMR_USER,
                        AMR_FIELD_CUSTOMER,
                        AMR_FIELD_METER,
                        AMR_PL_GROUP,
                        AMR_VC_TYPE,
                        AMR_PORT_INFO,
                        amr_poll_range,
                        amr_region
                    WHERE
                        AMR_PL_GROUP.PL_REGION_ID = amr_region.id AND
                        AMR_USER.USER_ENABLE=1 AND
                        amr_vc_type.id=amr_poll_range.evc_type AND
                        AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID AND
                        AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP AND
                        AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID AND
                        AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID AND
                        AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE AND
                        AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID
                        {tag_condition}
                        {region_condition}
                """
                tag_condition = "AND AMR_FIELD_ID.TAG_ID IS NOT NULL"
                region_condition = "AND amr_region.REGION_NAME = 'default_region_id'"
                region_results = fetch_data(ptt_pivot_connection, region_query)
                region_options = [str(region[1]) for region in region_results]
                tag_results = fetch_data(ptt_pivot_connection, tag_query, params={"region_id": selected_region})
                tag_options = [str(tag[0]) for tag in tag_results]
                tag_options.sort()
                if selected_tag:
                    tag_condition = f"AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'"
                if selected_region:
                    region_condition = f"AND amr_region.REGION_NAME = '{selected_region}'"
                query = query.format(tag_condition=tag_condition, region_condition=region_condition)
                results = fetch_data(ptt_pivot_connection, query)
                df = pd.DataFrame(
                    results,
                    columns=[
                        "RUN",
                        "Region",
                        "Sitename",
                        "NoRun",
                        "METERID",
                        "VCtype",
                        "IPAddress",
                        "Port",
                        "evc_type",
                        "vc_name",
                        "poll_billing",
                        "poll_config",
                        "poll_billing_enable",
                        "poll_config_enable",
                        "modbus_id",
                        "port_id"
                    ],
                )
                
                if selected_row: 
                    # Convert the selected_row to an integer
                        
                        selected_row_index = int(selected_row) - 1
                        
                        # Check if the index is within the valid range
                        if 0 <= selected_row_index < len(df):
                            
                            row_data = df.iloc[selected_row_index]
                            print(row_data)
                            row_data_dict = row_data.to_dict()
                            
                            run = row_data_dict.get("RUN")
                            Region = row_data_dict.get("Region")
                            Sitename = row_data_dict.get("Sitename")
                            NoRun = row_data_dict.get("NoRun")
                            METERID = row_data_dict.get("METERID")
                            VCtype = row_data_dict.get("VCtype")
                            tcp_ip = row_data_dict.get("IPAddress")
                            tcp_port = int(row_data_dict.get("Port"))
                            evc_type = row_data_dict.get("evc_type")
                            poll_billing = row_data_dict.get("poll_billing")
                            poll_billing_enable = row_data_dict.get("poll_billing_enable")
                            poll_config = row_data_dict.get("poll_config")
                            poll_config_enable = row_data_dict.get("poll_config_enable")
                            modbus_id = int(row_data_dict.get("modbus_id"))
                            port_id = int(row_data_dict.get("port_id"))
                            poll_billing_list = [int(x) for x in poll_billing.split(',')]
                            poll_billing_enable_list = poll_billing_enable.split(',')
                            pairs_billing = [(poll_billing_list[i], poll_billing_list[i+1]) for i in range(0, len(poll_billing_list), 2)]

                            poll_config_list = [int(x) for x in poll_config.split(',')]
                            poll_config_enable_list = poll_config_enable.split(',')
                            pairs_config = [(poll_config_list[i], poll_config_list[i+1]) for i in range(0, len(poll_config_list), 2)]

                            # Prepare dataframes for billing and config
                            data= {'starting_address_i': [], 'quantity_i': [], 'adjusted_quantity_i': []}
    
                            df_pollRange = pd.DataFrame(data)
                            df_pollBilling = pd.DataFrame(data)

                            for index, value in enumerate(poll_billing_enable_list):
                                if value == '1' and index < len(pairs_billing):
                                    starting_address_i = int(pairs_billing[index][0])  # Convert to integer
                                    quantity_i = int(pairs_billing[index][1])          # Convert to integer
                                    adjusted_quantity_i = quantity_i - starting_address_i + 1
                                    data = {'starting_address_i': [starting_address_i], 
                                            'quantity_i': [quantity_i], 
                                            'adjusted_quantity_i': [adjusted_quantity_i]}
                                    df_2 = pd.DataFrame(data)
                                    df_pollBilling = pd.concat([df_pollBilling, df_2], ignore_index=True)
                                    #print("df_pollBilling", df_pollBilling)

                            # Process config pairs
                            #print("\nConfig Pairs:")
                            for index, value in enumerate(poll_config_enable_list):
                                if value == '1' and index < len(pairs_config):
                                    starting_address_i = int(pairs_config[index][0])  # Convert to integer
                                    quantity_i = int(pairs_config[index][1])          # Convert to integer
                                    adjusted_quantity_i = quantity_i - starting_address_i + 1
                                    data = {'starting_address_i': [starting_address_i], 
                                            'quantity_i': [quantity_i], 
                                            'adjusted_quantity_i': [adjusted_quantity_i]}
                                    df_2 = pd.DataFrame(data)
                                    df_pollRange = pd.concat([df_pollRange, df_2], ignore_index=True)
                                    # print("df_pollRange", df_pollRange)

                            dataframes = {
                                    'address_start': [],
                                    'finish': [],
                                    'TX': [],
                                    'RX': []
                                }
                            df_Modbus = pd.DataFrame(dataframes)
                            df_Modbusbilling = pd.DataFrame(dataframes)
                            
                            try:
                                sock_i = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                                sock_i.settimeout(30)  
                                
                                sock_i.connect((tcp_ip, tcp_port))
                                #print("Connected successfully.")
                                with open(FILE_PATH, 'a') as file:
                                    file.write(f"Connected successfully.\n")
                                    file.write(f"IP ADDRESS:{tcp_ip}\n")
                                    file.write(f"Port:{tcp_port}\n")
                                    file.write(f"modbus_id:{modbus_id}\n\n")
                                    
                            except ConnectionRefusedError:
                                abort(400, f"Error: Connection refused to {tcp_ip}:{tcp_port}!")
                            except TimeoutError:
                                abort(400, f"Error: Connection timed out to {tcp_ip}:{tcp_port}!")
                            except Exception as e:
                                abort(400, f"Error: {e}")
                            ##########

                            # wake up
                            slave_id_1 = 0x01
                            function_code_1 = 0x03
                            starting_address_1 = 0x0004
                            quantity_1 = 0x0002

                            request_Actaris= bytearray([
                                    slave_id_1,
                                    function_code_1,
                                    starting_address_1 >> 8,
                                    starting_address_1 & 0xFF,
                                    quantity_1 >> 8,
                                    quantity_1 & 0xFF,
                                ])

                            crc_1 = computeCRC(request_Actaris)
                            request_Actaris += crc_1
                            #######
                            #if int(evc_type) in [5, 8, 9, 10,13]:
                            if int(evc_type) in [5, 8, 9, 10]:
                                
                                for _ in range(2):  
                                    sock_i.send(request_Actaris)
                                    time.sleep(3)
                                response = sock_i.recv(4096)
                                
                            ######## 
                            if int(evc_type) == 12:
                                if int(tcp_port) != 2101:
                                    
                                    for _ in range(2):
                                        sock_i.send(request_Actaris)
                                        time.sleep(3)
                                    response = sock_i.recv(4096)
                            ### Wake up

                            
                            current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            with open(FILE_PATH, 'a') as file:
                                file.write(f"{current_time}:Start polling Configuration\n")
                                
                            for i in range(0, len(df_pollRange)):
                                
                                if int(tcp_port) == 2101 and int(port_id) in [15, 16]:
                                    sock_i = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

                                    sock_i.settimeout(5)
                                    sock_i.connect((tcp_ip, int(tcp_port)))

                                start_address = int(df_pollRange.loc[i,'starting_address_i'])
                                
                                adjusted_quantity = int(df_pollRange.loc[i,'adjusted_quantity_i'])
                                
                               
                                request_message_i = bytearray(
                                [modbus_id, function_code, start_address >> 8, start_address & 0xFF, adjusted_quantity >> 8, adjusted_quantity & 0xFF])
                                crc_i = computeCRC(request_message_i)
                                request_message_i += crc_i
                                
                                communication_traffic_i = []
                                
                                communication_traffic_i.append(request_message_i.hex())
                                config_safe_tx = f"config_TX: {communication_traffic_i[0]}"
                                
                                with open(FILE_PATH, 'a') as file:
                                    file.write(f"Poll NO {i}\n")
                                    file.write(f"{current_time}\n")
                                    file.write(f"TX-config:{communication_traffic_i[0]}\n")
                                # logger_info.info(config_safe_tx)
                                try:
                                    
                                    sock_i.send(request_message_i)
                                    
                                    time.sleep(1)
                                    response_i = sock_i.recv(1024)
                                        
                                    communication_traffic_i.append(response_i.hex())
                                    
                                    config_safe_RX = f"config_RX: {communication_traffic_i[1]}"
                                
                                    # print("RX.config",communication_traffic_i[1])
                                    with open(FILE_PATH, 'a') as file:
                                        file.write(f"RX-config:{communication_traffic_i[1]}\n\n")
                            
                                    
                                except TimeoutError:
                                    abort(400, f"Error: Connection timed out while waiting for response from {tcp_ip}:{tcp_port}!")
                                except Exception as e:
                                    abort(400, f"Error: {e}")
                                if response_i[1:2] != b'\x03':
                                    
                                    abort(400, f"Error: Unexpected response code from device {communication_traffic_i[1]} ,{response_i[1:2]}!")
                                else:
                                    pass
                                data = {
                                    'address_start': [int(start_address)],
                                    'finish': [int(start_address+adjusted_quantity)],
                                    'TX': [communication_traffic_i[0]],
                                    'RX': [communication_traffic_i[1]]
                                }
                                # print(data)
                                df_2 = pd.DataFrame(data)
                                df_Modbus = pd.concat([df_Modbus, df_2], ignore_index=True)


                            with open(FILE_PATH, 'a') as file:
                                file.write(f"Finish Polling Config\n\n")
                            ##############   billing
                            with open(FILE_PATH, 'a') as file:
                                file.write(f"{current_time}:Start polling Logged Data\n")

                            # if int(evc_type) == 12:
                            #     if int(tcp_port) != 2101:
                                    
                            #         for _ in range(2):
                            #             sock_i.send(request_Actaris)
                            #             print(sock_i)
                            #             time.sleep(0.5)
                            #         response = sock_i.recv(4096)
                            for i in range(0, len(df_pollBilling)):
                                
                                
                                if int(tcp_port) == 2101 and int(port_id) in [15, 16]:
                                    
                                    sock_i = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                                    sock_i.settimeout(5)
                                    sock_i.connect((tcp_ip, int(tcp_port)))
                                    
                                
                                start_address = int(df_pollBilling.loc[i,'starting_address_i'])
                                
                                adjusted_quantity = int(df_pollBilling.loc[i,'adjusted_quantity_i'])
                            
                                request_message_i = bytearray(
                                [modbus_id, function_code, start_address >> 8, start_address & 0xFF, adjusted_quantity >> 8, adjusted_quantity & 0xFF])
                                crc_i = computeCRC(request_message_i)
                                request_message_i += crc_i
                                

                                communication_traffic_i = []
                            
                                
                                communication_traffic_i.append(request_message_i.hex())
                                billing_safe_tx = f"billing_TX: {communication_traffic_i[0]}"
                                # print("tx.billing",communication_traffic_i[0])

                                
                                with open(FILE_PATH, 'a') as file:
                                    file.write(f"Poll NO {i}\n")
                                    file.write(f"{current_time}\n")
                                    file.write(f"TX-billing:{communication_traffic_i[0]}\n")
                                if int(evc_type) == 12:
                                    try:
                                        
                                        time.sleep(2)  
                                        sock_i.send(request_message_i)
                                        response_i = sock_i.recv(4096)

                                        communication_traffic_i.append(response_i.hex())
                                        with open(FILE_PATH, 'a') as file:
                                            file.write(f"RX-billing:{communication_traffic_i[1]}\n\n")
                                        billing_safe = f"billing_RX: {communication_traffic_i[1]}"
                                        # print("rx.billing",communication_traffic_i[1])
                                        
   
                                    except TimeoutError:
                                        abort(400, f"Error: Connection timed out while waiting for response from {tcp_ip}:{tcp_port}!")
                                    except Exception as e:
                                        abort(400, f"Error: {e}")

                                else:
                                    try:

                                        sock_i.send(request_message_i)
                                        time.sleep(1)
                                        response_i = sock_i.recv(4096)

                                        communication_traffic_i.append(response_i.hex())
                                        with open(FILE_PATH, 'a') as file:
                                            file.write(f"RX-billing:{communication_traffic_i[1]}\n\n")
                                        billing_safe = f"billing_RX: {communication_traffic_i[1]}"
                                    
                                    except TimeoutError:
                                        abort(400, f"Error: Connection timed out while waiting for response from {tcp_ip}:{tcp_port}!")
                                    except Exception as e:
                                        abort(400, f"Error: {e}")
                                
                                if response_i[1:2] != b'\x03':
                                    
                                    abort(400, f"Error: Unexpected response code from device {communication_traffic_i[1]} ,{response_i[1:2]}!")
                                else:
                                    pass
                            

                                data = {
                                    'address_start': [int(start_address)],
                                    'finish': [int(start_address+adjusted_quantity-1)],
                                    'TX': [communication_traffic_i[0]],
                                    'RX': [communication_traffic_i[1]]
                                }
                                # print(data)
                                df_2 = pd.DataFrame(data)
                                df_Modbusbilling = pd.concat([df_Modbusbilling, df_2], ignore_index=True)

                                
                            with open(FILE_PATH, 'a') as file:
                                file.write(f"{current_time}:Polling successfully.\n")  
                            
                            encode_data_dict = {} 
                            df_Modbus_all = pd.concat([df_Modbus,df_Modbusbilling], ignore_index=True)
                            df_Modbus_html = df_Modbus_all.to_html(classes="data", index=False)

                            ######
                            # query Mapping and encode data
                            # fetch mapping    
                            evc_type = evc_type
                            query = """
                            select amc.or_der as order1 , amc.address as address1, amc.description as desc1, amc.data_type as dtype1
                            from amr_mapping_config amc
                            where amc.evc_type = :evc_type AND address is not null 
                            order by order1
                            """
                            poll_results = fetch_data(ptt_pivot_connection,query, params={"evc_type": evc_type})
                            df_mapping = pd.DataFrame(poll_results, columns=['order', 'address', 'desc', 'data_type'])
                                    
                            list_of_values_configured = []
                            
                            for i in range(0, len(df_mapping)):
                                    
                                address = int(df_mapping.iloc[i,1])
                                
                                data_type = str(df_mapping.iloc[i,3])
                                
                                for j in range(0,len(df_Modbus)):
                                    address_start = int(df_Modbus.iloc[j,0])
                                    address_finish = int(df_Modbus.iloc[j,1])
                                    #print(address)
                                    if address >= address_start and address <= address_finish:
                                        # print(address_start, address_finish, df_Modbus.iloc[j,3])
                                        location_data = (address - address_start)*int(8/2)
                                        frameRx = (df_Modbus.iloc[j,3])
                                        #
                                        if data_type == "EVODate":
                                            raw_data = frameRx[location_data + 6: location_data + 18]
                                        else:
                                            raw_data = frameRx[location_data + 6: location_data + 14]

                                        #list_of_values_configured.append(convert_raw_to_value_n(data_type,raw_data))
                                        
                                        list_of_values_configured.append(convert_raw_to_value(data_type,raw_data, mode="normal" ))
                                        # print(list_of_values_configured)
                                        break
                                    elif address == 0:
                                        list_of_values_configured.append('0')
                                        break
                                # print(list_of_values_configured)

                            ### list_of_value_billing
                            evc_type = evc_type
                            query = """
                            SELECT amb.daily ,amb.or_der ,amb.address,amb.description,amb.data_type  FROM amr_mapping_billing amb WHERE amb.evc_type = :evc_type AND address is not null order by amb.daily
                            ,amb.or_der
                            """
                            poll_resultsbilling = fetch_data(ptt_pivot_connection,query, params={"evc_type": evc_type})
                            # print(poll_resultsbilling)
                            df_mappingbilling = pd.DataFrame(poll_resultsbilling, columns=['daily','or_der', 'address', 'description', 'data_type'])
                            
                            list_of_values_billing = []
                            for i in range(0, len(df_mappingbilling)):
                                    
                                address = int(df_mappingbilling.iloc[i,2])
                                data_type = str(df_mappingbilling.iloc[i,4])
                                # print(frameRx)
                                for j in range(0,len(df_Modbusbilling)):
                                    address_start = int(df_Modbusbilling.iloc[j,0])
                                    address_finish = int(df_Modbusbilling.iloc[j,1])
                                
                                    if address >= address_start and address <= address_finish:
                                        # print(address)
                                        location_data = (address - address_start)*int(8/2)
                                        # print(location_data)
                                        frameRx = (df_Modbusbilling.iloc[j,3])
                                        
                                        if data_type == "EVODate":
                                            raw_data = frameRx[location_data + 6: location_data + 18]
                                        else :
                                            raw_data = frameRx[location_data + 6: location_data + 14]
                                        
                                        list_of_values_billing.append(convert_raw_to_value(data_type, raw_data, mode="billing"))
                                        break
                            
                            ###############################config##################################################

                            len_query= f"SELECT or_der FROM amr_mapping_billing WHERE evc_type = 17 AND daily = 1"

                            value_config = pd.DataFrame(list_of_values_configured, columns=['Value'])
                            result_config = pd.concat([df_mapping['desc'], value_config], axis=1)
                            result_config = result_config.transpose()

                            amr_vc_type_table = f"""
                    
                            SELECT DESCRIPTION 
                            FROM amr_mapping_billing 
                            WHERE evc_type LIKE '{evc_type}' 
                            AND DAILY like '1' 
                            ORDER BY OR_DER

                            """ 
                            amr_vc_type_table_result = fetch_data(ptt_pivot_connection, amr_vc_type_table)
                            filtered_result = pd.DataFrame(amr_vc_type_table_result)
                            print(len(filtered_result))
                    
                            transposed_result = filtered_result.T
                            selected_data = transposed_result.iloc[0, 0:len(filtered_result)]  # Extract values to get a flat array
                            
                            print("selected_data",selected_data)

                                        # Convert to HTML with custom CSS for styling
                            result_config_html = """
                                    <style>
                                        table.data_config1 {
                                            width: 100%;
                                            border: 0.2px solid #000000;
                                            
                                        }
                                        table.data_config1 th, table.data td {
                                            
                                            padding: 8px;
                                            font-weight: bold;
                                            color: #000000;
                                        }
                                        table.data_config1 th {
                                            padding-top: 12px;
                                            padding-bottom: 12px;
                                            justify-content: center;
                                            background-color: #4CAF50;
                                            color: white;
                                            font-weight: bold;
                                            
                                        }
                                        table.data_config1 tr:nth-child(1) {
                                            background-color: #f2f2f2;
                                            font-weight: bold; /* ทำให้ตัวหนังสือเป็นตัวหนา */
                                            color: #000000; /* ปรับสีตัวหนังสือให้เข้มกว่าเดิม */
                                            
                                }
                                    </style>
                                    <h2>Config</h2>
                                    """ + result_config.to_html(classes="data_config1", index=False, header=False)

                                    # Print HTML
                                    

                            list_of_values_configured = [str(val) if not pd.isna(val) else None for val in list_of_values_configured]

                            session['list_of_values_configured'] = list_of_values_configured
                            # print("session config:",session['list_of_values_configured'])


                ##################################billing################################################## note1 ##

                            session['list_of_values_billing'] = list_of_values_billing
                            list_of_values_billing = [str(val) if not pd.isna(val) else None for val in list_of_values_billing]
                            
                            
                            # print("session billing:",session['list_of_values_billing'])
                            # Write the list to a file (if needed)
                            

                            # Create DataFrame from the list of values
                            value_billing = pd.DataFrame(list_of_values_billing, columns=['Value'])

                            # Transpose the DataFrame
                            value_billing = value_billing.transpose()

                            # Define the chunk size
                            chunk_size = len(filtered_result)

                            # Calculate the number of chunks needed
                            num_chunks = (len(value_billing.columns) + chunk_size - 1) // chunk_size

                            # Split the DataFrame into chunks
                            chunks = np.array_split(value_billing, num_chunks, axis=1)

                            # Convert each chunk to HTML table
                            result_billing_html = [(
                                ("<h2>Billing</h2>"
                                "<style>"
                                "table.data_billing {border: 0.2px solid #000000; width: 100%;}"  # สไตล์สำหรับตารางที่มี class="data_billing"
                                "table.data_billing th, table.data_billing td {font-size: 12px;}"  # สไตล์สำหรับ th และ td ใน data_billing
                               "table.data_billing th {background-color: #f2f2f2; font-weight: bold; color: #000000; border: 0.2px solid #000000; font-size: small; padding: 12px;}"  # เพิ่ม padding ให้กับ <th> เพื่อขยายขนาดช่อง
                                "table.data_billing td {padding: 8px; color: #000000;} "  # สามารถเพิ่มสไตล์เฉพาะ td ได้เช่นกัน
                                "table.data_billing td, table.data_billing th {width: 20%;}"
                                "</style>"
                                "<table class='data_billing'>"  # กำหนดให้ใช้ class="data_billing"
                                "<thead>"
                                "<tr>"
                                + ''.join([f"<th>{i}</th>" for i in selected_data]) +  # วนลูปสร้าง <th> สำหรับแต่ละค่าใน selected_data
                                "</tr>"
                                "</thead>"
                                "<tbody>" if idx == 0 else "") +
                                chunk.to_html(classes="data_billing", index=False, header=False).replace('\n', '').replace('[', '').replace(']', '').replace(',', '') +
                                ("</tbody></table>" if idx == num_chunks - 1 else "")
                            ) for idx, chunk in enumerate(chunks)]


                            # Join the HTML strings
                            result_billing_html = ''.join(result_billing_html)

                        else:
                            
                            print("Index out of range") 

                        columns_to_drop = [
                            "NoRun",
                            "vc_name",
                            "poll_billing",
                            "poll_config",
                            "poll_billing_enable",
                            "poll_config_enable",
                            "modbus_id",
                            "port_id"
                        ]

                        # Drop the columns from the DataFrame
                        df = df.drop(columns=columns_to_drop)

                        test_poll_type = "Select Poll Type"
                        if poll_type == '1':
                            test_poll_type = "Daily"
                        elif poll_type == '2':
                            test_poll_type = "Hourly"

            else:  ################ hourly
                query = """
                    SELECT
                        AMR_FIELD_METER.METER_STREAM_NO as RunNo,
                        AMR_PL_GROUP.PL_REGION_ID as region,
                        AMR_FIELD_ID.TAG_ID as Sitename,
                        AMR_FIELD_METER.METER_NO_STREAM as NoRun,
                        AMR_FIELD_METER.METER_ID as METERID,
                        AMR_VC_TYPE.VC_NAME as VCtype,
                        AMR_FIELD_ID.SIM_IP as IPAddress,
                        AMR_PORT_INFO.PORT_NO as port,
                        amr_poll_range_hourly.evc_type as evc_type,
                        amr_vc_type.vc_name as vc_name,
                        amr_poll_range_hourly.poll_hourly as poll_hourly,
                        amr_poll_range_hourly.poll_hourly_enable as poll_hourly_enable,
                        amr_field_meter.modbus_id as modbus_id,
                        AMR_PORT_INFO.ID as port_id
                    FROM
                        AMR_FIELD_ID,
                        AMR_USER,
                        AMR_FIELD_CUSTOMER,
                        AMR_FIELD_METER,
                        AMR_PL_GROUP,
                        AMR_VC_TYPE,
                        AMR_PORT_INFO,
                        amr_poll_range_hourly,
                        amr_region
                    WHERE
                        AMR_PL_GROUP.PL_REGION_ID = amr_region.id AND
                        AMR_USER.USER_ENABLE=1 AND
                        amr_vc_type.id=amr_poll_range_hourly.evc_type AND
                        AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID AND
                        AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP AND
                        AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID AND
                        AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID AND
                        AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE AND
                        AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID
                        {tag_condition}
                        {region_condition}
                """
                print("hourly")
                tag_condition = "AND AMR_FIELD_ID.TAG_ID IS NOT NULL"
                region_condition = "AND amr_region.REGION_NAME = 'default_region_id'"
                region_results = fetch_data(ptt_pivot_connection, region_query)
                region_options = [str(region[1]) for region in region_results]
                tag_results = fetch_data(ptt_pivot_connection, tag_query, params={"region_id": selected_region})
                tag_options = [str(tag[0]) for tag in tag_results]
                tag_options.sort()
                if selected_tag:
                    tag_condition = f"AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'"
                if selected_region:
                    region_condition = f"AND amr_region.REGION_NAME = '{selected_region}'"
                query = query.format(tag_condition=tag_condition, region_condition=region_condition)
                results = fetch_data(ptt_pivot_connection, query)
           
                df = pd.DataFrame(
                    results,
                    columns=[
                        "RUN",
                        "Region",
                        "Sitename",
                        "NoRun",
                        "METERID",
                        "VCtype",
                        "IPAddress",
                        "Port",
                        "evc_type",
                        "vc_name",
                        "poll_hourly",
                        "poll_hourly_enable",
                        "modbus_id",
                        "port_id"
                    ],
                )

                if selected_row: 
                        # Convert the selected_row to an integer
                        selected_row_index = int(selected_row) - 1
                        print("selected_row_index",selected_row_index)
                        # Check if the index is within the valid range
                        if 0 <= selected_row_index < len(df):
                            
                            row_data = df.iloc[selected_row_index]
                            
                            row_data_dict = row_data.to_dict()

                            run = row_data_dict.get("RUN")
                            print("run",run)
                            Region = row_data_dict.get("Region")
                            Sitename = row_data_dict.get("Sitename")
                            NoRun = row_data_dict.get("NoRun")
                            METERID = row_data_dict.get("METERID")
                            VCtype = row_data_dict.get("VCtype")
                            tcp_ip = row_data_dict.get("IPAddress")
                            tcp_port = int(row_data_dict.get("Port"))
                            evc_type = row_data_dict.get("evc_type")
                            poll_hourly = row_data_dict.get("poll_hourly")
                            print(poll_hourly)
                            poll_hourly_enable = row_data_dict.get("poll_hourly_enable")
                            print(poll_hourly_enable)
                            
                            modbus_id = int(row_data_dict.get("modbus_id"))
                            port_id = int(row_data_dict.get("port_id"))

                            poll_hourly_list = [int(x) for x in poll_hourly.split(',')]
                            poll_hourly_enable_list = poll_hourly_enable.split(',')
                            pairs_hourly = [(poll_hourly_list[i], poll_hourly_list[i+1]) for i in range(0, len(poll_hourly_list), 2)]

                            # Prepare dataframes for billing and config
                            data= {'starting_address_i': [], 'quantity_i': [], 'adjusted_quantity_i': []}
                            
                            df_pollhourly = pd.DataFrame(data)

                            # Populate the billing dataframe
                            print("\nhourly Pairs:")
                            for index, value in enumerate(poll_hourly_enable_list):
                                if value == '1' and index < len(pairs_hourly):
                                    starting_address_i = int(pairs_hourly[index][0])  # Convert to integer
                                    quantity_i = int(pairs_hourly[index][1])          # Convert to integer
                                    adjusted_quantity_i = quantity_i - starting_address_i + 1
                                    data = {'starting_address_i': [starting_address_i], 
                                            'quantity_i': [quantity_i], 
                                            'adjusted_quantity_i': [adjusted_quantity_i]}
                                    df_2 = pd.DataFrame(data)
                                    df_pollhourly = pd.concat([df_pollhourly, df_2], ignore_index=True)
                                    print("df_pollhourly", df_pollhourly)
                            
                            dataframes = {
                                    'address_start': [],
                                    'finish': [],
                                    'TX': [],
                                    'RX': []
                                }
                            
                            df_Modbushourly = pd.DataFrame(dataframes)
                            
                            try:
                                sock_i = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                                sock_i.settimeout(30)  
                                
                                sock_i.connect((tcp_ip, tcp_port))
                                #print("Connected successfully.")
                                with open(FILE_PATH, 'a') as file:
                                    file.write(f"Connected successfully.\n")
                                    file.write(f"IP ADDRESS:{tcp_ip}\n")
                                    file.write(f"Port:{tcp_port}\n")
                                    file.write(f"modbus_id:{modbus_id}\n\n")
                                    
                            except ConnectionRefusedError:
                                abort(400, f"Error: Connection refused to {tcp_ip}:{tcp_port}!")
                            except TimeoutError:
                                abort(400, f"Error: Connection timed out to {tcp_ip}:{tcp_port}!")
                            except Exception as e:
                                abort(400, f"Error: {e}")
                            
                            slave_id_1 = 0x01
                            function_code_1 = 0x03
                            starting_address_1 = 0x0004
                            quantity_1 = 0x0002

                            request_Actaris= bytearray([
                                    slave_id_1,
                                    function_code_1,
                                    starting_address_1 >> 8,
                                    starting_address_1 & 0xFF,
                                    quantity_1 >> 8,
                                    quantity_1 & 0xFF,
                                ])

                            crc_1 = computeCRC(request_Actaris)
                            request_Actaris += crc_1
                          
                            if int(evc_type) in [5, 8, 9, 10]:

                                for _ in range(2):  
                                    sock_i.send(request_Actaris)
                                    time.sleep(3)
                                response = sock_i.recv(4096)

                            if int(evc_type) == 12:
                                if int(tcp_port) != 2101:
                                    for _ in range(2):
                                        sock_i.send(request_Actaris)
                                        time.sleep(3)
                                    response = sock_i.recv(4096)

                            current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            with open(FILE_PATH, 'a') as file:
                                file.write(f"{current_time}:Start polling Logged Data Hourly\n")

                            for i in range(0, len(df_pollhourly)):
                                
                                if int(tcp_port) == 2101 and int(port_id) in [15, 16]:

                                    sock_i = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                                    sock_i.settimeout(5)
                                    sock_i.connect((tcp_ip, int(tcp_port))) 
                                
                                start_address = int(df_pollhourly.loc[i,'starting_address_i'])
                                
                                adjusted_quantity = int(df_pollhourly.loc[i,'adjusted_quantity_i'])
                            
                                request_message_i = bytearray(
                                [modbus_id, function_code, start_address >> 8, start_address & 0xFF, adjusted_quantity >> 8, adjusted_quantity & 0xFF])
                                crc_i = computeCRC(request_message_i)
                                request_message_i += crc_i
                                
                                communication_traffic_i = []
                                
                                communication_traffic_i.append(request_message_i.hex())
                                billing_safe_tx = f"hourly_TX: {communication_traffic_i[0]}"
                                print("tx.hourly",communication_traffic_i[0])

                                with open(FILE_PATH, 'a') as file:
                                    file.write(f"Poll NO {i}\n")
                                    file.write(f"{current_time}\n")
                                    file.write(f"TX-hourly:{communication_traffic_i[0]}\n")

                                if int(evc_type) == 12:
                                    try:
                                        
                                        time.sleep(2)  
                                        sock_i.send(request_message_i)
                                        
                                        response_i = sock_i.recv(4096)
                                   
                                        communication_traffic_i.append(response_i.hex())
                                        
                                        billing_safe = f"hourly_RX: {communication_traffic_i[1]}"
                                        print("rx.hourly",communication_traffic_i[1])
                                        with open(FILE_PATH, 'a') as file:
                                            file.write(f"RX-hourly:{communication_traffic_i[1]}\n\n")
                                    
                                    except TimeoutError:
                                        abort(400, f"Error: Connection timed out while waiting for response from {tcp_ip}:{tcp_port}!")
                                    except Exception as e:
                                        abort(400, f"Error: {e}")

                                else:
                                    try:
                                        
                                        sock_i.send(request_message_i)
                                        time.sleep(1)
                                        response_i = sock_i.recv(4096)
                                                                           
                                        communication_traffic_i.append(response_i.hex())
                                        
                                        billing_safe = f"hourly_RX: {communication_traffic_i[1]}"
                                        print("rx.hourly",communication_traffic_i[1])
                                        with open(FILE_PATH, 'a') as file:
                                            file.write(f"RX-hourly:{communication_traffic_i[1]}\n\n")
                                    
                                    except TimeoutError:
                                        abort(400, f"Error: Connection timed out while waiting for response from {tcp_ip}:{tcp_port}!")
                                    except Exception as e:
                                        abort(400, f"Error: {e}")

                                if response_i[1:2] != b'\x03':
                                    
                                    abort(400, f"Error: Unexpected response code from device {communication_traffic_i[1]} ,{response_i[1:2]}!")
                                else:
                                    pass

                                # sock_i.close()
                                # print(communication_traffic_i)
                                data = {
                                    'address_start': [int(start_address)],
                                    'finish': [int(start_address+adjusted_quantity-1)],
                                    'TX': [communication_traffic_i[0]],
                                    'RX': [communication_traffic_i[1]]
                                }
                                # print(data)
                                df_2 = pd.DataFrame(data)
                                df_Modbushourly = pd.concat([df_Modbushourly, df_2], ignore_index=True)
                                # print(df_Modbushourly)
                                
                                amr_vc_type_table = f"""
                    
                                SELECT DESCRIPTION 
                                FROM amr_mapping_hourly
                                WHERE evc_type LIKE '{evc_type}' 
                                AND hourly like '1' 
                                ORDER BY OR_DER
                                """
                                
                                amr_vc_type_table_result = fetch_data(ptt_pivot_connection, amr_vc_type_table)
                                filtered_result = pd.DataFrame(amr_vc_type_table_result)
                                
                                transposed_result = filtered_result.T
                                selected_data = transposed_result.iloc[0, 0:5]  # Extract values to get a flat array
                                
                                print("selected_data",selected_data)

                                evc_type = evc_type
                                query = """
                                SELECT amb.hourly ,amb.or_der ,amb.address,amb.description,amb.data_type  FROM amr_mapping_hourly amb WHERE amb.evc_type = :evc_type AND address is not null order by amb.hourly
                                ,amb.or_der
                                """
                                poll_resultsbilling = fetch_data(ptt_pivot_connection,query, params={"evc_type": evc_type})
                                # print(poll_resultsbilling)
                                df_mappingbilling = pd.DataFrame(poll_resultsbilling, columns=['hourly','or_der', 'address', 'description', 'data_type'])
                                
                                
                                #print(df_mappingbilling)
                                # print(df_Modbusbilling)   
                                
                                list_of_values_billing = []
                                for i in range(0, len(df_mappingbilling)):
                                        
                                    address = int(df_mappingbilling.iloc[i,2])
                                    data_type = str(df_mappingbilling.iloc[i,4])

                                    # print(i, df_mappingbilling.loc[i])
                                    for j in range(0,len(df_Modbushourly)):
                                        address_start = int(df_Modbushourly.iloc[j,0])
                                        address_finish = int(df_Modbushourly.iloc[j,1])
                                    
                                        if address >= address_start and address <= address_finish:
                                            # print(address)
                                            # print(address_start, address_finish)
                                            location_data = (address - address_start)*int(8/2)
                                            # print(location_data)
                                            frameRx = (df_Modbushourly.iloc[j,3])
                                        
                                            raw_data = frameRx[location_data + 6: location_data + 14]
                                        
                                            list_of_values_billing.append(convert_raw_to_value(data_type, raw_data, mode="hourly"))
                                            break
                                list_of_values_billing = [str(val) if not pd.isna(val) else None for val in list_of_values_billing]

                                # Write the list to a file (if needed)
                                session['list_of_values_billing'] = list_of_values_billing
                                                                
                                # Create DataFrame from the list of values
                                value_billing = pd.DataFrame(list_of_values_billing, columns=['Value'])

                                # Transpose the DataFrame
                                value_billing = value_billing.transpose()

                                # Define the chunk size
                                chunk_size = 5

                                # Calculate the number of chunks needed
                                num_chunks = (len(value_billing.columns) + chunk_size - 1) // chunk_size

                                # Split the DataFrame into chunks
                                chunks = np.array_split(value_billing, num_chunks, axis=1)

                                # Convert each chunk to HTML table
                                result_billing_html = [(
                                    ("<h2>Hourly</h2>"
                                    "<style>"
                                "table.data_billing {border: 0.2px solid #000000; width: 100%;}"  # สไตล์สำหรับตารางที่มี class="data_billing"
                                "table.data_billing th, table.data_billing td {font-size: 12px;}"  # สไตล์สำหรับ th และ td ใน data_billing
                               "table.data_billing th {background-color: #f2f2f2; font-weight: bold; color: #333333; border: 0.2px solid #000000; font-size: small; padding: 12px;}"  # เพิ่ม padding ให้กับ <th> เพื่อขยายขนาดช่อง
                                "table.data_billing td {padding: 8px; color: #000000;} "  # สามารถเพิ่มสไตล์เฉพาะ td ได้เช่นกัน
                                "table.data_billing td, table.data_billing th {width: 20%;}"
                                "</style>"
                                "<table class='data_billing'>"  # กำหนดให้ใช้ class="data_billing"
                                "<thead>"
                                "<tr>"
                                + ''.join([f"<th>{i}</th>" for i in selected_data]) +  # วนลูปสร้าง <th> สำหรับแต่ละค่าใน selected_data
                                "</tr>"
                                "</thead>"
                                "<tbody>" if idx == 0 else "") + 
                                    chunk.to_html(classes="data_billing", index=False, header=False).replace('\n', '').replace('[', '').replace(']', '').replace(',', '') +
                                ("</tbody></table>" if idx == num_chunks - 1 else "")
                                ) for idx, chunk in enumerate(chunks)]

                                # Join the HTML strings
                                result_billing_html = ''.join(result_billing_html)
                            with open(FILE_PATH, 'a') as file:
                                file.write(f"{current_time}:Polling successfully.\n")  
                else:   
                            print("Index out of range") 
                columns_to_drop = [
                "NoRun",
                "vc_name",
                "poll_hourly",
                "poll_hourly_enable",
                "modbus_id",
                "port_id"
                ]

                # Drop the columns from the DataFrame
                df = df.drop(columns=columns_to_drop)

                test_poll_type = "Select Poll Type"
                if poll_type == '1':
                    test_poll_type = "Daily"
                elif poll_type == '2':
                    test_poll_type = "Hourly"
                else:
                    test_poll_type = ""
    except Exception as e:
            abort(400, f"Error: {e}")
    
    ## Prevent Warning
    # check data is not none
    METERID = METERID if 'METERID' in locals() and METERID is not None else ""
    df_mapping = df_mapping if 'df_mapping' in locals() and df_mapping is not None else pd.DataFrame()
    result_config = result_config if 'result_config' in locals() and result_config is not None else pd.DataFrame()
    df_mappingbilling = df_mappingbilling if 'df_mappingbilling' in locals() and df_mappingbilling is not None else pd.DataFrame()
    modbus_id = modbus_id if 'modbus_id' in locals() and modbus_id is not None else 1
    port_id = modbus_id if 'port_id' in locals() and port_id is not None else 0
    slave_id = slave_id if 'slave_id' in locals() and slave_id is not None else 1
    result_billing_html = result_billing_html if 'result_billing_html' in locals() and result_billing_html is not None else ""
    result_config_html = result_config_html if 'result_config_html' in locals() and result_config_html is not None else ""
    df_Modbus_html = df_Modbus_html if 'df_Modbus_html' in locals() and df_Modbus_html is not None else ""
    starting_address_i = starting_address_i if 'starting_address_i' in locals() and starting_address_i is not None else 0
    quantity_i = quantity_i if 'quantity_i' in locals() and quantity_i is not None else 0
    communication_traffic_i = communication_traffic_i if 'communication_traffic_i' in locals() and communication_traffic_i is not None else 0
    run = run if 'run' in locals() and run is not None else 0
    billing_safe = billing_safe if 'billing_safe' in locals() and billing_safe is not None else ""
    billing_safe_tx= billing_safe_tx if 'billing_safe_tx' in locals() and billing_safe_tx is not None else ""
    billing_safe = billing_safe if 'billing_safe' in locals() and billing_safe is not None else ""
    config_safe_RX= config_safe_RX if 'config_safe_RX' in locals() and config_safe_RX is not None else ""
    poll_config_list = poll_config_list if 'poll_config_list' in locals() and isinstance(poll_config_list, list) else []
    poll_billing_list = poll_billing_list if 'poll_billing_list' in locals() and isinstance(poll_billing_list, list) else []
    poll_config_enable_list = poll_config_enable_list if 'poll_config_enable_list' in locals() and isinstance(poll_config_enable_list, list) else []
    poll_billing_enable_list = poll_billing_enable_list if 'poll_billing_enable_list' in locals() and isinstance(poll_billing_enable_list, list) else []
    test_poll_type = test_poll_type if 'test_poll_type' in locals() and test_poll_type is not None else ""

    if poll_type == "1": #Daily
        return render_template(
            "Manual poll.html",
            df=df, METERID=METERID, df_mapping=df_mapping, df_mappingbilling=df_mappingbilling, result_config=result_config, modbus_id=modbus_id, port_id=port_id,
            slave_id=slave_id, result_billing_html=result_billing_html, result_config_html=result_config_html, df_Modbus_html=df_Modbus_html,
            function_code=function_code, starting_address_i=starting_address_i, quantity_i=quantity_i, communication_traffic_i=communication_traffic_i,
            run=run, poll_config_list=poll_config_list, poll_billing_list=poll_billing_list,
            poll_config_enable_list=poll_config_enable_list, poll_billing_enable_list=poll_billing_enable_list,
            tables=[df.to_html(classes="data")],
            titles=df.columns.values,
            selected_tag=selected_tag,
            selected_region=selected_region,
            region_options=region_options,
            tag_options=tag_options,username=username,
            billing_safe=billing_safe, billing_safe_tx=billing_safe_tx, config_safe_RX=config_safe_RX,test_poll_type=test_poll_type,poll_type=poll_type
        )
    else:
        # Return something else here, for example:
        return render_template(
            "Manual poll.html",
            df=df, METERID=METERID,  df_mappingbilling=df_mappingbilling,  modbus_id=modbus_id, port_id=port_id,
            slave_id=slave_id, result_billing_html=result_billing_html,  
            function_code=function_code, starting_address_i=starting_address_i, quantity_i=quantity_i, communication_traffic_i=communication_traffic_i,
            run=run,
            tables=[df.to_html(classes="data")],
            titles=df.columns.values,
            selected_tag=selected_tag,
            selected_region=selected_region,
            region_options=region_options,
            tag_options=tag_options,username=username,
            billing_safe=billing_safe, billing_safe_tx=billing_safe_tx,test_poll_type=test_poll_type ,poll_type=poll_type
        )
      
def handle_actaris_action(i, address):
    return address

def handle_action_configuration(i, value, address):
    return value, address

def get_description_from_database(address):
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        query = "SELECT DESCRIPTION FROM AMR_ADDRESS_MAPPING1 WHERE ADDRESS = :address"
        params = {"address": address}
        result = fetch_data(ptt_pivot_connection,query, params)
        return result[0][0] if result else None


@app.route("/process_selected_rows", methods=["POST"])
def process_selected_rows():
    selected_rows = request.form.getlist("selected_rows")
    return "Selected rows processed successfully"


def get_type_value_from_database(address):
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        query = "SELECT TYPE_VALUE FROM AMR_ADDRESS_MAPPING1 WHERE ADDRESS = :address"
        result = fetch_data(ptt_pivot_connection,query, params={"address": address})
        if result:
            return result[0][0]  # Assuming TYPE_VALUE is the first column in the result
        return None


########### ping manul ############################
@app.route("/pingdata", methods=["POST"])
@login_required
def pingdata():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        try:
            selected_tag = session.get('selected_tag')
            selected_region = session.get('selected_region')
            query = f"""
                SELECT
                    AMR_FIELD_ID.SIM_IP as IPAddress,
                    AMR_PORT_INFO.PORT_NO as port
                FROM
                    AMR_FIELD_ID,
                    AMR_USER,
                    AMR_FIELD_CUSTOMER,
                    AMR_FIELD_METER,
                    AMR_PL_GROUP,
                    AMR_VC_TYPE,
                    AMR_PORT_INFO,
                    amr_poll_range_hourly,
                    amr_region
                WHERE
                AMR_PL_GROUP.PL_REGION_ID = amr_region.id AND
                    AMR_USER.USER_ENABLE=1 AND
                    amr_vc_type.id=amr_poll_range_hourly.evc_type AND
                    AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID AND
                    AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP AND
                    AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID AND
                    AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID AND
                    AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE AND
                    AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID AND
                    AMR_FIELD_ID.TAG_ID like '{selected_tag}' AND
                    amr_region.REGION_NAME like '{selected_region}'
            """
            #print(query)
            
            results = fetch_data(ptt_pivot_connection, query)
            print(results)
            df = pd.DataFrame(
                results,
                columns=[
                    "IPAddress",
                    "port"
                ],
            )
            # print(df)
            
            responses = []
            
            for index, row in df.iterrows():
                tcp_ip = row['IPAddress']
                tcp_port = int(row['port'])  # Convert port to integer
                
                sock_i = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock_i.settimeout(6)
                
                try:
                    sock_i.connect((tcp_ip, tcp_port))
                    print(f"Connected to {tcp_ip} on port {tcp_port}")
                    response = {
                        "status": "success",
                        "message": f"Ping successful to {tcp_ip}:{tcp_port}",
                    }
                except Exception as e:
                    print(f"Failed to connect to {tcp_ip} on port {tcp_port}: {e}")
                    response = {
                        "status": "error",
                        "message": f"Ping failed to {tcp_ip}:{tcp_port}: {e}",
                    }
                finally:
                    sock_i.close()
                
                responses.append(response)
            
            return jsonify(responses)
        
        except Exception as e:
            print(f"An error occurred: {e}")
            return jsonify({"status": "error", "message": "An error occurred while processing the request."}), 500

##########################################################



##################save_to_oracle_manualpoll ###############
@app.route("/save_to_oracle_manualpoll", methods=["POST"])
@login_required
def save_to_oracle_manualpoll():
    
    if 'username' not in session:
        return redirect(url_for('login'))
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        
        try:
                        
            selected_tag = session.get('selected_tag')
            selected_region = session.get('selected_region')
            poll_type = session.get('poll_type')
            selected_row = session.get('selected_row')
            print("selected_row:",selected_row)
            print("selected_tag:",selected_tag)
            print("selected_region:",selected_region)
            print("poll_type:",poll_type)


            region_query = """
                SELECT * FROM AMR_REGION 
            """
            
            tag_query = """
                SELECT DISTINCT AMR_FIELD_ID.TAG_ID
                FROM AMR_FIELD_ID,amr_region,AMR_PL_GROUP
                
                WHERE AMR_PL_GROUP.PL_REGION_ID = amr_region.id
                AND AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
                AND amr_region.REGION_NAME = :region_id
                AND AMR_FIELD_ID.tag_id NOT like '%.remove%'
                ORDER BY  TAG_ID
                """
            region_results = fetch_data(ptt_pivot_connection, region_query)
            region_options = [str(region[1]) for region in region_results]
            
            if poll_type == "1":  #### daily
                query = """
                    SELECT
                        AMR_FIELD_METER.METER_STREAM_NO as RunNo,
                        AMR_PL_GROUP.PL_REGION_ID as region,
                        AMR_FIELD_ID.TAG_ID as Sitename,
                        AMR_FIELD_METER.METER_NO_STREAM as NoRun,
                        AMR_FIELD_METER.METER_ID as METERID,
                        AMR_VC_TYPE.VC_NAME as VCtype,
                        AMR_FIELD_ID.SIM_IP as IPAddress,
                        AMR_PORT_INFO.PORT_NO as port,
                        amr_poll_range.evc_type as evc_type,
                        amr_vc_type.vc_name as vc_name,
                        amr_poll_range.poll_billing as poll_billing,
                        amr_poll_range.poll_config as poll_config,
                        amr_poll_range.poll_billing_enable as poll_billing_enable,
                        amr_poll_range.poll_config_enable as poll_config_enable,
                        amr_field_meter.modbus_id as modbus_id,
                        AMR_PORT_INFO.ID as port_id
                    FROM
                        AMR_FIELD_ID,
                        AMR_USER,
                        AMR_FIELD_CUSTOMER,
                        AMR_FIELD_METER,
                        AMR_PL_GROUP,
                        AMR_VC_TYPE,
                        AMR_PORT_INFO,
                        amr_poll_range,
                        amr_region
                    WHERE
                    AMR_PL_GROUP.PL_REGION_ID = amr_region.id AND
                        AMR_USER.USER_ENABLE=1 AND
                        amr_vc_type.id=amr_poll_range.evc_type AND
                        AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID AND
                        AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP AND
                        AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID AND
                        AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID AND
                        AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE AND
                        AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID
                        {tag_condition}
                        {region_condition}
                """
                tag_condition = "AND AMR_FIELD_ID.TAG_ID IS NOT NULL"
                region_condition = "AND amr_region.REGION_NAME = 'default_region_id'"
                region_results = fetch_data(ptt_pivot_connection, region_query)
                region_options = [str(region[1]) for region in region_results]
                tag_results = fetch_data(ptt_pivot_connection, tag_query, params={"region_id": selected_region})
                tag_options = [str(tag[0]) for tag in tag_results]
                tag_options.sort()
                if selected_tag:
                    tag_condition = f"AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'"
                if selected_region:
                    region_condition = f"AND amr_region.REGION_NAME = '{selected_region}'"
                query = query.format(tag_condition=tag_condition, region_condition=region_condition)
                
                results = fetch_data(ptt_pivot_connection, query)
                
                df = pd.DataFrame(
                    results,
                    columns=[
                        "RUN",
                        "Region",
                        "Sitename",
                        "NoRun",
                        "METERID",
                        "VCtype",
                        "IPAddress",
                        "Port",
                        "evc_type",
                        "vc_name",
                        "poll_billing",
                        "poll_config",
                        "poll_billing_enable",
                        "poll_config_enable",
                        "modbus_id",
                        "port_id" 
                    ],
                )
                
                #print(df)
                #print("Select Row", selected_row)
                if selected_row:
                        # Convert the selected_row to an integer
                        selected_row_index = int(selected_row) - 1
                        # Check if the index is within the valid range
                        if 0 <= selected_row_index < len(df):
                            
                            row_data = df.iloc[selected_row_index]
                            
                            row_data_dict = row_data.to_dict()
                            
                            run = row_data_dict.get("RUN")
                            Region = row_data_dict.get("Region")
                            Sitename = row_data_dict.get("Sitename")
                            NoRun = row_data_dict.get("NoRun")
                            METERID = row_data_dict.get("METERID")
                            VCtype = row_data_dict.get("VCtype")
                            tcp_ip = row_data_dict.get("IPAddress")
                            tcp_port = int(row_data_dict.get("Port"))
                            evc_type = row_data_dict.get("evc_type")
                            poll_billing = row_data_dict.get("poll_billing")
                            #print(poll_billing)
                            poll_billing_enable = row_data_dict.get("poll_billing_enable")
                            #print(poll_billing_enable)
                            poll_config = row_data_dict.get("poll_config")
                            poll_config_enable = row_data_dict.get("poll_config_enable")
                            modbus_id = int(row_data_dict.get("modbus_id"))
                            port_id = int(row_data_dict.get("port_id"))
                            
                            #session['list_of_values_configured']came from storage `manulalpoll`
                            list_of_values_configured = session['list_of_values_configured']
                            # list_of_values_configured = read_and_print_file(file_databaes_config)
                            #print("dataconfig",list_of_values_configured)

                            query = """
                            select amc.or_der as order1 , amc.address as address1, amc.description as desc1, amc.data_type as dtype1
                            from amr_mapping_config amc
                            where amc.evc_type = :evc_type AND address is not null 
                            order by order1
                            """
                            poll_results = fetch_data(ptt_pivot_connection,query, params={"evc_type": evc_type})
                            df_mapping = pd.DataFrame(poll_results, columns=['order', 'address', 'desc', 'data_type'])
                            
                            # สร้าง current_datetime
                            current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7)))
                            # จัดรูปแบบเป็น 'DD-MM-YYYY HH24:MI:SS.FF6'
                            current_datetime_formatted = current_datetime.strftime('%d-%m-%Y %H:%M:%S.%f')[:-3]  # ตัดมิลลิวินาทีให้เหลือ 6 หลัก
                            current_datetime_upper = current_datetime_formatted.upper()

                            date_system = datetime.datetime.now().strftime('%d-%m-%Y') 

                            sql_text_config_delete = f"""delete from AMR_CONFIGURED_DATA where METER_ID = '{METERID}' AND METER_STREAM_NO = '{run}' AND DATA_DATE = TO_DATE('{date_system}', 'DD-MM-YYYY')"""
                            
                            sql_text_config_insert = "insert into AMR_CONFIGURED_DATA (DATA_DATE, METER_ID,METER_STREAM_NO, AMR_VC_TYPE,TIME_CREATE, "
                            for i in range(0, len(df_mapping)):  
                                    
                                sql_text_config_insert+=f" AMR_CONFIG{i+1},"
                            sql_text_config_insert+=" CREATED_BY) values ("

                            sql_text_config_insert+=f"TO_DATE('{date_system}', 'DD-MM-YYYY'), '{METERID}','{run}','{evc_type}','{current_datetime_upper}',"
                                        
                            for i in range(0, len(df_mapping)):
                                value = f"'{str(list_of_values_configured[i])}',"
                                if value.strip() == 'NULL,' :  # Remove the single quotes and the trailing comma
                                    value = "'',"
                                sql_text_config_insert += value

                            sql_text_config_insert+="'')"  
                            #print(sql_text_config_insert)
                            
                            with cx_Oracle.connect(username, password, f"{hostname}:{port}/{service_name}") as connection:
                                with connection.cursor() as cursor:
                                    cursor.execute(sql_text_config_delete)  
                                    cursor.execute(sql_text_config_insert)  
                                connection.commit()     
                            #     print("Insert data 'config' successful")

                            ## Billing
                            list_of_values_billing = session['list_of_values_billing']
            #                 list_of_values_billing = read_and_print_file(file_databaes_billing)

                            ##################### TUL2024-12-26
                            list_cut = []
                            day_polled = 0
                            for i in range(0, len(list_of_values_billing), 5):
                                values_subset = list_of_values_billing[i:i+5]
                                
                                if not is_valid_date(values_subset[0]):
                                    continue
                                list_cut.extend(list_of_values_billing[i:i + 5])
                                day_polled += 1
                                    
                            #print(list_cut)
                            list_of_values_billing = list_cut.copy()

                            ############################################################################### TULA
                            current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))).strftime('%d-%b-%y %I.%M.%S.%f %p ASIA/BANGKOK')
                            current_datetime_upper = current_datetime.upper()
                            date_system = datetime.datetime.now().strftime('%d-%m-%Y')   
                            
                            full_sql_text = ""
                            #for i in range(0, len(df_mappingbilling), 5):    
                            #        values_subset = list_of_values_billing[i:i+5]
                            for i in range(0, day_polled):
                                
                                values_subset = list_of_values_billing[(i*5):(i*5)+5]
                                date_polled = values_subset[0]
                                corrected_polled = verifyNumericReturnNULL(values_subset[1])
                                uncorrected_polled = verifyNumericReturnNULL(values_subset[2])
                                avr_pf_polled = verifyNumericReturnNULL(values_subset[3])
                                avr_tf_polled = verifyNumericReturnNULL(values_subset[4])
                            
                                
                                # query for checked 
                                sql_billing_DB = f"""SELECT DATA_DATE, CORRECTED_VOL, UNCORRECTED_VOL, AVR_PF, AVR_TF, METER_ID, METER_STREAM_NO 
                                                                            FROM amr_billing_data
                                                                            WHERE DATA_DATE = TO_DATE('{date_polled}', 'DD-MM-YYYY')
                                                                            AND METER_ID = '{METERID}' 
                                                                            AND METER_STREAM_NO = '{run}'"""
                                

                                billing_DB = fetch_data(ptt_pivot_connection,sql_billing_DB)
                                if billing_DB:
                                    # already have data check
                                    date_db = billing_DB[0][0]
                                    corrected_db = verifyNumericReturnNULL(billing_DB[0][1])
                                    uncorrected_db = verifyNumericReturnNULL(billing_DB[0][2])
                                    avr_pf_db = verifyNumericReturnNULL(billing_DB[0][3])
                                    avr_tf_db = verifyNumericReturnNULL(billing_DB[0][4])

                                    if (corrected_polled == corrected_db and 
                                        uncorrected_polled == uncorrected_db and 
                                        (avr_pf_polled == avr_pf_db or (avr_pf_polled is None and avr_pf_db is None)) and 
                                        (avr_tf_polled == avr_tf_db or (avr_tf_polled is None and avr_tf_db is None))):

                                        # case 0 have billing match = do nothing
                                        pass
                                    else :
                                        # case 4 not match  =  delete from billing -> insert both into error
                                        full_sql_text = full_sql_text + create_SQL_text_delete_Billing(METERID, run, date_polled) + "\n"
                                        full_sql_text = full_sql_text + create_SQL_text_insert_Billing_error(METERID, run, current_datetime_upper, date_polled,corrected_polled, uncorrected_polled, avr_pf_polled, avr_tf_polled) + "\n"
                                        full_sql_text = full_sql_text + create_SQL_text_insert_Billing_error(METERID, run, current_datetime_upper, date_polled,corrected_db, uncorrected_db, avr_pf_db, avr_tf_db) + "\n"
                                        
                                else:
                                    # not found Check from Error 
                                    sql_billing_Error = f"""SELECT DATA_DATE, METER_ID, METER_STREAM_NO 
                                                                            FROM amr_billing_data_error
                                                                            WHERE DATA_DATE = TO_DATE('{date_polled}', 'DD-MM-YYYY')
                                                                            AND METER_ID = '{METERID}' 
                                                                            AND METER_STREAM_NO = '{run}'"""
                                    billing_Error = fetch_data(ptt_pivot_connection, sql_billing_Error)
                                
                                    if billing_Error:
                                        # Check if already not insert
                                        # case 2 data already in error =  skip # case 3 new error data = insert into error
                                        # combine to insert if not exist
                                        full_sql_text = full_sql_text + create_SQL_text_insert_Billing_error(METERID, run, current_datetime_upper, date_polled,corrected_polled, uncorrected_polled, avr_pf_polled, avr_tf_polled) + "\n"
                                    else:
                                        # case  1 new data = insert into billing
                                        full_sql_text = full_sql_text + create_SQL_text_insert_Billing(METERID, run, current_datetime_upper, date_polled,corrected_polled, uncorrected_polled, avr_pf_polled, avr_tf_polled) + "\n"
                            
                            sql_update_repeat99 = f"update AMR_ERROR set REPEAT = 99 WHERE TAG_ID ='{selected_tag}' \
                                AND TO_CHAR(DATA_DATE, 'DD-MM-YYYY') = TO_CHAR(TRUNC(SYSDATE), 'DD-MM-YYYY');"

                            full_sql_text = full_sql_text + sql_update_repeat99
                            #print(full_sql_text)
                            if full_sql_text: 
                                with cx_Oracle.connect(username, password, f"{hostname}:{port}/{service_name}") as connection:
                                    with connection.cursor() as cursor:
                                        for sql_statement in full_sql_text.split(";"):
                                            if sql_statement.strip():
                                                cursor.execute(sql_statement.strip())    
                                        connection.commit()
                                        #print("Insert data billing successful")
                                        
            # ### hourly      
            else:
                            
                query = """
                    SELECT
                        AMR_FIELD_METER.METER_STREAM_NO as RunNo,
                        AMR_PL_GROUP.PL_REGION_ID as region,
                        AMR_FIELD_ID.TAG_ID as Sitename,
                        AMR_FIELD_METER.METER_NO_STREAM as NoRun,
                        AMR_FIELD_METER.METER_ID as METERID,
                        AMR_VC_TYPE.VC_NAME as VCtype,
                        AMR_FIELD_ID.SIM_IP as IPAddress,
                        AMR_PORT_INFO.PORT_NO as port,
                        amr_poll_range_hourly.evc_type as evc_type,
                        amr_vc_type.vc_name as vc_name,
                        amr_poll_range_hourly.poll_hourly as poll_hourly,
                        amr_poll_range_hourly.poll_hourly_enable as poll_hourly_enable,
                        amr_field_meter.modbus_id as modbus_id,
                        AMR_PORT_INFO.ID as port_id
                    FROM
                        AMR_FIELD_ID,
                        AMR_USER,
                        AMR_FIELD_CUSTOMER,
                        AMR_FIELD_METER,
                        AMR_PL_GROUP,
                        AMR_VC_TYPE,
                        AMR_PORT_INFO,
                        amr_poll_range_hourly,
                        amr_region
                    WHERE
                        AMR_PL_GROUP.PL_REGION_ID = amr_region.id AND
                        AMR_USER.USER_ENABLE=1 AND
                        amr_vc_type.id=amr_poll_range_hourly.evc_type AND
                        AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID AND
                        AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP AND
                        AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID AND
                        AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID AND
                        AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE AND
                        AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID
                        {tag_condition}
                        {region_condition}
                """
                print("hourly")
                tag_condition = "AND AMR_FIELD_ID.TAG_ID IS NOT NULL"
                region_condition = "AND amr_region.REGION_NAME = 'default_region_id'"
                region_results = fetch_data(ptt_pivot_connection, region_query)
                region_options = [str(region[1]) for region in region_results]
                tag_results = fetch_data(ptt_pivot_connection, tag_query, params={"region_id": selected_region})
                tag_options = [str(tag[0]) for tag in tag_results]
                tag_options.sort()
                if selected_tag:
                    tag_condition = f"AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'"
                if selected_region:
                    region_condition = f"AND amr_region.REGION_NAME = '{selected_region}'"
                query = query.format(tag_condition=tag_condition, region_condition=region_condition)
                results = fetch_data(ptt_pivot_connection, query)
           
                df = pd.DataFrame(
                    results,
                    columns=[
                        "RUN",
                        "Region",
                        "Sitename",
                        "NoRun",
                        "METERID",
                        "VCtype",
                        "IPAddress",
                        "Port",
                        "evc_type",
                        "vc_name",
                        "poll_hourly",
                        "poll_hourly_enable",
                        "modbus_id",
                        "port_id"
                    ],
                )

                if selected_row:                         
                        # Convert the selected_row to an integer
                        selected_row_index = int(selected_row) - 1
                        
                        # Check if the index is within the valid range
                        if 0 <= selected_row_index < len(df):
                            
                            row_data = df.iloc[selected_row_index]
                            
                            row_data_dict = row_data.to_dict()
                            
                            run = int(row_data_dict.get("RUN"))
                            print("run",run)
                            Region = row_data_dict.get("Region")
                            Sitename = row_data_dict.get("Sitename")
                            NoRun = row_data_dict.get("NoRun")
                            METERID = row_data_dict.get("METERID")
                            VCtype = row_data_dict.get("VCtype")
                            tcp_ip = row_data_dict.get("IPAddress")
                            tcp_port = int(row_data_dict.get("Port"))
                            evc_type = int(row_data_dict.get("evc_type"))
                            poll_hourly = row_data_dict.get("poll_hourly")
                            
                            poll_hourly_enable = row_data_dict.get("poll_hourly_enable")
                            
                            
                            modbus_id = int(row_data_dict.get("modbus_id"))
                            port_id = int(row_data_dict.get("port_id"))

                            list_of_valueshourly = session['list_of_values_billing'] 
                            print("session billing:",session['list_of_values_billing'])
                            processed_values_hourly = []
                            for item in list_of_valueshourly:
                                if " on " in item:
                                    date_part,time_part  = item.split(" on ")
                                    print(time_part)
                                    hour_part = time_part.split(":")[0].strip()  # Extract the hour part
                                    
                                    date_part = pd.to_datetime(date_part.strip(), format="%d-%m-%Y").strftime('%Y-%m-%d')  # Convert to YYYY-MM-DD
                                    processed_values_hourly.append(hour_part.zfill(2))  # Zero-pad hour part
                                    processed_values_hourly.append(date_part)  # Add the date to the list
                                    print("processed_values_hourly1",processed_values_hourly)
                                else:
                                    processed_values_hourly.append(item)
                                    print("processed_values_hourly2",processed_values_hourly)

                            
                            current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))).strftime('%d-%b-%y %I.%M.%S.%f %p ASIA/BANGKOK')
                            current_datetime_upper = current_datetime.upper()
                            evc_type = evc_type
                            query = """
                            SELECT amh.hourly ,amh.or_der ,amh.address,amh.description,amh.data_type  FROM amr_mapping_hourly amh WHERE amh.evc_type = :evc_type AND address is not null order by amh.hourly
                            ,amh.or_der
                            """
                            poll_resultsbilling = fetch_data(ptt_pivot_connection,query, params={"evc_type": evc_type})
                            # print(poll_resultsbilling)
                            df_mappingbilling = pd.DataFrame(poll_resultsbilling, columns=['hourly','or_der', 'address', 'description', 'data_type'])
                            # print(df_mappingbilling)
                            sql_texts = []
                            for i in range(0, len(processed_values_hourly), 6):
                                
                                values_subset = processed_values_hourly[i:i+6]
                                original_date_str = values_subset[1]  
                                date_obj = datetime.datetime.strptime(original_date_str, '%Y-%m-%d')
                                formatted_date_str = date_obj.strftime('%d-%m-%Y')
                                
                                sql_text_billing_insert = f"""
                                

                                -- อัพเดทหรือเพิ่มข้อมูลใหม่ถ้าจำเป็น
                               -- MERGE statement to handle inserting new or updating existing records
                                MERGE INTO AMR_BILLING_HOURLY_DATA target
                                USING (
                                    SELECT '{METERID}' AS METER_ID, '{run}' AS METER_STREAM_NO, TO_DATE('{formatted_date_str}', 'DD-MM-YYYY') AS DATA_DATE, '{values_subset[0]}' AS DATA_HOUR,
                                        '{current_datetime_upper}' AS CREATED_TIME,
                                        {values_subset[2]} AS UNCORRECTED_VOL, {values_subset[3]} AS CORRECTED_VOL, {values_subset[4]} AS AVR_PF, {values_subset[5]} AS AVR_TF
                                    FROM DUAL
                                ) source
                                ON (
                                    target.METER_ID = source.METER_ID
                                    AND target.METER_STREAM_NO = source.METER_STREAM_NO
                                    AND target.DATA_DATE = source.DATA_DATE
                                    AND target.DATA_HOUR = source.DATA_HOUR
                                )
                                WHEN NOT MATCHED THEN
                                    INSERT (METER_ID, METER_STREAM_NO, DATA_DATE, DATA_HOUR, CREATED_TIME, UNCORRECTED_VOL, CORRECTED_VOL, AVR_PF, AVR_TF)
                                    VALUES (source.METER_ID, source.METER_STREAM_NO, source.DATA_DATE, source.DATA_HOUR, source.CREATED_TIME, source.UNCORRECTED_VOL, source.CORRECTED_VOL, source.AVR_PF, source.AVR_TF);

                                -- Delete records with mismatched values
                                DELETE FROM AMR_BILLING_HOURLY_DATA
                                WHERE 
                                    METER_ID = '{METERID}'
                                    AND METER_STREAM_NO = '{run}'
                                    AND DATA_DATE = TO_DATE('{formatted_date_str}', 'DD-MM-YYYY')
                                    AND DATA_HOUR = '{values_subset[0]}'
                                    AND (UNCORRECTED_VOL != {values_subset[2]} OR CORRECTED_VOL != {values_subset[3]});

                                -- Insert deleted/mismatched records into AMR_BILLING_HOURLY_DATA_ERROR
                                INSERT INTO AMR_BILLING_HOURLY_DATA_ERROR (METER_ID, METER_STREAM_NO, DATA_DATE, DATA_HOUR, CREATED_TIME, UNCORRECTED_VOL, CORRECTED_VOL, AVR_PF, AVR_TF)
                                SELECT 
                                    METER_ID, METER_STREAM_NO, DATA_DATE, DATA_HOUR, SYSDATE AS CREATED_TIME,
                                    UNCORRECTED_VOL, CORRECTED_VOL, AVR_PF, AVR_TF
                                FROM AMR_BILLING_HOURLY_DATA
                                WHERE 
                                    METER_ID = '{METERID}'
                                    AND METER_STREAM_NO = '{run}'
                                    AND DATA_DATE = TO_DATE('{formatted_date_str}', 'DD-MM-YYYY')
                                    AND DATA_HOUR = '{values_subset[0]}'
                                    AND (UNCORRECTED_VOL != {values_subset[2]} OR CORRECTED_VOL != {values_subset[3]});
                                    """

                                sql_texts.append(sql_text_billing_insert.strip())

                            full_sql_text = "\n".join(sql_texts)
                            print("full_sql_text", full_sql_text)

                            with ptt_pivot_connection.cursor() as cursor:
                                for sql_statement in full_sql_text.split(";"):
                                    if sql_statement.strip():
                                        cursor.execute(sql_statement.strip())
                                ptt_pivot_connection.commit()

                            print("Insert data billing successful")
        
            response = {"status": "success", "message": "Data updated successfully"}

        except ValueError as ve:
            response = {"status": "error", "message": str(ve)}
        except cx_Oracle.DatabaseError as e:
            (error,) = e.args
            print(f"Oracle Database Error {error.code}: {error.message}")
            traceback.print_exc()
            response = {
                "status": "error",
                "message": f"Database Error: {error.code} - {error.message}",
            }
        except Exception as e:
            print(f"Error: {e}")
            traceback.print_exc()
            response = {
                "status": "error",
                "message": f"An error occurred while updating data: {str(e)}",
            }

        return jsonify(response)
############ /Manualpoll_data  #####################



@app.route('/get_tag')
def get_tag():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
    region_name = request.args.get('region')

    # Use the selected REGION_NAME to fetch associated tag_id values
    tag_query = """
    SELECT DISTINCT tag_id
    FROM VW_ASGS_AMR_BILLING_DATA
    WHERE REGION_NAME = :region_name
    ORDER BY tag_id
    """
    tag_results = fetch_data(ptt_pivot_connection,tag_query, {'region_name': region_name})

    # Return the tag_id values as JSON
    return jsonify(tag_results)
############ / View Billing Data  #####################

@app.route('/Manualpoll')
@login_required
def index():
    # if 'username' not in session:
    #     return redirect(url_for('login'))
    global tcp_ip, tcp_port
    return render_template('index.html', slave_id=0, function_code=0, starting_address=0, quantity=0, data_list=[], is_16bit=False, communication_traffic=communication_traffic)

@app.route("/write_evc_old",methods=["GET"])
@login_required
def Manualpoll_data_old():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
       
        region_query = """
            SELECT * FROM AMR_REGION 
        """
        tag_query = """
            SELECT DISTINCT TAG_ID
            FROM AMR_FIELD_ID, AMR_PL_GROUP
            WHERE AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID
            AND AMR_PL_GROUP.PL_REGION_ID = :region_id
        """
        run_query = """
            SELECT DISTINCT METER_STREAM_NO
            FROM AMR_FIELD_ID , amr_field_meter
            WHERE amr_field_id.meter_id = amr_field_meter.meter_id
            AND amr_field_id.tag_id = :tag_id
        """
        region_results = fetch_data(ptt_pivot_connection,region_query)
        region_options = [str(region[0]) for region in region_results]

        query = """
            SELECT
                AMR_FIELD_METER.METER_STREAM_NO as RunNo,
                AMR_PL_GROUP.PL_REGION_ID as region,
                AMR_FIELD_ID.TAG_ID as Sitename,
                AMR_FIELD_METER.METER_NO_STREAM as NoRun,
                AMR_FIELD_METER.METER_ID as METERID,
                AMR_VC_TYPE.VC_NAME as VCtype,
                AMR_FIELD_ID.SIM_IP as IPAddress,
                AMR_PORT_INFO.PORT_NO as port
            FROM
                AMR_FIELD_ID,
                AMR_USER,
                AMR_FIELD_CUSTOMER,
                AMR_FIELD_METER,
                AMR_PL_GROUP,
                AMR_VC_TYPE,
                AMR_PORT_INFO
            WHERE
                AMR_USER.USER_ENABLE=1 AND
                AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID AND
                AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP AND
                AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID AND
                AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID AND
                AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE AND
                AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID
                {tag_condition}
                {region_condition}
                {run_condition}
        """

        tag_condition = "AND AMR_FIELD_ID.TAG_ID IS NOT NULL"
        region_condition = "AND amr_pl_group.pl_region_id = 'default_region_id'"
        run_condition = "AND AMR_FIELD_METER.METER_STREAM_NO IS NOT NULL"

        selected_tag = request.args.get("tag_dropdown")
        selected_region = request.args.get("region_dropdown")
        selected_run = request.args.get("run_dropdown")

        region_results = fetch_data(ptt_pivot_connection,region_query)
        region_options = [str(region[0]) for region in region_results]

        tag_results = fetch_data(ptt_pivot_connection,tag_query, params={"region_id": selected_region})
        tag_options = [str(tag[0]) for tag in tag_results]

        run_results = fetch_data(ptt_pivot_connection,run_query, params={"tag_id": selected_tag})
        run_options = [str(run[0]) for run in run_results]

        # Sort the tag options alphabetically
        tag_options.sort()

        
        if selected_tag:
            tag_condition = f"AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'"
        if selected_region:
            region_condition = f"AND amr_pl_group.pl_region_id = '{selected_region}'"
        if selected_run:
            run_condition = f"AND AMR_FIELD_METER.METER_STREAM_NO = {selected_run}"


        query = query.format(tag_condition=tag_condition, region_condition=region_condition, run_condition=run_condition)

        results = fetch_data(ptt_pivot_connection,query)
        
        df = pd.DataFrame(
            results,
            columns=[
                "RUN",
                "Region",
                "Sitename",
                "NoRun",
                "METERID",
                "VCtype",
                "IPAddress",
                "Port",
            ],
        )
    tcp_ip = df.get(["IPAddress"]).values.tolist()
    if tcp_ip:
        ip_str = str(tcp_ip).strip("['']")
        print(ip_str)
    else:
        ip_str = [''] 


    tcp_port = df.get(["Port"]).values.tolist()
    if tcp_port:
        Port_str = str(tcp_port).strip("['']")
    else:
        Port_str = [''] 
    
    return render_template(
        "write_evc_old.html",
        tables=[df.to_html(classes="data")],
        titles=df.columns.values,
        selected_tag=selected_tag,
        selected_region=selected_region,
        selected_run=selected_run,
        region_options=region_options,
        tag_options=tag_options,
        run_options=run_options,
        df=df,ip_str=ip_str,tcp_port=tcp_port,Port_str=Port_str,tcp_ip=tcp_ip
    )


@app.route('/write_evc_old', methods=['POST'])
@login_required
def read_data_old_1():
    if 'username' not in session:
        return redirect(url_for('login'))
    try:
        global change_to_32bit_counter, communication_traffic
        
        # Fetch form data
        slave_id = int(request.form['slave_id'])
        function_code = int(request.form['function_code'])
        starting_address = int(request.form['starting_address'])
        quantity = int(request.form['quantity'])
        tcp_ip = request.form['tcp_ip']
        tcp_port = int(request.form['tcp_port'])
        is_16bit = request.form.get('is_16bit') == 'true'

        # Adjust quantity based on data format
        if not is_16bit and change_to_32bit_counter > 0:
            quantity *= 2
            change_to_32bit_counter -= 1

        
        request_data = bytearray()
        for i in range(quantity // 2):
            data_i = float(request.form.get(f'data_{i}'))  
            request_data.extend(struct.pack('>f', data_i)) 

        
        request_message = format_tx_message(slave_id, function_code, starting_address, quantity, request_data)
       
        # Connect to Modbus TCP server
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.settimeout(20) 
            sock.connect((tcp_ip, tcp_port))
            communication_traffic = []

            communication_traffic.append({"direction": "TX", "data": request_message.hex()})
            sock.send(request_message)
            response = sock.recv(1024)

            communication_traffic.append({"direction": "RX", "data": response.hex()})
            if response[1:2] != b'\x10':
            
                    abort(400, f"Error: Unexpected response code from device {communication_traffic[1]} ,{response[1:2]}!")
            else:
                    pass

            data = response[3:]
            
            
            session['tcp_ip'] = tcp_ip
            session['tcp_port'] = tcp_port

    
        with connect_to_ptt_pivot_db() as ptt_pivot_connection:
            print("Active Connection:", active_connection)
        
            region_query = """
                SELECT * FROM AMR_REGION 
            """
            tag_query = """
                SELECT DISTINCT TAG_ID
                FROM AMR_FIELD_ID, AMR_PL_GROUP
                WHERE AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID
                AND AMR_PL_GROUP.PL_REGION_ID = :region_id
            """
            run_query = """
                SELECT DISTINCT METER_STREAM_NO
                FROM AMR_FIELD_ID , amr_field_meter
                WHERE amr_field_id.meter_id = amr_field_meter.meter_id
                AND amr_field_id.tag_id = :tag_id
            """
            region_results = fetch_data(ptt_pivot_connection,region_query)
            region_options = [str(region[0]) for region in region_results]

            query = """
                SELECT
                    AMR_FIELD_METER.METER_STREAM_NO as RunNo,
                    AMR_PL_GROUP.PL_REGION_ID as region,
                    AMR_FIELD_ID.TAG_ID as Sitename,
                    AMR_FIELD_METER.METER_NO_STREAM as NoRun,
                    AMR_FIELD_METER.METER_ID as METERID,
                    AMR_VC_TYPE.VC_NAME as VCtype,
                    AMR_FIELD_ID.SIM_IP as IPAddress,
                    AMR_PORT_INFO.PORT_NO as port
                FROM
                    AMR_FIELD_ID,
                    AMR_USER,
                    AMR_FIELD_CUSTOMER,
                    AMR_FIELD_METER,
                    AMR_PL_GROUP,
                    AMR_VC_TYPE,
                    AMR_PORT_INFO
                WHERE
                    AMR_USER.USER_ENABLE=1 AND
                    AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID AND
                    AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP AND
                    AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID AND
                    AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID AND
                    AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE AND
                    AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID
                    {tag_condition}
                    {region_condition}
                    {run_condition}
            """

            tag_condition = "AND AMR_FIELD_ID.TAG_ID IS NOT NULL"
            region_condition = "AND amr_pl_group.pl_region_id = 'default_region_id'"
            run_condition = "AND AMR_FIELD_METER.METER_STREAM_NO IS NOT NULL"

            selected_tag = request.args.get("tag_dropdown")
            selected_region = request.args.get("region_dropdown")
            selected_run = request.args.get("run_dropdown")

            region_results = fetch_data(ptt_pivot_connection,region_query)
            region_options = [str(region[0]) for region in region_results]

            tag_results = fetch_data(ptt_pivot_connection,tag_query, params={"region_id": selected_region})
            tag_options = [str(tag[0]) for tag in tag_results]

            run_results = fetch_data(ptt_pivot_connection,run_query, params={"tag_id": selected_tag})
            run_options = [str(run[0]) for run in run_results]

            # Sort the tag options alphabetically
            tag_options.sort()

            if selected_tag:
                tag_condition = f"AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'"
            if selected_region:
                region_condition = f"AND amr_pl_group.pl_region_id = '{selected_region}'"
            if selected_run:
                run_condition = f"AND AMR_FIELD_METER.METER_STREAM_NO = {selected_run}"


            query = query.format(tag_condition=tag_condition, region_condition=region_condition, run_condition=run_condition)

            results = fetch_data(ptt_pivot_connection,query)
            
            df = pd.DataFrame(
                results,
                columns=[
                    "RUN",
                    "Region",
                    "Sitename",
                    "NoRun",
                    "METERID",
                    "VCtype",
                    "IPAddress",
                    "Port",
                ],
            )
        tcp_ip = df.get(["IPAddress"]).values.tolist()
        if tcp_ip:
            ip_str = str(tcp_ip).strip("['']")
            print(ip_str)
        else:
            ip_str = [''] 


        tcp_port = df.get(["Port"]).values.tolist()
        if tcp_port:
            Port_str = str(tcp_port).strip("['']")
        else:
            Port_str = [''] 
        
    except TimeoutError:
            abort(400, f"Error: Connection timed out while waiting for response from {tcp_ip}:{tcp_port}!")
    except Exception as e:
            abort(400, f"Error: {e}")
    
    return render_template('write_evc_old.html',   df=df,

            slave_id=slave_id,
            function_code=function_code,
            starting_address=starting_address,
            quantity=quantity,
            is_16bit=is_16bit,
            communication_traffic=communication_traffic,
            data=data,
            tables=[df.to_html(classes="data")],
            titles=df.columns.values,
            selected_tag=selected_tag,
            selected_region=selected_region,
            region_options=region_options,
            tag_options=tag_options,ip_str=ip_str,tcp_port=tcp_port,Port_str=Port_str,tcp_ip=tcp_ip)
        
@app.route("/write_evc",methods=["GET"])
@login_required
def Manualpoll_data_write_evc():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        username = session['username']
        region_query = """
            SELECT * FROM AMR_REGION 
        """
        tag_query = """
            SELECT DISTINCT AMR_FIELD_ID.TAG_ID
            FROM AMR_FIELD_ID,amr_region,AMR_PL_GROUP
            
            WHERE AMR_PL_GROUP.PL_REGION_ID = amr_region.id
            AND AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
            AND amr_region.REGION_NAME like = :region_id
            AND AMR_FIELD_ID.tag_id NOT like '%.remove%'
            ORDER BY  TAG_ID
        
                    
        """
        run_query = """
            SELECT DISTINCT METER_STREAM_NO
            FROM AMR_FIELD_ID , amr_field_meter
            WHERE amr_field_id.meter_id = amr_field_meter.meter_id
            AND amr_field_id.tag_id = :tag_id
        """
        region_results = fetch_data(ptt_pivot_connection,region_query)
        region_options = [str(region[1]) for region in region_results]
      
        query = """
            SELECT
                AMR_FIELD_METER.METER_STREAM_NO as RunNo,
                AMR_PL_GROUP.PL_REGION_ID as region,
                AMR_FIELD_ID.TAG_ID as Sitename,
                AMR_FIELD_METER.METER_NO_STREAM as NoRun,
                AMR_FIELD_METER.METER_ID as METERID,
                AMR_VC_TYPE.VC_NAME as VCtype,
                AMR_VC_TYPE.id as EVCtype,
                AMR_FIELD_ID.SIM_IP as IPAddress,
                AMR_PORT_INFO.PORT_NO as port
            FROM
                AMR_FIELD_ID,
                AMR_USER,
                AMR_FIELD_CUSTOMER,
                AMR_FIELD_METER,
                AMR_PL_GROUP,
                AMR_VC_TYPE,
                AMR_PORT_INFO,
                amr_region
            WHERE
                AMR_USER.USER_ENABLE=1
                AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id AND
                AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID AND
                AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP AND
                AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID AND
                AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID AND
                AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE AND
                AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID
                {tag_condition}
                {region_condition}
                {run_condition}
        """
        tag_condition = "AND AMR_FIELD_ID.TAG_ID IS NOT NULL"
        region_condition = "AND amr_region.REGION_NAME = 'default_region_id'"
        run_condition = "AND AMR_FIELD_METER.METER_STREAM_NO IS NOT NULL"

        selected_tag = request.args.get("tag_dropdown")
        selected_region = request.args.get("region_dropdown")
        selected_run = request.args.get("run_dropdown")

        region_results = fetch_data(ptt_pivot_connection,region_query)
        region_options = [str(region[1]) for region in region_results]

        tag_results = fetch_data(ptt_pivot_connection,tag_query, params={"region_id": selected_region})
        tag_options = [str(tag[0]) for tag in tag_results]

        run_results = fetch_data(ptt_pivot_connection,run_query, params={"tag_id": selected_tag})
        run_options = [str(run[0]) for run in run_results]

        # Sort the tag options alphabetically
        tag_options.sort()

        if selected_tag:
            tag_condition = f"AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'"
        if selected_region:
            region_condition = f"AND amr_region.REGION_NAME = '{selected_region}'"
        if selected_run:
            run_condition = f"AND AMR_FIELD_METER.METER_STREAM_NO = {selected_run}"


        query = query.format(tag_condition=tag_condition, region_condition=region_condition, run_condition=run_condition)

        results = fetch_data(ptt_pivot_connection,query)
       
        
        df = pd.DataFrame(
            results,
            columns=[
                "RUN",
                "Region",
                "Sitename",
                "NoRun",
                "METERID",
                "VCtype",
                "EVCtype",
                "IPAddress",
                "Port",
            ],
        )

    
    evc_type = df.get(["EVCtype"]).values.tolist()
    if evc_type:
        type_str = str(evc_type).strip("['']")
       
    else:
        type_str = [''] 
    
    session['type_str'] = type_str
    tcp_ip = df.get(["IPAddress"]).values.tolist()
    if tcp_ip:
        ip_str = str(tcp_ip).strip("['']")
        
    else:
        ip_str = [''] 
    session['ip_str'] = ip_str

    tcp_port = df.get(["Port"]).values.tolist()
    if tcp_port:
        Port_str = str(tcp_port).strip("['']")
    else:
        Port_str = [''] 
    session['Port_str'] = Port_str
    return render_template(
        "evc.html",
        tables=[df.to_html(classes="data")],
        titles=df.columns.values,
        selected_tag=selected_tag,
        selected_region=selected_region,
        selected_run=selected_run,
        region_options=region_options,
        tag_options=tag_options,
        run_options=run_options,
        df=df,ip_str=ip_str,tcp_port=tcp_port,Port_str=Port_str,tcp_ip=tcp_ip,type_str=type_str,username=username
    )


@app.route('/write_evc', methods=['POST'])
@login_required
def read_data_write_evc():
        
        if 'username' not in session:
            return redirect(url_for('login'))
        try:
            global change_to_32bit_counter, communication_traffic
            username = session['username']
            # Fetch form data
            slave_id = int(1)
            function_code = int(16)
            starting_addresses = []
            for key in request.form:
                if key.startswith('starting_address_'):
                    starting_address = int(request.form[key])
                    if starting_address:
                        starting_addresses.append(int(starting_address))
            #print(starting_addresses)
            
            quantity = int(2)
            tcp_ip = session['ip_str']
            
            tcp_port = int(session['Port_str'])
            
            is_16bit = request.form.get('is_16bit') == 'true'
            evc_type = session['type_str']
            print(evc_type)
            # Adjust quantity based on data format
            if not is_16bit and change_to_32bit_counter > 0:
                quantity *= 2
                change_to_32bit_counter -= 1

            
            request_data = bytearray()
            for i in range(quantity // 2):
                data_i = float(request.form.get(f'data_{i}'))
                print("data_i",data_i)  
                request_data.extend(struct.pack('>f', data_i)) 

             
            request_message = format_tx_message(slave_id, function_code, starting_address, quantity, request_data)
            print(request_message)
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(20) 
            sock.connect((tcp_ip, tcp_port))
            
            
            # Connect to Modbus TCP server
            if int(evc_type) in [5, 8, 9, 10,12]:
                for _ in range(2):
                    sock.send(request_message)
                    print("Ataris",sock)
                    time.sleep(1)
                try:
                    response = sock.recv(1024)
                except socket.timeout:
                    abort(400, f"Error: Connection timed out while waiting for response from {tcp_ip}:{tcp_port}!")
                except Exception as e:
                    abort(400, f"Error while receiving data: {e}")

            communication_traffic = []
        
            communication_traffic.append({"direction": "TX", "data": request_message.hex()})
            print("TX",communication_traffic)
           
            sock.send(request_message)
            time.sleep(1)
            try:
                response = sock.recv(1024)
            except socket.timeout:
                abort(400, f"Error: Connection timed out while waiting for response from {tcp_ip}:{tcp_port}!")
            except Exception as e:
                abort(400, f"Error while receiving data: {e}")

            communication_traffic.append({"direction": "RX", "data": response.hex()})
            print("RX",communication_traffic)
            if response[1:2] != b'\x10':
            
                    abort(400, f"Error: Unexpected response code from device {communication_traffic[1]} ,{response[1:2]}!")
            else:
                    pass
            data = response[3:]
            

            session['tcp_ip'] = tcp_ip
            session['tcp_port'] = tcp_port

            with connect_to_ptt_pivot_db() as ptt_pivot_connection:
                print("Active Connection:", active_connection)
            
                region_query = """
                    SELECT * FROM AMR_REGION 
                """
                tag_query = """
                     SELECT DISTINCT AMR_FIELD_ID.TAG_ID
            FROM AMR_FIELD_ID,amr_region,AMR_PL_GROUP
            
            WHERE AMR_PL_GROUP.PL_REGION_ID = amr_region.id
            AND AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
            AND amr_region.REGION_NAME like = :region_id
            AND AMR_FIELD_ID.tag_id NOT like '%.remove%'
            ORDER BY  TAG_ID"""

                run_query = """
                    SELECT DISTINCT METER_STREAM_NO
                    FROM AMR_FIELD_ID , amr_field_meter
                    WHERE amr_field_id.meter_id = amr_field_meter.meter_id
                    AND amr_field_id.tag_id = :tag_id
                """
                region_results = fetch_data(ptt_pivot_connection,region_query)
                region_options = [str(region[1]) for region in region_results]

                query = """
                    SELECT
                        AMR_FIELD_METER.METER_STREAM_NO as RunNo,
                        AMR_PL_GROUP.PL_REGION_ID as region,
                        AMR_FIELD_ID.TAG_ID as Sitename,
                        AMR_FIELD_METER.METER_NO_STREAM as NoRun,
                        AMR_FIELD_METER.METER_ID as METERID,
                        AMR_VC_TYPE.VC_NAME as VCtype,
                        AMR_FIELD_ID.SIM_IP as IPAddress,
                        AMR_PORT_INFO.PORT_NO as port
                    FROM
                        AMR_FIELD_ID,
                        AMR_USER,
                        AMR_FIELD_CUSTOMER,
                        AMR_FIELD_METER,
                        AMR_PL_GROUP,
                        AMR_VC_TYPE,
                        AMR_PORT_INFO,
                        amr_region
                    WHERE
                    AMR_PL_GROUP.PL_REGION_ID = amr_region.id AND
                        AMR_USER.USER_ENABLE=1 AND
                        AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID AND
                        AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP AND
                        AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID AND
                        AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID AND
                        AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE AND
                        AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID
                        {tag_condition}
                        {region_condition}
                        {run_condition}
                """
                
                tag_condition = "AND AMR_FIELD_ID.TAG_ID IS NOT NULL"
                region_condition = "AND amr_region.REGION_NAME = 'default_region_id'"
                run_condition = "AND AMR_FIELD_METER.METER_STREAM_NO IS NOT NULL"

                selected_tag = request.args.get("tag_dropdown")
                selected_region = request.args.get("region_dropdown")
                selected_run = request.args.get("run_dropdown")

                region_results = fetch_data(ptt_pivot_connection,region_query)
                region_options = [str(region[1]) for region in region_results]

                tag_results = fetch_data(ptt_pivot_connection,tag_query, params={"region_id": selected_region})
                tag_options = [str(tag[0]) for tag in tag_results]

                run_results = fetch_data(ptt_pivot_connection,run_query, params={"tag_id": selected_tag})
                run_options = [str(run[0]) for run in run_results]

                # Sort the tag options alphabetically
                tag_options.sort()

                if selected_tag:
                    tag_condition = f"AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'"
                if selected_region:
                    region_condition = f"AND amr_region.REGION_NAME = '{selected_region}'"
                if selected_run:
                    run_condition = f"AND AMR_FIELD_METER.METER_STREAM_NO = {selected_run}"


                query = query.format(tag_condition=tag_condition, region_condition=region_condition, run_condition=run_condition)

                results = fetch_data(ptt_pivot_connection,query)
                
                df = pd.DataFrame(
                    results,
                    columns=[
                        "RUN",
                        "Region",
                        "Sitename",
                        "NoRun",
                        "METERID",
                        "VCtype",
                        "IPAddress",
                        "Port",
                    ],
                )
            tcp_ip = df.get(["IPAddress"]).values.tolist()
            if tcp_ip:
                ip_str = str(tcp_ip).strip("['']")
                print(ip_str)
            else:
                ip_str = [''] 

            tcp_port = df.get(["Port"]).values.tolist()
            if tcp_port:
                Port_str = str(tcp_port).strip("['']")
            else:
                Port_str = [''] 
        
        except TimeoutError:
            abort(400, f"Error: Connection timed out while waiting for response from {tcp_ip}:{tcp_port}!")
        except Exception as e:
            print(f"Exception occurred: {e}")
            abort(400, f"Error: {e}")
        
        return render_template('evc.html',   df=df,

                slave_id=slave_id,
                function_code=function_code,
                starting_address=starting_address,
                quantity=quantity,
                is_16bit=is_16bit,
                communication_traffic=communication_traffic,
                data=data,
                tables=[df.to_html(classes="data")],
                titles=df.columns.values,
                selected_tag=selected_tag,
                selected_region=selected_region,
                region_options=region_options,
                tag_options=tag_options,ip_str=ip_str,tcp_port=tcp_port,Port_str=Port_str,tcp_ip=tcp_ip,username=username)
            
@app.route('/get_tag_asgs')
def get_tags_asgs():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        region_name = request.args.get('region')

        # Use the selected REGION_NAME to fetch associated tag_id values
        tag_query = """
        SELECT DISTINCT tag_id
        FROM VW_AMR_BILLING_DATA
        WHERE REGION_NAME = :region_name
        ORDER BY tag_id
        """
        tag = fetch_data(ptt_pivot_connection, tag_query, {'region_name': region_name})

        # Return the tag_id values as JSON
        return jsonify(tag)

@app.route("/get_runs", methods=["GET"])
def get_runs():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        selected_tag = request.args.get("selected_tag")

        run_query = """
            SELECT DISTINCT METER_STREAM_NO
            FROM AMR_FIELD_ID , amr_field_meter
            WHERE amr_field_id.meter_id = amr_field_meter.meter_id
            AND amr_field_id.tag_id = :tag_id
        """

        run_results = fetch_data(ptt_pivot_connection,run_query, params={"tag_id": selected_tag})
        run_options = [str(run[0]) for run in run_results]
        run_options.sort()
        return jsonify({"run_options": run_options})
    
@app.route('/manual_write')
@login_required
def write_test():
    if 'username' not in session:
        return redirect(url_for('login'))

    global tcp_ip, tcp_port
    return render_template('write_test_ptt.html', slave_id=0, function_code=0, starting_address=0, quantity=0, data_list=[], is_16bit=False, communication_traffic=communication_traffic)
@app.route('/manual_write', methods=['POST'])
@login_required
def read_data_write():
    if 'username' not in session:
        return redirect(url_for('login'))
    try:
        global change_to_32bit_counter, communication_traffic
        
        # Fetch form data
        slave_id = int(request.form['slave_id'])
        function_code = int(request.form['function_code'])
        starting_address = int(request.form['starting_address'])
        quantity = int(request.form['quantity'])
        tcp_ip = request.form['tcp_ip']
        tcp_port = int(request.form['tcp_port'])
        is_16bit = request.form.get('is_16bit') == 'true'

        # Adjust quantity based on data format
        if not is_16bit and change_to_32bit_counter > 0:
            quantity *= 2
            change_to_32bit_counter -= 1

        
        request_data = bytearray()
        for i in range(quantity // 2):
            data_i = float(request.form.get(f'data_{i}'))  
            request_data.extend(struct.pack('>f', data_i)) 

        
        request_message = format_tx_message(slave_id, function_code, starting_address, quantity, request_data)
       
        # Connect to Modbus TCP server
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.settimeout(20) 
            sock.connect((tcp_ip, tcp_port))
            communication_traffic = []

            communication_traffic.append({"direction": "TX", "data": request_message.hex()})
            sock.send(request_message)
            time.sleep(3)
            response = sock.recv(1024)

            communication_traffic.append({"direction": "RX", "data": response.hex()})
            if response[1:2] != b'\x10':
            
                    abort(400, f"Error: Unexpected response code from device {communication_traffic[1]} ,{response[1:2]}!")
            else:
                    pass

            data = response[3:]
            
            
         
            
            
            session['tcp_ip'] = tcp_ip
            session['tcp_port'] = tcp_port

    except TimeoutError:
            abort(400, f"Error: Connection timed out while waiting for response from {tcp_ip}:{tcp_port}!")
    except Exception as e:
            abort(400, f"Error: {e}")
    
    return render_template('write_test_ptt.html',  
            slave_id=slave_id,
            function_code=function_code,
            starting_address=starting_address,
            quantity=quantity,
            is_16bit=is_16bit,
            communication_traffic=communication_traffic,
            data=data,
            
           )
        
# update polling
def update_sql(connection, sql_statement):
    with connection.cursor() as cursor:
        
        cursor.execute(sql_statement)
    connection.commit()

def insert_address_range_to_oracle(
    connection, poll_config, poll_billing, enable_config, enable_billing, evc_type
):
    
    with connection.cursor() as cursor:
        sql_insert = """
            INSERT INTO AMR_POLL_RANGE (POLL_CONFIG, POLL_BILLING, POLL_CONFIG_ENABLE, POLL_BILLING_ENABLE, EVC_TYPE)
            VALUES (:1, :2, :3, :4, :5)
        """

        # Convert enable_config and enable_billing to comma-separated strings
        enable_config_str = ",".join(map(str, enable_config))
        enable_billing_str = ",".join(map(str, enable_billing))

        data_to_insert = (
            poll_config,
            poll_billing,
            enable_config_str,
            enable_billing_str,
            evc_type,
        )

        cursor.execute( sql_insert, data_to_insert)

    connection.commit()

@app.route("/polling_route")
@login_required
def polling_route():
    if 'username' not in session:
        return redirect(url_for('login'))
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        username = session['username']
        print("Active Connection:", active_connection)
        # Fetch type options for the dropdown
        type_query = "SELECT VC_NAME FROM AMR_VC_TYPE ORDER BY VC_NAME"
        type_results = fetch_data(ptt_pivot_connection,type_query)
        type_options = [str(type[0]) for type in type_results]
        # print(type_results)

        # Define the base query for fetching polling data
        base_query = """
        SELECT
            apr.evc_type,
            apr.poll_config,
            apr.poll_billing,
            apr.poll_config_enable,
            apr.poll_billing_enable
        FROM
            amr_poll_range apr
        JOIN
            amr_vc_type avt ON apr.evc_type = avt.id
        {type_condition}
        """

        # Get selected type from the dropdown
        selected_type = request.args.get("type_dropdown")

        # Define type condition based on the selected type
        type_condition = f"AND avt.VC_NAME = '{selected_type}'" if selected_type else ""

        # Check if a type is selected before executing the query
        if selected_type:
            # Modify the base query with the selected conditions
            query = base_query.format(type_condition=type_condition)

            # Fetch data using the modified query
            results = fetch_data(ptt_pivot_connection,query)
            # print(results)

            columns = [
                "evc_type",
                "poll_config",
                "poll_billing",
                "poll_config_enable",
                "poll_billing_enable",
            ]
            df = pd.DataFrame(results, columns=columns)

            
            poll_config_list = df.get(["poll_config"]).values.tolist()
            list_config = str(poll_config_list[0]).strip("[]'").split(",")
            # print("===", poll_config_list)
            
            poll_billing_list = df.get(["poll_billing"]).values.tolist()
            list_billing = str(poll_billing_list[0]).strip("[]'").split(",")
            
            poll_config_enable_list = df.get(["poll_config_enable"]).values.tolist()
            list_enable_config = str(poll_config_enable_list[0]).strip("[]'").split(",")
        
            poll_billing_enable_list = df.get(["poll_billing_enable"]).values.tolist()
            list_enable_billing = str(poll_billing_enable_list[0]).strip("[]'").split(",")
            
            return render_template(
                "polling.html",
                tables=[df.to_html(classes="data", index=False)],
                titles=columns,
                selected_type=selected_type,
                type_options=type_options,
                list_config=list_config,
                list_billing=list_billing,
                list_enable_config=list_enable_config,
                list_enable_billing=list_enable_billing,username=username
            )
        else:
        # Render the HTML template without the table if no type is selected
            return render_template(
                "polling.html",
                tables=[],
                titles=[],
                selected_type=None,
                type_options=type_options,
                list_config=[],
                list_billing=[],
                list_enable_config=[],
                list_enable_billing=[],
                username=username
            )
    
MAX_ADDRESS_LENGTH = 249

@app.route("/update_polling_data", methods=["POST"])
@login_required
def update_polling_data(): 
    if 'username' not in session:
        return redirect(url_for('login'))
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        selected_type = request.form.get("selected_type")
        print(selected_type)
        type_id_query = f"SELECT ID FROM AMR_VC_TYPE WHERE VC_NAME LIKE '{selected_type}'"
        results = fetch_data(ptt_pivot_connection,type_id_query)
        print(results)
        type_id = str(results[0]).strip("',()")
        print(type_id)
        type_query = "SELECT VC_NAME FROM AMR_VC_TYPE ORDER BY VC_NAME"
        type_results = fetch_data(ptt_pivot_connection,type_query)
        type_options = [str(type[0]) for type in type_results]
        
        # Update configuration data
        poll_config_all = ""
        enable_config = ""
        for i in range(0, 5):
            start_key = f"start_config{i + 1}"
            end_key = f"end_config{i + 1}"
            enable_key = f"enable_config[{i}]"
            
            start_value = request.form.get(start_key)
            end_value = request.form.get(end_key)
            enable_value = 1 if request.form.get(enable_key) == "on" else 0

            address_range = f"{start_value},{end_value}"
            if len(poll_config_all + address_range) <= MAX_ADDRESS_LENGTH:
                if i > 0:
                    poll_config_all += ","
                poll_config_all += address_range

            if i == 0:
                enable_config = str(enable_value)
            else:
                enable_config +=  "," + str(enable_value)
                
        print("poll_config:", poll_config_all)
        print("poll_config_enable:", enable_config)
        
        # Update billing data
        poll_billing_all = ""
        enable_billing = ""
        for i in range(0, 10):
            start_key = f"start{i + 1}"
            end_key = f"end{i + 1}"
            enable_key = f"enable[{i}]"
            
            start_value = request.form.get(start_key)
            end_value = request.form.get(end_key)
            enable_value = 1 if request.form.get(enable_key) == "on" else 0
            
            address_range = f"{start_value},{end_value}"
            if len(poll_billing_all + address_range) <= MAX_ADDRESS_LENGTH:
                if i > 0:
                    poll_billing_all += ","
                poll_billing_all += address_range

            if i == 0:
                enable_billing = str(enable_value)
            else:
                enable_billing += "," + str(enable_value)
        
        print("poll_billing:", poll_billing_all)
        print("poll_config_enable:", enable_billing)

        update_query = f"""
        UPDATE amr_poll_range
        SET 
            poll_config = '{poll_config_all}',
            poll_billing = '{poll_billing_all}',
            poll_config_enable = '{enable_config}',
            poll_billing_enable = '{enable_billing}'
        WHERE evc_type = '{type_id}'
        """
        print("Update Query:", update_query)
        
        update_sql(ptt_pivot_connection, update_query)
        # with connection.cursor() as cursor:
        
        #     cursor.execute(update_query)
        # connection.commit()
        # After updating the data, you may redirect to the polling route or perform any other necessary actions
        if execute_query(connect_to_ptt_pivot_db(), update_query):
                flash('polling successfully', 'success') 
                
                return render_template('polling.html',type_options=type_options)
        else:
                flash('Failed to polling', 'error')
    return redirect("/polling_route",type_options=type_options)

@app.route("/add_polling_route")
@login_required
def add_polling_route():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template("add_polling.html")

MAX_ADDRESS_LENGTH = 249


@app.route("/polling_hourly")
@login_required
def polling_hourly():
    if 'username' not in session:
        return redirect(url_for('login'))
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        # Fetch type options for the dropdown
        username = session['username']
        type_query = "SELECT VC_NAME FROM AMR_VC_TYPE ORDER BY VC_NAME"
        type_results = fetch_data(ptt_pivot_connection,type_query)
        type_options = [str(type[0]) for type in type_results]
        # print(type_results)

        # Define the base query for fetching polling data
        base_query = """
        SELECT
            apr.evc_type,
            apr.poll_hourly,
            apr.poll_hourly_enable
        FROM
            amr_poll_range_hourly apr
        JOIN
            amr_vc_type avt ON apr.evc_type = avt.id
        {type_condition}
        """

        # Get selected type from the dropdown
        selected_type = request.args.get("type_dropdown")
        print(selected_type)
        # Define type condition based on the selected type
        type_condition = f"AND avt.VC_NAME = '{selected_type}'" if selected_type else ""

        # Check if a type is selected before executing the query
        if selected_type:
            # Modify the base query with the selected conditions
            query = base_query.format(type_condition=type_condition)

            # Fetch data using the modified query
            results = fetch_data(ptt_pivot_connection,query)
            # print(results)

            columns = [
                "evc_type",

                "poll_hourly",
                "poll_hourly_enable",
            ]
            df = pd.DataFrame(results, columns=columns)

            poll_billing_list = df.get(["poll_hourly"]).values.tolist()
            list_billing = str(poll_billing_list[0]).strip("[]'").split(",")
        
            poll_billing_enable_list = df.get(["poll_hourly_enable"]).values.tolist()
            list_enable_billing = str(poll_billing_enable_list[0]).strip("[]'").split(",")
            
            return render_template(
                "polling_hourly.html",
                tables=[df.to_html(classes="data", index=False)],
                titles=columns,
                selected_type=selected_type,
                type_options=type_options,
                list_billing=list_billing,
                list_enable_billing=list_enable_billing,username=username
            )
        else:
        # Render the HTML template without the table if no type is selected
            return render_template(
                "polling_hourly.html",
                tables=[],
                titles=[],
                selected_type=None,
                type_options=type_options,
                list_billing=[],
                list_enable_billing=[],username=username
            )
    
@app.route("/update_polling_hourly", methods=["POST"])
@login_required
def update_polling_hourly(): 
    if 'username' not in session:
        return redirect(url_for('login'))
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        selected_type = request.form.get("selected_type")
        print("tets",selected_type)
        type_id_query = f"SELECT ID FROM AMR_VC_TYPE WHERE VC_NAME LIKE '{selected_type}' ORDER BY VC_NAME"
        results = fetch_data(ptt_pivot_connection,type_id_query)
        print(results)
        type_id = str(results[0]).strip("',()")
        print(type_id)
        type_query = "SELECT VC_NAME FROM AMR_VC_TYPE ORDER BY VC_NAME"
        type_results = fetch_data(ptt_pivot_connection,type_query)
        type_options = [str(type[0]) for type in type_results]
        
        # Update billing data
        poll_billing_all = ""
        enable_billing = ""
        for i in range(0, 12):
            start_key = f"start{i + 1}"
            end_key = f"end{i + 1}"
            enable_key = f"enable[{i}]"
            
            start_value = request.form.get(start_key)
            end_value = request.form.get(end_key)
            enable_value = 1 if request.form.get(enable_key) == "on" else 0
            
            address_range = f"{start_value},{end_value}"
            if len(poll_billing_all + address_range) <= MAX_ADDRESS_LENGTH:
                if i > 0:
                    poll_billing_all += ","
                poll_billing_all += address_range

            if i == 0:
                enable_billing = str(enable_value)
            else:
                enable_billing += "," + str(enable_value)
        
        print("poll_billing:", poll_billing_all)

        update_query = f"""
        UPDATE amr_poll_range_hourly
        SET 
            poll_hourly = '{poll_billing_all}',
            poll_hourly_enable = '{enable_billing}'
        WHERE evc_type = '{type_id}'
        """
        print("Update Query:", update_query)
        
        update_sql(ptt_pivot_connection, update_query)
        # with connection.cursor() as cursor:
        
        #     cursor.execute(update_query)
        # connection.commit()
        # After updating the data, you may redirect to the polling route or perform any other necessary actions
        if execute_query(connect_to_ptt_pivot_db(), update_query):
                flash('polling_hourly successfully', 'success') 
                
                return render_template('polling_hourly.html',type_options=type_options)
        else:
                flash('Failed to polling', 'error')
    return redirect("/polling_hourly",type_options=type_options)

@app.route('/mapping_config')  
@login_required
def mapping_config_route():
    if 'username' not in session:
        return redirect(url_for('login'))
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        # SQL query to fetch options for the dropdown
        username = session['username']
        type_query = "SELECT VC_NAME FROM AMR_VC_TYPE ORDER BY VC_NAME"
        type_results = fetch_data(ptt_pivot_connection,type_query)
        type_options = [str(type[0]) for type in type_results]

        # SQL query to fetch data based on selected type
        base_query = """
        SELECT
            TO_CHAR(address),
            description,
            data_type,
            evc_type,
            or_der          
        FROM
            amr_mapping_config, amr_vc_type
        WHERE
            amr_mapping_config.evc_type = amr_vc_type.id
            AND amr_vc_type.VC_NAME LIKE '{selected_type}'
        ORDER BY or_der
        """
        
        selected_type = request.args.get("type_dropdown")
        selected_type = f"{selected_type}" if selected_type else ""

        if selected_type:
            query = base_query.format(selected_type=selected_type)
            results = fetch_data(ptt_pivot_connection,query)

            columns = [
                "address",
                "description",
                "data_type",
                "evc_type",
                "or_der",            
            ]
            df = pd.DataFrame(results, columns=columns)

            address_list = df.get(["address"]).values.tolist()
            list_address = str(address_list[0]).strip("[]'").split(",")
            print("map:", df)
                
            description_list = df.get(["description"]).values.tolist()
            list_description = str(description_list[0]).strip("[]'").split(",")
                
            data_type_list = df.get(["data_type"]).values.tolist()
            list_data_type = str(data_type_list[0]).strip("[]'").split(",")       
            
            evc_type_list = df.get(["evc_type"]).values.tolist()
            list_evc_type = str(evc_type_list[0]).strip("[]'").split(",")
            
            or_der_list = df.get(["or_der"]).values.tolist()
            list_or_der = str(or_der_list[0]).strip("[]'").split(",")

            return render_template(
                'mapping_config.html', 
                type_options=type_options, 
                selected_type=selected_type, 
                table=df.to_html(index=False),
                list_address=df["address"].tolist(),
                list_description=df["description"].tolist(),
                list_data_type=df["data_type"].tolist(),
                list_evc_type=df["evc_type"].tolist(),
                list_or_der=df["or_der"].tolist(),
                username=username          
            )

        return render_template('mapping_config.html', type_options=type_options,username=username)

def checkStrNone(stringcheck):
    if stringcheck == "None": return ""
    return stringcheck
    
@app.route('/update_mapping_config_route', methods=['POST'])
@login_required
def update_mapping_config():
    
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        
        if 'username' not in session:
            return redirect(url_for('login'))
    
        selected_type = request.form.get('selected_type')
        type_query = "SELECT VC_NAME FROM AMR_VC_TYPE ORDER BY VC_NAME"
        type_results = fetch_data(ptt_pivot_connection,type_query)
        type_options = [str(type[0]) for type in type_results]

        # Fetch type_id from the database
        type_id_query = f"SELECT ID FROM AMR_VC_TYPE WHERE VC_NAME LIKE '{selected_type}'"
        results = fetch_data(ptt_pivot_connection,type_id_query)
        type_id = str(results[0]).strip("',()")
        print("type:", type_id)
        
        query = """
            SELECT
                apr.evc_type,
                apr.poll_config
            FROM
                amr_poll_range apr
            JOIN
            amr_vc_type avt ON apr.evc_type = avt.id
            {type_condition}
            """
        type_condition = f"AND avt.VC_NAME = '{selected_type}'" if selected_type else ""
        
        # Check if a type is selected before executing the query
        if type_condition:
            query = query.format(type_condition=type_condition)
            results = fetch_data(ptt_pivot_connection,query)
            # print(results)

            columns = [
                "evc_type",
                "poll_config",
                
            ]
            df = pd.DataFrame(results, columns=columns)
            
            poll_config_list = df.get(["poll_config"]).values.tolist()
            print(poll_config_list)
            list_config = str(poll_config_list[0]).strip("[]'").split(",")
            print("start:", list_config)     

        description_VC_TYPE = [] 

        for j in range(0,QUANTITY_CONFIG_DATA):  
            i = f"{j:02d}"
            address_key = f"list_address{i}"
            description_key = f"list_description{i}"
            data_type_key = f"list_data_type{i}"
            evc_type_key = f"list_evc_type{i}"
            or_der_key = f"list_or_der{i}"

            address_value = checkStrNone(request.form.get(address_key))
            description_value = checkStrNone(request.form.get(description_key))
            data_type_value = checkStrNone(request.form.get(data_type_key))
            evc_type_value = request.form.get(evc_type_key)
            or_der_value = request.form.get(or_der_key)

            valid = False
            
            #check config
            print(address_value)
            # if textbox = None ไม่ต้อง check address แต่ต้องappend array ด้วย blank เพราะเดี๋ยวจะไม่ครบใน sql command
            if address_value == "": 
                description_VC_TYPE.append("")
                continue
            k = 0
            for k in range(0,QUANTITY_RANGE_CONFIG_LIST,2):
            
                address_value_int = int(address_value)
                address_check_low = int(list_config[k])
                address_check_high = int(list_config[k+1])
                k+=2
                print("low = ", address_check_low, ", high = ", address_check_high, "; value = ", address_value_int)
                if address_check_low  <= address_value_int :
                    if address_value_int <= address_check_high:
                        valid = True; print("Result = True"); break
                print("Result = False")
            
            if valid == False:
                
                flash(f"Error: error length {address_value}.", 'error')
                return render_template('mapping_config.html')

            description_VC_TYPE.append(checkStrNone(description_value))
            # print("---", description_value)


            # Update SQL query based on your table structure
            update_query = f"""
            UPDATE AMR_MAPPING_CONFIG
            SET
                ADDRESS = '{address_value}',
                DESCRIPTION = '{description_value}',
                DATA_TYPE = '{data_type_value}',
                OR_DER = '{or_der_value}'        
            WHERE evc_type = '{evc_type_value}' and or_der = '{or_der_value}'
            """
            update_sql(ptt_pivot_connection, update_query)

        update_vc_info_query = f"""
        UPDATE AMR_VC_CONFIGURED_INFO
        SET
            CONFIG1 = '{description_VC_TYPE[0]}',
            CONFIG2 = '{description_VC_TYPE[1]}',
            CONFIG3 = '{description_VC_TYPE[2]}',
            CONFIG4 = '{description_VC_TYPE[3]}',
            CONFIG5 = '{description_VC_TYPE[4]}',
            CONFIG6 = '{description_VC_TYPE[5]}',
            CONFIG7 = '{description_VC_TYPE[6]}',
            CONFIG8 = '{description_VC_TYPE[7]}',
            CONFIG9 = '{description_VC_TYPE[8]}',
            CONFIG10 = '{description_VC_TYPE[9]}',
            CONFIG11 = '{description_VC_TYPE[10]}',
            CONFIG12 = '{description_VC_TYPE[11]}',
            CONFIG13 = '{description_VC_TYPE[12]}',
            CONFIG14 = '{description_VC_TYPE[13]}',
            CONFIG15 = '{description_VC_TYPE[14]}',
            CONFIG16 = '{description_VC_TYPE[15]}',
            CONFIG17 = '{description_VC_TYPE[16]}',
            CONFIG18 = '{description_VC_TYPE[17]}',
            CONFIG19 = '{description_VC_TYPE[18]}',
            CONFIG20 = '{description_VC_TYPE[19]}'    
        WHERE 
            VC_TYPE = '{evc_type_value}'
            
        """
        
        update_sql(ptt_pivot_connection, update_vc_info_query)
        # print(update_vc_info_query)
        if execute_query(connect_to_ptt_pivot_db(), update_vc_info_query):
            flash('mapping_config successfully', 'success') 
            return render_template('mapping_config.html',type_options=type_options)
        else:
            flash('Failed to mapping_config', 'error')
            return render_template('mapping_config.html',type_options=type_options)
    return redirect("/mapping_config",type_options=type_options)

@app.route('/mapping_billing')  
@login_required
def mapping_billing_route():
    if 'username' not in session:
        return redirect(url_for('login'))
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        username = session['username']
        # SQL query to fetch options for the dropdown
        type_query = "SELECT VC_NAME FROM AMR_VC_TYPE ORDER BY VC_NAME"
        type_results = fetch_data(ptt_pivot_connection,type_query)
        type_options = [str(type[0]) for type in type_results]

        # SQL query to fetch data based on selected type
        base_query = """
        SELECT
            address,
            description,
            data_type,
            evc_type,
            or_der,
            daily
            
        FROM
            amr_mapping_billing, amr_vc_type
        WHERE
            amr_mapping_billing.evc_type = amr_vc_type.id
            AND amr_vc_type.VC_NAME LIKE :1
            AND daily = 1
           
        """
        selected_type = request.args.get("type_dropdown") or ""
        if selected_type:
            type_id_query = "SELECT ID FROM AMR_VC_TYPE WHERE VC_NAME LIKE :1"
            results = fetch_data(ptt_pivot_connection,type_id_query, (selected_type,))
            
            interval = ""

            
            if results:
                type_id = str(results[0][0])

                query = base_query
                results = fetch_data(ptt_pivot_connection,query, (selected_type,))
                
                print("type:", type_id)
                max_daily_query = "SELECT MAX(daily) FROM amr_mapping_billing WHERE evc_type LIKE :1"
                max_daily_result = fetch_data(ptt_pivot_connection,max_daily_query, (type_id,))
                max_daily_value = str(max_daily_result[0][0]) if max_daily_result and max_daily_result[0][0] else ""
                print("max_day:", max_daily_value)

                #if max_daily_value <=
                columns = [
                    "address",
                    "description",
                    "data_type",
                    "evc_type",
                    "or_der",
                    "daily",
                ]
                df = pd.DataFrame(results, columns=columns)
                # print("dd", df)
                
                # Extracting lists directly from DataFrame columns
                list_address = df["address"].tolist()
                list_description = df["description"].tolist()
                list_data_type = df["data_type"].tolist()
                list_evc_type = df["evc_type"].tolist()
                list_or_der = df["or_der"].tolist()
                list_daily = df["daily"].tolist()

                
                query_interval = f"""SELECT address 
                    FROM amr_mapping_billing 
                    WHERE evc_type = {type_id} AND or_der = 1 
                    
                    ORDER BY daily
                    FETCH FIRST 2 ROWS ONLY """
                result_interval = fetch_data(ptt_pivot_connection,query_interval)
                print("query_interval",result_interval)
                
                default_value = 0
                if len(result_interval) < 2:
                    # กำหนดค่าแทนในกรณีที่ข้อมูลไม่ครบ
                    value1 = value2 = default_value  # ใส่ค่าที่คุณต้องการแทนที่นี่
                else:
                    value1 = result_interval[0][0] if result_interval[0][0] is not None else default_value
                    value2 = result_interval[1][0] if result_interval[1][0] is not None else default_value

                # คำนวณค่า interval
                interval = value2 - value1

                return render_template(
                    'mapping_billing.html', 
                    type_options=type_options, 
                    selected_type=selected_type, 
                    max_daily_value=max_daily_value,
                    table=df.to_html(index=False),
                    list_address=list_address,
                    list_description=list_description,
                    list_data_type=list_data_type,
                    list_evc_type=list_evc_type,
                    list_or_der=list_or_der,
                    list_daily=list_daily,
                    interval=interval,
                    username=username
                    
                )

        return render_template('mapping_billing.html', type_options=type_options,username=username)

@app.route('/update_mapping_billing_route', methods=['POST'])
@login_required
def update_mapping_billing():
    if 'username' not in session:
        return redirect(url_for('login'))
    try:
        with connect_to_ptt_pivot_db() as ptt_pivot_connection:
            selected_type = request.form.get('selected_type')
            type_query = "SELECT VC_NAME FROM AMR_VC_TYPE ORDER BY VC_NAME"
            type_results = fetch_data(ptt_pivot_connection,type_query)
            type_options = [str(type[0]) for type in type_results]
            # Fetch type_id from the database
            type_id_query = f"SELECT ID FROM AMR_VC_TYPE WHERE VC_NAME = '{selected_type}'"
            results = fetch_data(ptt_pivot_connection, type_id_query)
            type_id = str(results[0][0])

            # Fetch the maximum daily value
            max_daily_query = f"SELECT MAX(daily) FROM amr_mapping_billing WHERE evc_type = {type_id}"
            max_daily_result = fetch_data(ptt_pivot_connection, max_daily_query)
            max_daily_query_value = int(max_daily_result[0][0])

            # fine num of data
            or_der_query = f"""SELECT or_der FROM amr_mapping_billing WHERE evc_type = {type_id} AND daily = 1 ORDER by or_der"""
            or_der_result = [row[0] for row in fetch_data(ptt_pivot_connection, or_der_query)]

            print(len(or_der_result),"or_der_result")
            dataframes = {
                'address': [],
                'description': [],
                'data_type': [],
                'evc_type': [],
                'or_der': [],
                'daily': [],
            }
            df_data = pd.DataFrame(dataframes)

            max_daily_new = int(request.form.get('max_day'))

            address_value_array = []
            description_value_array = []
            data_type_value_array = []
            or_der_value_array = []

            for i in range(0, len(or_der_result)): #QUANTITY_BILLING_PER_DAY = 5
                i_str = f"{i:02d}"
                address_value = request.form.get(f"list_address{i_str}")
                description_value = request.form.get(f"list_description{i_str}")
                data_type_value = request.form.get(f"list_data_type{i_str}")
                or_der_value = request.form.get(f"list_or_der{i_str}")
                #print("or_der_value",data_type_value)
                address_value_array.append(int(address_value))
                description_value_array.append(description_value)
                data_type_value_array.append(data_type_value)
                or_der_value_array.append(or_der_value)

            interval = int(request.form.get('interval'))
            
            if max_daily_new <= max_daily_query_value:
                if max_daily_new < max_daily_query_value:
                    delete_query = f"""
                    DELETE FROM AMR_MAPPING_BILLING
                    WHERE evc_type = '{type_id}' AND DAILY > {max_daily_new}
                    """
                    update_sql(ptt_pivot_connection, delete_query)
                    print(delete_query,"delete_query")
                for j in range(max_daily_new):
                    for i in range(len(or_der_result)):
                        update_query = f"""
                        UPDATE AMR_MAPPING_BILLING
                        SET 
                            ADDRESS = '{int(address_value_array[i]) + (j * interval)}',
                            DESCRIPTION = '{description_value_array[i]}',
                            data_type = '{data_type_value_array[i]}'
                        WHERE 
                            evc_type = '{type_id}' and 
                            or_der = {i + 1} and 
                            daily = {j + 1}
                        """
                        update_sql(ptt_pivot_connection, update_query)
            else:
                for j in range(max_daily_query_value):
                    for i in range(len(or_der_result)):
                        update_billing = f"""
                        UPDATE AMR_MAPPING_BILLING
                        SET 
                            ADDRESS = '{int(address_value_array[i]) + (j * interval)}',
                            DESCRIPTION = '{description_value_array[i]}',
                            data_type = '{data_type_value_array[i]}'
                        WHERE 
                            evc_type = '{type_id}' and 
                            or_der = {i + 1} and 
                            daily = {j + 1}
                        """
                        update_sql(ptt_pivot_connection, update_billing)

                for j in range(max_daily_new - max_daily_query_value):
                    for i in range(len(or_der_result)):
                        new_address = int(address_value_array[i]) + ((max_daily_query_value + j) * interval)
                        new_description = description_value_array[i]
                        new_data_type = data_type_value_array[i]
                        new_evc_type = type_id
                        new_or_der = or_der_result[i]
                        
                        new_daily = max_daily_query_value + j + 1

                        insert_query = f"""
                        INSERT INTO AMR_MAPPING_BILLING (ADDRESS, DESCRIPTION, DATA_TYPE, evc_type, OR_DER, DAILY)
                        VALUES ('{new_address}', '{new_description}', '{new_data_type}', '{new_evc_type}', '{new_or_der}', '{new_daily}')
                        """
                        # print(insert_query)
                        update_sql(ptt_pivot_connection, insert_query)

            flash('Mapping billing successfully updated', 'success')
            return render_template('mapping_billing.html',type_options=type_options)

    except Exception as e:
        flash(f'Failed to update mapping billing: {str(e)}', 'error')
        return render_template('mapping_billing.html',type_options=type_options)

@app.route('/mapping_hourly')  
@login_required
def mapping_hourly_route():
    if 'username' not in session:
        return redirect(url_for('login'))
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        # SQL query to fetch options for the dropdown
        username = session['username']
        type_query = "SELECT VC_NAME FROM AMR_VC_TYPE  ORDER BY VC_NAME"
        type_results = fetch_data(ptt_pivot_connection,type_query)
        type_options = [str(type[0]) for type in type_results]

        # SQL query to fetch data based on selected type
        base_query = """
        SELECT
            address,
            description,
            data_type,
            evc_type,
            or_der,
            hourly
            
        FROM
            amr_mapping_hourly, amr_vc_type
        WHERE
            amr_mapping_hourly.evc_type = amr_vc_type.id
            AND amr_vc_type.VC_NAME LIKE :1
            AND hourly = 1
        ORDER BY
            or_der
        """

        selected_type = request.args.get("type_dropdown") or ""
        if selected_type:
            type_id_query = "SELECT ID FROM AMR_VC_TYPE WHERE VC_NAME LIKE :1"
            results = fetch_data(ptt_pivot_connection,type_id_query, (selected_type,))
            
            interval = ""
    
            if results:
                type_id = str(results[0][0])

                query = base_query
                results = fetch_data(ptt_pivot_connection,query, (selected_type,))
                
                #print("type:", type_id)
                max_daily_query = "SELECT MAX(hourly) FROM amr_mapping_hourly WHERE evc_type LIKE :1 "
                max_daily_result = fetch_data(ptt_pivot_connection,max_daily_query, (type_id,))
                max_daily_value = str(max_daily_result[0][0]) if max_daily_result and max_daily_result[0][0] else ""
                #print("max_day:", max_daily_value)

                #if max_daily_value <=
                columns = [
                    "address",
                    "description",
                    "data_type",
                    "evc_type",
                    "or_der",
                    "hourly",
                ]
                df = pd.DataFrame(results, columns=columns)
                print("dd", df)
                
                # Extracting lists directly from DataFrame columns
                list_address = df["address"].tolist()
                list_description = df["description"].tolist()
                list_data_type = df["data_type"].tolist()
                list_evc_type = df["evc_type"].tolist()
                list_or_der = df["or_der"].tolist()
                list_daily = df["hourly"].tolist()

                query_interval = f"""SELECT address 
                    FROM amr_mapping_hourly 
                    WHERE evc_type = {type_id} 
                    AND or_der = 1 
                    
                    ORDER BY hourly
                    FETCH FIRST 2 ROWS ONLY """

                result_interval = fetch_data(ptt_pivot_connection,query_interval)
                print("query_interval",result_interval)
                default_value = 0
                if len(result_interval) < 2:
                    # กำหนดค่าแทนในกรณีที่ข้อมูลไม่ครบ
                    value1 = value2 = default_value  # ใส่ค่าที่คุณต้องการแทนที่นี่
                else:
                    value1 = result_interval[0][0] if result_interval[0][0] is not None else default_value
                    value2 = result_interval[1][0] if result_interval[1][0] is not None else default_value

                # คำนวณค่า interval
                interval = value2 - value1

                return render_template(
                    'mapping_hourly.html', 
                    type_options=type_options, 
                    selected_type=selected_type, 
                    max_daily_value=max_daily_value,
                    table=df.to_html(index=False),
                    list_address=list_address,
                    list_description=list_description,
                    list_data_type=list_data_type,
                    list_evc_type=list_evc_type,
                    list_or_der=list_or_der,
                    list_daily=list_daily,
                    interval=interval,username=username
                )

        return render_template('mapping_hourly.html', type_options=type_options,username=username)
 
@app.route('/update_mapping_hourly_route', methods=['POST'])
@login_required
def update_mapping_hourly():
    if 'username' not in session:
        return redirect(url_for('login'))
    try:
        with connect_to_ptt_pivot_db() as ptt_pivot_connection:
            selected_type = request.form.get('selected_type')
            type_query = "SELECT VC_NAME FROM AMR_VC_TYPE ORDER BY VC_NAME"
            type_results = fetch_data(ptt_pivot_connection,type_query)
            type_options = [str(type[0]) for type in type_results]
            # Fetch type_id from the database
            type_id_query = f"SELECT ID FROM AMR_VC_TYPE WHERE VC_NAME = '{selected_type}'"
            results = fetch_data(ptt_pivot_connection, type_id_query)
            type_id = str(results[0][0])

            # Fetch the maximum daily value
            max_daily_query = f"SELECT MAX(hourly) FROM amr_mapping_hourly WHERE evc_type = {type_id}"
            max_daily_result = fetch_data(ptt_pivot_connection, max_daily_query)
            max_daily_query_value = int(max_daily_result[0][0])

            dataframes = {
                'address': [],
                'description': [],
                'data_type': [],
                'evc_type': [],
                'or_der': [],
                'hourly': [],
            }
            df_data = pd.DataFrame(dataframes)

            max_daily_new = int(request.form.get('max_day'))

            address_value_array = []
            description_value_array = []
            data_type_value_array = []
            or_der_value_array = []

            for i in range(0, QUANTITY_BILLING_PER_DAY): #QUANTITY_BILLING_PER_DAY = 5
                i_str = f"{i:02d}"
                address_value = request.form.get(f"list_address{i_str}")
                description_value = request.form.get(f"list_description{i_str}")
                data_type_value = request.form.get(f"list_data_type{i_str}")
                or_der_value = request.form.get(f"list_or_der{i_str}")

                address_value_array.append(int(address_value))
                description_value_array.append(description_value)
                data_type_value_array.append(data_type_value)
                or_der_value_array.append(or_der_value)

            interval = int(request.form.get('interval'))

            if max_daily_new <= max_daily_query_value:
                if max_daily_new < max_daily_query_value:
                    delete_query = f"""
                    DELETE FROM AMR_MAPPING_hourly
                    WHERE evc_type = '{type_id}' AND hourly > {max_daily_new}
                    """
                    update_sql(ptt_pivot_connection, delete_query)

                for j in range(max_daily_new):
                    for i in range(QUANTITY_BILLING_PER_DAY):
                        update_query = f"""
                        UPDATE AMR_MAPPING_hourly
                        SET 
                            ADDRESS = '{int(address_value_array[i]) + (j * interval)}',
                            DESCRIPTION = '{description_value_array[i]}',
                            data_type = '{data_type_value_array[i]}'
                        WHERE 
                            evc_type = '{type_id}' and 
                            or_der = {i + 1} and 
                            hourly = {j + 1}
                        """
                        update_sql(ptt_pivot_connection, update_query)
            else:
                for j in range(max_daily_query_value):
                    for i in range(QUANTITY_BILLING_PER_DAY):
                        update_billing = f"""
                        UPDATE AMR_MAPPING_hourly
                        SET 
                            ADDRESS = '{int(address_value_array[i]) + (j * interval)}',
                            DESCRIPTION = '{description_value_array[i]}',
                            data_type = '{data_type_value_array[i]}'
                        WHERE 
                            evc_type = '{type_id}' and 
                            or_der = {i + 1} and 
                            hourly = {j + 1}
                        """
                        update_sql(ptt_pivot_connection, update_billing)

                for j in range(max_daily_new - max_daily_query_value):
                    for i in range(QUANTITY_BILLING_PER_DAY):
                        new_address = int(address_value_array[i]) + ((max_daily_query_value + j) * interval)
                        new_description = description_value_array[i]
                        new_data_type = data_type_value_array[i]
                        new_evc_type = type_id
                        new_or_der = or_der_value_array[i]
                        new_daily = max_daily_query_value + j + 1

                        insert_query = f"""
                        INSERT INTO AMR_MAPPING_hourly (ADDRESS, DESCRIPTION, DATA_TYPE, evc_type, OR_DER, hourly)
                        VALUES ('{new_address}', '{new_description}', '{new_data_type}', '{new_evc_type}', '{new_or_der}', '{new_daily}')
                        """
                        update_sql(ptt_pivot_connection, insert_query)

            flash('Mapping billing successfully updated', 'success')
            return render_template('mapping_hourly.html',type_options=type_options)

    except Exception as e:
        flash(f'Failed to update mapping billing: {str(e)}', 'error')
        return render_template('mapping_hourly.html',type_options=type_options)

@app.route("/add_mapping_route")
@login_required
def add_mapping_route():
    return render_template("add_mapping.html")


@app.route("/submit_form", methods=["POST"])
@login_required
def submit_form():
    cursor = None
    connection = None
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
    try:
        data_list = []
        for i in range(1, 21):
            address = request.form[f"address{i}"]
            description = request.form[f"description{i}"]
            type_value = request.form.get(f"type_value{i}")
            evc_type = request.form[f"evc_type{i}"]
            or_der = request.form[f"or_der{i}"]
            data_type = request.form[f"data_type{i}"]

            data_list.append(
                (address, description, type_value, evc_type, or_der, data_type)
            )

        dsn_tns = cx_Oracle.makedsn(host, port, service_name=service)
        connection = cx_Oracle.connect(user=username, password=password, dsn=dsn_tns)

        cursor = connection.cursor()

        sql_merge = """
            MERGE INTO AMR_MAPPING_CONFIG dst
            USING (
                SELECT
                    :address as address,
                    :description as description,
                    :type_value as type_value,
                    :evc_type as evc_type,
                    :or_der as or_der,
                    :data_type as data_type
                FROM dual
            ) src
            ON (dst.address = src.address)
            WHEN MATCHED THEN
                UPDATE SET
                    dst.description = src.description,
                    dst.type_value = src.type_value,
                    dst.evc_type = src.evc_type,
                    dst.or_der = src.or_der,
                    dst.data_type = src.data_type
            WHEN NOT MATCHED THEN
                INSERT (
                    address,
                    description,
                    type_value,
                    evc_type,
                    or_der,
                    data_type
                ) VALUES (
                    src.address,
                    src.description,
                    src.type_value,
                    src.evc_type,
                    src.or_der,
                    src.data_type
                )
        """

        cursor.executemany(sql_merge, data_list)

        # Commit the changes to the database
        connection.commit()

        return "Data saved successfully"
    except Exception as e:
        return f"Error occurred: {str(e)}"
    finally:
        if cursor is not None:
        # Close the cursor
            cursor.close()

    if connection is not None:
        # Close the connection
        connection.close()


@app.route("/add_actraris_route")
@login_required
def add_actraris_route():
    return render_template("add_actraris.html")


@app.route("/new_form", methods=["POST"])
def submit_new_form():
    cursor = None
    connection = None
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
    try:
        data_list = []
        for i in range(1, 18):
            address = request.form.get(f"address{i}")
            description = request.form.get(f"description{i}")
            type_value = request.form.get(f"type_value{i}")
            evc_type = request.form.get(f"evc_type{i}")
            or_der = request.form.get(f"or_der{i}")
            data_type = request.form.get(f"data_type{i}")

            data_list.append(
                (address, description, type_value, evc_type, or_der, data_type)
            )

        dsn_tns = cx_Oracle.makedsn(host, port, service_name=service)
        connection = cx_Oracle.connect(user=username, password=password, dsn=dsn_tns)

        cursor = connection.cursor()

        sql_merge = """
            MERGE INTO AMR_ADDRESS_MAPPING1 dst
            USING (
                SELECT
                    :address as address,
                    :description as description,
                    :type_value as type_value,
                    :evc_type as evc_type,
                    :or_der as or_der,
                    :data_type as data_type
                FROM dual
            ) src
            ON (dst.address = src.address)
            WHEN MATCHED THEN
                UPDATE SET
                    dst.description = src.description,
                    dst.type_value = src.type_value,
                    dst.evc_type = src.evc_type,
                    dst.or_der = src.or_der,
                    dst.data_type = src.data_type
            WHEN NOT MATCHED THEN
                INSERT (
                    address,
                    description,
                    type_value,
                    evc_type,
                    or_der,
                    data_type
                ) VALUES (
                    src.address,
                    src.description,
                    src.type_value,
                    src.evc_type,
                    src.or_der,
                    src.data_type
                )
        """

        cursor.executemany(sql_merge, data_list)

        connection.commit()

        # Commit the changes to the database
        connection.commit()

        return "Data saved successfully"
    except Exception as e:
        return f"Error occurred: {str(e)}"
    finally:
        if cursor is not None:
        # Close the cursor
            cursor.close()

    if connection is not None:
        # Close the connection
        connection.close()


@app.route('/add_site', methods=['GET', 'POST'])
@login_required
def add_site():

    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        username=session['username']
        
        query_all_type = "SELECT id,vc_name from amr_vc_type"
        all_type_result = fetch_data(ptt_pivot_connection,query_all_type)
        user_update = session['username']
        max_id_query = "SELECT MAX(ID) + 1 FROM amr_field_id"
        max_id_result = fetch_data(ptt_pivot_connection,max_id_query)
        max_id_value = str(max_id_result[0][0]) if max_id_result and max_id_result[0][0] is not None else ""
        print("max_id_value", max_id_result)
    
        AMR0 = "AMR0"
        CUST0 = "CUST0"
        MET0 = "MET0"
        PROT0 = "PROT0"
    
        sql_commands = []
            
        if request.method == 'POST':
            phase = request.form['phase']
            site_name = request.form['site_name']
            factory_name = request.form['factory_name']
            region = request.form['region']
            rmiu_type = request.form['rmiu_type']
            power_indicator = request.form['power_indicator']
            modbus_id = request.form['modbus_id']
            ip_address = request.form['ip_address']
            ready_to_billing = request.form.get('ready_to_billing')  # Check if the checkbox is checked
            auto_ping = request.form.get('auto_ping')  # Check if the checkbox is checked
            billing_date = request.form['billing_date']
            show_sg_co2_n2 = request.form['show_sg_co2_n2']
            amount_of_meter = request.form['amount_of_meter']
            initial_username = request.form['initial_username']
            
            port_1 = request.form['port_1']
            port1 = request.form['port1']
            print(port1)
            auto_1 = request.form['auto_1']
            
            port_2 = request.form['port_2']
            port2 = request.form['port2']
            print(port2)
            auto_2 = request.form['auto_2']
            
            port_3 = request.form['port_3']
            port3 = request.form['port3']
            auto_3 = request.form['auto_3']
            
            port_4 = request.form['port_4']
            port4 = request.form['port4']
            auto_4 = request.form['auto_4']
            
            port_5 = request.form['port_5']
            port5 = request.form['port5']
            auto_5 = request.form['auto_5']
            
            port_6 = request.form['port_6']
            port6 = request.form['port6']
            auto_6 = request.form['auto_6']
            current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))).strftime('%d-%b-%y %I.%M.%S.%f %p ASIA/BANGKOK').upper()
            initial_password = request.form['initial_password']
            hashed_password = md5_hash(initial_password)
            amr_user = f"""
            INSERT INTO AMR_USER (ID, DESCRIPTION, USER_NAME, PASSWORD, USER_LEVEL, USER_GROUP, USER_ENABLE,TIME_CREATE,UPDATED_BY)
            VALUES ({max_id_value}, '{site_name}', '{initial_username}', '{hashed_password}', '3', '{MET0+max_id_value}', '1','{current_datetime}','{user_update}')
            """
            update_sql(ptt_pivot_connection, amr_user)
            
            amr_field_id = f"""
            INSERT INTO AMR_FIELD_ID (
                                    ID, 
                                    TAG_ID, 
                                    FIELD_ID, 
                                    CUST_ID, 
                                    METER_ID, 
                                    PROTOCOL_ID, 
                                    RTU_MODBUS_ID,
                                    RMIU_AUTO_ENABLE,
                                    PING_ENABLE,
                                    RMIU_TYPE, 
                                    SIM_IP,
                                    RMIU_POLL_REPEAT1,
                                    RMIU_POLL_REPEAT2,
                                    AMR_PHASE,
                                    TIME_CREATE,
                                    UPDATED_BY
                                    )
            VALUES ({max_id_value}, '{site_name}', 
                                    '{AMR0+max_id_value}', 
                                    '{CUST0+max_id_value}', 
                                    '{MET0+max_id_value}', 
                                    '{PROT0+max_id_value}', 
                                    '{modbus_id}',
                                    '{ready_to_billing}',
                                    '{auto_ping}',
                                    '{rmiu_type}', 
                                    '{ip_address}',
                                    '{0}',
                                    '{0}',
                                    '{phase}',
                                    '{current_datetime}',
                                    '{user_update}'
                                    )
            """
            update_sql(ptt_pivot_connection, amr_field_id)
            
            amr_field_customer= f"""
            INSERT INTO AMR_FIELD_CUSTOMER (
                                            ID, 
                                            CUST_ID,
                                            CUST_NAME,
                                            CUST_FACTORY_NAME,
                                            METER_RUN,
                                            TIME_CREATE,
                                            UPDATED_BY
                                            )
            VALUES ({max_id_value}, '{CUST0+max_id_value}',
                                    '{site_name}',
                                    '{factory_name}',
                                    '{amount_of_meter}',
                                    '{current_datetime}',
                                    '{user_update}'            
                                    )
            """
            update_sql(ptt_pivot_connection, amr_field_customer)
            
            amr_pl_group = f""" 
            INSERT INTO AMR_PL_GROUP (ID, PL_REGION_ID, FIELD_ID,TIME_CREATE,UPDATED_BY) 
            VALUES ({max_id_value}, '{region}', '{AMR0+max_id_value}','{current_datetime}','{user_update}')
            """
            update_sql(ptt_pivot_connection, amr_pl_group)
            
            for i in range(1, int(amount_of_meter) + 1):
                modbus_id = '2' if request.form['port' + str(i)] == '16' else '1'
                
                amr_field_meter = f"""
                INSERT INTO AMR_FIELD_METER (METER_ID, 
                                            METER_STREAM_NO, 
                                            METER_NO_STREAM,
                                            METER_STREAM_TYPE,
                                            METER_PORT_NO,
                                            METER_AUTO_ENABLE,
                                            METER_POLL_REPEAT1,
                                            METER_POLL_REPEAT2,
                                            METER_POLL_REPEAT3,
                                            METER_POLL_REPEAT4,
                                            METER_POLL_REPEAT5,
                                            METER_POLL_REPEAT6,
                                            METER_POLL_REPEAT7,
                                            MODBUS_ID,
                                            TIME_CREATE,
                                            UPDATED_BY
                                            )
                VALUES ('{MET0+max_id_value}', '{i}',
                                            '{amount_of_meter}',
                                            '{request.form['port_' + str(i)]}',
                                            '{request.form['port' + str(i)]}',
                                            '{request.form['auto_' + str(i)]}',
                                            '{0}',
                                            '{0}',
                                            '{0}',
                                            '{0}',
                                            '{0}',
                                            '{0}',
                                            '{0}',
                                            '{modbus_id}',
                                            '{current_datetime}',
                                            '{user_update}'
                )         
                """
                update_sql(ptt_pivot_connection, amr_field_meter)
                
                amr_field_profile = f"""
                INSERT INTO AMR_FIELD_PROFILE (METER_ID, 
                                                METER_STREAM_NO,
                                                WRITE_CONFIG_ENABLE,
                                                WRITE_CONFIG_REPEAT1,
                                                WRITE_CONFIG_REPEAT2,
                                                TIME_CREATE,
                                                UPDATED_BY
                                                )
                VALUES ('{MET0+max_id_value}', '{i}',
                                                '{auto_ping}',
                                                '{0}',
                                                '{0}',
                                                '{current_datetime}',
                                                '{user_update}'
                                                )
                """
                update_sql(ptt_pivot_connection, amr_field_profile)
                
                amr_field_protocol = f"""
                INSERT INTO AMR_FIELD_PROTOCOL (PROTOCOL_ID, PROTOCOL_STREAM_NO, PROTOCOL_NO_STREAM, TIME_CREATE,UPDATED_BY)
                VALUES ('{PROT0+max_id_value}', '{i}', '{amount_of_meter}','{current_datetime}','{user_update}')
                """
                # print("amr_field_protocol", amr_field_protocol)
                update_sql(ptt_pivot_connection, amr_field_protocol)
                
            if execute_query(connect_to_ptt_pivot_db(), amr_field_protocol):
                flash('AddSite successfully', 'success') 
              
                return render_template('add_site.html',max_id_value=max_id_value,all_type_result=all_type_result,username=username)
            else:
                flash('Failed to AddSite', 'error')  
    return render_template('add_site.html', max_id_value=max_id_value,all_type_result=all_type_result,username=username)


@app.route('/get_tags_by_region', methods=['GET'])
@login_required
def get_tags_by_region():
    selected_region = request.args.get("region")
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        tag_query = """
       SELECT DISTINCT AMR_FIELD_ID.TAG_ID
        FROM AMR_FIELD_ID,amr_region,AMR_PL_GROUP
        WHERE AMR_PL_GROUP.PL_REGION_ID = amr_region.id
        AND AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
        AND amr_region.REGION_NAME = :region_id
        AND AMR_FIELD_ID.tag_id NOT like '%.remove%'
        ORDER BY  TAG_ID
        """
        tag_results = fetch_data(ptt_pivot_connection, tag_query, params={"region_id": selected_region})
        tags = [str(tag[0]) for tag in tag_results]
    return jsonify(tags=tags)

@app.route('/edit_site', methods=['GET', 'POST'])
@login_required
def edit_site():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:

        query_all_type = "SELECT id,vc_name from amr_vc_type"
        all_type_result = fetch_data(ptt_pivot_connection,query_all_type)
        user_update = session['username']

        region_query = "SELECT * FROM AMR_REGION"
        selected_tag = request.args.get("tag_dropdown")
        selected_region = request.args.get("region_dropdown")

        # Fetch unique region values
        region_results = fetch_data(ptt_pivot_connection, region_query)
        region_options = [str(region[1]) for region in region_results]

        # Fetch tag options based on the selected region
       
        tag_query = """
        SELECT DISTINCT AMR_FIELD_ID.TAG_ID
        FROM AMR_FIELD_ID,amr_region,AMR_PL_GROUP
        
        WHERE AMR_PL_GROUP.PL_REGION_ID = amr_region.id
        AND AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
        AND amr_region.REGION_NAME = :region_id
        AND AMR_FIELD_ID.tag_id NOT like '%.remove%'
        ORDER BY  TAG_ID

        """
        tag_results = fetch_data(ptt_pivot_connection, tag_query, params={"region_id": selected_region})
        tag_options = [str(tag[0]) for tag in tag_results]
        print(tag_options)
        show_tag_query = """
        SELECT DISTINCT
            AMR_FIELD_ID.ID,
            AMR_FIELD_ID.TAG_ID,
            AMR_FIELD_ID.SIM_IP,
            AMR_FIELD_ID.RTU_MODBUS_ID,
            AMR_FIELD_ID.AMR_PHASE,
            AMR_FIELD_CUSTOMER.CUST_FACTORY_NAME,
            AMR_USER.USER_NAME,
            AMR_USER.PASSWORD,
            AMR_FIELD_METER.METER_ID,
            AMR_FIELD_METER.METER_STREAM_NO,
            AMR_FIELD_METER.METER_NO_STREAM,
            AMR_FIELD_METER.METER_STREAM_TYPE,
            AMR_FIELD_METER.METER_PORT_NO,
            AMR_FIELD_METER.METER_AUTO_ENABLE,
            AMR_FIELD_PROFILE.METER_ID,
            AMR_FIELD_PROTOCOL.PROTOCOL_ID,
            AMR_FIELD_PROTOCOL.PROTOCOL_NO_STREAM
        FROM 
            AMR_FIELD_ID
            JOIN AMR_FIELD_CUSTOMER ON AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID
            JOIN AMR_USER ON AMR_FIELD_ID.ID = AMR_USER.ID
            JOIN AMR_FIELD_METER ON AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID
            JOIN AMR_FIELD_PROFILE ON AMR_FIELD_ID.METER_ID = AMR_FIELD_PROFILE.METER_ID
            JOIN AMR_FIELD_PROTOCOL ON AMR_FIELD_ID.PROTOCOL_ID = AMR_FIELD_PROTOCOL.PROTOCOL_ID
            JOIN AMR_PL_GROUP ON AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID 
            JOIN AMR_REGION ON AMR_PL_GROUP.PL_REGION_ID = AMR_REGION.ID
        WHERE 
            AMR_REGION.REGION_NAME = :region_id AND
            AMR_FIELD_ID.TAG_ID = :tag_id
        ORDER BY METER_STREAM_NO
        """
        
        data = []
        if selected_region is not None and selected_tag is not None:
            data = fetch_data(ptt_pivot_connection, show_tag_query, params={"region_id": selected_region, "tag_id": selected_tag})

        columns = [
            'id', 'tag_id', 'sim_ip', 'rtu_modbus_id', 'amr_phase',
            'cust_factory_name', 'user_name', 'password', 'meter_id',
            'meter_stream_no', 'meter_no_stream', 'meter_stream_type',
            'meter_port_no', 'meter_auto_enable', 'meter_id',
            'protocol_id', 'protocol_no_stream'
        ]
        df = pd.DataFrame(data, columns=columns)
        
        if not df.empty:
            list_id = df["id"].iloc[0]
            list_tag_id = df["tag_id"].iloc[0]
            list_cust_factory_name = df["cust_factory_name"].iloc[0]
            list_amr_phase = df["amr_phase"].iloc[0]
            list_rtu_modbus_id = df["rtu_modbus_id"].iloc[0]
            list_sim_ip = df["sim_ip"].iloc[0]
            list_user_name = df["user_name"].iloc[0]
            list_password = df["password"].iloc[0]
            list_meter_stream_type = df["meter_stream_type"].tolist()
            list_meter_stream_no = df["meter_stream_no"].tolist()
            list_meter_port_no = df["meter_port_no"].tolist()
            list_meter_auto_enable = df["meter_auto_enable"].tolist()

            list_meter_stream_type = []
            for meter_stream_type in df["meter_stream_type"]:
                meter_stream_type_list = f"""SELECT vc_name FROM amr_vc_type WHERE id = '{meter_stream_type}' ORDER BY id"""
                type = fetch_data(ptt_pivot_connection, meter_stream_type_list)
                list_meter_stream_type.append(type[0][0] if type else None)

            list_meter_port_no = []
            for meter_port_no in df["meter_port_no"]:
                meter_port_no_list =f"""SELECT port_no FROM amr_port_info WHERE id = '{meter_port_no}' ORDER BY id"""
                port = fetch_data(ptt_pivot_connection, meter_port_no_list)
                list_meter_port_no.append(port[0][0] if port else None)

        else:
            list_id = None
            list_tag_id = None
            list_cust_factory_name = None
            list_amr_phase = None
            list_rtu_modbus_id = None
            list_sim_ip = None
            list_user_name = None
            list_password = None
            list_meter_stream_no = None
            list_meter_auto_enable = None
            list_meter_stream_type = None
            list_meter_port_no = None

        html_table = df.to_html(index=False) if not df.empty else ""

        return render_template('edit_site.html', 
                                region_options=region_options, 
                                tag_options=tag_options, 
                                selected_region=selected_region,
                                selected_tag=selected_tag,
                                list_id=list_id,
                                list_tag_id=list_tag_id,
                                list_cust_factory_name=list_cust_factory_name,
                                list_amr_phase=list_amr_phase,
                                list_rtu_modbus_id=list_rtu_modbus_id,
                                list_sim_ip=list_sim_ip,
                                list_user_name=list_user_name,
                                list_password=list_password,
                                list_meter_stream_no=list_meter_stream_no,
                                list_meter_auto_enable=list_meter_auto_enable,
                                list_meter_stream_type=list_meter_stream_type,
                                list_meter_port_no=list_meter_port_no,
                                html_table=html_table, all_type_result=all_type_result,user_update=user_update)
        
        
@app.route('/update_edit_site_route', methods=['POST'])
@login_required
def update_edit_site():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        region_query = "SELECT * FROM AMR_REGION"
       
        # Fetch unique region values
        region_results = fetch_data(ptt_pivot_connection, region_query)
        region_options = [str(region[1]) for region in region_results]
        MET = 'MET0'
        user_update = session['username']
        id = request.form.get("id")
        tag_id = request.form.get("tag_id")
        cust_factory_name = request.form.get("cust_factory_name")
        amr_phase = request.form.get("amr_phase")
        rtu_modbus_id = int(request.form.get("rtu_modbus_id"))
        sim_ip = request.form.get("sim_ip")
        user_name = request.form.get("user_name")
        password = request.form.get("password")
        
        current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))).strftime('%d-%b-%y %I.%M.%S.%f %p ASIA/BANGKOK').upper()
      
        if password:
            print("password1",password)
            hashed_password = md5_hash(password)
            update_user = f"""
                UPDATE AMR_USER 
                SET 
                    description = '{tag_id}',
                    user_name = '{user_name}',
                    password = '{hashed_password}',
                    UPDATED_BY = '{user_update}',
                    TIME_CREATE = '{current_datetime}'
                WHERE id = '{id}'
            """
            # print("update_user:", update_user)
            update_sql(ptt_pivot_connection, update_user)
        else: 
            print("password2",password)
            update_user = f"""
                UPDATE AMR_USER 
                SET 
                    description = '{tag_id}',
                    user_name = '{user_name}',
                    UPDATED_BY = '{user_update}',
                    TIME_CREATE = '{current_datetime}'
                WHERE id = '{id}'
            """
            update_sql(ptt_pivot_connection, update_user)
            
            
        update_field_id = f"""
            UPDATE AMR_FIELD_ID
            SET
                tag_id = '{tag_id}',
                sim_ip = '{sim_ip}',
                amr_phase = '{amr_phase}',
                rtu_modbus_id = '{rtu_modbus_id}',
                UPDATED_BY = '{user_update}',
                TIME_CREATE = '{current_datetime}'
            WHERE id = '{id}'
        """
        # print("update_field_id", update_field_id)
        update_sql(ptt_pivot_connection, update_field_id)
        
        update_field_customer = f"""
            UPDATE AMR_FIELD_CUSTOMER
            SET
                cust_factory_name = '{cust_factory_name}',
                UPDATED_BY = '{user_update}',
                TIME_CREATE = '{current_datetime}'
            WHERE id = '{id}'
        """
        update_sql(ptt_pivot_connection, update_field_customer)
        
        meter_stream_type = request.form.getlist("list_meter_stream_type")      
        meter_stream_no = request.form.get("meter_stream_no")
        meter_port_no = request.form.getlist("list_meter_port_no")
        
        for i, (stream_type, port_no) in enumerate(zip(meter_stream_type, meter_port_no)):
            stream_no = meter_stream_no[i % len(meter_stream_no)]  
            update_field_meter = f"""
                UPDATE AMR_FIELD_METER
                SET
                    meter_stream_type = '{stream_type}',
                    meter_port_no = '{port_no}',
                    UPDATED_BY = '{user_update}',
                    TIME_CREATE = '{current_datetime}'
                WHERE meter_id = '{MET+id}'
                AND meter_stream_no = '{i + 1}'
            """
            # print("update_field_meter", update_field_meter)
            update_sql(ptt_pivot_connection, update_field_meter)
        if execute_query(connect_to_ptt_pivot_db(), update_field_meter):
                flash('edit_site successfully', 'success') 
                
                return render_template('edit_site.html',region_options=region_options)
        else:
                flash('Failed to edit_site', 'error')
    return redirect(url_for('edit_site'))

@app.route('/remove_site', methods=['GET', 'POST'])
@login_required
def remove_site():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:

        query_all_type = "SELECT id,vc_name from amr_vc_type"
        all_type_result = fetch_data(ptt_pivot_connection,query_all_type)


        region_query = "SELECT * FROM AMR_REGION"
        selected_tag = request.args.get("tag_dropdown")
        selected_region = request.args.get("region_dropdown")

        # Fetch unique region values
        region_results = fetch_data(ptt_pivot_connection, region_query)
        region_options = [str(region[1]) for region in region_results]
        username = session['username']
        # Fetch tag options based on the selected region
        tag_options = []
        if selected_region:
            tag_query = """
            SELECT DISTINCT AMR_FIELD_ID.TAG_ID
            FROM AMR_FIELD_ID,amr_region,AMR_PL_GROUP
            
            WHERE AMR_PL_GROUP.PL_REGION_ID = amr_region.id
            AND AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
            AND amr_region.REGION_NAME = :region_id
            AND AMR_FIELD_ID.tag_id NOT like '%.remove%'
            ORDER BY  TAG_ID

            """
            tag_results = fetch_data(ptt_pivot_connection, tag_query, params={"region_id": selected_region})
            tag_options = [str(tag[0]) for tag in tag_results]

        show_tag_query = """
        SELECT DISTINCT
            AMR_FIELD_ID.ID,
            AMR_FIELD_ID.TAG_ID,
            AMR_FIELD_ID.SIM_IP,
            AMR_FIELD_ID.RTU_MODBUS_ID,
            AMR_FIELD_ID.AMR_PHASE,
            AMR_FIELD_CUSTOMER.CUST_FACTORY_NAME,
            AMR_USER.USER_NAME,
            AMR_USER.PASSWORD,
            AMR_FIELD_METER.METER_ID,
            AMR_FIELD_METER.METER_STREAM_NO,
            AMR_FIELD_METER.METER_NO_STREAM,
            AMR_FIELD_METER.METER_STREAM_TYPE,
            AMR_FIELD_METER.METER_PORT_NO,
            AMR_FIELD_METER.METER_AUTO_ENABLE,
            AMR_FIELD_PROFILE.METER_ID,
            AMR_FIELD_PROTOCOL.PROTOCOL_ID,
            AMR_FIELD_PROTOCOL.PROTOCOL_NO_STREAM
        FROM 
            AMR_FIELD_ID
            JOIN AMR_FIELD_CUSTOMER ON AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID
            JOIN AMR_USER ON AMR_FIELD_ID.ID = AMR_USER.ID
            JOIN AMR_FIELD_METER ON AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID
            JOIN AMR_FIELD_PROFILE ON AMR_FIELD_ID.METER_ID = AMR_FIELD_PROFILE.METER_ID
            JOIN AMR_FIELD_PROTOCOL ON AMR_FIELD_ID.PROTOCOL_ID = AMR_FIELD_PROTOCOL.PROTOCOL_ID
            JOIN AMR_PL_GROUP ON AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID 
            JOIN AMR_REGION ON AMR_PL_GROUP.PL_REGION_ID = AMR_REGION.ID
        WHERE 
            amr_region.REGION_NAME = :region_id AND
            AMR_FIELD_ID.TAG_ID = :tag_id
        ORDER BY METER_STREAM_NO
        """
        
        data = []
        if selected_region is not None and selected_tag is not None:
            data = fetch_data(ptt_pivot_connection, show_tag_query, params={"region_id": selected_region, "tag_id": selected_tag})

        columns = [
            'id', 'tag_id', 'sim_ip', 'rtu_modbus_id', 'amr_phase',
            'cust_factory_name', 'user_name', 'password', 'meter_id',
            'meter_stream_no', 'meter_no_stream', 'meter_stream_type',
            'meter_port_no', 'meter_auto_enable', 'meter_id',
            'protocol_id', 'protocol_no_stream'
        ]
        df = pd.DataFrame(data, columns=columns)
        
        if not df.empty:
            list_id = df["id"].iloc[0]
            list_tag_id = df["tag_id"].iloc[0]
            list_cust_factory_name = df["cust_factory_name"].iloc[0]
            list_amr_phase = df["amr_phase"].iloc[0]
            list_rtu_modbus_id = df["rtu_modbus_id"].iloc[0]
            list_sim_ip = df["sim_ip"].iloc[0]
            list_user_name = df["user_name"].iloc[0]
            list_password = df["password"].iloc[0]
            list_meter_stream_type = df["meter_stream_type"].tolist()
            list_meter_stream_no = df["meter_stream_no"].tolist()
            list_meter_port_no = df["meter_port_no"].tolist()
            list_meter_auto_enable = df["meter_auto_enable"].tolist()

            list_meter_stream_type = []
            for meter_stream_type in df["meter_stream_type"]:
                meter_stream_type_list = f"""SELECT vc_name FROM amr_vc_type WHERE id = '{meter_stream_type}' ORDER BY id"""
                type = fetch_data(ptt_pivot_connection, meter_stream_type_list)
                list_meter_stream_type.append(type[0][0] if type else None)

            list_meter_port_no = []
            for meter_port_no in df["meter_port_no"]:
                meter_port_no_list =f"""SELECT port_no FROM amr_port_info WHERE id = '{meter_port_no}' ORDER BY id"""
                port = fetch_data(ptt_pivot_connection, meter_port_no_list)
                list_meter_port_no.append(port[0][0] if port else None)

        else:
            list_id = None
            list_tag_id = None
            list_cust_factory_name = None
            list_amr_phase = None
            list_rtu_modbus_id = None
            list_sim_ip = None
            list_user_name = None
            list_password = None
            list_meter_stream_no = None
            list_meter_auto_enable = None
            list_meter_stream_type = None
            list_meter_port_no = None

        html_table = df.to_html(index=False) if not df.empty else ""

        return render_template('remove_site.html', 
                                region_options=region_options, 
                                tag_options=tag_options, 
                                selected_region=selected_region,
                                selected_tag=selected_tag,
                                list_id=list_id,
                                list_tag_id=list_tag_id,
                                list_cust_factory_name=list_cust_factory_name,
                                list_amr_phase=list_amr_phase,
                                list_rtu_modbus_id=list_rtu_modbus_id,
                                list_sim_ip=list_sim_ip,
                                list_user_name=list_user_name,
                                list_password=list_password,
                                list_meter_stream_no=list_meter_stream_no,
                                list_meter_auto_enable=list_meter_auto_enable,
                                list_meter_stream_type=list_meter_stream_type,
                                list_meter_port_no=list_meter_port_no,
                                html_table=html_table, all_type_result=all_type_result,username=username)
        
        
@app.route('/update_remove_site_route', methods=['POST'])
@login_required
def update_remove_site():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        print("Active Connection:", active_connection)
        
        MET = 'MET0'
        region_query = "SELECT * FROM AMR_REGION"
       
        # Fetch unique region values
        region_results = fetch_data(ptt_pivot_connection, region_query)
        region_options = [str(region[1]) for region in region_results]
        user_update = session['username']
        id = request.form.get("id")
        tag_id = request.form.get("tag_id")
        cust_factory_name = request.form.get("cust_factory_name")
        amr_phase = request.form.get("amr_phase")
        rtu_modbus_id = int(request.form.get("rtu_modbus_id"))
        sim_ip = request.form.get("sim_ip")
        user_name = request.form.get("user_name")
      
        current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))).strftime('%d-%b-%y %I.%M.%S.%f %p ASIA/BANGKOK').upper()
      
        hashed_password = md5_hash(password)
        user_remove = f"{tag_id}.remove"
        update_user = f"""
            UPDATE AMR_USER 
            SET 
                description = '{user_remove}',
                
                UPDATED_BY = '{user_update}',
                TIME_CREATE = '{current_datetime}',
                USER_ENABLE = '0'
            WHERE id = '{id}'
        """
        # print("update_user:", update_user)
        update_sql(ptt_pivot_connection, update_user)
        
        update_field_id = f"""
            UPDATE AMR_FIELD_ID
            SET
                tag_id = '{user_remove}',
                UPDATED_BY = '{user_update}',
                TIME_CREATE = '{current_datetime}',
                RMIU_AUTO_ENABLE = '0'
            WHERE id = '{id}'
        """
        # print("update_field_id", update_field_id)
        update_sql(ptt_pivot_connection, update_field_id)
        
        # update_field_customer = f"""
        #     UPDATE AMR_FIELD_CUSTOMER
        #     SET
        #         cust_factory_name = '{cust_factory_name}'
        #     WHERE id = '{id}'
        # """
        # # print("update_field_customer", update_field_customer)
        # update_sql(ptt_pivot_connection, update_field_customer)
        
        meter_stream_type = request.form.getlist("list_meter_stream_type")      
        meter_stream_no = request.form.get("meter_stream_no")

        meter_port_no = request.form.getlist("list_meter_port_no")
        
        meter_stream_no_count = len(meter_stream_no)

        # ใช้ meter_stream_no_count เป็นจำนวนรอบใน for loop
        for i in range(meter_stream_no_count):
            update_field_meter = f"""
                UPDATE AMR_FIELD_METER
                SET
                    METER_AUTO_ENABLE = '0',
                    METER_POLL_REPEAT1 = '0',
                    METER_POLL_REPEAT2 = '0',
                    METER_POLL_REPEAT3 = '0',
                    METER_POLL_REPEAT4 = '0',
                    METER_POLL_REPEAT5 = '0',
                    METER_POLL_REPEAT6 = '0',
                    UPDATED_BY = '{user_update}',
                    TIME_CREATE = '{current_datetime}'
                WHERE meter_id = '{MET + id}'
                AND meter_stream_no = '{i + 1}'
            """
            update_sql(ptt_pivot_connection, update_field_meter)
        if execute_query(connect_to_ptt_pivot_db(), update_field_meter):
                flash('remove_site successfully', 'success') 
                
                return render_template('remove_site.html',region_options=region_options)
        else:
                flash('Failed to remove_site', 'error')
    return redirect(url_for('remove_site'))

@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('login'))

# @app.route('/popup')
# def popup():
#     print("test")
#     return render_template('popup.html')

@app.route('/allsite', methods=['GET'])
@login_required
def allsite():
    #dateAllSite = session.get("selected_date_allsite", datetime.datetime.now().strftime('%d-%m-%Y'))
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        username = session['username']
        # date_system = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%d-%m-%Y')
        #date_system = datetime.datetime.now().strftime('%d-%m-%Y')
        date_system = session.get("selected_date_allsite", datetime.datetime.now().strftime('%Y-%m-%d'))
        region_query = """
        SELECT ID, REGION_NAME FROM AMR_REGION  ORDER BY ID
        """
        region_results = fetch_data(ptt_pivot_connection, region_query)
        result_data = pd.DataFrame(region_results, columns=['ID', 'REGION_NAME'])
        region_names = result_data['REGION_NAME'].tolist()
        
        region_data = {}  # Dictionary to store data per region
        all_tag_ids_success = []
        all_tag_ids_error = []
        all_tag_ids_manual = []
        for region in region_names:
            error_data_db = f""" 
            SELECT DISTINCT amr_error.meter_id, amr_error.tag_id
            FROM amr_region, amr_pl_group, amr_field_id, amr_error 
            WHERE amr_region.id = amr_pl_group.PL_REGION_ID 
            AND amr_pl_group.field_id = amr_field_id.field_id
            AND amr_field_id.meter_id = amr_error.meter_id
            AND amr_region.region_name LIKE '{region}'
            AND TRUNC(amr_error.DATA_DATE) = TO_DATE('{date_system}', 'YYYY-MM-DD')
            AND amr_error.REPEAT = {DAY_ROUND_REPEAT}
            """
            #print(error_data_db)
            error_data_db_results = fetch_data(ptt_pivot_connection, error_data_db)
            result_error_data_db = pd.DataFrame(error_data_db_results, columns=['meter_id', 'tag_id'])
            error_content = result_error_data_db.shape[0]

            manual_data_db = f""" 
            SELECT DISTINCT amr_error.meter_id, amr_error.tag_id
            FROM amr_region, amr_pl_group, amr_field_id, amr_error 
            WHERE amr_region.id = amr_pl_group.PL_REGION_ID 
            AND amr_pl_group.field_id = amr_field_id.field_id
            AND amr_field_id.meter_id = amr_error.meter_id
            AND amr_region.region_name LIKE '{region}'
            AND TRUNC(amr_error.DATA_DATE) = TO_DATE('{date_system}', 'YYYY-MM-DD')
            AND amr_error.REPEAT = 99
            """
            #print(manual_data_db)
            manual_data_db_results = fetch_data(ptt_pivot_connection, manual_data_db)            
            result_manual_data_db = pd.DataFrame(manual_data_db_results, columns=['meter_id', 'tag_id'])
            manual_content = result_manual_data_db.shape[0]

            success_data_db = f""" 
            SELECT DISTINCT amr_field_id.tag_id, amr_configured_data.meter_id
            FROM amr_region, amr_pl_group, amr_field_id, amr_configured_data
            WHERE amr_region.id = amr_pl_group.PL_REGION_ID 
            AND amr_pl_group.field_id = amr_field_id.field_id
            AND amr_field_id.meter_id = amr_configured_data.meter_id
            AND amr_region.region_name LIKE '{region}'
            AND amr_configured_data.data_date = TO_DATE('{date_system}', 'YYYY-MM-DD')
            """
            #print(success_data_db)
            success_data_db_results = fetch_data(ptt_pivot_connection, success_data_db)
            result_success_data_db = pd.DataFrame(success_data_db_results, columns=['tag_id', 'meter_id'])
            
            # Exclude error tag_ids from success_data
            filter_data_db = pd.concat([result_error_data_db, result_manual_data_db], ignore_index=True)
            df_data_cleaned = result_success_data_db[~result_success_data_db['tag_id'].isin(filter_data_db['tag_id'])]              
  
            success_content = df_data_cleaned.shape[0]

            # Sort and combine tag IDs
            tag_ids_success = sorted(df_data_cleaned['tag_id'].tolist())
            tag_ids_error = sorted(result_error_data_db['tag_id'].tolist())
            tag_ids_manual = sorted(result_manual_data_db['tag_id'].tolist())

            combined_tags = sorted(tag_ids_success + tag_ids_error+tag_ids_manual)
            all_tag_ids_success.append(tag_ids_success)
            all_tag_ids_error.append(tag_ids_error)
            all_tag_ids_manual.append(tag_ids_manual)

            # Store data in the dictionary by region
            region_data[region] = {
                'combined_tags': combined_tags,
                'tag_ids_success': tag_ids_success,
                'tag_ids_error': tag_ids_error,
                'tag_ids_manual': tag_ids_manual,
                'error_content': error_content,
                'success_content': success_content,
                'manual_content': manual_content
            }
        #HARD CODE
        all_tag_ids_success_flat = list(itertools.chain(*all_tag_ids_success))
        all_tag_ids_error_flat = list(itertools.chain(*all_tag_ids_error))
        all_tag_ids_manual_flat = list(itertools.chain(*all_tag_ids_manual))
        # Count total success tags
        total_success_tags = len(all_tag_ids_success_flat)
        total_manual_tags = len(all_tag_ids_manual_flat)
        total_error_tags = len(all_tag_ids_error_flat)
        
        allsite = total_success_tags + total_error_tags + total_manual_tags
        total_error_tags =  total_error_tags + total_manual_tags

    return render_template('allsite.html',username=username, region_data=region_data,total_success_tags=total_success_tags,total_error_tags=total_error_tags,allsite=allsite)

##################### success ################################

@app.route('/success')
@login_required
def success():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        username = session['username']
        date_system = session.get("selected_date_success", datetime.datetime.now().strftime('%Y-%m-%d'))
        data_result = get_autopoll_statistic(ptt_pivot_connection, "Success", date_system)

        error_content = data_result['error_content']
        #manual_content = data_result['manual_content']
        success_content = data_result['success_content']
        allsite = data_result['allsite']
        tag_ids = data_result['tag_ids']
        tags_manuals = data_result['tags_manuals']
    return render_template('success.html',username=username, tag_ids=tag_ids,error_content=error_content,success_content=success_content,allsite=allsite)


@app.route('/success_user_group')
@login_required
def success_user_group():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        #date_system = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%d-%m-%Y')
        date_system = datetime.datetime.now().strftime('%d-%m-%Y')
        # Adjust the query to use bind variables
        
        selected_region = session['username']
        region_query = f"""
        SELECT amr_region.REGION_NAME from amr_user,amr_region WHERE amr_user.user_group=amr_region.id AND amr_user.user_name like '{selected_region}'
        """
        
        # Fetch unique region values
        region_results = fetch_data(ptt_pivot_connection, region_query)
        df_detail = pd.DataFrame(region_results, columns=['REGION_NAME'])

        region_options = df_detail['REGION_NAME'].iloc[0]
        print("region_options",region_options)
        
        error_data_db = f""" SELECT DISTINCT amr_error.meter_id, amr_error.tag_id
        FROM amr_region, amr_pl_group, amr_field_id, amr_error 
        WHERE amr_region.id = amr_pl_group.PL_REGION_ID 
        AND amr_pl_group.field_id = amr_field_id.field_id
        AND amr_field_id.meter_id = amr_error.meter_id
        AND amr_region.region_name like '{region_options}'
        AND TRUNC(amr_error.DATA_DATE) = TO_DATE('{date_system}', 'DD-MM-YYYY') AND REPEAT = {DAY_ROUND_REPEAT}"""
        error_data_db_results  =  fetch_data(ptt_pivot_connection, error_data_db)
        result_error_data_db = pd.DataFrame(error_data_db_results, columns=['meter_id', 'tag_id'])
        #print(result_error_data_db)
        
        error_content = result_error_data_db.shape[0]
        
    ############################################

        success_data_db = f"""SELECT DISTINCT  amr_field_id.tag_id,amr_configured_data.meter_id
        FROM amr_region, amr_pl_group, amr_field_id, amr_configured_data
        WHERE amr_region.id = amr_pl_group.PL_REGION_ID 
        AND amr_pl_group.field_id = amr_field_id.field_id
        AND amr_field_id.meter_id = amr_configured_data.meter_id
        AND amr_region.region_name like '{region_options}'
        AND amr_configured_data.data_date = TO_DATE('{date_system}', 'DD-MM-YYYY')"""
        success_data_db_results  =  fetch_data(ptt_pivot_connection, success_data_db)
        result_success_data_db = pd.DataFrame(success_data_db_results, columns=['tag_id', 'meter_id'])
        # print("result_success_data_db",result_success_data_db)
        
        
        df_data_cleaned = result_success_data_db[~result_success_data_db['tag_id'].isin(result_error_data_db['tag_id'])]
        success_content = df_data_cleaned.shape[0]
        # print("success_content",success_content)
        tag_ids_success = sorted(df_data_cleaned['tag_id'].tolist())
        tag_ids_error = sorted(result_error_data_db['tag_id'].tolist())
        combined_tags = sorted(tag_ids_success + tag_ids_error)
        allsite = len(combined_tags)
    #########################################################################
    return render_template('success_user_group.html',allsite=allsite,combined_tags=combined_tags,tag_ids_success=tag_ids_success,tag_ids_error=tag_ids_error,region_options=region_options,error_content=error_content,success_content=success_content,selected_region=selected_region)

##############################################################


##################### error ################################
@app.route('/hourly_data')
@login_required
def hourly_data():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        # date_system = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%d-%m-%Y')
        date_system = datetime.datetime.now().strftime('%d-%m-%Y')
        # Adjust the query to use bind variables
        data_autopoll = f"""
        SELECT DISTINCT AMR_BILLING_HOURLY_DATA.meter_id, amr_field_id.tag_id
        FROM AMR_BILLING_HOURLY_DATA, amr_field_id
        WHERE AMR_BILLING_HOURLY_DATA.meter_id = amr_field_id.meter_id
        AND AMR_BILLING_HOURLY_DATA.data_date =  TO_DATE('{date_system}', 'DD-MM-YYYY')
        """

        results_datapoll = fetch_data(ptt_pivot_connection,data_autopoll)
        df_data_autopoll = pd.DataFrame(results_datapoll, columns=['meter_id', 'tag_id'])

        data_autopoll_error = f""" SELECT DISTINCT TAG_ID FROM AMR_HOURLY_ERROR where TRUNC(DATA_DATE) =  TO_DATE('{date_system}', 'DD-MM-YYYY')"""
        results_datapoll_error = fetch_data(ptt_pivot_connection,data_autopoll_error)
        df_data_autopoll_error = pd.DataFrame(results_datapoll_error, columns=['tag_id'])

        df_data_autopoll_cleaned = df_data_autopoll[~df_data_autopoll['tag_id'].isin(df_data_autopoll_error['tag_id'])]

        error_content = df_data_autopoll_error.shape[0]
        success_content = df_data_autopoll_cleaned.shape[0]
        
    
        tag_ids = sorted(df_data_autopoll_error['tag_id'].tolist())
        
    return render_template('hourly_data.html', tag_ids=tag_ids,error_content=error_content,success_content=success_content)


@app.route('/error_data_report')
@login_required
def error_data_report():
    # เชื่อมต่อกับฐานข้อมูล
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        # ดึงข้อมูลวันที่ปัจจุบัน - 1 วัน
        date_system = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%d-%m-%Y')
        
        # SQL Query ดึงข้อมูลจาก AMR_ERROR
        data_autopoll_error = f"""SELECT DISTINCT TAG_ID, meter_stream_no, error_desc 
                                  FROM AMR_ERROR 
                                  WHERE TRUNC(DATA_DATE) = TO_DATE('{date_system}', 'DD-MM-YYYY') ORDER BY TAG_ID"""
        
        # ดึงข้อมูลจากฐานข้อมูล
        results_datapoll_error = fetch_data(ptt_pivot_connection, data_autopoll_error)
        
        # แปลงข้อมูลเป็น DataFrame
        df_data_autopoll_error = pd.DataFrame(results_datapoll_error, columns=['tag_id', 'meter_stream_no', 'error_desc'])

        # จัดการกับ tag_id ให้มีค่าเพียงตัวเดียว
        previous_tag_id = None
        for index, row in df_data_autopoll_error.iterrows():
            if row['tag_id'] == previous_tag_id:
                df_data_autopoll_error.at[index, 'tag_id'] = ''
            else:
                previous_tag_id = row['tag_id']

        # ใช้ BytesIO เพื่อสร้างไฟล์ Excel ชั่วคราวในหน่วยความจำ
        output = BytesIO()
        
        # เขียนข้อมูล DataFrame ลงในไฟล์ Excel
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df_data_autopoll_error.to_excel(writer, index=False, sheet_name='ErrorData')

            # ปรับแต่งไฟล์ Excel ที่เขียนลงไป
            worksheet = writer.sheets['ErrorData']
            # ตั้งค่าในเซลล์ A1
            worksheet.write('A1', f'DATA REPORT ERROR - {date_system}', writer.book.add_format({'bold': True, 'font_size': 14, 'bg_color': '#C0C0C0', 'align': 'center'}))
            worksheet.merge_range('A1:C1', f'DATA REPORT ERROR - {date_system}', writer.book.add_format({'bold': True, 'font_size': 14, 'bg_color': '#C0C0C0', 'align': 'center'}))
            # ตั้งค่าหัวตาราง
            header_format = writer.book.add_format({'bold': True, 'bg_color': '#C0C0C0', 'border': 1, 'align': 'center'})
            worksheet.set_row(1, 20)  # เพิ่มความสูงให้หัวตาราง
            for col_num, value in enumerate(df_data_autopoll_error.columns.values):
                worksheet.write(1, col_num, value, header_format)

            # ขยายความกว้างคอลัมน์สำหรับข้อมูล
            worksheet.set_column('A:A', 20)  # tag_id
            worksheet.set_column('B:B', 15)  # meter_stream_no

            # Calculate the maximum width for the error_desc column
            max_length = max(df_data_autopoll_error['error_desc'].astype(str).apply(len).max(), len('Error Description')) + 2  # Add extra padding
            worksheet.set_column('C:C', max_length)  # error_desc

            # เพิ่มการตั้งค่ารูปแบบที่เหลือให้กับข้อมูล
            cell_format = writer.book.add_format({'border': 1, 'align': 'center'})  # รูปแบบของเซลล์
            for row in range(2, len(df_data_autopoll_error) + 2):  # เริ่มต้นที่แถวที่ 2
                worksheet.set_row(row, 20)  # ตั้งความสูงให้กับแต่ละแถว
                for col in range(len(df_data_autopoll_error.columns)):
                    worksheet.write(row, col, df_data_autopoll_error.iat[row-2, col], cell_format)

        # เลื่อน pointer ของ BytesIO ไปที่จุดเริ่มต้น
        output.seek(0)

        # ส่งไฟล์ Excel กลับเป็นไฟล์แนบ
        return send_file(
            output,
            as_attachment=True,
            download_name='error_data_report.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

@app.route('/error_data_report_pdf')
def error_data_report_pdf():
    # เชื่อมต่อกับฐานข้อมูล
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        # ดึงข้อมูลวันที่ปัจจุบัน - 1 วัน
        date_system = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%d-%m-%Y')
        
        # SQL Query ดึงข้อมูลจาก AMR_ERROR
        data_autopoll_error = f"""SELECT DISTINCT TAG_ID, meter_stream_no, error_desc 
                                  FROM AMR_ERROR 
                                  WHERE TRUNC(DATA_DATE) = TO_DATE('{date_system}', 'DD-MM-YYYY') AND REPEAT = {DAY_ROUND_REPEAT} ORDER BY TAG_ID"""
        
        # ดึงข้อมูลจากฐานข้อมูล
        results_datapoll_error = fetch_data(ptt_pivot_connection, data_autopoll_error)
        
        # แปลงข้อมูลเป็น DataFrame
        df_data_autopoll_error = pd.DataFrame(results_datapoll_error, columns=['tag_id', 'meter_stream_no', 'error_desc'])

        # จัดการกับ tag_id ให้มีค่าเพียงตัวเดียว
        previous_tag_id = None
        for index, row in df_data_autopoll_error.iterrows():
            if row['tag_id'] == previous_tag_id:
                df_data_autopoll_error.at[index, 'tag_id'] = ''
            else:
                previous_tag_id = row['tag_id']

        # สร้าง HTML ภายในฟังก์ชัน
        html_content = f"""
        <html>
        <head>
            <style>
                table {{
                    width: 100%;
                    border-collapse: collapse;
                }}
                th, td {{
                    border: 1px solid black;
                    padding: 8px;
                    text-align: center;
                    color: #000000;
                }}
                th {{
                    background-color: #C0C0C0;
                }}
                td {{
                    white-space: nowrap;  /* Prevent text from wrapping */
                }}
                h1 {{
                    text-align: center;
                }}
            </style>
        </head>
        <body>
            <h1>DATA REPORT ERROR - {date_system}</h1>
            <table>
                <tr>
                    <th>Tag ID</th>
                    <th>Meter Stream No</th>
                    <th>Error Description</th>
                </tr>"""

        for index, row in df_data_autopoll_error.iterrows():
            tag_id = row['tag_id'] if row['tag_id'] else '-'  # แทนที่ค่าที่ว่างด้วย '-'
            html_content += f"""
            <tr>
                <td>{tag_id}</td>
                <td>{row['meter_stream_no']}</td>
                <td>{row['error_desc']}</td>
            </tr>"""

        html_content += """
                </table>
            </body>
        </html>"""

        # สร้าง PDF ในหน่วยความจำ
        pdf_file = io.BytesIO()
        pisa_status = pisa.CreatePDF(html_content, dest=pdf_file)

        # ตรวจสอบข้อผิดพลาด
        if pisa_status.err:
            return "Error generating PDF", 500

        # รีเซ็ต pointer ไปที่จุดเริ่มต้นของ BytesIO
        pdf_file.seek(0)

        # ส่ง PDF กลับเป็นไฟล์แนบ
        return send_file(
            pdf_file,
            as_attachment=True,
            download_name='error_data_report.pdf',
            mimetype='application/pdf'
        )

@app.route('/error_data_user_group')
@login_required
def error_data_user_group():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        #date_system = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%d-%m-%Y')
        date_system = datetime.datetime.now().strftime('%d-%m-%Y')
        # Adjust the query to use bind variables
        
        selected_region = session['username']
        region_query = f"""
        SELECT amr_region.REGION_NAME from amr_user,amr_region WHERE amr_user.user_group=amr_region.id AND amr_user.user_name like '{selected_region}'
        """
        
        # Fetch unique region values
        region_results = fetch_data(ptt_pivot_connection, region_query)
        df_detail = pd.DataFrame(region_results, columns=['REGION_NAME'])

        region_options = df_detail['REGION_NAME'].iloc[0]
        print("region_options",region_options)
        # region_query = """
        # SELECT ID,REGION_NAME FROM AMR_REGION 
        # """
        # region_results = fetch_data(ptt_pivot_connection, region_query)
        # result_data = pd.DataFrame(region_results, columns=['ID', 'REGION_NAME'])
        # region_names = result_data['REGION_NAME'].tolist()
        # print(result_data['REGION_NAME'])
    ##########################################
        
        error_data_db = f""" SELECT DISTINCT amr_error.meter_id, amr_error.tag_id
        FROM amr_region, amr_pl_group, amr_field_id, amr_error 
        WHERE amr_region.id = amr_pl_group.PL_REGION_ID 
        AND amr_pl_group.field_id = amr_field_id.field_id
        AND amr_field_id.meter_id = amr_error.meter_id
        AND amr_region.region_name like '{region_options}'
        AND TRUNC(amr_error.DATA_DATE) = TO_DATE('{date_system}', 'DD-MM-YYYY')"""
        error_data_db_results  =  fetch_data(ptt_pivot_connection, error_data_db)
        result_error_data_db = pd.DataFrame(error_data_db_results, columns=['meter_id', 'tag_id'])
        error_content = result_error_data_db.shape[0]
        
    ############################################

        success_data_db = f"""SELECT DISTINCT  amr_field_id.tag_id,amr_configured_data.meter_id
        FROM amr_region, amr_pl_group, amr_field_id, amr_configured_data
        WHERE amr_region.id = amr_pl_group.PL_REGION_ID 
        AND amr_pl_group.field_id = amr_field_id.field_id
        AND amr_field_id.meter_id = amr_configured_data.meter_id
        AND amr_region.region_name like '{region_options}'
        AND amr_configured_data.data_date = TO_DATE('{date_system}', 'DD-MM-YYYY')"""
        success_data_db_results  =  fetch_data(ptt_pivot_connection, success_data_db)
        result_success_data_db = pd.DataFrame(success_data_db_results, columns=['tag_id', 'meter_id'])
        # print("result_success_data_db",result_success_data_db)
        
        
        df_data_cleaned = result_success_data_db[~result_success_data_db['tag_id'].isin(result_error_data_db['tag_id'])]
        success_content = df_data_cleaned.shape[0]
        # print("success_content",success_content)
        tag_ids_success = sorted(df_data_cleaned['tag_id'].tolist())
        tag_ids_error = sorted(result_error_data_db['tag_id'].tolist())
        combined_tags = sorted(tag_ids_success + tag_ids_error)
        allsite = len(combined_tags)
    #########################################################################      
    return render_template('error_data_user_group.html',allsite=allsite,combined_tags=combined_tags,tag_ids_success=tag_ids_success,tag_ids_error=tag_ids_error,region_options=region_options,error_content=error_content,success_content=success_content,selected_region=selected_region)


@app.route('/error_data')
@login_required
def error_data():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        username = session['username']
        date_system = session.get("selected_date_errorData", datetime.datetime.now().strftime('%Y-%m-%d'))
        data_result = get_autopoll_statistic(ptt_pivot_connection, "Failed", date_system)
        error_content = data_result['error_content']
        #manual_content = data_result['manual_content']
        success_content = data_result['success_content']
        allsite = data_result['allsite']
        tag_ids = data_result['tag_ids']
        tags_manuals = data_result['tags_manuals']
    return render_template('error_data.html',username=username, tag_ids=tag_ids,error_content=error_content,success_content=success_content,allsite=allsite,tags_manuals=tags_manuals)

##############################################################
@app.route('/show_error_data', methods=['POST'])
@login_required
def show_error_data():
    data = request.get_json()
    tag_id = data.get('tag_id')
    print(tag_id)
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        date_system = datetime.datetime.now().strftime('%d-%m-%Y')
        # date_system = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%d-%m-%Y')
        
        show_error_data_query = f"""
        SELECT TAG_ID, METER_ID, METER_STREAM_NO, EVC_TYPE, ERROR_DESC 
        FROM AMR_ERROR 
        WHERE TRUNC(DATA_DATE) = TO_DATE('{date_system}', 'DD-MM-YYYY') 
        AND TAG_ID LIKE '{tag_id}' 
        
        """
        result = fetch_data(ptt_pivot_connection, show_error_data_query)
        
        # Convert result to a DataFrame
        result_data = pd.DataFrame(result, columns=['TAG_ID', 'METER_ID', 'METER_STREAM_NO', 'EVC_TYPE', 'ERROR_DESC'])
        
        # Convert DataFrame to JSON (orient='records' ensures each row is a JSON object)
        result_json = result_data.to_json(orient='records')
        
    # Return JSON response
    return result_json
    

@app.route('/DailySummary',methods=['GET'])
@login_required
def DailySummary():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        
        selected_region = request.args.get('region_dropdown', '')
        print("selected_region",selected_region)
        selected_date = request.args.get('date_dropdown', '')
        print("selected_date",selected_date)
        username = session['username']
        
        region_query = """
        SELECT ID,REGION_NAME FROM AMR_REGION 
        """
        region_results = fetch_data(ptt_pivot_connection, region_query)
        result_data = pd.DataFrame(region_results, columns=['ID', 'REGION_NAME'])
        region_names = result_data['REGION_NAME'].tolist()
        # print(result_data['REGION_NAME'])
    ##########################################
        
        error_data_db = f""" SELECT DISTINCT amr_error.meter_id, amr_error.tag_id
        FROM amr_region, amr_pl_group, amr_field_id, amr_error 
        WHERE amr_region.id = amr_pl_group.PL_REGION_ID 
        AND amr_pl_group.field_id = amr_field_id.field_id
        AND amr_field_id.meter_id = amr_error.meter_id
        AND amr_region.region_name like '{selected_region}'
        AND TRUNC(amr_error.DATA_DATE) = TO_DATE('{selected_date}', 'DD-MM-YYYY') AND amr_error.REPEAT = {DAY_ROUND_REPEAT}"""
        error_data_db_results  =  fetch_data(ptt_pivot_connection, error_data_db)
        result_error_data_db = pd.DataFrame(error_data_db_results, columns=['meter_id', 'tag_id'])

        error_content = result_error_data_db.shape[0]
        
    ############################################

        success_data_db = f"""SELECT DISTINCT  amr_field_id.tag_id,amr_configured_data.meter_id
        FROM amr_region, amr_pl_group, amr_field_id, amr_configured_data
        WHERE amr_region.id = amr_pl_group.PL_REGION_ID 
        AND amr_pl_group.field_id = amr_field_id.field_id
        AND amr_field_id.meter_id = amr_configured_data.meter_id
        AND amr_region.region_name like '{selected_region}'
        AND amr_configured_data.data_date = TO_DATE('{selected_date}', 'DD-MM-YYYY')"""
        success_data_db_results  =  fetch_data(ptt_pivot_connection, success_data_db)
        result_success_data_db = pd.DataFrame(success_data_db_results, columns=['tag_id', 'meter_id'])
        # print("result_success_data_db",result_success_data_db)
        
        
        df_data_cleaned = result_success_data_db[~result_success_data_db['tag_id'].isin(result_error_data_db['tag_id'])]
        success_content = df_data_cleaned.shape[0]
        # print("success_content",success_content)
        tag_ids_success = sorted(df_data_cleaned['tag_id'].tolist())
        tag_ids_error = sorted(result_error_data_db['tag_id'].tolist())
        combined_tags = sorted(tag_ids_success + tag_ids_error)
    #########################################################################
    return render_template('DailySummary.html',username=username,combined_tags=combined_tags,tag_ids_success=tag_ids_success,tag_ids_error=tag_ids_error,region_names=region_names,error_content=error_content,success_content=success_content,selected_region=selected_region,selected_date=selected_date)


@app.route('/DailySummary_user_group',methods=['GET'])
@login_required
def DailySummary_user_group():
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        username = session['username']
        selected_date = request.args.get('date_dropdown', '')
        print("selected_date",selected_date)
        
        selected_region = session['username']
        region_query = f"""
        SELECT amr_region.REGION_NAME from amr_user,amr_region WHERE amr_user.user_group=amr_region.id AND amr_user.user_name like '{selected_region}'
        """

        # Fetch unique region values
        region_results = fetch_data(ptt_pivot_connection, region_query)
        df_detail = pd.DataFrame(region_results, columns=['REGION_NAME'])

        region_options = df_detail['REGION_NAME'].iloc[0]
        print("region_options",region_options)
        # region_query = """
        # SELECT ID,REGION_NAME FROM AMR_REGION 
        # """
        # region_results = fetch_data(ptt_pivot_connection, region_query)
        # result_data = pd.DataFrame(region_results, columns=['ID', 'REGION_NAME'])
        # region_names = result_data['REGION_NAME'].tolist()
        # print(result_data['REGION_NAME'])
    ##########################################
        
        
        error_data_db = f""" SELECT DISTINCT amr_error.meter_id, amr_error.tag_id
        FROM amr_region, amr_pl_group, amr_field_id, amr_error 
        WHERE amr_region.id = amr_pl_group.PL_REGION_ID 
        AND amr_pl_group.field_id = amr_field_id.field_id
        AND amr_field_id.meter_id = amr_error.meter_id
        AND amr_region.region_name like '{region_options}'
        AND TRUNC(amr_error.DATA_DATE) = TO_DATE('{selected_date}', 'DD-MM-YYYY')"""
        error_data_db_results  =  fetch_data(ptt_pivot_connection, error_data_db)
        result_error_data_db = pd.DataFrame(error_data_db_results, columns=['meter_id', 'tag_id'])       
        error_content = result_error_data_db.shape[0]
        
    ############################################

        success_data_db = f"""SELECT DISTINCT  amr_field_id.tag_id,amr_configured_data.meter_id
        FROM amr_region, amr_pl_group, amr_field_id, amr_configured_data
        WHERE amr_region.id = amr_pl_group.PL_REGION_ID 
        AND amr_pl_group.field_id = amr_field_id.field_id
        AND amr_field_id.meter_id = amr_configured_data.meter_id
        AND amr_region.region_name like '{region_options}'
        AND amr_configured_data.data_date = TO_DATE('{selected_date}', 'DD-MM-YYYY')"""
        success_data_db_results  =  fetch_data(ptt_pivot_connection, success_data_db)
        result_success_data_db = pd.DataFrame(success_data_db_results, columns=['tag_id', 'meter_id'])
        # print("result_success_data_db",result_success_data_db)
        
        
        df_data_cleaned = result_success_data_db[~result_success_data_db['tag_id'].isin(result_error_data_db['tag_id'])]
        success_content = df_data_cleaned.shape[0]
        # print("success_content",success_content)
        tag_ids_success = sorted(df_data_cleaned['tag_id'].tolist())
        tag_ids_error = sorted(result_error_data_db['tag_id'].tolist())
        combined_tags = sorted(tag_ids_success + tag_ids_error)
    ######################################################################### 
    return render_template('DailySummary_user_group.html',username=username,combined_tags=combined_tags,tag_ids_success=tag_ids_success,tag_ids_error=tag_ids_error,region_options=region_options,error_content=error_content,success_content=success_content,selected_region=selected_region,selected_date=selected_date)

@app.route('/DailySummary_errordatashow', methods=['POST'])
@login_required
def DailySummary_errordatashow():
    data = request.json
    tag_id = data.get('tag_id')
    
    selected_date = data.get('selected_date')
    
    
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        # date_system = datetime.datetime.now().strftime('%d-%m-%Y')
        date_system = selected_date
        print("date_system",date_system)
        show_error_data_query = f"""
        SELECT TAG_ID, METER_ID, METER_STREAM_NO, EVC_TYPE, ERROR_DESC 
        FROM AMR_ERROR 
        WHERE TRUNC(DATA_DATE) = TO_DATE('{date_system}', 'DD-MM-YYYY') 
        AND TAG_ID LIKE '{tag_id}' 
        
        """
        # Assuming you have a function to execute the query and fetch results
        result = fetch_data(ptt_pivot_connection, show_error_data_query)
        
        # Convert result to a DataFrame
        result_data = pd.DataFrame(result, columns=['TAG_ID', 'METER_ID', 'METER_STREAM_NO', 'EVC_TYPE', 'ERROR_DESC'])
        
        # Convert DataFrame to JSON (orient='records' ensures each row is a JSON object)
        result_json = result_data.to_json(orient='records')
        
        # Log the DataFrame for debugging (optional)
        print(result_data)
        
    # Return JSON response
    return result_json

@app.route('/site_detail')
@login_required
def site_detail():
    selected_region = request.args.get('region_dropdown', '')

    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        region_query = """
        SELECT region_name FROM AMR_REGION
        """
        username = session['username']
        # Fetch unique region values
        region_results = fetch_data(ptt_pivot_connection, region_query)
        region_options = [str(region[0]) for region in region_results]
        
        site_detail_data = f"""
            SELECT 
                AMR_FIELD_ID.id,
                AMR_FIELD_ID.TAG_ID AS Sitename,
                AMR_FIELD_ID.SIM_IP AS IPAddress,
                amr_rmiu_type.rmiu_name,
                MAX(CASE WHEN AMR_FIELD_METER.METER_STREAM_NO = 1 THEN AMR_VC_TYPE.VC_NAME ELSE NULL END) AS VCtype_Run1,
                MAX(CASE WHEN AMR_FIELD_METER.METER_STREAM_NO = 2 THEN AMR_VC_TYPE.VC_NAME ELSE NULL END) AS VCtype_Run2,
                MAX(CASE WHEN AMR_FIELD_METER.METER_STREAM_NO = 3 THEN AMR_VC_TYPE.VC_NAME ELSE NULL END) AS VCtype_Run3,
                MAX(CASE WHEN AMR_FIELD_METER.METER_STREAM_NO = 4 THEN AMR_VC_TYPE.VC_NAME ELSE NULL END) AS VCtype_Run4
            FROM
                AMR_FIELD_ID,
                AMR_USER,
                AMR_FIELD_CUSTOMER,
                AMR_FIELD_METER,
                AMR_PL_GROUP,
                AMR_VC_TYPE,
                AMR_PORT_INFO,
                amr_rmiu_type,
                amr_region
            WHERE
                AMR_USER.USER_ENABLE = 1 AND
                AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID AND
                AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP AND
                AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID AND
                AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID AND
                AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE AND
                amr_rmiu_type.id = amr_field_id.rmiu_type AND
                AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID AND
                amr_region.id = amr_pl_group.pl_region_id AND
                amr_region.region_name LIKE '{selected_region}'
            GROUP BY
                AMR_FIELD_ID.id,
                AMR_FIELD_ID.TAG_ID,
                AMR_FIELD_ID.SIM_IP,
                amr_rmiu_type.rmiu_name
            ORDER BY
                AMR_FIELD_ID.TAG_ID ASC """
                
        site_detail_results = fetch_data(ptt_pivot_connection, site_detail_data)
        df_site_detail = pd.DataFrame(site_detail_results, columns=['ID', 'SITE', 'IPAddress', 'RMIU', 'RUN1', 'RUN2', 'RUN3', 'RUN4'])
        df_site_detail = df_site_detail.fillna('-')  # Replace None values with '-'
        
    site_detail_list = df_site_detail.to_dict('records')
    return render_template('sitedetail.html',username=username, region_options=region_options, site_detail_list=site_detail_list,selected_region=selected_region)

@app.route('/site_detail_user_group')
@login_required
def site_detail_user_group():
    selected_region = session['username']  
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        region_query = f"""
        SELECT amr_region.REGION_NAME from amr_user,amr_region WHERE amr_user.user_group=amr_region.id AND amr_user.user_name like '{selected_region}'
        """
        
        # Fetch unique region values
        region_results = fetch_data(ptt_pivot_connection, region_query)
        df_detail = pd.DataFrame(region_results, columns=['REGION_NAME'])

        region_options = df_detail['REGION_NAME'].iloc[0]
        print("region_options",region_options)
        site_detail_data = f"""
            SELECT 
                AMR_FIELD_ID.id,
                AMR_FIELD_ID.TAG_ID AS Sitename,
                AMR_FIELD_ID.SIM_IP AS IPAddress,
                amr_rmiu_type.rmiu_name,
                MAX(CASE WHEN AMR_FIELD_METER.METER_STREAM_NO = 1 THEN AMR_VC_TYPE.VC_NAME ELSE NULL END) AS VCtype_Run1,
                MAX(CASE WHEN AMR_FIELD_METER.METER_STREAM_NO = 2 THEN AMR_VC_TYPE.VC_NAME ELSE NULL END) AS VCtype_Run2,
                MAX(CASE WHEN AMR_FIELD_METER.METER_STREAM_NO = 3 THEN AMR_VC_TYPE.VC_NAME ELSE NULL END) AS VCtype_Run3,
                MAX(CASE WHEN AMR_FIELD_METER.METER_STREAM_NO = 4 THEN AMR_VC_TYPE.VC_NAME ELSE NULL END) AS VCtype_Run4
            FROM
                AMR_FIELD_ID,
                AMR_USER,
                AMR_FIELD_CUSTOMER,
                AMR_FIELD_METER,
                AMR_PL_GROUP,
                AMR_VC_TYPE,
                AMR_PORT_INFO,
                amr_rmiu_type,
                amr_region
            WHERE
                AMR_USER.USER_ENABLE = 1 AND
                AMR_FIELD_ID.FIELD_ID = AMR_PL_GROUP.FIELD_ID AND
                AMR_FIELD_ID.METER_ID = AMR_USER.USER_GROUP AND
                AMR_FIELD_ID.CUST_ID = AMR_FIELD_CUSTOMER.CUST_ID AND
                AMR_FIELD_ID.METER_ID = AMR_FIELD_METER.METER_ID AND
                AMR_VC_TYPE.ID = AMR_FIELD_METER.METER_STREAM_TYPE AND
                amr_rmiu_type.id = amr_field_id.rmiu_type AND
                AMR_FIELD_METER.METER_PORT_NO = AMR_PORT_INFO.ID AND
                amr_region.id = amr_pl_group.pl_region_id AND
                amr_region.region_name LIKE '{region_options}'
            GROUP BY
                AMR_FIELD_ID.id,
                AMR_FIELD_ID.TAG_ID,
                AMR_FIELD_ID.SIM_IP,
                amr_rmiu_type.rmiu_name
            ORDER BY
                AMR_FIELD_ID.TAG_ID ASC """
                
        site_detail_results = fetch_data(ptt_pivot_connection, site_detail_data)
        df_site_detail = pd.DataFrame(site_detail_results, columns=['ID', 'SITE', 'IPAddress', 'RMIU', 'RUN1', 'RUN2', 'RUN3', 'RUN4'])
        df_site_detail = df_site_detail.fillna('-')  # Replace None values with '-'
        
    site_detail_list = df_site_detail.to_dict('records')
    return render_template('sitedetail_user_group.html', region_options=region_options, site_detail_list=site_detail_list,selected_region=selected_region)

@app.route('/autopoll_Scheduler')
@login_required
def autopoll_Scheduler():
    desired_names = ['Actaris(G1)_repeat1','Actaris(G1)_repeat2','Actaris(G1)_repeat3'
                     ,'Actaris(G2)_repeat1','Actaris(G2)_repeat2','Actaris(G2)_repeat3'
                     ,'Actaris(G3)_repeat1','Actaris(G3)_repeat2','Actaris(G3)_repeat3'
                     ,'Actaris(G4)_repeat1','Actaris(G4)_repeat2','Actaris(G4)_repeat3'
                     ,'Actaris(u)_repeat1','Actaris(u)_repeat2','Actaris(u)_repeat3'
                     ,'Elster-280_repeat1','Elster-280_repeat2','Elster-280_repeat3'
                     , 'Hourly(G1)','Hourly(G2)','Hourly(G3)','Hourly(G4)'
                     ,'Hourly(u)','Hourly(elster)','data_autopoll_repeat1','data_autopoll_repeat2','data_autopoll_repeat3'
                     
                     ]  # Change as needed
    tasks = get_all_tasks(desired_names)
    
    # Sort tasks according to desired_names order
    task_dict = {task['name']: task for task in tasks}
    sorted_tasks = [task_dict[name] for name in desired_names if name in task_dict]
    
    return render_template('autopollScheduler.html', tasks=sorted_tasks)

@app.route('/edit/<task_name>', methods=['GET', 'POST'])
@login_required
def edit_task(task_name):
    if request.method == 'POST':
        new_name = request.form.get('name')
        trigger_type = request.form.get('trigger_type')

        days_of_month = request.form.getlist('days_of_month')
        months_of_year = request.form.getlist('months_of_year')

        print(f'Received Days of Month: {days_of_month}')  # Debug output
        print(f'Received Months of Year: {months_of_year}')  # Debug output

        try:
            if trigger_type == '1':
                new_start_time_str = request.form.get('one_time_start')
                if new_start_time_str:
                    new_start_time = datetime.datetime.strptime(new_start_time_str, '%Y-%m-%dT%H:%M')
                    update_task(task_name, new_name, new_start_time, trigger_type)
                else:
                    flash('Start time for one-time trigger is required.', 'error')

            elif trigger_type == '2':
                new_start_time_str = request.form.get('daily_start')
                if new_start_time_str:
                    new_start_time = datetime.datetime.strptime(new_start_time_str, '%Y-%m-%dT%H:%M')
                    days_interval = int(request.form.get('days_interval', 1))
                    update_task(task_name, new_name, new_start_time, trigger_type, days_interval=days_interval)
                else:
                    flash('Start time for daily trigger is required.', 'error')

            elif trigger_type == '3':
                new_start_time_str = request.form.get('weekly_start')
                if new_start_time_str:
                    new_start_time = datetime.datetime.strptime(new_start_time_str, '%Y-%m-%dT%H:%M')
                    weeks_interval = int(request.form.get('weeks_interval', 1))
                    days_of_week = request.form.getlist('days_of_week')
                    update_task(task_name, new_name, new_start_time, trigger_type, weeks_interval=weeks_interval, days_of_week=days_of_week)
                else:
                    flash('Start time for weekly trigger is required.', 'error')

            elif trigger_type == '4':
                new_start_time_str = request.form.get('monthly_start')
                if new_start_time_str:
                    new_start_time = datetime.datetime.strptime(new_start_time_str, '%Y-%m-%dT%H:%M')
                    days_of_month = request.form.getlist('days_of_month')
                    months_of_year = request.form.getlist('months_of_year')
                    update_task(task_name, new_name, new_start_time, trigger_type, days_of_month=days_of_month, months_of_year=months_of_year)
                else:
                    flash('Start time for monthly trigger is required.', 'error')

            else:
                flash('Invalid trigger type.', 'error')

            flash('Task updated successfully!', 'success')
            return redirect(url_for('autopoll_Scheduler'))

        except ValueError as e:
            flash(str(e), 'error')
    return render_template('edit_task.html', task_name=task_name)


@app.route('/Numeric_Prediction', methods=['GET'])
@login_required
def Numeric_Prediction():
    selected_date = request.args.get("date_picker")
    username = session['username']
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
        region_query = """
            SELECT * FROM AMR_REGION 
        """
        tag_query = """
        SELECT DISTINCT AMR_FIELD_ID.TAG_ID
        FROM AMR_FIELD_ID,amr_region,AMR_PL_GROUP
        
        WHERE AMR_PL_GROUP.PL_REGION_ID = amr_region.id
        AND AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID
        AND amr_region.REGION_NAME = :region_id
        AND AMR_FIELD_ID.tag_id NOT like '%.remove%'
        ORDER BY  TAG_ID
        """
        run_query = """
            SELECT DISTINCT METER_STREAM_NO
            FROM AMR_FIELD_ID , amr_field_meter
            WHERE amr_field_id.meter_id = amr_field_meter.meter_id
            AND amr_field_id.tag_id = :tag_id
        """
        region_results = fetch_data(ptt_pivot_connection, region_query)
        region_options = [str(region[1]) for region in region_results]
        selected_tag = request.args.get("tag_dropdown")
        selected_region = request.args.get("region_dropdown")
        selected_run = request.args.get("run_dropdown")

        run_results = fetch_data(ptt_pivot_connection, run_query, params={"tag_id": selected_tag})
        run_options = [str(run[0]) for run in run_results]

        query = f"""
            SELECT AMR_PL_GROUP.PL_REGION_ID,
                AMR_FIELD_ID.TAG_ID,
                AMR_FIELD_ID.METER_ID,
                AMR_BILLING_DATA.DATA_DATE,
                AMR_BILLING_DATA.CORRECTED_VOL as CORRECTED,
                AMR_BILLING_DATA.UNCORRECTED_VOL as UNCORRECTED,
                AMR_BILLING_DATA.METER_STREAM_NO  
            FROM amr_billing_data, AMR_PL_GROUP, AMR_FIELD_ID,amr_region
            WHERE AMR_BILLING_DATA.DATA_DATE BETWEEN TRUNC(SYSDATE) - 180 AND TRUNC(SYSDATE)
            AND AMR_PL_GROUP.PL_REGION_ID = amr_region.id
            AND AMR_PL_GROUP.FIELD_ID = AMR_FIELD_ID.FIELD_ID 
            AND AMR_BILLING_DATA.METER_ID = AMR_FIELD_ID.METER_ID
            AND AMR_BILLING_DATA.METER_STREAM_NO = '{selected_run}'
            AND AMR_FIELD_ID.TAG_ID = '{selected_tag}'
            AND amr_region.REGION_NAME = '{selected_region}'
            ORDER BY AMR_BILLING_DATA.DATA_DATE DESC
        """
        results = fetch_data(ptt_pivot_connection, query)
        df = pd.DataFrame(results, columns=['PL_REGION_ID', 'TAG_ID', 'METER_ID', 'DATA_DATE', 'CORRECTED', 'UNCORRECTED', 'METER_STREAM_NO'])
        
        corrected_graph_html = None
        uncorrected_graph_html = None

        if selected_date:
            date_obj = datetime.datetime.strptime(selected_date, "%Y-%m-%d")
            formatted_date = date_obj.strftime("%d-%b-%Y").upper()
            print(formatted_date)
            df['DATA_DATE'] = pd.to_datetime(df['DATA_DATE'])
            
            df['data_date_ordinal'] = df['DATA_DATE'].map(pd.Timestamp.toordinal)

            X = df[['data_date_ordinal']]
            degree = 2
            poly = PolynomialFeatures(degree)

            def plot_graph(y, y_label, title):
                X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
                X_poly = poly.fit_transform(X_train)
                X_test_poly = poly.transform(X_test)
                model = LinearRegression()
                model.fit(X_poly, y_train)
                y_pred = model.predict(X_test_poly)
                mse = mean_squared_error(y_test, y_pred, squared=False)
                r2 = model.score(X_test_poly, y_test)

                forecast_date = pd.to_datetime(formatted_date, format="%d-%b-%Y")
                forecast_date_ordinal = forecast_date.toordinal()
                new_forecast_date_ordinal = poly.transform(np.array([[forecast_date_ordinal]]))
                predicted_value = model.predict(new_forecast_date_ordinal)

                X_plot = pd.date_range(start=df['DATA_DATE'].min(), end=df['DATA_DATE'].max() + pd.DateOffset(days=180), freq='D')
                X_plot_ordinal = X_plot.to_series().map(pd.Timestamp.toordinal).values.reshape(-1, 1)
                X_plot_poly = poly.transform(X_plot_ordinal)
                y_plot = model.predict(X_plot_poly)

                trace_original = go.Scatter(x=df['DATA_DATE'], y=y, mode='markers', name='Data')
                trace_fit = go.Scatter(x=X_plot, y=y_plot, mode='lines', name='Predicted Value')
                trace_forecast = go.Scatter(x=[forecast_date], y=[predicted_value[0]], mode='markers+text',
                                            text=[f"{predicted_value[0]:.2f}"], textposition='top center',
                                            name='Forecasted Value', marker=dict(size=12, color='green'))

                layout = go.Layout(title=title,
                                   xaxis=dict(title='Date', tickformat='%d-%b-%Y'),
                                   yaxis=dict(title=y_label),
                                   annotations=[dict(
                                       x=forecast_date,
                                       y=predicted_value[0],
                                       xref="x", yref="y",
                                       text=f"Predicted Value: {predicted_value[0]:.2f} (Date: {forecast_date})",
                                       showarrow=True,
                                       arrowhead=7
                                   ),
                                   dict(
                                       x=0.95,
                                       y=0.05,
                                       xref='paper', yref='paper',
                                       text=f"Mean Squared Error: {mse:.2f}<br>R-squared: {r2:.2f}<br>Predicted Value: {predicted_value[0]:.2f} (Date: {forecast_date})<br>Site: {selected_tag}<br>Run: {selected_run}",
                                       showarrow=False,
                                       font=dict(size=10),
                                       align="right",
                                       bgcolor="white",
                                       opacity=0.5
                                   )])

                fig = go.Figure(data=[trace_original, trace_fit, trace_forecast], layout=layout)
                return fig, mse, r2

            corrected_fig, mse_corrected, r2_corrected = plot_graph(df['CORRECTED'], 'Corrected Volume', 'Corrected Volume Polynomial Regression Forecast')
            uncorrected_fig, mse_uncorrected, r2_uncorrected = plot_graph(df['UNCORRECTED'], 'Uncorrected Volume', 'Uncorrected Volume Polynomial Regression Forecast')

            corrected_graph_html = pio.to_html(corrected_fig, full_html=False)
            uncorrected_graph_html = pio.to_html(uncorrected_fig, full_html=False)
        else:
            mse_corrected = None
            r2_corrected = None
            mse_uncorrected = None
            r2_uncorrected = None

    return render_template(
        'Numeric_Prediction.html', 
        region_options=region_options, 
        run_options=run_options, 
        selected_date=selected_date,
        selected_tag=selected_tag,
        selected_region=selected_region,
        selected_run=selected_run,
        corrected_graph_html=corrected_graph_html,
        uncorrected_graph_html=uncorrected_graph_html,
        mse_corrected=mse_corrected,
        r2_corrected=r2_corrected,
        mse_uncorrected=mse_uncorrected,
        r2_uncorrected=r2_uncorrected,username=username
    )

@app.route("/evc_management", methods=[ "GET","POST"])
@login_required
def evc_management():
        # Query the data again for rendering the template
    
    username = session['username']
    if request.method == "POST":
        input_field = request.form["input_field"]
        print(input_field)

        try:
            with connect_to_ptt_pivot_db() as ptt_pivot_connection:
                region_query = "SELECT vc_name, id FROM amr_vc_type "
                results = fetch_data(ptt_pivot_connection, region_query)
                print(results)

                df = pd.DataFrame(results, columns=['vc_name', 'id'])
                max_id = df['id'].max()
                id_max = max_id + 1

                print("Max ID:", id_max)
                region_insert = f"INSERT INTO AMR_VC_TYPE (vc_name, id) VALUES ('{input_field}', '{id_max}')"
                
                update_sql(ptt_pivot_connection, region_insert)
                print("Data inserted successfully")
                
                for i in range(1, 6):
                    mapping_billing_insert = f"""INSERT INTO amr_mapping_billing(ADDRESS,
                                                DESCRIPTION,
                                                DATA_TYPE,
                                                EVC_TYPE,
                                                OR_DER,
                                                DAILY) VALUES ('0','0','0','{id_max}','{i}','1')"""
                                                
                    mapping_hourly_insert = f"""INSERT INTO amr_mapping_hourly(DESCRIPTION,
                                                HOURLY,
                                                EVC_TYPE,
                                                OR_DER,
                                                ADDRESS,
                                                DATA_TYPE) VALUES ('0','1','{id_max}','{i}','0','0')"""

                    update_sql(ptt_pivot_connection, mapping_hourly_insert)
                    update_sql(ptt_pivot_connection, mapping_billing_insert)
                    print("Data inserted successfully")
                    
                for i in range(1, 21):
                    mapping_config_insert = f"""INSERT INTO amr_mapping_config(ADDRESS,
                                                DESCRIPTION,
                                                DATA_TYPE,
                                                EVC_TYPE,
                                                OR_DER) VALUES ('0','0','0','{id_max}','{i}')"""

                    update_sql(ptt_pivot_connection, mapping_config_insert)
                    print("Data inserted successfully")
                    
                poll_range_insert = f"""INSERT INTO amr_poll_range 
                                    (POLL_CONFIG,POLL_BILLING,POLL_CONFIG_ENABLE,POLL_BILLING_ENABLE,EVC_TYPE) 
                                    VALUES 
                                    ('0,0,0,0,0,0,0,0,0,0', '0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0', '0,0,0,0,0', '0,0,0,0,0,0,0,0,0,0', '{id_max}')"""     
                                       
                poll_range_hourly_insert = f"""INSERT INTO amr_poll_range_hourly
                                (POLL_HOURLY,POLL_HOURLY_ENABLE,EVC_TYPE) 
                                VALUES 
                                ('0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0','0,0,0,0,0,0,0,0,0,0', '{id_max}')"""
                
                update_sql(ptt_pivot_connection, poll_range_hourly_insert)
                update_sql(ptt_pivot_connection, poll_range_insert)
                print("Data inserted successfully")

                flash('EVC added successfully', 'success')
        except Exception as e:
            print(f"An error occurred: {e}")
            flash('Failed to add EVC', 'error')
        
        return redirect(url_for('evc_management'))
    
    with connect_to_ptt_pivot_db() as ptt_pivot_connection:
            region_query = "SELECT vc_name, id FROM amr_vc_type ORDER BY vc_name"
            results = fetch_data(ptt_pivot_connection, region_query)
            df = pd.DataFrame(results, columns=['vc_name', 'id'])
            
    return render_template('evc_management.html', df=df,username=username)


def send_reset_email(username, email):
    sender_email = ""  # อีเมลผู้ส่ง
    sender_password = " "        # รหัสผ่านอีเมลผู้ส่ง  (ไปสร้างapp password ใน Google Account ผู้ส่งก่อน เพื่อได้รหัสapp มาใส่เเทนรหัสจริงตรงนี้)
    receiver_email = email                   # อีเมลผู้รับ
    # สร้างลิงก์สำหรับรีเซ็ตรหัสผ่าน (ลิงก์จำลองในตัวอย่างนี้)
    reset_link = f"https://tsoamr.pttplc.com/reset_password/{username}"

    # เนื้อหาอีเมล
    subject = "Password Reset Request"
    body = f"Hello {username},\n\nYou requested to reset your password. Please click the following link to reset it:\n{reset_link}\n\nIf you didn't request this, please ignore this email."
    
    # สร้างอีเมล
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    try:
        # เริ่มต้นเซสชัน SMTP และส่งอีเมล
        server = smtplib.SMTP('smtp.gmail.com', 587)  # ตัวอย่างนี้ใช้ Gmail
        server.starttls()
        server.login(sender_email, sender_password)
        text = msg.as_string()
        server.sendmail(sender_email, receiver_email, text)
        server.quit()
        print("Reset password email sent successfully!")
    except Exception as e:
        print(f"Error: {e}")

@app.route("/forget_password", methods=['GET', 'POST'])
def forget_password():
    if request.method == 'POST':
        # รับค่าจากฟอร์ม
        with connect_to_ptt_pivot_db() as ptt_pivot_connection:
            username = request.form.get('username')
            email = request.form.get('email')
            print("username:", username)
            print("email:", email)
            
            # SQL query สำหรับตรวจสอบ username
            query = f"""SELECT user_name FROM amr_user WHERE user_name like '{username}'"""
            results = fetch_data(ptt_pivot_connection, query)
            print(results)
            
            # ตรวจสอบว่ามีข้อมูลหรือไม่
            if results:
                # ถ้ามีข้อมูล
                error_message = None 

                # ส่งอีเมลรีเซ็ตรหัสผ่าน
                send_reset_email(username, email)

                # แจ้งให้ผู้ใช้ทราบว่าอีเมลถูกส่งไปแล้ว
                flash("Password reset link has been sent to your email.", "success")
                return redirect(url_for('login'))
            else:
                # ถ้าไม่มีข้อมูล
                error_message = "Invalid username or email"
            
        return render_template('forget_password.html', error_message=error_message)
    
    # ถ้าเป็น method GET แสดงหน้าฟอร์มลืมรหัสผ่าน
    return render_template('forget_password.html')

@app.route("/reset_password/<username>", methods=['GET', 'POST'])

def reset_password(username):
    if request.method == 'POST':
        new_password = request.form.get('new_password')
        hashed_password = md5_hash(new_password)
        current_datetime = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7))).strftime('%d-%b-%y %I.%M.%S.%f %p ASIA/BANGKOK').upper()
        # ทำการอัปเดตรหัสผ่านในฐานข้อมูล (โค้ดนี้ต้องมีการเชื่อมต่อกับฐานข้อมูล)
        with connect_to_ptt_pivot_db() as ptt_pivot_connection:
            query = f"""UPDATE amr_user SET password = '{hashed_password}',TIME_CREATE = '{current_datetime}',UPDATED_BY = '{username}' WHERE user_name = '{username}'"""
            update_sql(ptt_pivot_connection, query)  # ฟังก์ชันสำหรับรันคำสั่ง SQL
        
        flash("Password has been reset successfully!", "success")
        return redirect(url_for('login'))  # เปลี่ยนไปที่หน้าเข้าสู่ระบบ
    return render_template('reset_password.html', username=username)

if __name__ == "__main__":
    app.run(debug=True)
